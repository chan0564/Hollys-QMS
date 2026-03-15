import pandas as pd
import io
import os
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import streamlit as st

# ==========================================
# 1. 계산 관련 로직 (calc_ 및 helper)
# ==========================================

def calc_annual_leave(hire_date_str, leave_records_df, emp_id, today=None):
    """
    파트타이머 연차 발생/사용/잔여 계산 로직
    """
    from datetime import date as dclass
    if today is None:
        today = dclass.today()

    try:
        if not hire_date_str or str(hire_date_str).lower() in ["nan", "none", ""]:
            return 0, 0, 0, {}
        hire = pd.to_datetime(hire_date_str).date()
    except Exception:
        return 0, 0, 0, {}

    emp_id_str = str(emp_id).replace("nan", "").strip()
    emp_df = leave_records_df[leave_records_df["사번"].astype(str).str.strip() == emp_id_str].copy()
    emp_df["날짜_dt"] = pd.to_datetime(emp_df["날짜"], errors="coerce")
    emp_df = emp_df.dropna(subset=["날짜_dt"])

    try:
        rd = relativedelta(today, hire)
        months_worked = rd.months + rd.years * 12
        years_worked  = rd.years
    except (AssertionError, Exception):
        # Python 3.14 등 특정 환경에서 relativedelta 내부 오류 발생 시 대비
        months_worked = 0
        years_worked = 0

    accrued = 0
    detail  = {}

    if years_worked >= 1:
        year1_accrued = 0
        for m in range(1, 12):
            month_point = hire + relativedelta(months=m)
            if month_point > today: break
            window_start = hire + relativedelta(months=m-1)
            window_end   = hire + relativedelta(months=m) - timedelta(days=1)
            has_unpaid = any(
                window_start <= r["날짜_dt"].date() <= window_end
                for _, r in emp_df[emp_df["유형"] == "무급휴가"].iterrows()
            )
            if not has_unpaid:
                year1_accrued += 1
        accrued += year1_accrued
        detail["입사 1년 미만 (월 발생)"] = year1_accrued

        for yr in range(1, years_worked + 1):
            anniversary = hire + relativedelta(years=yr)
            if anniversary <= today:
                accrued += 15
                detail[f"입사 {yr}년 (연간 15개)"] = 15
    else:
        year1_accrued = 0
        for m in range(1, months_worked + 1):
            month_point = hire + relativedelta(months=m)
            if month_point > today: break
            window_start = hire + relativedelta(months=m-1)
            window_end   = hire + relativedelta(months=m) - timedelta(days=1)
            has_unpaid = any(
                window_start <= r["날짜_dt"].date() <= window_end
                for _, r in emp_df[emp_df["유형"] == "무급휴가"].iterrows()
            )
            if not has_unpaid:
                year1_accrued += 1
        accrued = year1_accrued
        detail[f"입사 {months_worked}개월 (월 발생)"] = year1_accrued

    # 미래 날짜를 포함하여 해당 직원의 모든 '연차' 기록을 사용량으로 집계
    used = len(emp_df[emp_df["유형"] == "연차"])
    remaining = accrued - used
    
    return accrued, used, remaining, detail

def calc_ccp(q1, q2, q21, q3, q4, q5):
    """
    HACCP CCP 결정도 판단 로직
    """
    if q1 == "예":
        if q2 == "예":
            if q3 == "아니오":
                if q4 == "아니오":
                    return "CCP"
                else:
                    if q5 == "예":      return "CP"
                    if q5 == "아니오":  return "CCP"
            else:
                return "CP"
        else:
            if q21 == "예":
                if q3 == "아니오":
                    if q4 == "아니오":
                        return "CCP"
                    else:
                        if q5 == "예":      return "CP"
                        if q5 == "아니오":  return "CCP"
                else:
                    return "CP"
            else:
                return "CP"
    return ""

def active_questions(q1, q2, q21, q3, q4):
    """CCP 결정도에서 현재 답변 상태에 따라 활성화되어야 할 질문 집합 반환"""
    active = {"Q1"}
    if q1 == "아니오":
        active.add("Q2")
        if q2 == "아니오":
            active.add("Q21")
        elif q2 == "예":
            active.add("Q3")
            if q3 == "아니오":
                active.add("Q4")
                if q4 == "예":
                    active.add("Q5")
    return active

def calc_next_filter(row):
    """필터 차기 점검일 계산"""
    try:
        last_date = pd.to_datetime(row['점검일자']).date()
        cycle = int(row['주기_개월'])
        return (last_date + relativedelta(months=cycle)).strftime('%Y-%m-%d')
    except:
        return ""

def calc_next_calib(row):
    """계측기 차기 검교정일 계산"""
    try:
        last_date = pd.to_datetime(row['검교정일자']).date()
        cycle = int(str(row['주기']).replace('개월',''))
        return (last_date + relativedelta(months=cycle)).strftime('%Y-%m-%d')
    except:
        return ""

# ==========================================
# 2. 엑셀 출력 로직 (export_*)
# ==========================================

def export_full_excel(sel_code, sel_name, df_b, df_r, df_h, img_path):
    import openpyxl as _opx
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as gcl
    wb  = _opx.Workbook(); ws = wb.active; ws.title = "제품규격서"
    ws.page_setup.paperSize = 9; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.5; ws.page_margins.right  = 0.5
    ws.page_margins.top  = 0.8; ws.page_margins.bottom = 0.8
    thin = Side(style="thin", color="B0C4D8")
    med  = Side(style="medium", color="1F4E79")
    bd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    bdm  = Border(left=med,  right=med,  top=med,  bottom=med)
    hfill = PatternFill("solid", start_color="D6E4F0")   # 옅은 파란 (항목 헤더)
    sfill = PatternFill("solid", start_color="D6E4F0")   # 섹션헤더도 동일
    tfill = PatternFill("solid", start_color="1F4E79")   # 타이틀만 진한 네이비
    hfont = Font(name="맑은 고딕", bold=True, color="1F4E79", size=10)
    sfont = Font(name="맑은 고딕", bold=True, color="1F4E79", size=11)  # 섹션헤더 글자 파란색
    tfont = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=14)
    nfont = Font(name="맑은 고딕", size=10)
    nalign = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    calign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, w in enumerate([22, 30, 20, 22, 18, 18], 1):
        ws.column_dimensions[gcl(ci)].width = w
    cur = 1
    ws.merge_cells(f"A{cur}:F{cur}")
    ws[f"A{cur}"] = f"제품 규격서  |  {sel_name}  ({sel_code})"
    ws[f"A{cur}"].font = tfont; ws[f"A{cur}"].fill = tfill
    ws[f"A{cur}"].alignment = calign; ws.row_dimensions[cur].height = 36; cur += 1
    ws.merge_cells(f"A{cur}:D{cur}")
    ws[f"A{cur}"] = "■ 1. 기본 정보"; ws[f"A{cur}"].font = sfont
    ws[f"A{cur}"].fill = sfill; ws[f"A{cur}"].alignment = calign
    ws.merge_cells(f"E{cur}:F{cur}")
    ws[f"E{cur}"] = "■ 제품 사진"; ws[f"E{cur}"].font = sfont
    ws[f"E{cur}"].fill = sfill; ws[f"E{cur}"].alignment = calign
    ws.row_dimensions[cur].height = 20; cur += 1
    basic_start = cur
    for _, br in df_b.iterrows():
        ws[f"A{cur}"] = str(br.iloc[0]) if len(br) > 0 else ""
        ws[f"A{cur}"].font = hfont; ws[f"A{cur}"].fill = hfill
        ws[f"A{cur}"].border = bd; ws[f"A{cur}"].alignment = calign
        ws.merge_cells(f"B{cur}:D{cur}")
        ws[f"B{cur}"] = str(br.iloc[1]) if len(br) > 1 and pd.notna(br.iloc[1]) else ""
        ws[f"B{cur}"].font = nfont; ws[f"B{cur}"].border = bd; ws[f"B{cur}"].alignment = nalign
        ws.row_dimensions[cur].height = 20; cur += 1
    basic_end = cur - 1
    if basic_end >= basic_start:
        ws.merge_cells(f"E{basic_start}:F{basic_end}")
        ws[f"E{basic_start}"].border = bdm; ws[f"E{basic_start}"].alignment = calign
    if img_path and os.path.exists(img_path):
        try:
            from openpyxl.drawing.image import Image as OI
            img_o = OI(img_path); img_o.width = 130
            img_o.height = min(int(20 * (basic_end - basic_start + 1) * 1.33), 220)
            ws.add_image(img_o, f"E{basic_start}")
        except Exception: pass
    cur += 1
    ws.merge_cells(f"A{cur}:F{cur}")
    ws[f"A{cur}"] = "■ 2. 배합비 (BOM)"; ws[f"A{cur}"].font = sfont
    ws[f"A{cur}"].fill = sfill; ws[f"A{cur}"].alignment = calign
    ws.row_dimensions[cur].height = 20; cur += 1
    for ci2, col in enumerate(df_r.columns, 1):
        c = ws.cell(row=cur, column=ci2, value=col)
        c.font = hfont; c.fill = hfill; c.border = bd; c.alignment = calign
    ws.row_dimensions[cur].height = 20; cur += 1
    for _, rr in df_r.iterrows():
        for ci2, val in enumerate(rr, 1):
            c = ws.cell(row=cur, column=ci2, value=str(val) if pd.notna(val) else "")
            c.font = nfont; c.border = bd
            c.alignment = calign if ci2 >= 4 else nalign
        ws.row_dimensions[cur].height = 20; cur += 1
    cur += 1
    ws.merge_cells(f"A{cur}:F{cur}")
    ws[f"A{cur}"] = "■ 3. 완제품 규격 (위해요소)"; ws[f"A{cur}"].font = sfont
    ws[f"A{cur}"].fill = sfill; ws[f"A{cur}"].alignment = calign
    ws.row_dimensions[cur].height = 20; cur += 1
    for ci2, col in enumerate(df_h.columns, 1):
        c = ws.cell(row=cur, column=ci2, value=col)
        c.font = hfont; c.fill = hfill; c.border = bd; c.alignment = calign
    ws.row_dimensions[cur].height = 20; cur += 1
    for _, hr in df_h.iterrows():
        for ci2, val in enumerate(hr, 1):
            c = ws.cell(row=cur, column=ci2, value=str(val) if pd.notna(val) else "")
            c.font = hfont if ci2 == 1 else nfont
            c.border = bd
            c.alignment = calign if ci2 >= 2 else nalign
            if ci2 == 1: c.fill = hfill
        ws.row_dimensions[cur].height = 20; cur += 1
    out = io.BytesIO(); wb.save(out); return out.getvalue()

def export_trend_excel(data, prod_name, cat_name):
    import openpyxl as _opx
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as gcl
    wb2 = _opx.Workbook(); ws2 = wb2.active; ws2.title="품질트렌드"
    thin2=Side(style="thin",color="B0C4D8"); bd2=Border(left=thin2,right=thin2,top=thin2,bottom=thin2)
    hfill2=PatternFill("solid",start_color="D6E4F0"); tfill2=PatternFill("solid",start_color="1F4E79")
    hfont2=Font(name="맑은 고딕",bold=True,color="1F4E79",size=10)
    tfont2=Font(name="맑은 고딕",bold=True,color="FFFFFF",size=13)
    nfont2=Font(name="맑은 고딕",size=10)
    calign2=Alignment(horizontal="center",vertical="center")
    for ci2, w in enumerate([15,12,10,30],1): ws2.column_dimensions[gcl(ci2)].width = w
    ws2.merge_cells("A1:D1")
    ws2["A1"] = f"[{prod_name}] {cat_name} 품질 데이터 보관 대장"
    ws2["A1"].font = tfont2; ws2["A1"].fill = tfill2
    ws2["A1"].alignment = calign2; ws2.row_dimensions[1].height = 30
    for ci2, col in enumerate(data.columns, 1):
        c = ws2.cell(row=2, column=ci2, value=col)
        c.font = hfont2; c.fill = hfill2; c.border = bd2; c.alignment = calign2
    ws2.row_dimensions[2].height = 20
    pass_fill=PatternFill("solid",start_color="D4EDDA")
    fail_fill=PatternFill("solid",start_color="F8D7DA")
    for ri2, row in data.iterrows():
        판정_x = str(row.get("판정",""))
        for ci2, col in enumerate(data.columns,1):
            c = ws2.cell(row=ri2+3, column=ci2, value=str(row[col]) if pd.notna(row[col]) else "")
            c.font = nfont2; c.border = bd2; c.alignment = calign2
            if col == "판정":
                if 판정_x=="PASS": c.fill = pass_fill
                elif 판정_x=="FAIL": c.fill = fail_fill
        ws2.row_dimensions[ri2+3].height = 18
    out2 = io.BytesIO(); wb2.save(out2); return out2.getvalue()

def export_org_excel(df_view):
    from datetime import date
    df_view = df_view.fillna("") 
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet('HACCP조직도')
        
        worksheet.set_paper(9) 
        worksheet.fit_to_pages(1, 1) 
        worksheet.center_horizontally() 
        worksheet.set_margins(left=0.5, right=0.5, top=0.5, bottom=0.5) 
        
        title_fmt = workbook.add_format({'bold': True, 'font_size': 20, 'align': 'center', 'valign': 'vcenter'})
        doc_info_title_fmt = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bold': True, 'border': 1, 'bg_color': '#F2F2F2'})
        doc_info_val_fmt = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
        
        header_fmt = workbook.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})
        name_fmt = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        section_fmt = workbook.add_format({'bg_color': '#E2EFDA', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True, 'font_size': 12})
        
        worksheet.set_column('A:A', 3)
        worksheet.set_column('B:G', 16) 
        
        worksheet.merge_range('B2:E4', "HACCP팀 조직도", title_fmt)
        
        worksheet.write('F2', "문서번호", doc_info_title_fmt)
        worksheet.write('G2', "HLSHS-101-A01", doc_info_val_fmt)
        worksheet.write('F3', "제정일자", doc_info_title_fmt)
        worksheet.write('G3', "2018-11-05", doc_info_val_fmt)
        worksheet.write('F4', "개정일자", doc_info_title_fmt)
        worksheet.write('G4', date.today().strftime('%Y-%m-%d'), doc_info_val_fmt)
        
        team_leaders = df_view[df_view['HACCP 직책'] == 'HACCP 팀장']['이름'].tolist()
        leader_name = ", ".join(team_leaders) if team_leaders else "(공란)"
        worksheet.merge_range('D6:E6', "HACCP 팀장", header_fmt)
        worksheet.merge_range('D7:E8', leader_name, name_fmt)
        
        teams = ["생산/시설관리팀", "품질관리팀", "업무지원팀"]
        col_map = {"생산/시설관리팀": 1, "품질관리팀": 3, "업무지원팀": 5} 
        
        max_team_members = 0
        for t in teams:
            c_idx = col_map[t]
            worksheet.merge_range(10, c_idx, 10, c_idx+1, t, header_fmt)
            worksheet.set_row(10, 25)
            
            members = df_view[df_view['HACCP 직책'] == t]
            team_heads = members[members['직급'].str.contains('팀장', na=False)]
            team_normals = members[~members['직급'].str.contains('팀장', na=False)]
            sorted_members = pd.concat([team_heads, team_normals])
            
            if len(sorted_members) > max_team_members: max_team_members = len(sorted_members)
            
            row_idx = 11
            for _, row in sorted_members.iterrows():
                info = f"{row['이름']}\n({row['직급']})"
                worksheet.merge_range(row_idx, c_idx, row_idx, c_idx+1, info, name_fmt)
                worksheet.set_row(row_idx, 35)
                row_idx += 1
            
            while row_idx <= 11 + 5: 
                worksheet.merge_range(row_idx, c_idx, row_idx, c_idx+1, "", name_fmt)
                worksheet.set_row(row_idx, 35)
                row_idx += 1

        ccp_start_row = 11 + max(max_team_members, 6) + 2
        
        worksheet.merge_range(ccp_start_row, 1, ccp_start_row, 6, "CCP 모니터링 담당", section_fmt)
        worksheet.set_row(ccp_start_row, 25)
        
        ccps = ["CCP-1", "CCP-2", "CCP-3", "CCP-4", "CCP-5", "CCP-6"]
        for i, ccp in enumerate(ccps):
            c_idx = i + 1  
            
            worksheet.write(ccp_start_row + 1, c_idx, ccp, header_fmt)
            worksheet.set_row(ccp_start_row + 1, 20)
            
            assigned = []
            for _, row in df_view.iterrows():
                if ccp in str(row['모니터링 CCP']):
                    assigned.append(f"{row['이름']}")
        
            row_idx = ccp_start_row + 2
            for name in assigned:
                worksheet.write(row_idx, c_idx, name, name_fmt)
                worksheet.set_row(row_idx, 25)
                row_idx += 1
            
            end_row = ccp_start_row + 2 + 5 
            while row_idx < end_row:
                worksheet.write(row_idx, c_idx, "", name_fmt)
                worksheet.set_row(row_idx, 25)
                row_idx += 1

    return output.getvalue()

def export_health_excel(df_view):
    from datetime import date
    df_view = df_view.fillna("") 
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet('보건증현황')
        
        title_format = workbook.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
        header_format = workbook.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})
        normal_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
        alert_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_color': 'red', 'bold': True})
        
        worksheet.merge_range('A1:G1', "건강진단결과(보건증) 현황 관리표", title_format)
        worksheet.set_row(0, 40)
        worksheet.write('A2', f"출력일자: {date.today().strftime('%Y-%m-%d')}")
        
        headers = list(df_view.columns)
        for col_num, header in enumerate(headers):
            worksheet.write(2, col_num, header, header_format)
        
        for row_num, row in df_view.iterrows():
            for col_num, col_name in enumerate(headers):
                val = row[col_name]
                if isinstance(val, (date, datetime)): val = val.strftime('%Y-%m-%d')
                    
                is_alert = False
                if col_name in ['D-Day', '상태']:
                    if val != "" and row['D-Day'] != "":
                        try:
                            if float(row['D-Day']) <= 30:
                                is_alert = True
                        except ValueError:
                            pass
                
                if is_alert:
                    worksheet.write(row_num + 3, col_num, val, alert_format)
                else:
                    worksheet.write(row_num + 3, col_num, val, normal_format)
        
        worksheet.set_column('A:B', 15); worksheet.set_column('C:C', 20)
        worksheet.set_column('D:E', 18); worksheet.set_column('F:G', 15)
        
    return output.getvalue()

def export_facility_list(df_list):
    df_list = df_list.fillna("")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
        wb = writer.book
        ws = wb.add_worksheet('설비목록표')
        ws.set_paper(9)
        ws.fit_to_pages(1, 0)
        title_fmt = wb.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
        header_fmt = wb.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})
        val_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
        ws.merge_range('A1:K2', "제조위생설비 목록표", title_fmt)
        headers = list(df_list.columns)
        for col_num, header in enumerate(headers):
            ws.write(3, col_num, header, header_fmt)
            ws.set_column(col_num, col_num, 12 if col_num != 10 else 35) 
        for r_idx, row in df_list.iterrows():
            for c_idx, col_name in enumerate(headers):
                idx = df_list.index.get_loc(r_idx)
                ws.write(idx + 4, c_idx, str(row[col_name]), val_fmt)
                ws.set_row(idx + 4, 25)
    return output.getvalue()

def export_facility_excel(f_data, r_data, img_path):
    from datetime import date
    f_data = f_data.fillna("") 
    r_data = r_data.fillna("")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
        wb = writer.book
        ws = wb.add_worksheet('설비이력카드')
        
        title_fmt = wb.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
        doc_fmt = wb.add_format({'align': 'right', 'valign': 'vcenter', 'bold': True})
        header_fmt = wb.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})
        val_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        val_left_fmt = wb.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter', 'text_wrap': True})
        section_fmt = wb.add_format({'bold': True, 'font_size': 12, 'valign': 'vcenter'})
        
        ws.set_paper(9)
        ws.fit_to_pages(1, 1)
        ws.center_horizontally()
        ws.set_margins(0.5, 0.5, 0.5, 0.5)
        
        ws.set_column('A:A', 12)
        ws.set_column('B:B', 15)
        ws.set_column('C:C', 12)
        ws.set_column('D:D', 15)
        ws.set_column('E:E', 12)
        ws.set_column('F:F', 20)
        
        ws.merge_range('A1:F2', "제조위생설비 이력카드", title_fmt)
        ws.write('F3', "문서번호", doc_fmt)
        ws.write('F4', "HLSPP-203-F02", wb.add_format({'align': 'right', 'valign': 'vcenter'}))
        
        ws.merge_range('A6:D6', "1. 기계기본정보", section_fmt)
        ws.merge_range('E6:F6', "2. 기계사진", section_fmt)
        
        ws.write('A7', '설비명', header_fmt)
        ws.write('B7', f_data['설비명'], val_fmt)
        ws.write('C7', '설비번호', header_fmt)
        ws.write('D7', f_data['설비번호'], val_fmt)
        
        ws.write('A8', '사용용도', header_fmt)
        ws.write('B8', f_data['사용용도'], val_fmt)
        ws.write('C8', '전압', header_fmt)
        ws.write('D8', f_data['전압'], val_fmt)
        
        ws.write('A9', '구입년월', header_fmt)
        ws.write('B9', f_data['구입년월'], val_fmt)
        ws.write('C9', '제조회사명', header_fmt)
        ws.write('D9', f_data['제조회사명'], val_fmt)
        
        ws.write('A10', '설치장소', header_fmt)
        ws.write('B10', f_data['설치장소'], val_fmt)
        ws.write('C10', '관리부서', header_fmt)
        ws.write('D10', f_data['관리부서'], val_fmt)
        
        ws.merge_range('A11:A12', '관리자', header_fmt)
        ws.write('B11', '정', header_fmt)
        ws.merge_range('C11:D11', f_data['관리자_정'], val_fmt)
        
        ws.write('B12', '부', header_fmt)
        ws.merge_range('C12:D12', f_data['관리자_부'], val_fmt)
        
        ws.merge_range('A13:A16', '특이사항', header_fmt)
        ws.merge_range('B13:D16', f_data['특이사항'], val_left_fmt)
        
        ws.merge_range('E7:F16', '', val_fmt)
        if os.path.exists(img_path):
            try: ws.insert_image('E7', img_path, {'x_offset': 5, 'y_offset': 5, 'object_position': 1, 'x_scale': 0.25, 'y_scale': 0.25})
            except Exception: pass
        
        for r in range(6, 16): ws.set_row(r, 22)
        
        ws.merge_range('A18:F18', "3. 수리이력사항", section_fmt)
        ws.write('A19', '수리일자', header_fmt)
        ws.merge_range('B19:C19', '수리사항', header_fmt)
        ws.merge_range('D19:E19', '수리처', header_fmt)
        ws.write('F19', '비고', header_fmt)
        
        r_idx = 19
        for _, row in r_data.iterrows():
            val_date = row.get('수리일자', '')
            if isinstance(val_date, (date, datetime)): val_date = val_date.strftime('%Y-%m-%d')
            ws.write(r_idx, 0, str(val_date), val_fmt)
            ws.merge_range(r_idx, 1, r_idx, 2, str(row.get('수리사항', '')), val_fmt)
            ws.merge_range(r_idx, 3, r_idx, 4, str(row.get('수리처', '')), val_fmt)
            ws.write(r_idx, 5, str(row.get('비고', '')), val_fmt)
            ws.set_row(r_idx, 22)
            r_idx += 1
            
        while r_idx < 35:
            ws.write(r_idx, 0, "", val_fmt)
            ws.merge_range(r_idx, 1, r_idx, 2, "", val_fmt)
            ws.merge_range(r_idx, 3, r_idx, 4, "", val_fmt)
            ws.write(r_idx, 5, "", val_fmt)
            ws.set_row(r_idx, 22)
            r_idx += 1
            
    return output.getvalue()

def export_filter_excel(df_view):
    from datetime import date
    df_view = df_view.fillna("")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
        wb = writer.book
        ws = wb.add_worksheet('필터점검계획표')
        
        ws.set_paper(9) # A4
        ws.fit_to_pages(1, 0)
        
        title_fmt = wb.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
        header_fmt = wb.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})
        val_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
        
        ws.merge_range('A1:H2', f"{date.today().year}년 필터 점검 계획 및 기록표", title_fmt)
        
        headers = ["설치장소", "필터명", "내용(교체/세척)", "주기(개월)", "점검일자", "차기점검일자", "상태", "비고"]
        for c, h in enumerate(headers):
            ws.write(3, c, h, header_fmt)
        
        ws.set_column('A:A', 18)
        ws.set_column('B:B', 20)
        ws.set_column('C:C', 15)
        ws.set_column('D:D', 12)
        ws.set_column('E:F', 15)
        ws.set_column('G:G', 10)
        ws.set_column('H:H', 20)
        
        r_idx = 4
        for _, row in df_view.iterrows():
            ws.write(r_idx, 0, str(row.get('설치장소', '')), val_fmt)
            ws.write(r_idx, 1, str(row.get('필터명', '')), val_fmt)
            ws.write(r_idx, 2, str(row.get('내용', '')), val_fmt)
            ws.write(r_idx, 3, str(row.get('주기_개월', '')), val_fmt)
            
            d1 = row.get('점검일자', '')
            if isinstance(d1, (datetime, pd.Timestamp, date)): d1 = d1.strftime('%Y-%m-%d')
            ws.write(r_idx, 4, str(d1), val_fmt)
            
            d2 = row.get('차기점검일자', '')
            if isinstance(d2, (datetime, pd.Timestamp, date)): d2 = d2.strftime('%Y-%m-%d')
            ws.write(r_idx, 5, str(d2), val_fmt)
            
            ws.write(r_idx, 6, str(row.get('상태', '')), val_fmt)
            ws.write(r_idx, 7, str(row.get('비고', '')), val_fmt)
            ws.set_row(r_idx, 25)
            r_idx += 1
            
    return output.getvalue()

def export_cleaning_excel(df_view):
    import re
    df_view = df_view.fillna("")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
        wb = writer.book
        ws = wb.add_worksheet('세척소독 기준표')

        title_fmt = wb.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
        header_fmt = wb.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})
        val_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        left_fmt = wb.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter', 'text_wrap': True})

        ws.merge_range('A1:H2', "SSOP 11분류 세척소독 기준표", title_fmt)

        headers = ['대분류', '관리대상/설비명', '부위', '세척·소독 방법', '주기', '사용세제/도구', '책임자', '현장 사진']
        for c, h in enumerate(headers):
            ws.write(3, c, h, header_fmt)

        ws.set_column('A:A', 15); ws.set_column('B:C', 15)
        ws.set_column('D:D', 35)
        ws.set_column('E:G', 15); ws.set_column('H:H', 22)

        r_idx = 4
        for _, row in df_view.sort_values(by="대분류").iterrows():
            ws.write(r_idx, 0, str(row.get('대분류', '')), val_fmt)
            ws.write(r_idx, 1, str(row.get('설비명', '')), val_fmt)
            ws.write(r_idx, 2, str(row.get('부위', '')), val_fmt)
            ws.write(r_idx, 3, str(row.get('세척소독방법', '')), left_fmt)
            ws.write(r_idx, 4, str(row.get('주기', '')), val_fmt)
            ws.write(r_idx, 5, str(row.get('사용도구', '')), val_fmt)
            ws.write(r_idx, 6, str(row.get('책임자', '')), val_fmt)

            img_path = str(row.get('사진파일', ''))
            ws.write(r_idx, 7, "", val_fmt)
            
            real_img_path = img_path
            if row.get('대분류') == "5. 식품 제조시설":
                match = re.search(r'\[(.*?)\]', str(row.get('설비명', '')))
                if match:
                    f_no = match.group(1)
                    real_img_path = f"fac_photo_{f_no}.png"
                    
            if real_img_path and os.path.exists(real_img_path):
                try: ws.insert_image(r_idx, 7, real_img_path, {'x_offset': 5, 'y_offset': 5, 'x_scale': 0.12, 'y_scale': 0.12, 'positioning': 1})
                except Exception: pass
                
            ws.set_row(r_idx, 75)
            r_idx += 1
    return output.getvalue()

def export_rm_excel(code, name, df_b2, df_s2, df_sup2, img_p):
    import io as _io2
    import openpyxl as _opx
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as gcl
    wb2 = _opx.Workbook(); ws2 = wb2.active; ws2.title="원부자재규격서"
    ws2.page_setup.paperSize = 9   # A4
    ws2.page_setup.fitToWidth = 1
    ws2.page_setup.fitToPage = True
    ws2.print_options.horizontalCentered = True
    ws2.page_margins.left  = 0.5; ws2.page_margins.right  = 0.5
    ws2.page_margins.top   = 0.8; ws2.page_margins.bottom = 0.8

    thin  = Side(style="thin",  color="B0C4D8")
    med   = Side(style="medium", color="1F4E79")
    bd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    bdm   = Border(left=med,  right=med,  top=med,  bottom=med)
    hfill = PatternFill("solid", start_color="D6E4F0")
    sfill = PatternFill("solid", start_color="D6E4F0")
    tfill = PatternFill("solid", start_color="1F4E79")
    hfont = Font(name="맑은 고딕", bold=True, color="1F4E79",  size=10)
    sfont = Font(name="맑은 고딕", bold=True, color="1F4E79",  size=11)
    tfont = Font(name="맑은 고딕", bold=True, color="FFFFFF",  size=14)
    nfont = Font(name="맑은 고딕", size=10)
    nalign = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    calign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, w in enumerate([20, 28, 15, 28, 15, 15], 1):
        ws2.column_dimensions[gcl(ci)].width = w

    cur = 1
    ws2.merge_cells(f"A{cur}:F{cur}")
    ws2[f"A{cur}"] = f"원·부자재 규격서  |  {name}  ({code})"
    ws2[f"A{cur}"].font = tfont; ws2[f"A{cur}"].fill = tfill
    ws2[f"A{cur}"].alignment = calign; ws2.row_dimensions[cur].height = 36; cur += 1

    ws2.merge_cells(f"A{cur}:D{cur}")
    ws2[f"A{cur}"] = "■ 기본 정보"; ws2[f"A{cur}"].font = sfont
    ws2[f"A{cur}"].fill = sfill; ws2[f"A{cur}"].alignment = calign
    ws2.merge_cells(f"E{cur}:F{cur}")
    ws2[f"E{cur}"] = "■ 제품 사진"; ws2[f"E{cur}"].font = sfont
    ws2[f"E{cur}"].fill = sfill; ws2[f"E{cur}"].alignment = calign
    ws2.row_dimensions[cur].height = 20; cur += 1

    basic_start = cur
    for _, br in df_b2.iterrows():
        ws2[f"A{cur}"] = br["항목"]; ws2[f"A{cur}"].font = hfont
        ws2[f"A{cur}"].fill = hfill; ws2[f"A{cur}"].border = bd; ws2[f"A{cur}"].alignment = calign
        ws2.merge_cells(f"B{cur}:D{cur}")
        ws2[f"B{cur}"] = str(br["내용"]) if pd.notna(br["내용"]) else ""
        ws2[f"B{cur}"].font = nfont; ws2[f"B{cur}"].border = bd; ws2[f"B{cur}"].alignment = nalign
        ws2.row_dimensions[cur].height = 20; cur += 1
    basic_end = cur - 1

    if basic_end >= basic_start:
        ws2.merge_cells(f"E{basic_start}:F{basic_end}")
        ws2[f"E{basic_start}"].border = bdm; ws2[f"E{basic_start}"].alignment = calign
    if img_p and os.path.exists(img_p):
        try:
            from openpyxl.drawing.image import Image as OPXImg
            img_o = OPXImg(img_p)
            row_h_px = 20 * (basic_end - basic_start + 1) * 1.33
            img_o.width = 130; img_o.height = min(int(row_h_px), 200)
            ws2.add_image(img_o, f"E{basic_start}")
        except Exception: pass
    cur += 1

    ws2.merge_cells(f"A{cur}:F{cur}")
    ws2[f"A{cur}"] = "■ 품질 규격"; ws2[f"A{cur}"].font = sfont
    ws2[f"A{cur}"].fill = sfill; ws2[f"A{cur}"].alignment = calign
    ws2.row_dimensions[cur].height = 20; cur += 1
    for ci2, col in enumerate(df_s2.columns, 1):
        c = ws2.cell(row=cur, column=ci2, value=col)
        c.font = hfont; c.fill = hfill; c.border = bd; c.alignment = calign
    ws2.row_dimensions[cur].height = 20; cur += 1
    for _, sr in df_s2.iterrows():
        row_h = 20
        for ci2, val in enumerate(sr, 1):
            v_str = str(val) if pd.notna(val) else ""
            if len(v_str) > 30: row_h = 35
            c = ws2.cell(row=cur, column=ci2, value=v_str)
            c.font = nfont; c.border = bd
            c.alignment = calign if ci2 in (2,5,6) else nalign
            if ci2 == 1: c.fill = hfill; c.font = hfont
        ws2.row_dimensions[cur].height = row_h; cur += 1
    cur += 1

    if not df_sup2.empty:
        ws2.merge_cells(f"A{cur}:F{cur}")
        ws2[f"A{cur}"] = "■ 원료 정보"; ws2[f"A{cur}"].font = sfont
        ws2[f"A{cur}"].fill = sfill; ws2[f"A{cur}"].alignment = calign
        ws2.row_dimensions[cur].height = 20; cur += 1
        for ci2, col in enumerate(df_sup2.columns, 1):
            c = ws2.cell(row=cur, column=ci2, value=col)
            c.font = hfont; c.fill = hfill; c.border = bd; c.alignment = calign
        ws2.row_dimensions[cur].height = 20; cur += 1
        for _, sr in df_sup2.iterrows():
            for ci2, val in enumerate(sr, 1):
                c = ws2.cell(row=cur, column=ci2, value=str(val) if pd.notna(val) else "")
                c.font = nfont; c.border = bd; c.alignment = nalign
            ws2.row_dimensions[cur].height = 20; cur += 1

    out = _io2.BytesIO(); wb2.save(out); return out.getvalue()

def export_flowchart_excel(df_input, cat_nm):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_input.to_excel(writer, sheet_name='공정흐름도', index=False)
    return output.getvalue()

def export_ccp_excel(df_d, cat_nm):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_d.to_excel(writer, sheet_name='CCP결정도', index=False)
    return output.getvalue()

def export_hazard_excel(df_data, cat_name):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_data.to_excel(writer, sheet_name='공정위해요소분석', index=False)
    return output.getvalue()

def export_rm_hazard_excel(df_data, rm_name, rm_code):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_data.to_excel(writer, sheet_name='원부자재분석', index=False)
    return output.getvalue()

def make_leave_excel(emp_row, records, acc, us, rem, unpd, detail_d):
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.page import PageMargins

    wb = Workbook(); ws = wb.active; ws.title = "연차기록"
    BK = Side(style="thin", color="000000")
    def bdr(): return Border(top=BK,bottom=BK,left=BK,right=BK)
    def fill(h): return PatternFill("solid",fgColor=h)
    def fnt(bold=False,sz=10,fc="000000"): return Font(name="맑은 고딕",bold=bold,size=sz,color=fc)
    def aln(h="center",v="center",wrap=True): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
    def wc(r,c,val="",bold=False,sz=10,fc="000000",bg=None,h="center"):
        cell=ws.cell(row=r,column=c,value=str(val) if val is not None else "")
        cell.font=fnt(bold,sz,fc); cell.alignment=aln(h); cell.border=bdr()
        if bg: cell.fill=fill(bg)
    def mc(r1,c1,r2,c2): ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)

    # 제목
    mc(1,1,1,6); wc(1,1,"파트타이머 연차 관리표",bold=True,sz=14,fc="FFFFFF",bg="1F4E79"); ws.row_dimensions[1].height=28
    mc(2,1,2,6); wc(2,1,"커피클럽 로스팅센터  |  ㈜케이지할리스에프앤비",sz=9,fc="595959",bg="EBF3FB"); ws.row_dimensions[2].height=13

    # 직원 정보
    info_items = [("이름",emp_row["이름"]),("사번",emp_row["사번"]),
                  ("직급",emp_row["직급"]),("입사일",emp_row.get("입사일",""))]
    for i,(lbl,val) in enumerate(info_items):
        c=i*2+1
        wc(3,c,lbl,bold=True,bg="D6E4F0"); wc(3,c+1,val,h="left")
    ws.row_dimensions[3].height=16

    # 요약
    mc(4,1,4,2); wc(4,1,"발생 연차",bold=True,bg="2E75B6",fc="FFFFFF")
    mc(4,3,4,4); wc(4,3,"사용 연차",bold=True,bg="E67E22",fc="FFFFFF")
    mc(4,5,4,6); wc(4,5,"잔여 연차",bold=True,bg="27AE60" if rem>=0 else "E74C3C",fc="FFFFFF")
    mc(5,1,5,2); wc(5,1,acc,bold=True,sz=14)
    mc(5,3,5,4); wc(5,3,us, bold=True,sz=14)
    mc(5,5,5,6); wc(5,5,rem,bold=True,sz=14)
    ws.row_dimensions[4].height=16; ws.row_dimensions[5].height=22

    # 헤더
    for ci,hdr in enumerate(["No","날짜","유형","비고"],1):
        bg="1F4E79"
        ws.cell(row=6,column=ci,value=hdr).font=fnt(bold=True,fc="FFFFFF")
        ws.cell(row=6,column=ci).fill=fill(bg)
        ws.cell(row=6,column=ci).border=bdr()
        ws.cell(row=6,column=ci).alignment=aln()
    ws.row_dimensions[6].height=16

    r=7
    for ni,(_, rec) in enumerate(records.iterrows(),1):
        bg="F5F5F5" if r%2==0 else "FFFFFF"
        wc(r,1,ni,bg=bg); wc(r,2,str(rec.get("날짜","")),bg=bg)
        type_bg = "FCE4D6" if rec.get("유형")=="연차" else "EDE7F6"
        wc(r,3,str(rec.get("유형","")),bg=type_bg)
        wc(r,4,str(rec.get("비고","")),bg=bg,h="left")
        ws.row_dimensions[r].height=18; r+=1

    ws.column_dimensions["A"].width=5; ws.column_dimensions["B"].width=14
    ws.column_dimensions["C"].width=12; ws.column_dimensions["D"].width=30

    ws.page_setup.paperSize=9; ws.page_setup.fitToPage=True
    ws.page_margins=PageMargins(left=0.5,right=0.5,top=0.7,bottom=0.7)

    out=_io.BytesIO(); wb.save(out); return out.getvalue()

def export_outbound_excel(df_view, doc_title_str, mode):
    import io as _io
    output = _io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        wb = writer.book
        ws = wb.add_worksheet('출고내역')
        
        title_fmt = wb.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#DDEBF7'})
        header_fmt = wb.add_format({'bold': True, 'border': 1, 'align': 'center', 'bg_color': '#F2F2F2'})
        data_fmt = wb.add_format({'border': 1, 'align': 'center'})
        num_fmt = wb.add_format({'border': 1, 'align': 'right', 'num_format': '#,##0'})
        text_left_fmt = wb.add_format({'align': 'left', 'valign': 'vcenter', 'font_size': 11})
        text_bold_fmt = wb.add_format({'align': 'left', 'valign': 'vcenter', 'bold': True, 'font_size': 12})
        
        if mode == "당일 전체 출고조회":
            ws.merge_range('A1:J2', doc_title_str, title_fmt)
            headers = ["출고일", "출고시간", "차량번호", "기사명", "출하처", "유형", "제품명", "규격", "수량", "비고"]
            for i, h in enumerate(headers):
                ws.write(3, i, h, header_fmt)
                
            ws.set_column('A:B', 12); ws.set_column('C:F', 15); ws.set_column('G:G', 25); ws.set_column('H:J', 15)
            
            for ri, (_, row) in enumerate(df_view.iterrows(), 4):
                ws.write(ri, 0, str(row["출고일"]), data_fmt)
                ws.write(ri, 1, str(row["출고시간"]), data_fmt)
                ws.write(ri, 2, str(row.get("차량번호", "")), data_fmt)
                ws.write(ri, 3, str(row.get("기사명", "")), data_fmt)
                ws.write(ri, 4, str(row["출하처"]), data_fmt)
                ws.write(ri, 5, str(row["유형"]), data_fmt)
                ws.write(ri, 6, str(row["제품명"]), data_fmt)
                ws.write(ri, 7, str(row["규격"]), data_fmt)
                try: qty_val = int(float(row["수량"]))
                except: qty_val = 0
                ws.write(ri, 8, qty_val, num_fmt)
                ws.write(ri, 9, str(row["비고"]), data_fmt)
        else:
            # 개별 배차 모드 (수기 작성용 명세서 양식)
            ws.set_column('A:A', 6); ws.set_column('B:C', 15); ws.set_column('D:D', 25)
            ws.set_column('E:G', 12)
            
            ws.merge_range('A1:G2', "출고 명세서", title_fmt)
            
            # 상단부
            r_date = df_view.iloc[0]["출고일"]
            r_time = df_view.iloc[0]["출고시간"]
            r_dest = df_view.iloc[0]["출하처"]
            
            ws.write('A4', f"출고일자 : {r_date}", text_bold_fmt)
            ws.write('D4', f"출고시간 : {r_time}", text_bold_fmt)
            
            # 데이터 헤더
            headers = ["No", "유형", "제품명", "규격", "수량", "비고"]
            for i, h in enumerate(headers):
                ws.write(5, i, h, header_fmt)
            
            # 제품 데이터 쓰기
            current_row = 6
            for idx, (_, row) in enumerate(df_view.iterrows(), 1):
                ws.write(current_row, 0, idx, data_fmt)
                ws.write(current_row, 1, str(row["유형"]), data_fmt)
                ws.write(current_row, 2, str(row["제품명"]), data_fmt)
                ws.write(current_row, 3, str(row["규격"]), data_fmt)
                try: qty_val = int(float(row["수량"]))
                except: qty_val = 0
                ws.write(current_row, 4, qty_val, num_fmt)
                ws.write(current_row, 5, str(row["비고"]), data_fmt)
                current_row += 1
            
            # 하단부 여백 및 수기 입력 공간
            current_row += 2
            ws.write(current_row, 0, f"▶ 출하처 : {r_dest}", text_bold_fmt)
            current_row += 2
            ws.write(current_row, 0, "기사명 : _____________________", text_left_fmt)
            ws.write(current_row, 3, "차량번호 : _____________________", text_left_fmt)
            current_row += 2
            ws.write(current_row, 0, "차량위생상태 : ( 양호 / 불량 )", text_left_fmt)
            ws.write(current_row, 3, "인수자 서명 : ____________________", text_left_fmt)
            
    return output.getvalue()

def export_calib_excel(df_calib):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        wb = writer.book
        ws = wb.add_worksheet('계측기기검교정대장')
        
        title_fmt = wb.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
        header_fmt = wb.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})
        val_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
        
        ws.merge_range('A1:G2', "계측기기 검교정 관리 대장", title_fmt)
        
        headers = ["구분", "계측기기명", "관리번호", "주기(개월)", "최종 검교정일", "차기 검교정일", "비고"]
        for c, h in enumerate(headers):
            ws.write(3, c, h, header_fmt)
        
        ws.set_column('A:A', 12); ws.set_column('B:C', 20); ws.set_column('D:D', 12); ws.set_column('E:F', 15); ws.set_column('G:G', 25)
        
        r_idx = 4
        for _, row in df_calib.iterrows():
            ws.write(r_idx, 0, str(row.get('구분', '')), val_fmt)
            ws.write(r_idx, 1, str(row.get('계측기기명', '')), val_fmt)
            ws.write(r_idx, 2, str(row.get('관리번호', '')), val_fmt)
            ws.write(r_idx, 3, str(row.get('주기', '')), val_fmt)
            ws.write(r_idx, 4, str(row.get('검교정일자', '')), val_fmt)
            ws.write(r_idx, 5, str(row.get('차기_검교정일자', '')), val_fmt)
            ws.write(r_idx, 6, str(row.get('비고', '')), val_fmt)
            ws.set_row(r_idx, 22)
            r_idx += 1
    return output.getvalue()

def export_production_daily_excel(df_daily, date_str):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        wb = writer.book
        ws = wb.add_worksheet('일일생산기록')
        
        title_fmt = wb.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
        header_fmt = wb.add_format({'bg_color': '#E7E6E6', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True, 'font_size': 10})
        val_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 9})
        pass_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 9, 'bg_color': '#D9EAD3', 'font_color': '#274E13'})
        fail_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 9, 'bg_color': '#F4CCCC', 'font_color': '#990000'})
        
        ws.merge_range('A1:M2', f"일일 생산 품질 점검 기록표 ({date_str})", title_fmt)
        
        headers = ["No", "유형", "제품명", "생산량", "소비기한", "규격", "질소(%)", "수분(%)", "색도", "추출시간", "판정", "비고", "상세"]
        for c, h in enumerate(headers):
            ws.write(3, c, h, header_fmt)
        
        ws.set_column('A:A', 5); ws.set_column('B:B', 10); ws.set_column('C:C', 25); ws.set_column('D:D', 10)
        ws.set_column('E:F', 15); ws.set_column('G:J', 10); ws.set_column('K:K', 10); ws.set_column('L:M', 20)
        
        r_idx = 4
        for ni, (_, row) in enumerate(df_daily.iterrows(), 1):
            ws.write(r_idx, 0, ni, val_fmt)
            ws.write(r_idx, 1, str(row.get('유형', '')), val_fmt)
            ws.write(r_idx, 2, str(row.get('제품명', '')), val_fmt)
            ws.write(r_idx, 3, str(row.get('생산량', '')), val_fmt)
            ws.write(r_idx, 4, str(row.get('소비기한', '')), val_fmt)
            ws.write(r_idx, 5, str(row.get('규격', '')), val_fmt)
            ws.write(r_idx, 6, str(row.get('질소(%)', '')), val_fmt)
            ws.write(r_idx, 7, str(row.get('수분(%)', '')), val_fmt)
            ws.write(r_idx, 8, str(row.get('색도(Agtron)', '')), val_fmt)
            ws.write(r_idx, 9, str(row.get('추출시간(sec)', '')), val_fmt)
            
            p_val = str(row.get('판정', ''))
            f = val_fmt
            if p_val == "PASS": f = pass_fmt
            elif p_val == "FAIL": f = fail_fmt
            ws.write(r_idx, 10, p_val, f)
            
            ws.write(r_idx, 11, str(row.get('비고', '')), val_fmt)
            ws.write(r_idx, 12, str(row.get('추출시간_상세', '')), val_fmt)
            ws.set_row(r_idx, 22)
            r_idx += 1
    return output.getvalue()

def export_verify_plan_excel(df_v):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        wb = writer.book
        ws = wb.add_worksheet('연간검증계획표')
        title_fmt = wb.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
        header_fmt = wb.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})
        val_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
        
        ws.merge_range('A1:G2', "연간 HACCP 검증 계획표", title_fmt)
        headers = ["계획일자", "검증종류", "검증항목", "세부내용", "검증방법", "상태", "담당자"]
        for c, h in enumerate(headers):
            ws.write(3, c, h, header_fmt)
        
        ws.set_column('A:A', 15); ws.set_column('B:B', 12); ws.set_column('C:C', 20); ws.set_column('D:D', 30)
        ws.set_column('E:G', 15)
        
        r_idx = 4
        for _, row in df_v.iterrows():
            ws.write(r_idx, 0, str(row.get('계획일자', '')), val_fmt)
            ws.write(r_idx, 1, str(row.get('검증종류', '')), val_fmt)
            ws.write(r_idx, 2, str(row.get('검증항목', '')), val_fmt)
            ws.write(r_idx, 3, str(row.get('세부내용', '')), val_fmt)
            ws.write(r_idx, 4, str(row.get('검증방법', '')), val_fmt)
            ws.write(r_idx, 5, str(row.get('상태', '')), val_fmt)
            ws.write(r_idx, 6, str(row.get('담당자', '')), val_fmt)
            ws.set_row(r_idx, 22); r_idx += 1
    return output.getvalue()

def export_rm_list_excel(df_rm):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        wb = writer.book
        ws = wb.add_worksheet('원부자재목록')
        title_fmt = wb.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
        header_fmt = wb.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})
        val_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
        
        ws.merge_range('A1:H2', "원·부자재 리스트", title_fmt)
        headers = ["코드", "원자자명", "유형", "규격", "원산지", "제조원", "검사주기", "비고"]
        for c, h in enumerate(headers):
            ws.write(3, c, h, header_fmt)
        
        ws.set_column('A:B', 15); ws.set_column('C:E', 12); ws.set_column('F:H', 20)
        
        r_idx = 4
        for _, row in df_rm.iterrows():
            ws.write(r_idx, 0, str(row.get('원자재코드', '')), val_fmt)
            ws.write(r_idx, 1, str(row.get('원자재명', '')), val_fmt)
            ws.write(r_idx, 2, str(row.get('유형', '')), val_fmt)
            ws.write(r_idx, 3, str(row.get('규격', '')), val_fmt)
            ws.write(r_idx, 4, str(row.get('원산지', '')), val_fmt)
            ws.write(r_idx, 5, str(row.get('제조원', '')), val_fmt)
            ws.write(r_idx, 6, str(row.get('검사주기', '')), val_fmt)
            ws.write(r_idx, 7, str(row.get('비고', '')), val_fmt)
            ws.set_row(r_idx, 22); r_idx += 1
    return output.getvalue()