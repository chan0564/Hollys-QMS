import streamlit as st
import pandas as pd
import os
import plotly.express as px
from datetime import datetime, date
import base64
import io
import re

# ==========================================
# 0. 다크모드 원천 차단 (라이트 모드 고정 설정 파일 자동 생성)
# ==========================================
if not os.path.exists('.streamlit'):
    os.makedirs('.streamlit')
if not os.path.exists('.streamlit/config.toml'):
    with open('.streamlit/config.toml', 'w', encoding='utf-8') as f:
        f.write('[theme]\nbase="light"\n')

# ==========================================
# 1. 페이지 기본 설정 및 디자인
# ==========================================
st.set_page_config(page_title="Hollys QMS Premium", layout="wide", initial_sidebar_state="expanded")

# ==========================================
# 로그인 기능
# ==========================================
# def check_login():
#     if "logged_in" not in st.session_state:
#         st.session_state.logged_in = False

#     if not st.session_state.logged_in:
#         # 로그인 화면
#         st.markdown("""
#         <style>
#         .login-wrap {
#             max-width: 420px;
#             margin: 80px auto;
#             background: white;
#             border-radius: 16px;
#             padding: 48px 40px;
#             box-shadow: 0 4px 24px rgba(0,0,0,0.10);
#             text-align: center;
#         }
#         .login-logo {
#             color: #D11031;
#             font-size: 36px;
#             font-weight: 900;
#             letter-spacing: 2px;
#             margin-bottom: 6px;
#         }
#         .login-sub {
#             color: #888;
#             font-size: 13px;
#             margin-bottom: 32px;
#         }
#         </style>
#         <div class="login-wrap">
#             <div class="login-logo">HOLLYS</div>
#             <div class="login-sub">Roasting Center QMS</div>
#         </div>
#         """, unsafe_allow_html=True)

#         col1, col2, col3 = st.columns([1, 2, 1])
#         with col2:
#             st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
#             pw = st.text_input("비밀번호", type="password", placeholder="비밀번호를 입력하세요", key="login_pw")
#             if st.button("로그인", use_container_width=True, type="primary"):
#                 # 에러가 나지 않도록 try-except 문으로 안전하게 감싸줍니다.
#                 try:
#                     correct_pw = st.secrets["PASSWORD"]
#                 except:
#                     correct_pw = "hollys!24124" # Secrets가 없으면 기본 비밀번호 적용
                
#                 if pw == correct_pw:
#                     st.session_state.logged_in = True
#                     st.rerun()
#                 else:
#                     st.error("비밀번호가 올바르지 않습니다.")
#         st.stop()

# check_login()
st.session_state.logged_in = True  # 임시 로그인 상태 유지

# [수정] 표를 가려버리던 악성 CSS 요소 제거 및 깔끔한 원복
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700&display=swap');
    
    html, body, [class*="css"]  { font-family: 'Noto Sans KR', sans-serif; }
    .stApp { background-color: #F8F9FA !important; }
    
    /* ── 기본 버튼: 옅은 회색 ── */
    .stButton > button {
        background-color: #EBEBEB !important;
        color: #333333 !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        border: 1px solid #D5D5D5 !important;
        padding: 0.5rem 1rem !important;
        box-shadow: 0 1px 3px rgba(0,0,0,0.08) !important;
        transition: all 0.15s ease !important;
        width: 100%;
    }
    .stButton > button:hover {
        background-color: #DCDCDC !important;
        color: #111111 !important;
        box-shadow: 0 2px 6px rgba(0,0,0,0.12) !important;
        transform: translateY(-1px) !important;
    }

    /* ── primary 버튼 (type="primary") : 옅은 파란색 ── */
    .stButton > button[kind="primary"] {
        background-color: #D6E8FF !important;
        color: #1A5FAD !important;
        border: 1px solid #A8CFFF !important;
        box-shadow: 0 1px 4px rgba(26,95,173,0.15) !important;
    }
    .stButton > button[kind="primary"]:hover {
        background-color: #C0DAFF !important;
        color: #134A8E !important;
        border-color: #80B8FF !important;
        box-shadow: 0 2px 8px rgba(26,95,173,0.22) !important;
    }

    /* ── form submit 버튼도 동일하게 ── */
    .stFormSubmitButton > button {
        background-color: #D6E8FF !important;
        color: #1A5FAD !important;
        border: 1px solid #A8CFFF !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        padding: 0.5rem 1rem !important;
        box-shadow: 0 1px 4px rgba(26,95,173,0.15) !important;
        transition: all 0.15s ease !important;
        width: 100%;
    }
    .stFormSubmitButton > button:hover {
        background-color: #C0DAFF !important;
        color: #134A8E !important;
        border-color: #80B8FF !important;
    }

    .metric-card { background-color: #FFFFFF; padding: 24px; border-radius: 12px; box-shadow: 0 4px 10px rgba(0,0,0,0.08); border-left: 6px solid #D11031; margin-bottom: 20px; }
    .section-title { font-size: 1.2rem; font-weight: 700; color: #D11031; margin-top: 20px; border-bottom: 2px solid #D11031; padding-bottom: 5px; margin-bottom: 15px; }
    
    /* 캘린더 상태 뱃지 */
    .badge { padding: 4px 12px; border-radius: 12px; font-size: 0.85rem; font-weight: bold; vertical-align: middle; display: inline-block; }
    .badge-upcoming { background-color: #E2F0FD; color: #0D6EFD; border: 1px solid #0D6EFD; }
    .badge-overdue  { background-color: #FFF0F2; color: #D11031; border: 1px solid #D11031; }
    .badge-done     { background-color: #E6F4EA; color: #28A745; border: 1px solid #28A745; }
    .badge-warning  { background-color: #FFF3CD; color: #856404; border: 1px solid #856404; }
    </style>
    """, unsafe_allow_html=True)

# ==========================================
# ✨ 1.5 인트로 스플래시 화면 (HELLO, HOLLYS) 확실히 3초 노출
# ==========================================
def get_base64_of_local_file(file_path):
    if os.path.exists(file_path):
        with open(file_path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    return ""

splash_logo_path = "image_27e084.jpg" if os.path.exists("image_27e084.jpg") else ("image_d70fe2.png" if os.path.exists("image_d70fe2.png") else "")
splash_logo_base64 = get_base64_of_local_file(splash_logo_path)
mime_type = "image/jpeg" if "jpg" in splash_logo_path.lower() else "image/png"
splash_img_tag = f'<img src="data:{mime_type};base64,{splash_logo_base64}" width="150" style="margin-bottom: 20px;">' if splash_logo_base64 else ''

# [수정] 마크다운이 코드블록(회색박스)으로 오해하지 않도록 문자열 안의 모든 들여쓰기를 제거했습니다.
if 'welcomed' not in st.session_state:
    st.session_state.welcomed = True
    
    splash_html = f"""
<style>
@keyframes splashFadeOut {{
0% {{ opacity: 1; visibility: visible; }}
80% {{ opacity: 1; visibility: visible; }}
100% {{ opacity: 0; visibility: hidden; display: none; z-index: -1; }}
}}
.splash-container {{
position: fixed; top: 0; left: 0; width: 100vw; height: 100vh;
background-color: #FFFFFF; z-index: 9999999;
display: flex; flex-direction: column; justify-content: center; align-items: center;
animation: splashFadeOut 3s forwards;
}}
.splash-container h1 {{ color: #212529; font-size: 5rem; letter-spacing: 5px; font-weight: 900; margin: 0; }}
.splash-container p {{ color: #6C757D; font-size: 1.5rem; margin-top: 15px; }}
</style>
<div class="splash-container">
{splash_img_tag}
<h1>HELLO, HOLLYS</h1>
<p>Digital QMS Premium System</p>
</div>
"""
    st.markdown(splash_html, unsafe_allow_html=True)

# ==========================================
# 2. 로고 및 타이틀
# ==========================================
import base64

# 👇 아까 성공하신 로고 파일 이름 그대로 적어주세요.
logo_filename = "image_1ab965.png" 

if os.path.exists(logo_filename):
    with open(logo_filename, "rb") as f:
        logo_b64 = base64.b64encode(f.read()).decode()
    
    # 💡 margin-right: 15px; 이 부분이 로고와 글자 사이의 간격입니다! (숫자를 5px 등으로 줄이면 더 가까워집니다)
    st.markdown(f"""
        <div style="display: flex; align-items: center; margin-top: -10px; margin-bottom: 20px;">
            <img src="data:image/png;base64,{logo_b64}" width="120" style="margin-right: 15px;">
            <div>
                <h1 style='color: #212529; margin: 0; font-size: 2.5rem;'>Hollys Roasting Center QMS</h1>
                <p style='color: #6C757D; font-size: 1.1rem; margin: 5px 0 0 0;'>품질관리팀 전용 디지털 거버넌스 시스템</p>
            </div>
        </div>
    """, unsafe_allow_html=True)
else:
    st.error(f"'{logo_filename}' 파일을 찾을 수 없습니다!")
    st.markdown("<h1 style='color: #212529; margin-top: -15px;'>Hollys Roasting Center QMS</h1>", unsafe_allow_html=True)
    st.markdown("<p style='color: #6C757D; font-size: 1.1rem; margin-top: -15px; margin-bottom: 20px;'>품질관리팀 전용 디지털 거버넌스 시스템</p>", unsafe_allow_html=True)
# ==========================================
# 3. 데이터베이스 로드 함수 및 상수
# ==========================================
DATA_FILE = "qc_data.csv"
SPEC_FILE = "qc_specs.csv"
CLEAN_FILE = "cleaning_specs.csv"
VERIFY_FILE = "verify_plan.csv"
OTHER_SCHED_FILE = "other_schedule.csv"
HEALTH_CERT_FILE = "health_cert.csv" 
EMPLOYEE_FILE = "employees.csv"
FACILITY_FILE = "facilities.csv"
REPAIR_FILE = "repairs.csv"
ANNUAL_LEAVE_FILE = "annual_leave_records.csv"   # 파트타이머 연차 기록
CALIB_LIST_FILE = "calib_list.csv"
CALIB_REPORT_FILE = "calib_report.csv"
FILTER_PLAN_FILE = "filter_plan.csv"
NOTICE_BOARD_FILE = "notice_board.csv" 
POLL_BOARD_FILE = "poll_board.csv" 
HACCP_REVISION_FILE = "haccp_revision.csv"
FLOW_FILE = "haccp_flowchart.csv"
FLOW_CATEGORY_FILE = "haccp_flow_categories.csv"
CCP_DECISION_DIR = "ccp_decisions"
HAZARD_FILE      = "hazard_analysis.csv"
RM_HAZARD_FILE   = "rm_hazard_analysis.csv"  # 원부자재 위해요소분석
RM_FILE          = "raw_materials.csv"        # 원·부자재 등록
RM_SPEC_DIR      = "rm_specs"                 # 원·부자재 규격서 디렉토리
OUTBOUND_FILE    = "outbound_records.csv"     # 출고 기록 데이터
INVENTORY_ADJ_FILE = "inventory_adj.csv"      # 재고 임의 조정 로그
HISTORY_LOG_FILE = "history_log.csv"          # 시스템 작업 히스토리 로그

def log_history(action, target, detail=""):
    """작성 일시, 작업자(세션 정보), 대상 메뉴/기능, 작업 내용, 세부 사항을 CSV에 기록"""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    worker = st.session_state.get("worker_name", "시스템/관리자")
    
    new_entry = pd.DataFrame([{
        "일시": now,
        "작업자": worker,
        "대상": target,
        "작업내용": action,
        "상세": detail
    }])
    
    if os.path.exists(HISTORY_LOG_FILE):
        try:
            new_entry.to_csv(HISTORY_LOG_FILE, mode='a', header=False, index=False, encoding='utf-8-sig')
        except:
            pass
    else:
        new_entry.to_csv(HISTORY_LOG_FILE, index=False, encoding='utf-8-sig')

def load_ccp_decision(cat_id: str):
    os.makedirs(CCP_DECISION_DIR, exist_ok=True)
    safe_id = str(cat_id).replace("/","_").replace("\\","_").replace(" ","_")
    fpath = os.path.join(CCP_DECISION_DIR, f"ccp_decision_{safe_id}.csv")
    cols = ["단계명","위해유형","위해요소",
            "Q1_답변","Q1_비고","Q2_답변","Q2_비고","Q21_답변","Q21_비고",
            "Q3_답변","Q3_비고","Q4_답변","Q4_비고","Q5_답변","Q5_비고",
            "CCP결정","CCP번호","비고"]
    if os.path.exists(fpath):
        try:
            df = pd.read_csv(fpath, dtype=str).fillna("")
            for c in cols:
                if c not in df.columns: df[c] = ""
            return df[cols]
        except Exception: pass
    return pd.DataFrame(columns=cols)

def save_ccp_decision(cat_id: str, df):
    os.makedirs(CCP_DECISION_DIR, exist_ok=True)
    safe_id = str(cat_id).replace("/","_").replace("\\","_").replace(" ","_")
    fpath = os.path.join(CCP_DECISION_DIR, f"ccp_decision_{safe_id}.csv")
    df.to_csv(fpath, index=False, encoding="utf-8-sig")
    log_history("CCP 결정도 업데이트", f"공정: {cat_id}", f"데이터 {len(df)}건 저장")

def load_haccp_revisions():
    if os.path.exists(HACCP_REVISION_FILE):
        try: 
            return pd.read_csv(HACCP_REVISION_FILE, dtype=str)
        except Exception: pass
    df_rev = pd.DataFrame(columns=["개정일자", "분류", "문서명", "개정번호", "개정사유"])
    df_rev.to_csv(HACCP_REVISION_FILE, index=False, encoding='utf-8-sig')
    return df_rev
# ──────────────────────────────────────────────────────────────
# STEP 2. 로드 함수 (load_haccp_revisions() 함수 바로 아래에 추가)
# ──────────────────────────────────────────────────────────────
def load_flowchart():
    """공정 흐름도 단계 데이터 로드. 없으면 기본 커피 제조 공정 샘플 생성."""
    if os.path.exists(FLOW_FILE):
        try:
            return pd.read_csv(FLOW_FILE, dtype=str)
        except Exception:
            pass
    # 기본 샘플 데이터 (Hollys 커피 캡슐 기준)
    default_steps = [
        {"순서": "1",  "단계명": "원료 입고",        "설명": "생두, 부자재 입고 및 수량 확인",        "유형": "공정",  "CCP여부": "N", "CCP번호": ""},
        {"순서": "2",  "단계명": "원료 검사",        "설명": "관능검사, 이물 확인, 서류 검토",         "유형": "검사",  "CCP여부": "N", "CCP번호": ""},
        {"순서": "3",  "단계명": "원료 보관",        "설명": "온·습도 관리 창고 보관 (15℃ 이하)",     "유형": "공정",  "CCP여부": "N", "CCP번호": ""},
        {"순서": "4",  "단계명": "로스팅",           "설명": "로스터기 가열 처리 (열처리 CCP)",        "유형": "공정",  "CCP여부": "Y", "CCP번호": "CCP-1"},
        {"순서": "5",  "단계명": "냉각",             "설명": "로스팅 후 신속 냉각",                   "유형": "공정",  "CCP여부": "N", "CCP번호": ""},
        {"순서": "6",  "단계명": "분쇄",             "설명": "제품 규격에 맞는 입도 분쇄",             "유형": "공정",  "CCP여부": "N", "CCP번호": ""},
        {"순서": "7",  "단계명": "품질 검사",        "설명": "색도·수분·질소 측정 및 합불 판정",      "유형": "검사",  "CCP여부": "Y", "CCP번호": "CCP-2"},
        {"순서": "8",  "단계명": "충전/포장",        "설명": "캡슐 충전, 질소 치환, 밀봉",            "유형": "공정",  "CCP여부": "N", "CCP번호": ""},
        {"순서": "9",  "단계명": "금속 검출",        "설명": "금속 검출기 통과 (이물 CCP)",           "유형": "검사",  "CCP여부": "Y", "CCP번호": "CCP-3"},
        {"순서": "10", "단계명": "완제품 검사",      "설명": "외관, 중량, 표시사항 최종 확인",         "유형": "검사",  "CCP여부": "N", "CCP번호": ""},
        {"순서": "11", "단계명": "완제품 보관",      "설명": "온·습도 관리 창고 보관",                "유형": "공정",  "CCP여부": "N", "CCP번호": ""},
        {"순서": "12", "단계명": "출하/납품",        "설명": "주문 확인 및 배송 출고",                "유형": "공정",  "CCP여부": "N", "CCP번호": ""},
    ]
    df_flow = pd.DataFrame(default_steps)
    df_flow.to_csv(FLOW_FILE, index=False, encoding="utf-8-sig")
    log_history("기본 공정 흐름도 생성", "HACCP 공정 흐름도", "최초 샘플 데이터 생성")
    return df_flow
def load_flow_categories():
    """공정 흐름도 카테고리(제품군/라인) 목록 로드"""
    if os.path.exists(FLOW_CATEGORY_FILE):
        try:
            return pd.read_csv(FLOW_CATEGORY_FILE, dtype=str)
        except Exception:
            pass
    default_cats = pd.DataFrame([
        {"cat_id": "main", "cat_name": "메인 공정 (커피 원두)", "description": "핵심 제조 공정 흐름"},
    ])
    default_cats.to_csv(FLOW_CATEGORY_FILE, index=False, encoding="utf-8-sig")
    return default_cats


def load_flowchart_by_cat(cat_id: str):
    """카테고리 ID에 해당하는 공정 단계 CSV 로드. 없으면 기본값 생성."""
    safe_id = str(cat_id).replace("/", "_").replace("\\", "_").replace(" ", "_")
    fpath = f"haccp_flowchart_{safe_id}.csv"
    if os.path.exists(fpath):
        try:
            return pd.read_csv(fpath, dtype=str)
        except Exception:
            pass
    # 기본 카테고리(main)는 샘플 데이터 자동 생성
    if cat_id == "main":
        default_steps = [
            {"순서": "1",  "단계명": "원료 입고",    "설명": "생두 입고 및 수량 확인",             "유형": "공정", "CCP여부": "N", "CCP번호": "", "lane": "원재료", "merge_from": "", "merge_label": "", "merge_from_2": "", "merge_label_2": "", "merge_from_3": "", "merge_label_3": ""},
            {"순서": "2",  "단계명": "원료 검사",    "설명": "관능검사, 이물 확인",                "유형": "검사", "CCP여부": "N", "CCP번호": "", "lane": "원재료", "merge_from": "", "merge_label": "", "merge_from_2": "", "merge_label_2": "", "merge_from_3": "", "merge_label_3": ""},
            {"순서": "3",  "단계명": "원료 보관",    "설명": "온·습도 관리 창고 (15℃ 이하)",      "유형": "공정", "CCP여부": "N", "CCP번호": "", "lane": "원재료", "merge_from": "", "merge_label": "", "merge_from_2": "", "merge_label_2": "", "merge_from_3": "", "merge_label_3": ""},
            {"순서": "4",  "단계명": "선 별",        "설명": "이물·불량두 제거",                   "유형": "공정", "CCP여부": "N", "CCP번호": "", "lane": "원재료", "merge_from": "", "merge_label": "", "merge_from_2": "", "merge_label_2": "", "merge_from_3": "", "merge_label_3": ""},
            {"순서": "5",  "단계명": "계량/배합",    "설명": "블렌딩 배합비 계량",                 "유형": "공정", "CCP여부": "N", "CCP번호": "", "lane": "원재료", "merge_from": "", "merge_label": "", "merge_from_2": "", "merge_label_2": "", "merge_from_3": "", "merge_label_3": ""},
            {"순서": "6",  "단계명": "가열(로스팅)", "설명": "로스터기 가열 처리 (열처리 CCP)",    "유형": "공정", "CCP여부": "Y", "CCP번호": "CCP-1", "lane": "원재료", "merge_from": "", "merge_label": "", "merge_from_2": "", "merge_label_2": "", "merge_from_3": "", "merge_label_3": ""},
            {"순서": "7",  "단계명": "방 냉",        "설명": "로스팅 후 신속 냉각",                "유형": "공정", "CCP여부": "N", "CCP번호": "", "lane": "원재료", "merge_from": "", "merge_label": "", "merge_from_2": "", "merge_label_2": "", "merge_from_3": "", "merge_label_3": ""},
            {"순서": "8",  "단계명": "내포장",       "설명": "내포장재(PE) 충전 및 밀봉",          "유형": "공정", "CCP여부": "N", "CCP번호": "", "lane": "원재료", "merge_from": "포장재", "merge_label": "내포장재(PE)", "merge_from_2": "기타", "merge_label_2": "질소", "merge_from_3": "", "merge_label_3": ""},
            {"순서": "9",  "단계명": "X-ray이물검출","설명": "이물 검출기 통과 (이물 CCP)",        "유형": "검사", "CCP여부": "Y", "CCP번호": "CCP-2", "lane": "원재료", "merge_from": "", "merge_label": "", "merge_from_2": "", "merge_label_2": "", "merge_from_3": "", "merge_label_3": ""},
            {"순서": "10", "단계명": "외포장",       "설명": "외포장재(박스) 포장",                "유형": "공정", "CCP여부": "N", "CCP번호": "", "lane": "원재료", "merge_from": "포장재", "merge_label": "외포장재(박스)", "merge_from_2": "", "merge_label_2": "", "merge_from_3": "", "merge_label_3": ""},
            {"순서": "11", "단계명": "보 관",        "설명": "완제품 창고 온·습도 관리",           "유형": "공정", "CCP여부": "N", "CCP번호": "", "lane": "원재료", "merge_from": "", "merge_label": "", "merge_from_2": "", "merge_label_2": "", "merge_from_3": "", "merge_label_3": ""},
            {"순서": "12", "단계명": "출 고",        "설명": "주문 확인 및 배송 출고",             "유형": "공정", "CCP여부": "N", "CCP번호": "", "lane": "원재료", "merge_from": "", "merge_label": "", "merge_from_2": "", "merge_label_2": "", "merge_from_3": "", "merge_label_3": ""},
            # 포장재 레인 (독립 흐름)
            {"순서": "1",  "단계명": "포장재 입고",  "설명": "내·외포장재 입고 및 검수",           "유형": "공정", "CCP여부": "N", "CCP번호": "", "lane": "포장재", "merge_from": "", "merge_label": "", "merge_from_2": "", "merge_label_2": "", "merge_from_3": "", "merge_label_3": ""},
            {"순서": "2",  "단계명": "포장재 보관",  "설명": "청결구역 보관 (온·습도 관리)",       "유형": "공정", "CCP여부": "N", "CCP번호": "", "lane": "포장재", "merge_from": "", "merge_label": "", "merge_from_2": "", "merge_label_2": "", "merge_from_3": "", "merge_label_3": ""},
            # 기타 레인 (질소 등)
            {"순서": "1",  "단계명": "질소 입고",    "설명": "식품용 질소 가스 입고·검수",         "유형": "공정", "CCP여부": "N", "CCP번호": "", "lane": "기타",   "merge_from": "", "merge_label": "", "merge_from_2": "", "merge_label_2": "", "merge_from_3": "", "merge_label_3": ""},
        ]
        df_f = pd.DataFrame(default_steps)
        df_f.to_csv(fpath, index=False, encoding="utf-8-sig")
        return df_f

    return pd.DataFrame(columns=["순서", "단계명", "설명", "유형", "CCP여부", "CCP번호", "lane", "merge_from", "merge_label", "merge_from_2", "merge_label_2", "merge_from_3", "merge_label_3"])

# 세척소독 대분류 11종
CLEAN_CATEGORIES = [
    "1. 종업원", "2. 위생복장", "3. 작업장 주변", "4. 작업장 내부(공통)", 
    "5. 식품 제조시설", "6. 보관시설", "7. 운반도구 및 용기", 
    "8. 모니터링 및 검사장비", "9. 환기시설", "10. 폐기물처리용기", "11. 세척 소독 도구"
]

def load_data():
    cols = ["생산일", "유형", "제품명", "생산량", "규격", "질소(%)", "수분(%)", "색도(Agtron)", "추출시간(sec)", "추출시간_상세", "날짜기록", "판정", "비고"]
    if os.path.exists(DATA_FILE): 
        try: 
            df = pd.read_csv(DATA_FILE)
            if "단위" in df.columns and "규격" not in df.columns:
                df = df.rename(columns={"단위": "규격"})
            for col in cols:
                if col not in df.columns:
                    if col == "생산량": df[col] = 0
                    else: df[col] = ""
            return df[cols]
        except Exception: pass
    return pd.DataFrame(columns=cols)

def load_outbound_records():
    """출고 기록 데이터 로드"""
    cols = ["출고일", "출고시간", "차량번호", "기사명", "출하처", "유형", "제품명", "규격", "수량", "비고"]
    if os.path.exists(OUTBOUND_FILE):
        try:
            df = pd.read_csv(OUTBOUND_FILE, dtype=str)
            for c in cols:
                if c not in df.columns: df[c] = ""
            return df[cols]
        except Exception: pass
    return pd.DataFrame(columns=cols)

def load_inventory_adj():
    cols = ["조정일시", "유형", "제품명", "규격", "기존재고", "변경재고", "방향", "차이", "사유"]
    if os.path.exists(INVENTORY_ADJ_FILE):
        try: return pd.read_csv(INVENTORY_ADJ_FILE, dtype=str)
        except Exception: pass
    return pd.DataFrame(columns=cols)

def load_specs():
    if os.path.exists(SPEC_FILE): 
        try:
            df = pd.read_csv(SPEC_FILE, dtype=str)
            if "제품코드" not in df.columns: df.insert(0, "제품코드", [f"P-{str(i+1).zfill(3)}" for i in range(len(df))])
            if "규격" not in df.columns:
                if "단위" in df.columns:
                    df = df.rename(columns={"단위": "규격"})
                else:
                    df.insert(3, "규격", "EA") 
            return df
        except Exception: pass
    return pd.DataFrame(columns=["제품코드", "제품명", "유형", "규격", "최소_질소", "최대_질소", "최소_수분", "최대_수분", "최소_색도", "최대_색도", "최소_추출", "최대_추출", "날짜유형"])

def load_cleaning_specs():
    if os.path.exists(CLEAN_FILE): 
        try:
            df = pd.read_csv(CLEAN_FILE, dtype=str)
            req_cols = ['ID', '대분류', '구역', '설비명', '부위', '세척소독방법', '주기', '사용도구', '책임자', '사진파일']
            for col in req_cols:
                if col not in df.columns:
                    if col == 'ID': df['ID'] = [f"C-{str(i).zfill(4)}" for i in range(len(df))]
                    elif col == '대분류': df['대분류'] = "5. 식품 제조시설"  
                    elif col == '설비명': df['설비명'] = "기본대상"
                    elif col == '부위': df['부위'] = df.get('관리부위', "")
                    elif col == '세척소독방법': df['세척소독방법'] = df.get('작업방법', "")
                    elif col == '주기': df['주기'] = df.get('청소주기', "")
                    elif col == '사용도구': df['사용도구'] = df.get('세제_도구', "")
                    else: df[col] = ""
            return df[req_cols]
        except Exception: pass
    df_clean = pd.DataFrame(columns=['ID', '대분류', '구역', '설비명', '부위', '세척소독방법', '주기', '사용도구', '책임자', '사진파일'])
    df_clean.to_csv(CLEAN_FILE, index=False, encoding='utf-8-sig')
    return df_clean

def load_filter_plan():
    if os.path.exists(FILTER_PLAN_FILE):
        try:
            df = pd.read_csv(FILTER_PLAN_FILE, dtype=str).fillna("")
            if '설비명_위치' in df.columns:
                df.rename(columns={'설비명_위치': '설치장소', '필터종류': '필터명', '최근점검일': '점검일자', '차기점검일': '차기점검일자'}, inplace=True)
                if '내용' not in df.columns: df['내용'] = '교체'
            if '설치장소' in df.columns: return df
        except Exception: pass
    df_f = pd.DataFrame(columns=["설치장소", "필터명", "내용", "주기_개월", "점검일자", "차기점검일자", "상태", "비고"])
    df_f.to_csv(FILTER_PLAN_FILE, index=False, encoding='utf-8-sig')
    return df_f

def load_verify():
    if os.path.exists(VERIFY_FILE): 
        try: 
            df = pd.read_csv(VERIFY_FILE)
            if '담당자' not in df.columns: df['담당자'] = ""
            if '계획일자' in df.columns: return df
        except Exception: pass
    df_v = pd.DataFrame(columns=["계획일자", "검증종류", "검증항목", "세부내용", "검증방법", "상태", "담당자"])
    df_v.to_csv(VERIFY_FILE, index=False, encoding='utf-8-sig')
    return df_v

def load_other_sched():
    if os.path.exists(OTHER_SCHED_FILE): 
        try:
            df = pd.read_csv(OTHER_SCHED_FILE)
            if '담당자' not in df.columns: df['담당자'] = ""
            if '일자' in df.columns: return df
        except Exception: pass
    df_o = pd.DataFrame(columns=["일자", "일정명", "세부내용", "상태", "담당자"])
    df_o.to_csv(OTHER_SCHED_FILE, index=False, encoding='utf-8-sig')
    return df_o

def load_health_cert():
    if os.path.exists(HEALTH_CERT_FILE):
        try:
            df = pd.read_csv(HEALTH_CERT_FILE)
            if '검진일자' in df.columns: return df
        except Exception: pass
    df_hc = pd.DataFrame(columns=["직급", "이름", "연락처", "검진일자"])
    df_hc.to_csv(HEALTH_CERT_FILE, index=False, encoding='utf-8-sig')
    return df_hc

def load_employees():
    if os.path.exists(EMPLOYEE_FILE):
        try:
            df = pd.read_csv(EMPLOYEE_FILE, dtype=str)
            if 'HACCP 직책' not in df.columns: df['HACCP 직책'] = "해당없음"
            if '모니터링 CCP' not in df.columns: df['모니터링 CCP'] = ""
            if '기타' not in df.columns: df['기타'] = ""
            if '사번' in df.columns: return df
        except Exception: pass
    df_emp = pd.DataFrame(columns=["사번", "직급", "이름", "연락처", "입사일", "재직상태", "HACCP 직책", "모니터링 CCP", "기타"])
    df_emp.to_csv(EMPLOYEE_FILE, index=False, encoding='utf-8-sig')
    return df_emp

def load_facilities():
    if os.path.exists(FACILITY_FILE):
        try: return pd.read_csv(FACILITY_FILE, dtype=str)
        except Exception: pass
    df = pd.DataFrame(columns=["설비번호", "설비명", "사용용도", "전압", "구입년월", "제조회사명", "설치장소", "관리부서", "관리자_정", "관리자_부", "특이사항"])
    df.to_csv(FACILITY_FILE, index=False, encoding='utf-8-sig')
    return df

def load_repairs():
    if os.path.exists(REPAIR_FILE):
        try: return pd.read_csv(REPAIR_FILE, dtype=str)
        except Exception: pass
    df = pd.DataFrame(columns=["설비번호", "수리일자", "수리사항", "수리처", "비고"])
    df.to_csv(REPAIR_FILE, index=False, encoding='utf-8-sig')
    return df
def load_annual_leave():
    """파트타이머 연차 사용 기록 로드 (사번, 이름, 날짜, 유형, 비고)"""
    if os.path.exists(ANNUAL_LEAVE_FILE):
        try:
            df = pd.read_csv(ANNUAL_LEAVE_FILE, dtype=str).fillna("")
            for c in ["사번","이름","날짜","유형","비고"]:
                if c not in df.columns: df[c] = ""
            return df[["사번","이름","날짜","유형","비고"]]
        except Exception: pass
    df = pd.DataFrame(columns=["사번","이름","날짜","유형","비고"])
    df.to_csv(ANNUAL_LEAVE_FILE, index=False, encoding="utf-8-sig")
    return df

def save_annual_leave(df):
    df.to_csv(ANNUAL_LEAVE_FILE, index=False, encoding="utf-8-sig")

def calc_annual_leave(hire_date_str, leave_records_df, emp_id, today=None):
    """
    발생 규칙:
      - 입사 후 1년 미만: 매월(입사일 기준) 1개 발생, 단 해당 월에 무급휴가 사용 시 그 달은 1개 미발생
      - 입사 후 1년 도달 시점에 15개 일괄 발생 (그 이후는 매년 15개)
    반환: (발생, 사용, 잔여, 상세dict)
    """
    from datetime import date as dclass, timedelta
    from dateutil.relativedelta import relativedelta

    if today is None:
        today = dclass.today()

    try:
        hire = pd.to_datetime(hire_date_str).date()
    except Exception:
        return 0, 0, 0, {}

    # 이 직원의 무급휴가/연차 사용 기록
    emp_df = leave_records_df[leave_records_df["사번"] == emp_id].copy()
    emp_df["날짜_dt"] = pd.to_datetime(emp_df["날짜"], errors="coerce")
    emp_df = emp_df.dropna(subset=["날짜_dt"])

    # 무급휴가 날짜 집합 (년-월)
    unpaid_months = set()
    for _, r in emp_df[emp_df["유형"] == "무급휴가"].iterrows():
        unpaid_months.add((r["날짜_dt"].year, r["날짜_dt"].month))

    # ── 발생 연차 계산 ──────────────────────────────────────────
    months_worked = relativedelta(today, hire).months + relativedelta(today, hire).years * 12
    years_worked  = relativedelta(today, hire).years

    accrued = 0
    detail  = {}   # {기간설명: 발생수}

    if years_worked >= 1:
        # 1년 미만 구간: 최대 11개월 (무급 제외)
        year1_accrued = 0
        for m in range(1, 12):           # 1개월차 ~ 11개월차
            month_point = hire + relativedelta(months=m)
            if month_point > today: break
            ym = (month_point.year, month_point.month)
            # 해당 달(발생 기준월: 전달)에 무급휴가가 있으면 미발생
            prev_ym = (month_point.year, month_point.month)
            # 무급휴가는 발생 전 월(hire ~ month_point 사이 월)에 있으면 차감
            # 정확히: 발생 기준월 = hire + (m-1)개월 ~ hire + m개월 구간
            window_start = hire + relativedelta(months=m-1)
            window_end   = hire + relativedelta(months=m) - timedelta(days=1)
            has_unpaid = any(
                window_start <= r["날짜_dt"].date() <= window_end
                for _, r in emp_df[emp_df["유형"] == "무급휴가"].iterrows()
            )
            if not has_unpaid:
                year1_accrued += 1
        accrued += year1_accrued
        detail["입사 1년 미만 (월 발생)"] = year1_accrued

        # 1년 단위: 1년째, 2년째, ...
        for yr in range(1, years_worked + 1):
            anniversary = hire + relativedelta(years=yr)
            if anniversary <= today:
                accrued += 15
                detail[f"입사 {yr}년 (연간 15개)"] = 15
    else:
        # 1년 미만
        year1_accrued = 0
        for m in range(1, months_worked + 1):
            month_point = hire + relativedelta(months=m)
            if month_point > today: break
            window_start = hire + relativedelta(months=m-1)
            window_end   = hire + relativedelta(months=m) - timedelta(days=1)
            has_unpaid = any(
                window_start <= r["날짜_dt"].date() <= window_end
                for _, r in emp_df[emp_df["유형"] == "무급휴가"].iterrows()
            )
            if not has_unpaid:
                year1_accrued += 1
        accrued = year1_accrued
        detail[f"입사 {months_worked}개월 (월 발생)"] = year1_accrued

    # ── 사용 연차 (유형=='연차'인 레코드 수) ──────────────────────
    used = len(emp_df[emp_df["유형"] == "연차"])
    remaining = accrued - used

    return accrued, used, remaining, detail



def load_calib_list():
    if os.path.exists(CALIB_LIST_FILE):
        try:
            df = pd.read_csv(CALIB_LIST_FILE, dtype=str)
            if "차기_검교정일자" not in df.columns: df["차기_검교정일자"] = ""
            return df
        except Exception: pass
    return pd.DataFrame(columns=["관리번호", "검사_설비명", "측정범위", "주기", "구분", "검교정일자", "차기_검교정일자", "비고"])

def load_calib_reports():
    if os.path.exists(CALIB_REPORT_FILE):
        try: return pd.read_csv(CALIB_REPORT_FILE, dtype=str)
        except Exception: pass
    df = pd.DataFrame(columns=["설비명", "교정일자", "작성자", "검교정방법", "판정기준", "표준값", "측정값", "보정율/오차", "개선조치", "판정결과"])
    df.to_csv(CALIB_REPORT_FILE, index=False, encoding='utf-8-sig')
    return df

def load_notices():
    if os.path.exists(NOTICE_BOARD_FILE):
        try: return pd.read_csv(NOTICE_BOARD_FILE)
        except Exception: pass
    df_n = pd.DataFrame(columns=["작성일", "작성자", "제목", "내용", "중요공지"])
    df_n.to_csv(NOTICE_BOARD_FILE, index=False, encoding='utf-8-sig')
    return df_n

def load_polls():
    if os.path.exists(POLL_BOARD_FILE):
        try: return pd.read_csv(POLL_BOARD_FILE)
        except Exception: pass
    df_p = pd.DataFrame(columns=["ID", "작성일", "제목", "선택지", "투표현황", "참여자"])
    df_p.to_csv(POLL_BOARD_FILE, index=False, encoding='utf-8-sig')
    return df_p

def toggle_task_status(file_path, idx):
    try:
        temp_df = pd.read_csv(file_path)
        current_status = temp_df.loc[idx, '상태']
        temp_df.loc[idx, '상태'] = '예정' if current_status == '완료' else '완료'
        temp_df.to_csv(file_path, index=False, encoding='utf-8-sig')
    except Exception: pass

df = load_data()
df_specs = load_specs()
df_clean = load_cleaning_specs()

if 'selected_pcode' not in st.session_state:
    st.session_state.selected_pcode = None
    st.session_state.selected_pname = None
if 'is_edit_mode' not in st.session_state:
    st.session_state.is_edit_mode = False

# ==========================================
# 좌측 사이드바 메뉴 
# ==========================================
with st.sidebar:
    st.markdown("## 메인 메뉴")
    
    # --- 사이드바 상단 데이터 연결성 요약 ---
    df_p_count = load_specs(); df_rm_count = pd.read_csv(RM_FILE) if os.path.exists(RM_FILE) else pd.DataFrame()
    df_emp_raw = load_employees()
    # 재직 중인 직원 중 직급별 카운트
    if not df_emp_raw.empty and "재직상태" in df_emp_raw.columns:
        active_emp = df_emp_raw[df_emp_raw["재직상태"] == "재직"]
    else:
        active_emp = df_emp_raw

    f_count = active_emp[active_emp['직급'].isin(['팀장', '매니저', '선임매니저'])].shape[0] if not active_emp.empty else 0
    p_count = active_emp[active_emp['직급'] == '파트타이머'].shape[0] if not active_emp.empty else 0
    
    st.markdown(f"""
    <div style='background:white; padding:15px; border-radius:12px; border-left:5px solid #D11031; margin-bottom:20px; box-shadow:0 4px 12px rgba(0,0,0,0.08);'>
        <div style='font-size:13px; font-weight:600; color:#555; margin-bottom:12px;'>📊 품질 데이터 연결 현황</div>
        <div style='display:grid; grid-template-columns: repeat(2, 1fr); gap:10px;'>
            <div style='text-align:center; padding:8px; background:#FDF2F4; border-radius:8px;'>
                <div style='font-size:20px; font-weight:bold; color:#D11031;'>{len(df_p_count)}</div>
                <div style='font-size:11px; color:#666; font-weight:500;'>제품 규격</div>
            </div>
            <div style='text-align:center; padding:8px; background:#F0F7FF; border-radius:8px;'>
                <div style='font-size:20px; font-weight:bold; color:#2E75B6;'>{len(df_rm_count)}</div>
                <div style='font-size:11px; color:#666; font-weight:500;'>원부자재</div>
            </div>
            <div style='text-align:center; padding:8px; background:#F1F9F4; border-radius:8px;'>
                <div style='font-size:20px; font-weight:bold; color:#217346;'>{f_count}</div>
                <div style='font-size:11px; color:#666; font-weight:500;'>정직원</div>
            </div>
            <div style='text-align:center; padding:8px; background:#F7F4FD; border-radius:8px;'>
                <div style='font-size:20px; font-weight:bold; color:#7B5EA7;'>{p_count}</div>
                <div style='font-size:11px; color:#666; font-weight:500;'>PT</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    menu_selection = st.radio(
        "이동할 메뉴를 선택하세요:",
        ["대시보드 (메인)", "게시판", "캘린더", "제품 관리", "원·부자재 관리", "재고 관리", "출고 관리", "직원 관리", "설비 관리", "HACCP", "시스템 히스토리"]
    )

    sub_menu = None
    if menu_selection == "캘린더":
        st.divider()
        st.markdown("#### 캘린더 관리")
        sub_menu = st.radio("하위 메뉴 선택:", ["📅 달력 보기", "검증 계획표", "기타 일정"])

    elif menu_selection == "직원 관리":
        st.divider()
        st.markdown("#### 직원 관리 하위메뉴")
        sub_menu = st.radio("하위 메뉴 선택:", ["조직도 및 인원 관리", "보건증 현황관리", "파트타이머 연차 관리"])

    elif menu_selection == "제품 관리":
        st.divider()
        st.markdown("#### 제품 관리 상세")
        sub_menu = st.radio("하위 메뉴 선택:", ["제품 등록", "제품 규격 마스터", "데이터 히스토리", "제품별 데이터 관리"])

    elif menu_selection == "원·부자재 관리":
        st.divider()
        st.markdown("#### 원·부자재 관리")
        sub_menu = st.radio("하위 메뉴 선택:", ["원·부자재 등록", "원·부자재 규격서 마스터"])

    elif menu_selection == "재고 관리":
        st.divider()
        st.markdown("#### 📦 재고 관리")
        sub_menu = st.radio("하위 메뉴 선택:", ["현재고 현황 및 조정"])

    elif menu_selection == "출고 관리":
        st.divider()
        st.markdown("#### 🚚 출고 관리")
        sub_menu = st.radio("하위 메뉴 선택:", ["출고등록"])

    elif menu_selection == "설비 관리":
        st.divider()
        st.markdown("#### 설비 관리 하위메뉴")
        sub_menu = st.radio("하위 메뉴 선택:", ["제조위생설비이력관리", "필터 점검관리", "세척소독 기준", "계측기기 검교정"])

    elif menu_selection == "HACCP":
        st.divider()
        st.markdown("#### HACCP 관리")
        sub_menu = st.radio("하위 메뉴 선택:", ["HACCP 일지", "HACCP 기준서", "개정이력", "공정 흐름도", "공정별 CCP 결정도", "심각성 설정", "공정 위해요소분석", "원부자재 위해요소분석"])

    st.markdown("<br>" * 3, unsafe_allow_html=True)
    st.markdown(
        "<div style='font-size:10px;color:#aaa;text-align:center;padding-top:8px;'>제작자 : 정찬혁</div>",
        unsafe_allow_html=True
    )

st.divider()

# ==========================================
# 본문 렌더링
# ==========================================

# --- 1. 대시보드 (메인 화면) ---
if menu_selection == "대시보드 (메인)":
    
    c_left, c_right = st.columns([1, 1.2])
    
    with c_left:
        df_notices = load_notices()
        df_polls = load_polls()
        df_emp = load_employees()
        
        today = pd.to_datetime(date.today())
        active_emps = df_emp[df_emp["재직상태"] == "재직"] if not df_emp.empty and "재직상태" in df_emp.columns else pd.DataFrame()
        emp_names = active_emps['이름'].tolist() if not active_emps.empty else []
        
        has_recent_notice = False
        
        if not df_notices.empty:
            df_notices['작성일'] = pd.to_datetime(df_notices['작성일'], errors='coerce')
            recent_notices = df_notices[(df_notices['중요공지'] == True) | (df_notices['중요공지'] == 'True')]
            recent_notices = recent_notices[(today - recent_notices['작성일']).dt.days <= 7]
            
            if not recent_notices.empty:
                has_recent_notice = True
                st.markdown('<div class="section-title" style="color: #0D6EFD; border-bottom: 2px solid #0D6EFD;">📢 중요 공지사항 (최근 7일)</div>', unsafe_allow_html=True)
                for _, notice in recent_notices.sort_values(by="작성일", ascending=False).iterrows():
                    st.info(f"**[{notice['작성일'].strftime('%Y-%m-%d')}] {notice['제목']}** (작성자: {notice['작성자']})\n\n{notice['내용']}")

        if not df_polls.empty:
            df_polls['작성일'] = pd.to_datetime(df_polls['작성일'], errors='coerce')
            recent_polls = df_polls[(today - df_polls['작성일']).dt.days <= 7]
            
            if not recent_polls.empty:
                if not has_recent_notice:
                    st.markdown('<div class="section-title" style="color: #0D6EFD; border-bottom: 2px solid #0D6EFD;">📢 최근 중요 공지사항 및 투표</div>', unsafe_allow_html=True)
                
                for _, poll in recent_polls.sort_values(by="작성일", ascending=False).iterrows():
                    poll_id = poll['ID']
                    with st.container(border=True):
                        st.markdown(f"#### 📊 [익명 투표] {poll['제목']}")
                        st.caption(f"등록일: {poll['작성일'].strftime('%Y-%m-%d')}")
                        
                        options = [o.strip() for o in str(poll['선택지']).split(",")]
                        counts_str = str(poll['투표현황'])
                        voters = str(poll['참여자']).split(",") if pd.notna(poll['참여자']) and poll['참여자'] != "" else []
                        
                        current_counts = {}
                        for item in counts_str.split(","):
                            if ":" in item:
                                k, v = item.split(":")
                                current_counts[k.strip()] = int(v)
                        
                        c_vote, c_res = st.columns([1, 1])
                        
                        with c_vote:
                            current_voter = st.selectbox("참여자명 (투표 확인용 / 익명처리)", ["선택 안함"] + emp_names, key=f"dash_voter_{poll_id}")
                            
                            if current_voter == "선택 안함":
                                st.info("상단에서 본인 이름을 선택해야 투표할 수 있습니다.")
                            elif current_voter in voters:
                                st.success("✅ 이미 참여하셨습니다.")
                            else:
                                with st.form(f"dash_vote_form_{poll_id}"):
                                    selected_opt = st.radio("항목을 선택하세요", options)
                                    if st.form_submit_button("투표하기"):
                                        if selected_opt in current_counts:
                                            current_counts[selected_opt] += 1
                                        else:
                                            current_counts[selected_opt] = 1
                                            
                                        voters.append(current_voter)
                                        
                                        new_counts_str = ",".join([f"{k}:{v}" for k, v in current_counts.items()])
                                        new_voters_str = ",".join(voters)
                                        
                                        temp_polls = load_polls()
                                        temp_polls.loc[temp_polls['ID'] == poll_id, '투표현황'] = new_counts_str
                                        temp_polls.loc[temp_polls['ID'] == poll_id, '참여자'] = new_voters_str
                                        temp_polls.to_csv(POLL_BOARD_FILE, index=False, encoding='utf-8-sig')
                                        st.rerun()
                                        
                        with c_res:
                            st.markdown("**투표 현황**")
                            total_votes = sum(current_counts.values())
                            
                            if total_votes == 0:
                                st.caption("아직 참여한 인원이 없습니다.")
                            else:
                                for opt, count in current_counts.items():
                                    pct = (count / total_votes) * 100 if total_votes > 0 else 0
                                    st.markdown(f"{opt} ({count}표)")
                                    st.progress(int(pct))
                    
        st.markdown('<div class="section-title">📦 생산제품 내역</div>', unsafe_allow_html=True)
        
        unique_dates = sorted(df['생산일'].dropna().unique(), reverse=True) if not df.empty else []
        
        if unique_dates:
            selected_dash_date = st.selectbox("조회할 생산일 선택", unique_dates)
            st.markdown(f"**{selected_dash_date} 생산제품 내역**")
            
            dash_filtered = df[df['생산일'] == selected_dash_date].copy()
            if not dash_filtered.empty:
                
                # 👇 [🔥연동 포인트] 제품 규격(Specs) 마스터 데이터를 불러와서 매핑 딕셔너리 생성
                df_specs_dash = load_specs()
                if not df_specs_dash.empty and '제품명' in df_specs_dash.columns:
                    unit_mapping = df_specs_dash.set_index('제품명')['규격'].to_dict()
                else:
                    unit_mapping = {}
                
                dash_filtered['유형'] = dash_filtered['유형'].fillna("-").astype(str)
                dash_filtered['제품명'] = dash_filtered['제품명'].fillna("-").astype(str)
                
                # 👇 대시보드에 표시할 때 제품명 기준으로 최신 '규격'을 찾아 덮어쓰기! (없으면 기존 데이터 유지)
                dash_filtered['규격'] = dash_filtered['제품명'].map(unit_mapping).fillna(dash_filtered['규격']).fillna("-").astype(str)
                dash_filtered['생산량'] = pd.to_numeric(dash_filtered['생산량'], errors='coerce').fillna(0)
                
                # 제품별 생산량 합계 계산
                prod_summary = dash_filtered.groupby(['유형', '제품명', '규격'])['생산량'].sum().reset_index()
                
                # 대시보드에는 생산량이 0 초과인(실제 생산한) 제품만 깔끔하게 보여주기
                prod_summary = prod_summary[prod_summary['생산량'] > 0]
                
                if prod_summary.empty:
                    st.info(f"해당 일자에 입력된 생산량이 없습니다. (데이터 히스토리 메뉴에서 생산량을 입력해 주세요)")
                else:
                    st.dataframe(prod_summary, use_container_width=True, hide_index=True)
            else:
                st.info("해당 일자에 기록된 생산 내역이 없습니다.")
        else:
            st.info("데이터 히스토리에 저장된 생산 기록이 없습니다.")
            
        st.markdown('<div class="section-title">📦 전체 실시간 재고 요약 (대시보드)</div>', unsafe_allow_html=True)
        # 위에서 사용된 `df` (전체 QC Data) 변수를 그대로 활용하거나 로드
        df_out_dash = load_outbound_records()
        df_adj_dash = load_inventory_adj()
        
        # 전체 재고 집계 로직
        inv_dict_dash = {}
        for _, row in df.iterrows():
            if str(row.get("제품명", "")).strip() in ["", "-", "None", "nan"]: continue
            k = (str(row.get("유형", "None")).strip(), str(row.get("제품명", "")).strip(), str(row.get("규격", "None")).strip())
            if k not in inv_dict_dash: inv_dict_dash[k] = {"총생산량": 0, "총출고량": 0, "조정수량": 0}
            try: qty = int(float(str(row.get("생산량", "0")).replace(",","")))
            except: qty = 0
            inv_dict_dash[k]["총생산량"] += qty
            
        for _, row in df_out_dash.iterrows():
            if str(row.get("제품명", "")).strip() in ["", "-", "None", "nan"]: continue
            k = (str(row.get("유형", "None")).strip(), str(row.get("제품명", "")).strip(), str(row.get("규격", "None")).strip())
            if k not in inv_dict_dash: inv_dict_dash[k] = {"총생산량": 0, "총출고량": 0, "조정수량": 0}
            try: qty = int(float(str(row.get("수량", "0")).replace(",","")))
            except: qty = 0
            inv_dict_dash[k]["총출고량"] += qty
            
        for _, row in df_adj_dash.iterrows():
            k = (str(row.get("유형", "None")).strip(), str(row.get("제품명", "")).strip(), str(row.get("규격", "None")).strip())
            if k not in inv_dict_dash: inv_dict_dash[k] = {"총생산량": 0, "총출고량": 0, "조정수량": 0}
            try: diff = int(float(str(row.get("차이", "0")).replace(",","")))
            except: diff = 0
            inv_dict_dash[k]["조정수량"] += diff
            
        inv_list_dash = []
        for k, v in inv_dict_dash.items():
            curr_qty = v["총생산량"] - v["총출고량"] + v["조정수량"]
            # 대시보드에서는 재고가 0이 아닌 유의미한 품목만 노출 (또는 전체 노출)
            if curr_qty > 0 or v["총생산량"] > 0:
                inv_list_dash.append({
                    "제품명": f"[{k[0]}] {k[1]}",
                    "규격": k[2] if k[2] not in ["", "nan"] else "-",
                    "현재고": curr_qty
                })
        
        if not inv_list_dash:
            st.info("현재 파악된 품목의 전산 재고가 없습니다.")
        else:
            df_dash_inv = pd.DataFrame(inv_list_dash).sort_values(by="제품명").reset_index(drop=True)
            st.dataframe(
                df_dash_inv,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "현재고": st.column_config.NumberColumn(format="%d 개")
                }
            )

    with c_right:
        st.markdown('<div class="section-title">다가오는 캘린더 일정 (D-Day 연동)</div>', unsafe_allow_html=True)
        
        df_v = load_verify()
        df_o = load_other_sched()
        df_hc = load_health_cert()
        df_calib_c = load_calib_list()
        df_filter = load_filter_plan()
        
        agenda_list = []
        today_date = date.today()
        
        if not df_v.empty and '계획일자' in df_v.columns:
            for idx, row in df_v.iterrows():
                manager_str = f" / 담당자: {row.get('담당자', '')}" if pd.notna(row.get('담당자')) and str(row.get('담당자', '')).strip() != "" else ""
                agenda_list.append({
                    'Date': pd.to_datetime(row['계획일자'], errors='coerce'),
                    'Title': f"[검증] {row.get('검증항목', '')}",
                    'Desc': f"{row.get('세부내용', '')}{manager_str}",
                    'Status': row.get('상태', '예정'),
                    'Source': VERIFY_FILE,
                    'Idx': idx
                })
                
        if not df_o.empty and '일자' in df_o.columns:
            for idx, row in df_o.iterrows():
                manager_str = f" / 담당자: {row.get('담당자', '')}" if pd.notna(row.get('담당자')) and str(row.get('담당자', '')).strip() != "" else ""
                agenda_list.append({
                    'Date': pd.to_datetime(row['일자'], errors='coerce'),
                    'Title': f"[기타] {row.get('일정명', '')}",
                    'Desc': f"{row.get('세부내용', '')}{manager_str}",
                    'Status': row.get('상태', '예정'),
                    'Source': OTHER_SCHED_FILE,
                    'Idx': idx
                })
                
        if not df_hc.empty and '검진일자' in df_hc.columns:
            for idx, row in df_hc.iterrows():
                if pd.isna(row['검진일자']): continue
                exam_date = pd.to_datetime(row['검진일자'], errors='coerce')
                if pd.isnull(exam_date): continue
                
                exp_date = exam_date + pd.DateOffset(years=1) - pd.Timedelta(days=1)
                days_left = (exp_date.date() - today_date).days
                
                if days_left <= 30:
                    hc_status = "만료" if days_left < 0 else f"D-{days_left}"
                    alert_prefix = "[보건증 만료]" if days_left < 0 else "[보건증 갱신]"
                    
                    agenda_list.append({
                        'Date': exp_date,
                        'Title': f"🚨 {alert_prefix} {row.get('이름', '')} ({row.get('직급', '')})",
                        'Desc': f"연락처: {row.get('연락처', '')} / 만기일: {exp_date.strftime('%Y-%m-%d')}",
                        'Status': hc_status,
                        'Source': 'HEALTH',
                        'Idx': idx
                    })
                    
        if not df_calib_c.empty and '차기_검교정일자' in df_calib_c.columns:
            for idx, row in df_calib_c.iterrows():
                if pd.isna(row['차기_검교정일자']) or str(row['차기_검교정일자']).strip() == "": continue
                try:
                    calib_date = pd.to_datetime(row['차기_검교정일자']).date()
                except: continue
                
                days_left = (calib_date - today_date).days
                
                if days_left <= 30:
                    c_status = "만료" if days_left < 0 else f"D-{days_left}"
                    c_alert = "[검교정 만료]" if days_left < 0 else "[검교정 도래]"
                    
                    agenda_list.append({
                        'Date': pd.to_datetime(calib_date),
                        'Title': f"⚖️ {c_alert} {row.get('검사_설비명', '')} ({row.get('관리번호', '')})",
                        'Desc': f"구분: {row.get('구분', '')} / 주기: {row.get('주기', '')}개월 / 만기일: {calib_date.strftime('%Y-%m-%d')}",
                        'Status': c_status,
                        'Source': 'CALIB',
                        'Idx': idx
                    })

        if not df_filter.empty and '차기점검일자' in df_filter.columns:
            for idx, row in df_filter.iterrows():
                if pd.isna(row['차기점검일자']) or str(row['차기점검일자']).strip() == "": continue
                try:
                    f_date = pd.to_datetime(row['차기점검일자']).date()
                except: continue
                
                days_left = (f_date - today_date).days
                
                if days_left <= 30:
                    f_status = "만료" if days_left < 0 else f"D-{days_left}"
                    f_alert = "[필터 만료]" if days_left < 0 else "[필터 점검]"
                    
                    agenda_list.append({
                        'Date': pd.to_datetime(f_date),
                        'Title': f"🚰 {f_alert} {row.get('설치장소', '')} ({row.get('필터명', '')}) [{row.get('내용', '')}]",
                        'Desc': f"주기: {row.get('주기_개월', '')}개월 / 만기일: {f_date.strftime('%Y-%m-%d')}",
                        'Status': f_status,
                        'Source': 'FILTER',
                        'Idx': idx
                    })
        
        df_agenda = pd.DataFrame(agenda_list, columns=['Date', 'Title', 'Desc', 'Status', 'Source', 'Idx'])
        df_agenda = df_agenda.dropna(subset=['Date']).sort_values('Date')
        
        if df_agenda.empty:
            st.info("임박하거나 등록된 캘린더 일정이 없습니다. 좌측 메뉴에서 항목을 추가해 주세요.")
        else:
            for _, row in df_agenda.iterrows():
                d_val = row['Date'].date()
                status = row['Status']
                source = row['Source']
                
                if status == "완료":
                    disp_status = "완료"; badge_class = "badge-done"
                elif source == 'HEALTH':
                    disp_status = status
                    badge_class = "badge-overdue" if status == "만료" else "badge-warning"
                elif source == 'CALIB' or source == 'FILTER':
                    disp_status = status
                    badge_class = "badge-overdue" if "만료" in status else "badge-warning"
                else:
                    if d_val < today_date: disp_status = "미완료"; badge_class = "badge-overdue"
                    else: disp_status = "예정"; badge_class = "badge-upcoming"
                        
                d_str = row['Date'].strftime("%m/%d")
                
                with st.container(border=True):
                    c1, c2 = st.columns([4, 1.2])
                    with c1:
                        text_style = "text-decoration: line-through; color: #A0A0A0;" if disp_status == "완료" else ""
                        date_color = "#A0A0A0" if disp_status == "완료" else "#D11031"
                        st.markdown(f"<span style='font-size:1.2rem; font-weight:800; color:{date_color}; margin-right:15px;'>{d_str}</span> <span style='font-size:1.1rem; font-weight:700; {text_style}'>{row['Title']}</span> &nbsp;<span class='badge {badge_class}'>{disp_status}</span>", unsafe_allow_html=True)
                        st.markdown(f"<span style='color:#6C757D; font-size:0.95rem; margin-left:62px; {text_style}'>{row['Desc']}</span>", unsafe_allow_html=True)
                    with c2:
                        st.markdown("<div style='margin-top: 5px;'></div>", unsafe_allow_html=True)
                        if source not in ['HEALTH', 'CALIB', 'FILTER']:
                            btn_label = "⏪ 완료 취소" if disp_status == "완료" else "✅ 완료 처리"
                            if st.button(btn_label, key=f"btn_{source}_{row['Idx']}", use_container_width=True):
                                toggle_task_status(source, row['Idx'])
                                st.rerun()
                        elif source == 'HEALTH':
                            st.caption("보건증 메뉴에서 갱신")
                        elif source == 'CALIB':
                            st.caption("계측기기 메뉴에서 갱신")
                        elif source == 'FILTER':
                            st.caption("설비관리 > 필터 점검관리에서 점검일자 갱신 요망")

# --- 게시판 메뉴 ---
elif menu_selection == "게시판":
    st.markdown('<div class="section-title">📢 사내 게시판 및 투표</div>', unsafe_allow_html=True)
    
    tab_notice, tab_poll = st.tabs(["게시판 (공지사항)", "익명 투표 개설"])
    
    with tab_notice:
        st.write("자유롭게 의견을 나누고 공지사항을 등록할 수 있습니다. **'중요공지'** 체크 시 작성일로부터 7일간 대시보드 메인 화면에 고정 노출됩니다.")
        
        df_notices = load_notices()
        
        with st.expander("➕ 새 게시글 작성하기"):
            with st.form("notice_form"):
                n_col1, n_col2 = st.columns([3, 1])
                with n_col1:
                    n_title = st.text_input("제목")
                with n_col2:
                    n_author = st.text_input("작성자")
                    
                n_content = st.text_area("내용")
                n_important = st.checkbox("🚨 중요 공지로 등록 (대시보드 7일 노출)")
                
                if st.form_submit_button("게시글 등록"):
                    if n_title and n_author and n_content:
                        new_notice = pd.DataFrame({
                            "작성일": [str(date.today())],
                            "작성자": [n_author],
                            "제목": [n_title],
                            "내용": [n_content],
                            "중요공지": [n_important]
                        })
                        df_notices = pd.concat([df_notices, new_notice], ignore_index=True)
                        df_notices.to_csv(NOTICE_BOARD_FILE, index=False, encoding='utf-8-sig')
                        log_history("게시글 등록", "사내 게시판", f"제목: {n_title}")
                        st.success("게시글이 성공적으로 등록되었습니다!")
                        st.rerun()
                    else:
                        st.error("제목, 작성자, 내용을 모두 입력해 주세요.")
                        
        st.markdown("### 📋 전체 게시글 목록")
        if df_notices.empty:
            st.info("등록된 게시글이 없습니다.")
        else:
            display_notices = df_notices.sort_values(by="작성일", ascending=False).reset_index(drop=True)
            display_notices['작성일'] = pd.to_datetime(display_notices['작성일'], errors='coerce').apply(lambda x: x.date() if pd.notna(x) else None)
            
            cfg_notice = {
                "작성일": st.column_config.DateColumn("작성일", format="YYYY-MM-DD"),
                "중요공지": st.column_config.CheckboxColumn("중요공지 (체크)"),
                "내용": st.column_config.TextColumn("내용")
            }
            
            st.caption("표에서 데이터를 직접 더블클릭하여 내용을 수정하거나 삭제할 수 있습니다.")
            edited_notices = st.data_editor(display_notices, num_rows="dynamic", use_container_width=True, column_config=cfg_notice)
            
            if st.button("게시판 변경사항 저장"):
                edited_notices['작성일'] = edited_notices['작성일'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) else "")
                edited_notices.to_csv(NOTICE_BOARD_FILE, index=False, encoding='utf-8-sig')
                log_history("게시판 일괄 수정", "사내 게시판", "데이터 에디터를 통한 수정")
                st.success("게시판 내용이 안전하게 업데이트 되었습니다.")
                st.rerun()

    with tab_poll:
        st.write("의견 수렴을 위한 익명 투표를 개설할 수 있습니다. 등록된 투표는 대시보드 메인 화면에서 참여 및 결과를 확인할 수 있습니다.")
        
        df_polls = load_polls()
        df_emp = load_employees()
        active_emps = df_emp[df_emp["재직상태"] == "재직"] if not df_emp.empty and "재직상태" in df_emp.columns else pd.DataFrame()
        
        if active_emps.empty:
            st.warning("등록된 재직 직원이 없어 투표 기능을 사용할 수 없습니다. 직원 관리 메뉴에서 인원을 등록해주세요.")
        else:
            with st.expander("➕ 새 투표 만들기", expanded=True):
                with st.form("poll_create_form"):
                    p_title = st.text_input("투표 주제 (질문)")
                    p_opts = st.text_input("선택지 (쉼표 ',' 로 구분하여 입력. 예: 찬성, 반대, 기권)")
                    
                    if st.form_submit_button("투표 등록"):
                        if p_title and p_opts:
                            new_id = f"P-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                            opts_list = [o.strip() for o in p_opts.split(",") if o.strip()]
                            initial_counts = ",".join([f"{o}:0" for o in opts_list])
                            
                            new_p = pd.DataFrame([[new_id, str(date.today()), p_title, p_opts, initial_counts, ""]], columns=df_polls.columns)
                            df_polls = pd.concat([df_polls, new_p], ignore_index=True)
                            df_polls.to_csv(POLL_BOARD_FILE, index=False, encoding='utf-8-sig')
                            log_history("투표 등록", "사내 투표", f"주제: {p_title}")
                            st.success("투표가 등록되었습니다! 메인 대시보드에서 투표에 참여할 수 있습니다.")
                            st.rerun()
                        else:
                            st.error("주제와 선택지를 모두 입력하세요.")
                            
        if not df_polls.empty:
            st.divider()
            st.markdown("### 🗑️ 등록된 투표 관리 (삭제 전용)")
            for idx, row in df_polls.sort_values(by="작성일", ascending=False).iterrows():
                with st.container(border=True):
                    c1, c2 = st.columns([5, 1])
                    with c1:
                        st.markdown(f"**[{row['작성일']}] {row['제목']}**")
                    with c2:
                        if st.button("투표 삭제", key=f"del_poll_mgr_{row['ID']}"):
                            df_polls = df_polls[df_polls['ID'] != row['ID']]
                            df_polls.to_csv(POLL_BOARD_FILE, index=False, encoding='utf-8-sig')
                            st.rerun()

# --- 2. 캘린더 (하위 메뉴들) ---
elif menu_selection == "캘린더":
    
    if sub_menu == "📅 달력 보기":
        st.markdown('<div class="section-title">월간 달력</div>', unsafe_allow_html=True)

        # ── 연/월 선택 ──────────────────────────────────────────
        today = date.today()
        col_y, col_m, col_spacer = st.columns([1, 1, 3])
        with col_y:
            sel_year = st.selectbox("연도", list(range(today.year - 2, today.year + 3)), index=2, key="cal_year")
        with col_m:
            sel_month = st.selectbox("월", list(range(1, 13)), index=today.month - 1, key="cal_month")

        # ── 데이터 로드 ─────────────────────────────────────────
        df_v = load_verify()
        df_o = load_other_sched()

        # ── 날짜별 이벤트 맵 만들기 ─────────────────────────────
        # 검증계획 → {날짜: [(라벨, 색상, source)]}
        ev_map = {}
        for _, r in df_v.iterrows():
            try:
                d = pd.to_datetime(str(r["계획일자"])).date()
                label = str(r.get("검증항목", "검증"))[:14]
                status = str(r.get("상태", "예정"))
                color = "#27AE60" if status == "완료" else "#4A90D9"
                ev_map.setdefault(d, []).append((f"✅ {label}", color, "verify"))
            except Exception: pass

        for _, r in df_o.iterrows():
            try:
                d = pd.to_datetime(str(r["일자"])).date()
                label = str(r.get("일정명", "일정"))[:14]
                status = str(r.get("상태", "예정"))
                color = "#27AE60" if status == "완료" else "#F39C12"
                ev_map.setdefault(d, []).append((f"🗓 {label}", color, "other"))
            except Exception: pass

        # ── 달력 HTML 생성 ──────────────────────────────────────
        import calendar as cal_mod
        cal_mod.setfirstweekday(6)  # 일요일 시작
        month_cal = cal_mod.monthcalendar(sel_year, sel_month)
        days_header = ["일", "월", "화", "수", "목", "금", "토"]

        html_parts = ["""
        <style>
        .qms-cal-wrap { font-family: 'Malgun Gothic', sans-serif; border: 1px solid #ddd; border-radius:8px; overflow:hidden; }
        .qms-cal-head { display:grid; grid-template-columns: repeat(7,1fr); background:#2C3E50; color:white; }
        .qms-cal-head div { text-align:center; padding:12px 0; font-weight:bold; font-size:16px; }
        .qms-cal-head div:first-child { color:#FF6B6B; }
        .qms-cal-head div:last-child  { color:#74B9FF; }
        .qms-cal-body { display:grid; grid-template-columns: repeat(7,1fr); }
        .qms-cell { border:1px solid #eee; min-height:130px; padding:6px; background:white; vertical-align:top; }
        .qms-cell.other-month { background:#f8f8f8; }
        .qms-cell.today { background:#FFF9E6; border:2px solid #F39C12; }
        .qms-day-num { font-size:17px; font-weight:bold; margin-bottom:5px; color:#333; }
        .qms-cell:nth-child(7n+1) .qms-day-num { color:#e74c3c; }
        .qms-cell:nth-child(7n)   .qms-day-num { color:#3498db; }
        .qms-ev { font-size:12px; border-radius:4px; padding:2px 6px; margin:2px 0;
                  white-space:nowrap; overflow:hidden; text-overflow:ellipsis;
                  max-width:100%; display:block; color:white; }
        </style>
        <div class="qms-cal-wrap">
          <div class="qms-cal-head">
        """]
        for d in days_header:
            html_parts.append(f'<div>{d}</div>')
        html_parts.append('</div><div class="qms-cal-body">')

        for week in month_cal:
            for wd_idx, day in enumerate(week):
                if day == 0:
                    html_parts.append('<div class="qms-cell other-month"></div>')
                else:
                    cell_date = date(sel_year, sel_month, day)
                    is_today  = (cell_date == today)
                    cls = "qms-cell today" if is_today else "qms-cell"
                    html_parts.append(f'<div class="{cls}">')
                    html_parts.append(f'<div class="qms-day-num">{day}</div>')
                    for (ev_label, ev_color, _) in ev_map.get(cell_date, []):
                        html_parts.append(
                            f'<span class="qms-ev" style="background:{ev_color}" '
                            f'title="{ev_label}">{ev_label}</span>'
                        )
                    html_parts.append('</div>')

        html_parts.append('</div></div>')
        st.markdown("".join(html_parts), unsafe_allow_html=True)

        # ── 달력에서 직접 일정 등록 (기타 일정으로 저장) ──────────
        st.divider()
        st.markdown("##### ➕ 달력에서 일정 등록 (기타 일정으로 저장됩니다)")
        with st.form("cal_quick_add"):
            qa1, qa2 = st.columns(2)
            with qa1:
                qa_date    = st.date_input("날짜", value=today, key="qa_date")
                qa_title   = st.text_input("일정명", placeholder="예: 식약처 현장점검")
            with qa2:
                qa_desc    = st.text_input("세부 내용", placeholder="선택 입력")
                qa_manager = st.text_input("담당자", placeholder="선택 입력")
                qa_status  = st.selectbox("상태", ["예정", "완료"], key="qa_status")

            if st.form_submit_button("📅 달력에 등록", use_container_width=True):
                if qa_title:
                    df_o2 = load_other_sched()
                    new_row = pd.DataFrame([[str(qa_date), qa_title, qa_desc, qa_status, qa_manager]],
                                           columns=df_o2.columns)
                    df_o2 = pd.concat([df_o2, new_row], ignore_index=True)
                    df_o2.to_csv(OTHER_SCHED_FILE, index=False, encoding='utf-8-sig')
                    st.success(f"✅ '{qa_title}' 일정이 등록되었습니다! 달력에 바로 반영됩니다.")
                    st.rerun()
                else:
                    st.error("일정명을 입력해 주세요.")

        # ── 당월 일정 목록 (빠른 확인) ─────────────────────────
        st.divider()
        st.markdown(f"##### 📋 {sel_year}년 {sel_month}월 일정 목록")
        month_events = []
        for d, evs in sorted(ev_map.items()):
            if d.year == sel_year and d.month == sel_month:
                for (lbl, color, src) in evs:
                    src_label = "검증계획" if src == "verify" else "기타일정"
                    month_events.append({"날짜": str(d), "구분": src_label, "내용": lbl.lstrip("✅🗓 ")})
        if month_events:
            st.dataframe(pd.DataFrame(month_events), use_container_width=True, hide_index=True)
        else:
            st.info("이 달에 등록된 일정이 없습니다.")

    elif sub_menu == "검증 계획표":
        st.markdown('<div class="section-title">연간 검증 계획표 관리</div>', unsafe_allow_html=True)
        df_v = load_verify()
        
        with st.expander("➕ 새 검증 계획 등록"):
            with st.form("form_verify"):
                v1, v2 = st.columns(2)
                with v1:
                    v_date = st.date_input("계획 일자")
                    v_type = st.selectbox("검증 종류", ["일상검증", "정기검증", "내부검증", "외부검증", "특별검증"])
                    v_item = st.text_input("검증 항목 (예: 자가품질검사)")
                with v2:
                    v_desc = st.text_input("세부 내용")
                    v_method = st.selectbox("검증 방법", ["기록확인", "현장조사", "시험∙검사", "평가서", "기타"])
                    v_manager = st.text_input("담당자")
                    v_status = st.selectbox("초기 상태", ["예정", "완료"])
                    
                if st.form_submit_button("일정 등록하기"):
                    if v_item:
                        new_df = pd.DataFrame([[str(v_date), v_type, v_item, v_desc, v_method, v_status, v_manager]], columns=df_v.columns)
                        df_v = pd.concat([df_v, new_df], ignore_index=True)
                        df_v.to_csv(VERIFY_FILE, index=False, encoding='utf-8-sig')
                        st.success("캘린더 연동 완료!")
                        st.rerun()
                    else: st.error("항목을 입력하세요.")
        
        cfg_v = {"상태": st.column_config.SelectboxColumn("상태", options=["예정", "완료"])}
        edited_v = st.data_editor(df_v, num_rows="dynamic", use_container_width=True, column_config=cfg_v)
        if st.button("변경사항 서버에 저장", key="save_v"):
            edited_v.to_csv(VERIFY_FILE, index=False, encoding='utf-8-sig')
            st.rerun()

    elif sub_menu == "기타 일정":
        st.markdown('<div class="section-title">🗓️ 기타 일정 관리</div>', unsafe_allow_html=True)
        df_o = load_other_sched()
        
        with st.expander("➕ 새 일정 등록"):
            with st.form("form_other"):
                o1, o2 = st.columns(2)
                with o1:
                    o_date = st.date_input("일자")
                    o_title = st.text_input("일정명 (예: 식약처 심사)")
                with o2:
                    o_desc = st.text_input("세부 내용")
                    o_manager = st.text_input("담당자")
                    o_status = st.selectbox("상태", ["예정", "완료"])
                
                if st.form_submit_button("일정 등록하기"):
                    if o_title:
                        new_df = pd.DataFrame([[str(o_date), o_title, o_desc, o_status, o_manager]], columns=df_o.columns)
                        df_o = pd.concat([df_o, new_df], ignore_index=True)
                        df_o.to_csv(OTHER_SCHED_FILE, index=False, encoding='utf-8-sig')
                        st.success("일정 등록 완료!")
                        st.rerun()
                    else: st.error("일정명을 입력하세요.")
                        
        cfg_o = {"상태": st.column_config.SelectboxColumn("상태", options=["예정", "완료"])}
        edited_o = st.data_editor(df_o, num_rows="dynamic", use_container_width=True, column_config=cfg_o)
        if st.button("변경사항 서버에 저장", key="save_o"):
            edited_o.to_csv(OTHER_SCHED_FILE, index=False, encoding='utf-8-sig')
            st.rerun()

# --- 제품 관리 메뉴 (현장 측정 기록, 데이터 히스토리 통합) ---
elif menu_selection == "제품 관리":
    # ── 공통 스타일 ──
    TH_P  = "background:#1F4E79;color:#fff;font-weight:bold;padding:9px 8px;text-align:center;border:1px solid #1a3d60;font-size:11px;"
    TH_P2 = "background:#D6E4F0;color:#1F4E79;font-weight:bold;padding:8px 8px;text-align:center;border:1px solid #b0c4d8;font-size:11px;"
    TD_P  = "background:#fff;padding:8px 8px;border:1px solid #d0d7de;vertical-align:middle;font-size:11px;text-align:center;word-break:break-word;"
    TDG_P = "background:#EBF3FB;padding:8px 8px;border:1px solid #d0d7de;vertical-align:middle;font-size:11px;font-weight:bold;color:#1F4E79;"
    TYPE_COLOR = {"캡슐커피":"#D11031","스틱커피":"#2E75B6","원두커피":"#7B5EA7","생두(원료)":"#217346"}

    # ════════════════════════════════════════════════════════
    # 1. 간편 판정 규격
    # ════════════════════════════════════════════════════════
    if sub_menu == "제품 등록":
        st.markdown('<div class="section-title">📦 제품 등록</div>', unsafe_allow_html=True)
        st.write("생산 기록 등록 시 합불 판정의 기준이 되는 제품 규격을 관리합니다.")
        df_specs = load_specs()

        with st.expander("➕ 새 제품 등록하기", expanded=False):
            with st.form("spec_form"):
                s_c1, s_c2 = st.columns([1, 1.5])
                with s_c1:
                    new_p_code = st.text_input("제품코드 (예: HL-001)")
                    new_p_name = st.text_input("제품명 (예: 시그니처 캡슐)")
                    new_p_type = st.selectbox("유형", ["캡슐커피", "스틱커피", "원두커피", "생두(원료)"])
                    new_p_spec = st.text_input("규격 (여러 개 입력 시 콤마로 구분)", placeholder="예: 100g, 200g, 1kg")
                    date_type  = st.selectbox("날짜 관리", ["해당없음", "소비기한", "제조일자"])
                with s_c2:
                    analysis_items = st.multiselect("📌 분석 항목 설정",
                        ["질소", "수분", "색도", "추출시간"], default=["질소", "수분", "색도", "추출시간"])
                    st.caption("선택된 항목의 최소/최대 규격만 입력 창이 활성화됩니다.")
                    c_a1, c_a2 = st.columns(2)
                    with c_a1:
                        min_n2  = st.number_input("질소 최소 (%)",     value=97.5) if "질소"   in analysis_items else None
                        max_n2  = st.number_input("질소 최대 (%)",     value=99.5) if "질소"   in analysis_items else None
                        min_moi = st.number_input("수분 최소 (%)",     value=0.0)  if "수분"   in analysis_items else None
                        max_moi = st.number_input("수분 최대 (%)",     value=5.0)  if "수분"   in analysis_items else None
                    with c_a2:
                        min_col = st.number_input("색도 최소",         value=50.0) if "색도"   in analysis_items else None
                        max_col = st.number_input("색도 최대",         value=60.0) if "색도"   in analysis_items else None
                        min_ext = st.number_input("추출시간 최소 (초)", value=22.0) if "추출시간" in analysis_items else None
                        max_ext = st.number_input("추출시간 최대 (초)", value=28.0) if "추출시간" in analysis_items else None
                if st.form_submit_button("제품 등록"):
                    if new_p_code and new_p_name and new_p_spec:
                        # 규격을 쉼표로 구분된 하나의 문자열로 저장
                        clean_specs = ", ".join([s.strip() for s in new_p_spec.split(",") if s.strip()])
                        
                        new_row = pd.DataFrame([[
                            new_p_code, new_p_name, new_p_type, clean_specs,
                            min_n2  if min_n2  is not None else "N/A",
                            max_n2  if max_n2  is not None else "N/A",
                            min_moi if min_moi is not None else "N/A",
                            max_moi if max_moi is not None else "N/A",
                            min_col if min_col is not None else "N/A",
                            max_col if max_col is not None else "N/A",
                            min_ext if min_ext is not None else "N/A",
                            max_ext if max_ext is not None else "N/A",
                            date_type
                        ]], columns=df_specs.columns)
                        
                        # 동일 코드가 이미 있다면 업데이트, 없으면 추가
                        if new_p_code in df_specs["제품코드"].values:
                            df_specs = df_specs[df_specs["제품코드"] != new_p_code]
                        
                        df_specs = pd.concat([df_specs, new_row], ignore_index=True)
                        df_specs.to_csv(SPEC_FILE, index=False, encoding='utf-8-sig')
                        log_history("제품 신규 등록", "제품 관리", f"제품코드: {new_p_code}, 제품명: {new_p_name}, 규격: {clean_specs}")
                        st.success(f"[{new_p_code}] {new_p_name} 등록 완료!"); st.rerun()
                    else:
                        st.error("제품코드, 제품명, 규격을 모두 입력해 주세요.")

        st.write("")
        if df_specs.empty:
            st.info("등록된 제품이 없습니다.")
        else:
            # 표시용 그룹화 (이미 데이터상에 분리된 경우가 있을 수 있으므로 코드 기준 병합)
            df_display = df_specs.drop_duplicates(subset=["제품코드"]).copy()
            
            rows_html = ""
            for _, row in df_display.iterrows():
                p_code = row.get("제품코드","")
                # 같은 코드의 원본 행들에서 규격만 추출하여 합침
                all_specs = df_specs[df_specs["제품코드"] == p_code]["규격"].tolist()
                combined_specs = []
                for s in all_specs:
                    combined_specs.extend([item.strip() for item in str(s).split(",") if item.strip()])
                combined_specs = sorted(list(set(combined_specs))) # 중복 제거 및 정렬
                
                spec_count = len(combined_specs)
                if spec_count > 1:
                    spec_cell = (
                        f'<details style="cursor:pointer; font-size:11px;">'
                        f'<summary style="color:#1A5FAD; font-weight:600;">{combined_specs[0]} 외 {spec_count-1}건</summary>'
                        f'<ul style="margin:5px 0 0 15px; padding:0; list-style-type:circle; color:#666;">'
                        + "".join([f"<li>{s}</li>" for s in combined_specs]) +
                        f'</ul></details>'
                    )
                else:
                    spec_cell = combined_specs[0] if combined_specs else "-"

                tc = TYPE_COLOR.get(str(row.get("유형","")), "#555")
                def rng(mn_k, mx_k, r=row):
                    mv, xv = str(r.get(mn_k,"")), str(r.get(mx_k,""))
                    if mv in ("","N/A","nan") and xv in ("","N/A","nan"): return '<span style="color:#bbb;">-</span>'
                    return f'{mv} ~ {xv}'
                
                rows_html += (
                    f'<tr>'
                    f'<td style="{TD_P} font-weight:bold;color:#D11031;">{p_code}</td>'
                    f'<td style="{TD_P}text-align:left;font-weight:bold;">{row.get("제품명","")}</td>'
                    f'<td style="{TD_P}"><span style="background:{tc}22;color:{tc};padding:2px 8px;border-radius:10px;font-size:10px;font-weight:bold;">{row.get("유형","")}</span></td>'
                    f'<td style="{TD_P}">{spec_cell}</td>'
                    f'<td style="{TD_P}">{rng("최소_질소","최대_질소")}</td>'
                    f'<td style="{TD_P}">{rng("최소_수분","최대_수분")}</td>'
                    f'<td style="{TD_P}">{rng("최소_색도","최대_색도")}</td>'
                    f'<td style="{TD_P}">{rng("최소_추출","최대_추출")}</td>'
                    f'<td style="{TD_P}">{row.get("날짜유형","")}</td>'
                    f'</tr>'
                )
            head_cells = ""
            for h, w in [("코드","70px"),("제품명","160px"),("유형","80px"),("규격","50px"),
                         ("질소(%)","90px"),("수분(%)","90px"),("색도","90px"),("추출시간(초)","100px"),("날짜관리","70px")]:
                head_cells += f'<th style="{TH_P}width:{w};">{h}</th>'
            html_sp = (
                f'<style>.sp-tbl{{width:100%;border-collapse:collapse;}}'
                f'.sp-tbl tr:hover td{{background:#f5f8ff!important;}}</style>'
                f'<table class="sp-tbl"><thead><tr>{head_cells}</tr></thead><tbody>{rows_html}</tbody></table>'
            )
            st.markdown(html_sp, unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            with st.expander("✏️ 수정 / 🗑️ 삭제"):
                edited_specs = st.data_editor(df_specs, num_rows="dynamic", use_container_width=True)
                if st.button("제품 규격 저장", type="primary"):
                    edited_specs.to_csv(SPEC_FILE, index=False, encoding='utf-8-sig')
                    log_history("제품 규격 일괄 수정", "제품 관리", "데이터 에디터를 통한 수정")
                    st.success("제품 규격이 안전하게 저장되었습니다."); st.rerun()

    # ════════════════════════════════════════════════════════
    # 2. 상세 규격서 마스터
    # ════════════════════════════════════════════════════════
    elif sub_menu == "제품 규격 마스터":

        def export_full_excel(sel_code, sel_name, df_b, df_r, df_h, img_path):
            import openpyxl as _opx
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter as gcl
            wb  = _opx.Workbook(); ws = wb.active; ws.title = "제품규격서"
            ws.page_setup.paperSize = 9; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToPage = True
            ws.print_options.horizontalCentered = True
            ws.page_margins.left = 0.5; ws.page_margins.right  = 0.5
            ws.page_margins.top  = 0.8; ws.page_margins.bottom = 0.8
            thin = Side(style="thin", color="B0C4D8")
            med  = Side(style="medium", color="1F4E79")
            bd   = Border(left=thin, right=thin, top=thin, bottom=thin)
            bdm  = Border(left=med,  right=med,  top=med,  bottom=med)
            hfill = PatternFill("solid", start_color="D6E4F0")   # 옅은 파란 (항목 헤더)
            sfill = PatternFill("solid", start_color="D6E4F0")   # 섹션헤더도 동일
            tfill = PatternFill("solid", start_color="1F4E79")   # 타이틀만 진한 네이비
            hfont = Font(name="맑은 고딕", bold=True, color="1F4E79", size=10)
            sfont = Font(name="맑은 고딕", bold=True, color="1F4E79", size=11)  # 섹션헤더 글자 파란색
            tfont = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=14)
            nfont = Font(name="맑은 고딕", size=10)
            nalign = Alignment(horizontal="left",   vertical="center", wrap_text=True)
            calign = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for ci, w in enumerate([22, 30, 20, 22, 18, 18], 1):
                ws.column_dimensions[gcl(ci)].width = w
            cur = 1
            ws.merge_cells(f"A{cur}:F{cur}")
            ws[f"A{cur}"] = f"제품 규격서  |  {sel_name}  ({sel_code})"
            ws[f"A{cur}"].font = tfont; ws[f"A{cur}"].fill = tfill
            ws[f"A{cur}"].alignment = calign; ws.row_dimensions[cur].height = 36; cur += 1
            ws.merge_cells(f"A{cur}:D{cur}")
            ws[f"A{cur}"] = "■ 1. 기본 정보"; ws[f"A{cur}"].font = sfont
            ws[f"A{cur}"].fill = sfill; ws[f"A{cur}"].alignment = calign
            ws.merge_cells(f"E{cur}:F{cur}")
            ws[f"E{cur}"] = "■ 제품 사진"; ws[f"E{cur}"].font = sfont
            ws[f"E{cur}"].fill = sfill; ws[f"E{cur}"].alignment = calign
            ws.row_dimensions[cur].height = 20; cur += 1
            basic_start = cur
            for _, br in df_b.iterrows():
                ws[f"A{cur}"] = str(br.iloc[0]) if len(br) > 0 else ""
                ws[f"A{cur}"].font = hfont; ws[f"A{cur}"].fill = hfill
                ws[f"A{cur}"].border = bd; ws[f"A{cur}"].alignment = calign
                ws.merge_cells(f"B{cur}:D{cur}")
                ws[f"B{cur}"] = str(br.iloc[1]) if len(br) > 1 and pd.notna(br.iloc[1]) else ""
                ws[f"B{cur}"].font = nfont; ws[f"B{cur}"].border = bd; ws[f"B{cur}"].alignment = nalign
                ws.row_dimensions[cur].height = 20; cur += 1
            basic_end = cur - 1
            if basic_end >= basic_start:
                ws.merge_cells(f"E{basic_start}:F{basic_end}")
                ws[f"E{basic_start}"].border = bdm; ws[f"E{basic_start}"].alignment = calign
            if img_path and os.path.exists(img_path):
                try:
                    from openpyxl.drawing.image import Image as OI
                    img_o = OI(img_path); img_o.width = 130
                    img_o.height = min(int(20 * (basic_end - basic_start + 1) * 1.33), 220)
                    ws.add_image(img_o, f"E{basic_start}")
                except Exception: pass
            cur += 1
            ws.merge_cells(f"A{cur}:F{cur}")
            ws[f"A{cur}"] = "■ 2. 배합비 (BOM)"; ws[f"A{cur}"].font = sfont
            ws[f"A{cur}"].fill = sfill; ws[f"A{cur}"].alignment = calign
            ws.row_dimensions[cur].height = 20; cur += 1
            for ci2, col in enumerate(df_r.columns, 1):
                c = ws.cell(row=cur, column=ci2, value=col)
                c.font = hfont; c.fill = hfill; c.border = bd; c.alignment = calign
            ws.row_dimensions[cur].height = 20; cur += 1
            for _, rr in df_r.iterrows():
                for ci2, val in enumerate(rr, 1):
                    c = ws.cell(row=cur, column=ci2, value=str(val) if pd.notna(val) else "")
                    c.font = nfont; c.border = bd
                    c.alignment = calign if ci2 >= 4 else nalign
                ws.row_dimensions[cur].height = 20; cur += 1
            cur += 1
            ws.merge_cells(f"A{cur}:F{cur}")
            ws[f"A{cur}"] = "■ 3. 완제품 규격 (위해요소)"; ws[f"A{cur}"].font = sfont
            ws[f"A{cur}"].fill = sfill; ws[f"A{cur}"].alignment = calign
            ws.row_dimensions[cur].height = 20; cur += 1
            for ci2, col in enumerate(df_h.columns, 1):
                c = ws.cell(row=cur, column=ci2, value=col)
                c.font = hfont; c.fill = hfill; c.border = bd; c.alignment = calign
            ws.row_dimensions[cur].height = 20; cur += 1
            for _, hr in df_h.iterrows():
                for ci2, val in enumerate(hr, 1):
                    c = ws.cell(row=cur, column=ci2, value=str(val) if pd.notna(val) else "")
                    c.font = hfont if ci2 == 1 else nfont
                    c.border = bd
                    c.alignment = calign if ci2 >= 2 else nalign
                    if ci2 == 1: c.fill = hfill
                ws.row_dimensions[cur].height = 20; cur += 1
            out = io.BytesIO(); wb.save(out); return out.getvalue()

        df_specs = load_specs()

        if st.session_state.selected_pcode is None:
            st.markdown('<div class="section-title">📦 제품 규격서 디렉토리</div>', unsafe_allow_html=True)
            c_search, _ = st.columns([1, 2])
            with c_search:
                search_t7 = st.text_input("🔍 제품 검색 (이름 또는 코드)", placeholder="빠른 규격 검색...")
            st.write("상세 규격을 열람/수정할 제품의 **'상세 규격 열기 📝'** 버튼을 클릭하세요.")
            if df_specs.empty:
                st.info("등록된 제품이 없습니다. 제품 등록 메뉴에서 제품을 먼저 등록해 주세요.")
            else:
                view_df = df_specs[
                    df_specs['제품명'].str.contains(search_t7, case=False, na=False) |
                    df_specs['제품코드'].str.contains(search_t7, case=False, na=False)
                ] if search_t7 else df_specs
                if view_df.empty:
                    st.warning("일치하는 제품이 없습니다.")
                else:
                    cols4 = st.columns(4)
                    for idx, row in view_df.reset_index().iterrows():
                        p_code, p_name, p_type = row['제품코드'], row['제품명'], row['유형']
                        tc = TYPE_COLOR.get(str(p_type), "#555")
                        with cols4[idx % 4]:
                            with st.container(border=True):
                                st.markdown(
                                    f"<span style='background:{tc}22;color:{tc};padding:2px 8px;"
                                    f"border-radius:10px;font-size:10px;font-weight:bold;'>{p_type}</span>",
                                    unsafe_allow_html=True)
                                st.markdown(f"<h5 style='color:#D11031;margin:4px 0 0;'>{p_code}</h5>", unsafe_allow_html=True)
                                st.markdown(f"**{p_name}**")
                                if st.button("상세 규격 열기 📝", key=f"btn_{p_code}_{idx}", use_container_width=True):
                                    st.session_state.selected_pcode = p_code
                                    st.session_state.selected_pname = p_name
                                    st.session_state.is_edit_mode = False
                                    st.rerun()
        else:
            sel_code  = st.session_state.selected_pcode
            sel_name  = st.session_state.selected_pname
            safe_code = str(sel_code).replace("/","_").replace("\\","_")
            B_FILE   = f"spec_basic_{safe_code}.csv"
            R_FILE   = f"spec_recipe_{safe_code}.csv"
            H_FILE   = f"spec_hazard_{safe_code}.csv"
            IMG_FILE = f"spec_photo_{safe_code}.png"

            if os.path.exists(B_FILE):
                df_basic = pd.read_csv(B_FILE, dtype=str)
                if "항목1(고정)" in df_basic.columns: df_basic = None
            else: df_basic = None
            if df_basic is None:
                df_basic = pd.DataFrame([
                    ["제품명",sel_name],["제품유형",""],["품목제조보고 연월일",""],
                    ["포장단위",""],["작성 연월일",""],["포장방법",""],
                    ["보관/유통 주의사항",""],["포장재질",""],["소비기한",""],
                    ["표시사항",""],["제품용도",""],["섭취방법",""],
                    ["작성자",""],["기타",""]
                ], columns=["항목(고정)","내용"])
            df_recipe = pd.read_csv(R_FILE, dtype=str) if os.path.exists(R_FILE) else pd.DataFrame([["","","","",""]],columns=["1차원료","2차원료","3차원료","배합비율(%)","비고"])
            if os.path.exists(H_FILE): df_hazard = pd.read_csv(H_FILE, dtype=str)
            else:
                df_hazard = pd.DataFrame([
                    ["물리학적 위해요소","금속이물","불검출","불검출"],
                    ["화학적 위해요소","납","2.0 이하","2.0 이하"],
                    ["미생물학적 위해요소","세균수","-","-"],
                    ["성상","외관/풍미","고유의 색택과 향미","고유의 색택과 향미"]
                ], columns=["위해요소 구분","검사항목","법적규격","당사규격"])

            # 헤더 (뒤로/수정/엑셀)
            hd1, hd2, hd3 = st.columns([1, 1, 2])
            with hd1:
                if st.button("← 목록으로", use_container_width=True):
                    st.session_state.selected_pcode = None; st.rerun()
            with hd2:
                mode_label = "✏️ 수정 취소" if st.session_state.is_edit_mode else "✏️ 정보 수정"
                if st.button(mode_label, use_container_width=True):
                    st.session_state.is_edit_mode = not st.session_state.is_edit_mode; st.rerun()
            with hd3:
                excel_data = export_full_excel(sel_code, sel_name, df_basic, df_recipe, df_hazard, IMG_FILE)
                st.download_button("🖨️ 보관용 엑셀 다운로드", data=excel_data,
                    file_name=f"Hollys_Spec_{safe_code}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, type="primary")

            spec_row_p = df_specs[df_specs["제품코드"]==sel_code]
            p_type_v   = spec_row_p.iloc[0]["유형"] if not spec_row_p.empty else ""
            tc_v       = TYPE_COLOR.get(str(p_type_v), "#555")
            st.markdown(
                f"<div style='display:flex;align-items:center;gap:10px;margin:8px 0;'>"
                f"<span style='background:{tc_v}22;color:{tc_v};padding:3px 12px;"
                f"border-radius:12px;font-size:11px;font-weight:bold;'>{p_type_v}</span>"
                f"<span style='font-size:18px;font-weight:bold;color:#1F4E79;'>{sel_name}</span>"
                f"<span style='color:#888;font-size:13px;'>({sel_code})</span></div>",
                unsafe_allow_html=True)
            st.divider()

            HAZARD_COLOR = {"물리학적 위해요소":"#7B5EA7","화학적 위해요소":"#D11031","미생물학적 위해요소":"#217346","성상":"#2E75B6","기타":"#888"}

            if not st.session_state.is_edit_mode:
                tab_b, tab_r, tab_h = st.tabs(["📋 기본정보", "⚗️ 배합비 (BOM)", "⚠️ 완제품 규격"])
                with tab_b:
                    img_c, info_c = st.columns([1, 2])
                    with img_c:
                        if os.path.exists(IMG_FILE):
                            st.image(IMG_FILE, caption="제품 사진", use_container_width=True)
                        else:
                            st.markdown("<div style='border:2px dashed #ccc;height:160px;display:flex;align-items:center;justify-content:center;color:#aaa;border-radius:8px;font-size:13px;'>사진 없음</div>", unsafe_allow_html=True)
                    with info_c:
                        html_b2 = '<table style="width:100%;border-collapse:collapse;">'
                        for _, br in df_basic.iterrows():
                            html_b2 += f'<tr><td style="{TDG_P}width:140px;">{br.iloc[0]}</td><td style="{TD_P}text-align:left;">{br.iloc[1] if pd.notna(br.iloc[1]) else ""}</td></tr>'
                        html_b2 += "</table>"
                        st.markdown(html_b2, unsafe_allow_html=True)
                with tab_r:
                    if not df_recipe.empty:
                        html_r2 = f'<table style="width:100%;border-collapse:collapse;"><thead><tr>'
                        for col in df_recipe.columns:
                            html_r2 += f'<th style="{TH_P2}">{col}</th>'
                        html_r2 += '</tr></thead><tbody>'
                        for _, rr in df_recipe.iterrows():
                            html_r2 += '<tr>'
                            for ci2, v in enumerate(rr):
                                bg2 = "#EBF3FB" if ci2 == 0 else "#fff"
                                html_r2 += f'<td style="{TD_P}background:{bg2};">{v if pd.notna(v) else ""}</td>'
                            html_r2 += '</tr>'
                        html_r2 += '</tbody></table>'
                        st.markdown(html_r2, unsafe_allow_html=True)
                with tab_h:
                    if not df_hazard.empty:
                        row_cnt_h = {}
                        for _, hr in df_hazard.iterrows():
                            cat = str(hr.iloc[0]) if pd.notna(hr.iloc[0]) else ""
                            row_cnt_h[cat] = row_cnt_h.get(cat, 0) + 1
                        html_h2 = f'<table style="width:100%;border-collapse:collapse;"><thead><tr>'
                        for col in df_hazard.columns:
                            html_h2 += f'<th style="{TH_P2}">{col}</th>'
                        html_h2 += '</tr></thead><tbody>'
                        cat_written_h = {}
                        for _, hr in df_hazard.iterrows():
                            cat = str(hr.iloc[0]) if pd.notna(hr.iloc[0]) else ""
                            hc  = HAZARD_COLOR.get(cat, "#555")
                            html_h2 += '<tr>'
                            if cat not in cat_written_h:
                                span = row_cnt_h.get(cat, 1)
                                html_h2 += (
                                    f'<td style="{TDG_P}color:{hc};" rowspan="{span}">'
                                    f'<span style="background:{hc}22;color:{hc};padding:2px 8px;'
                                    f'border-radius:8px;font-size:10px;font-weight:bold;">{cat}</span></td>'
                                )
                                cat_written_h[cat] = True
                            for v in hr.iloc[1:]:
                                html_h2 += f'<td style="{TD_P}">{v if pd.notna(v) else ""}</td>'
                            html_h2 += '</tr>'
                        html_h2 += '</tbody></table>'
                        st.markdown(html_h2, unsafe_allow_html=True)
            else:
                with st.expander("🔄 다른 제품의 규격 양식 불러오기"):
                    st.caption("기존 다른 제품의 규격을 가져와 현재 제품에 덮어씁니다.")
                    c_copy1, c_copy2 = st.columns([3, 1])
                    with c_copy1:
                        other_products = df_specs[df_specs['제품코드'] != sel_code]['제품명'].tolist()
                        copy_target = st.selectbox("불러올 제품 선택", ["선택하세요"] + other_products)
                    with c_copy2:
                        st.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)
                        if st.button("양식 덮어쓰기", use_container_width=True):
                            if copy_target != "선택하세요":
                                target_code = df_specs[df_specs['제품명']==copy_target].iloc[0]['제품코드']
                                safe_tc = str(target_code).replace("/","_").replace("\\","_")
                                for src_f, dst_f in [(f"spec_basic_{safe_tc}.csv", B_FILE),
                                                     (f"spec_recipe_{safe_tc}.csv", R_FILE),
                                                     (f"spec_hazard_{safe_tc}.csv", H_FILE)]:
                                    if os.path.exists(src_f):
                                        tmp = pd.read_csv(src_f, dtype=str)
                                        if dst_f == B_FILE and len(tmp)>0 and tmp.iloc[0,0]=="제품명":
                                            tmp.iloc[0,1] = sel_name
                                        tmp.to_csv(dst_f, index=False, encoding='utf-8-sig')
                                st.success(f"✅ [{copy_target}]의 양식을 불러왔습니다!"); st.rerun()
                            else:
                                st.warning("불러올 제품을 먼저 선택해 주세요.")
                st.info("💡 각 표의 빈 셀을 더블클릭하여 수정하세요.")
                tab_ed_b, tab_ed_r, tab_ed_h, tab_ed_img = st.tabs(["1. 기본정보", "2. 배합비 (BOM)", "3. 완제품 규격", "4. 사진"])
                with tab_ed_b:
                    cfg_b = {"항목(고정)": st.column_config.Column(disabled=True)}
                    edited_b = st.data_editor(df_basic, use_container_width=True, hide_index=True, column_config=cfg_b, key="prod_ed_basic")
                    if st.button("기본정보 저장", type="primary", key="prod_save_basic"):
                        edited_b.to_csv(B_FILE, index=False, encoding='utf-8-sig'); st.success("저장 완료!"); st.rerun()
                with tab_ed_r:
                    edited_r = st.data_editor(df_recipe, num_rows="dynamic", use_container_width=True, hide_index=False, key="prod_ed_recipe")
                    if st.button("배합비 저장", type="primary", key="prod_save_recipe"):
                        edited_r.to_csv(R_FILE, index=False, encoding='utf-8-sig'); st.success("저장 완료!"); st.rerun()
                with tab_ed_h:
                    cfg_h = {"위해요소 구분": st.column_config.SelectboxColumn("구분", options=["물리학적 위해요소","화학적 위해요소","미생물학적 위해요소","성상","기타"])}
                    edited_h = st.data_editor(df_hazard, num_rows="dynamic", use_container_width=True, hide_index=False, column_config=cfg_h, key="prod_ed_hazard")
                    if st.button("완제품 규격 저장", type="primary", key="prod_save_hazard"):
                        edited_h.to_csv(H_FILE, index=False, encoding='utf-8-sig'); st.success("저장 완료!"); st.rerun()
                with tab_ed_img:
                    st.caption("제품 사진을 등록하면 규격서 보기 및 엑셀 출력에 자동 포함됩니다.")
                    upl_img = st.file_uploader("이미지 업로드 (PNG/JPG)", type=["png","jpg","jpeg"], key="prod_img_up")
                    if upl_img:
                        with open(IMG_FILE,"wb") as fimg: fimg.write(upl_img.getbuffer())
                        st.success("사진 저장 완료!"); st.rerun()
                    if os.path.exists(IMG_FILE):
                        ic1, ic2 = st.columns([1,3])
                        with ic1:
                            st.image(IMG_FILE, use_container_width=True)
                            if st.button("🗑️ 사진 삭제", key="prod_del_img"):
                                os.remove(IMG_FILE); st.rerun()

    # ════════════════════════════════════════════════════════
    # 3. 데이터 히스토리
    # ════════════════════════════════════════════════════════
    elif sub_menu == "데이터 히스토리":
        st.subheader("데이터 히스토리 및 일일 생산량 관리")
        df_specs = load_specs()
        top1, top2, top3 = st.columns([2, 3, 1.5])
        with top1:
            sel_date_input = st.date_input("📅 생산일", value=date.today(), key="hist_date_pick")
            selected_date  = str(sel_date_input)
        with top2:
            if df_specs.empty:
                st.info("먼저 '간편 판정 규격'에서 제품을 등록하세요.")
                add_prod = ""
            else:
                add_prod = st.selectbox("제품 선택", [""] + df_specs["제품명"].tolist(), key="hist_prod_pick")
        with top3:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            add_clicked = st.button("추가", use_container_width=True, type="primary")
        if add_clicked:
            if not add_prod:
                st.error("제품을 선택해 주세요.")
            else:
                sr = df_specs[df_specs["제품명"] == add_prod].iloc[0]
                # 등록된 규격 리스트 중 첫 번째를 기본값으로 사용
                prod_specs = [s.strip() for s in str(sr.get("규격","")).split(",") if s.strip()]
                default_spec = prod_specs[0] if prod_specs else ""
                
                new_row = {
                    "생산일": selected_date, "유형": str(sr.get("유형","")),
                    "제품명": add_prod, "생산량": 0, "규격": default_spec,
                    "질소(%)": None, "수분(%)": None, "색도(Agtron)": None,
                    "추출시간(sec)": None, "추출시간_상세": None,
                    "날짜기록": None, "판정": None, "비고": ""
                }
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                df = df.sort_values(by="생산일", ascending=False).reset_index(drop=True)
                df.to_csv(DATA_FILE, index=False, encoding='utf-8-sig')
                st.success(f"[{add_prod}] → {selected_date} 추가됐습니다!"); st.rerun()
        st.divider()
        mask        = df['생산일'] == selected_date
        filtered_df = df[mask].copy().reset_index()
        filtered_df['규격'] = filtered_df['규격'].fillna("").astype(str).replace("nan","")
        filtered_df['비고'] = filtered_df['비고'].fillna("").astype(str).replace("nan","")
        rec_count = len(filtered_df)
        st.markdown(f"**{selected_date} 기록** — 총 {rec_count}건")
        if rec_count == 0:
            st.info("이 날짜에 기록이 없습니다. 위에서 제품을 선택하고 추가하세요.")
        else:
            _css_imp = "!important"
            st.markdown(
                f"<style>"
                f'div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button {{'
                f"height:38px{_css_imp}; font-size:11px{_css_imp}; padding:0 4px{_css_imp}; white-space:nowrap; margin-top:0{_css_imp}; }}"
                f'div[data-testid="stHorizontalBlock"] div[data-testid="stNumberInput"] input,'
                f'div[data-testid="stHorizontalBlock"] div[data-testid="stTextInput"] input {{'
                f"height:38px{_css_imp}; font-size:11px{_css_imp}; padding:4px 6px{_css_imp}; }}"
                f"</style>",
                unsafe_allow_html=True)
            COLS = [0.3, 1.0, 0.8, 2.0, 0.9, 1.2, 0.7, 0.7, 0.8, 1.0, 0.7, 1.0, 1.0, 0.6]
            HDR  = ["#","생산일","유형","제품명","생산량","규격","질소(%)","수분(%)","색도","추출시간","판정","비고","측정기록","삭제"]
            hdr_cols = st.columns(COLS)
            for hc, ht in zip(hdr_cols, HDR):
                hc.markdown(f"<div style='background:#1F4E79;color:white;padding:6px 2px;font-size:11px;font-weight:bold;text-align:center;border-radius:3px;margin-bottom:2px;'>{ht}</div>", unsafe_allow_html=True)
            edited_rows = []
            for ri, frow in filtered_df.iterrows():
                orig_idx = frow["index"]
                row_bg   = "#f8f8f8" if ri % 2 == 0 else "#ffffff"
                def rcell(val, bg=row_bg, fc="#333"):
                    return f"<div style='padding:9px 3px;font-size:11px;text-align:center;border-bottom:1px solid #eee;background:{bg};color:{fc};'>{val}</div>"
                pname = str(frow.get("제품명",""))
                ptype = str(frow.get("유형",""))
                tc2   = TYPE_COLOR.get(ptype,"#555")
                판정  = str(frow.get("판정",""))
                판정_bg = "#d4edda" if 판정=="PASS" else ("#f8d7da" if 판정=="FAIL" else row_bg)
                판정_fc = "#155724" if 판정=="PASS" else ("#721c24" if 판정=="FAIL" else "#333")
                rc = st.columns(COLS)
                rc[0].markdown(rcell(ri+1), unsafe_allow_html=True)
                rc[1].markdown(rcell(str(frow.get("생산일",""))), unsafe_allow_html=True)
                rc[2].markdown(
                    f"<div style='padding:9px 3px;font-size:11px;text-align:center;border-bottom:1px solid #eee;background:{row_bg};'>"
                    f"<span style='background:{tc2}22;color:{tc2};padding:1px 6px;border-radius:8px;font-size:10px;font-weight:bold;'>{ptype}</span></div>",
                    unsafe_allow_html=True)
                rc[3].markdown(rcell(f"<b>{pname}</b>"), unsafe_allow_html=True)
                qty_val = frow.get("생산량",0)
                try: qty_val = int(float(str(qty_val)))
                except: qty_val = 0
                new_qty  = rc[4].number_input("", value=qty_val, min_value=0, step=1, key=f"qty_{ri}_{orig_idx}", label_visibility="collapsed")
                
                # 제품 규격 연동 selectbox
                cur_prod_specs_str = ""
                if not df_specs.empty and pname in df_specs["제품명"].values:
                    cur_prod_specs_str = str(df_specs[df_specs["제품명"] == pname].iloc[0].get("규격", ""))
                
                cur_specs_list = [s.strip() for s in cur_prod_specs_str.split(",") if s.strip()]
                if not cur_specs_list: cur_specs_list = ["-"]
                
                cur_spec_val = str(frow.get("규격","")).replace("nan","").strip()
                try: 
                    def_spec_idx = cur_specs_list.index(cur_spec_val) if cur_spec_val in cur_specs_list else 0
                except: 
                    def_spec_idx = 0
                
                new_spec = rc[5].selectbox("", cur_specs_list, index=def_spec_idx, key=f"spec_{ri}_{orig_idx}", label_visibility="collapsed")
                for ci2, col_key in enumerate(["질소(%)","수분(%)","색도(Agtron)","추출시간(sec)"]):
                    v = str(frow.get(col_key,""))
                    v = "" if v in ["None","nan"] else v
                    rc[6+ci2].markdown(rcell(v), unsafe_allow_html=True)
                rc[10].markdown(rcell(판정, bg=판정_bg, fc=판정_fc), unsafe_allow_html=True)
                new_note = rc[11].text_input("", value=str(frow.get("비고","")).replace("nan",""), key=f"note_{ri}_{orig_idx}", label_visibility="collapsed")
                cur_sel = st.session_state.get("hist_meas_prod","")
                is_open = (cur_sel == f"{pname}__{ri}")
                if 판정 in ("PASS","FAIL"): btn_label, btn_type = "입력완료","secondary"
                elif is_open:              btn_label, btn_type = "닫기",    "secondary"
                else:                      btn_label, btn_type = "입력",    "primary"
                if rc[12].button(btn_label, key=f"mbtn_{ri}_{orig_idx}", use_container_width=True, type=btn_type):
                    st.session_state["hist_meas_prod"] = "" if is_open else f"{pname}__{ri}"
                    st.rerun()
                if rc[13].button("삭제", key=f"del_{ri}_{orig_idx}", use_container_width=True):
                    st.session_state[f"del_confirm_{orig_idx}"] = True
                if st.session_state.get(f"del_confirm_{orig_idx}", False):
                    st.warning(f"**{pname}** 행을 삭제하시겠습니까?")
                    dc1, dc2, _ = st.columns([1,1,5])
                    if dc1.button("확인", key=f"del_ok_{orig_idx}", type="primary"):
                        df = df.drop(index=orig_idx).reset_index(drop=True)
                        df.to_csv(DATA_FILE, index=False, encoding='utf-8-sig')
                        st.session_state.pop(f"del_confirm_{orig_idx}", None)
                        st.session_state["hist_meas_prod"] = ""; st.rerun()
                    if dc2.button("취소", key=f"del_cancel_{orig_idx}"):
                        st.session_state.pop(f"del_confirm_{orig_idx}", None); st.rerun()
                edited_rows.append({**{c: frow.get(c) for c in df.columns}, "생산량": new_qty, "규격": new_spec, "비고": new_note})
            edited_filtered_df = pd.DataFrame(edited_rows)
            if st.button(f"{selected_date} 저장", type="primary"):
                df = df[~mask]
                df = pd.concat([df, edited_filtered_df], ignore_index=True)
                df = df.sort_values(by="생산일", ascending=False).reset_index(drop=True)
                df.to_csv(DATA_FILE, index=False, encoding='utf-8-sig')
                st.success(f"[{selected_date}] 저장됐습니다."); st.rerun()
            sel_key      = st.session_state.get("hist_meas_prod","")
            prod_list_td = [p for p in filtered_df["제품명"].dropna().unique() if str(p).strip() not in ["","None","nan"]]
            if sel_key:
                sel_prod = sel_key.split("__")[0] if "__" in sel_key else sel_key
                if sel_prod in prod_list_td and not df_specs.empty and sel_prod in df_specs["제품명"].values:
                    spec_row = df_specs[df_specs["제품명"]==sel_prod].iloc[0]
                    ptype_m  = str(spec_row.get("유형",""))
                    tc_m     = TYPE_COLOR.get(ptype_m,"#555")
                    st.markdown(
                        f"<div style='background:#EBF3FB;border-left:5px solid #2E75B6;padding:10px 16px;border-radius:6px;margin:8px 0;'>"
                        f"<b>📋 {sel_prod}</b> &nbsp;|&nbsp; {selected_date} &nbsp;|&nbsp;"
                        f"<span style='background:{tc_m}22;color:{tc_m};padding:1px 8px;border-radius:8px;font-size:11px;font-weight:bold;'>{ptype_m}</span>"
                        f"</div>", unsafe_allow_html=True)
                    with st.form(f"form_meas_{sel_prod}", clear_on_submit=False):
                        n2 = moisture = color = "N/A"
                        mc1, mc2, mc3 = st.columns(3)
                        with mc1:
                            if str(spec_row.get("최소_질소","")) not in ["","N/A","nan"]:
                                n2 = str(st.number_input("질소 (%)", value=0.0, step=0.01, key=f"n2_inp_{sel_prod}"))
                        with mc2:
                            if str(spec_row.get("최소_수분","")) not in ["","N/A","nan"]:
                                moisture = str(st.number_input("수분 (%)", value=0.0, step=0.01, key=f"moi_inp_{sel_prod}"))
                        with mc3:
                            if str(spec_row.get("최소_색도","")) not in ["","N/A","nan"]:
                                color = str(st.number_input("색도 (Agtron)", value=0.0, step=0.1, key=f"col_inp_{sel_prod}"))
                        st.caption("추출시간 (최대 5회 입력, 쉼표 또는 스페이스 구분)")
                        ext_input = st.text_input("추출시간 (초)", placeholder="예: 24.5, 25.0, 23.8", key=f"ext_inp_{sel_prod}")
                        if st.form_submit_button("판정 저장", type="primary"):
                            ext_times = []
                            if ext_input.strip():
                                for v in ext_input.replace(","," ").split():
                                    try: ext_times.append(float(v))
                                    except Exception: pass
                            ext_avg = str(round(sum(ext_times)/len(ext_times),2)) if ext_times else "N/A"
                            ext_detail_str = ", ".join([str(v) for v in ext_times]) if ext_times else ""
                            is_pass = True; fail_reason = []
                            def chk(val, mn, mx):
                                if val != "N/A" and str(mn) not in ["","N/A","nan"]:
                                    return float(mn) <= float(val) <= float(mx)
                                return True
                            if not chk(n2,      spec_row.get("최소_질소",""),  spec_row.get("최대_질소","")): is_pass=False; fail_reason.append("질소")
                            if not chk(moisture,spec_row.get("최소_수분",""),  spec_row.get("최대_수분","")): is_pass=False; fail_reason.append("수분")
                            if ext_times and ext_avg != "N/A":
                                if not chk(ext_avg,spec_row.get("최소_추출",""),spec_row.get("최대_추출","")): is_pass=False; fail_reason.append("추출시간")
                            final_status = "PASS" if is_pass else "FAIL"
                            um = (df["생산일"]==selected_date) & (df["제품명"]==sel_prod)
                            if um.any():
                                ix = df[um].index[-1]
                                df.at[ix,"질소(%)"] = n2; df.at[ix,"수분(%)"] = moisture
                                df.at[ix,"색도(Agtron)"] = color; df.at[ix,"추출시간(sec)"] = ext_avg
                                df.at[ix,"추출시간_상세"] = ext_detail_str; df.at[ix,"판정"] = final_status
                            else:
                                nr = {"생산일":selected_date,"유형":str(spec_row.get("유형","")),"제품명":sel_prod,"생산량":0,"단위":str(spec_row.get("단위","")),"질소(%)":n2,"수분(%)":moisture,"색도(Agtron)":color,"추출시간(sec)":ext_avg,"추출시간_상세":ext_detail_str,"날짜기록":None,"판정":final_status,"비고":""}
                                df = pd.concat([df, pd.DataFrame([nr])], ignore_index=True)
                            df.to_csv(DATA_FILE, index=False, encoding='utf-8-sig')
                            st.session_state["hist_meas_prod"] = ""
                            if is_pass: st.success(f"✅ [{sel_prod}] 판정: PASS")
                            else:       st.error(f"❌ [{sel_prod}] 판정: FAIL — {fail_reason}")
                            st.rerun()

    # ════════════════════════════════════════════════════════
    # 4. 제품별 데이터 관리
    # ════════════════════════════════════════════════════════
    elif sub_menu == "제품별 데이터 관리":
        st.markdown('<div class="section-title">📈 제품별 데이터 트렌드 관리</div>', unsafe_allow_html=True)
        if df.empty:
            st.warning("저장된 데이터 히스토리가 없습니다. '데이터 히스토리'에서 먼저 데이터를 추가해 주세요.")
        else:
            product_list = [p for p in df['제품명'].dropna().unique().tolist() if str(p).strip() not in ["","-","None"]]
            if not product_list:
                st.info("조회할 수 있는 제품 데이터가 없습니다.")
            else:
                fc1, fc2 = st.columns([2,3])
                with fc1:
                    selected_prod = st.selectbox("🔍 제품 선택", product_list)
                with fc2:
                    category_map = {"질소":"질소(%)","수분":"수분(%)","색도":"색도(Agtron)","추출시간":"추출시간(sec)"}
                    selected_cat = st.radio("📈 품질 지표 선택", list(category_map.keys()), horizontal=True)
                target_col = category_map[selected_cat]
                prod_df = df[df['제품명']==selected_prod].copy()
                prod_df['생산일'] = pd.to_datetime(prod_df['생산일'], errors='coerce')
                prod_df = prod_df.dropna(subset=['생산일']).sort_values('생산일')
                prod_df[target_col] = pd.to_numeric(prod_df[target_col], errors='coerce')
                chart_df = prod_df.dropna(subset=[target_col]).copy()
                chart_df['생산일_str'] = chart_df['생산일'].dt.strftime('%Y-%m-%d')
                st.divider()
                if chart_df.empty:
                    st.info(f"'{selected_prod}' 제품의 '{selected_cat}' 측정 데이터가 아직 없습니다.")
                else:
                    vals = chart_df[target_col]
                    m1,m2,m3,m4 = st.columns(4)
                    m1.metric("평균", f"{vals.mean():.2f}")
                    m2.metric("최솟값", f"{vals.min():.2f}")
                    m3.metric("최댓값", f"{vals.max():.2f}")
                    m4.metric("측정 횟수", f"{len(vals)}회")
                    c_chart, c_data = st.columns([3,2])
                    with c_chart:
                        st.markdown(f"**📉 {selected_prod} — {selected_cat} 트렌드**")
                        fig = px.line(chart_df, x='생산일_str', y=target_col, markers=True, text=target_col,
                            labels={'생산일_str':'생산 일자', target_col:f'{selected_cat} 측정값'},
                            color_discrete_sequence=['#D11031'])
                        fig.update_traces(textposition="top center")
                        fig.update_layout(margin=dict(l=20,r=20,t=30,b=20))
                        st.plotly_chart(fig, use_container_width=True)
                    with c_data:
                        st.markdown("**📋 상세 기록**")
                        display_df = chart_df[['생산일_str',target_col,'판정','비고']].copy()
                        display_df.rename(columns={'생산일_str':'생산일자'}, inplace=True)
                        html_td2 = f'<table style="width:100%;border-collapse:collapse;"><thead><tr>'
                        for col in display_df.columns:
                            html_td2 += f'<th style="{TH_P2}">{col}</th>'
                        html_td2 += '</tr></thead><tbody>'
                        for _, drow in display_df.iterrows():
                            판정_v = str(drow.get("판정",""))
                            p_bg = "#d4edda" if 판정_v=="PASS" else ("#f8d7da" if 판정_v=="FAIL" else "#fff")
                            p_fc = "#155724" if 판정_v=="PASS" else ("#721c24" if 판정_v=="FAIL" else "#333")
                            html_td2 += '<tr>'
                            for col, v in drow.items():
                                if col == "판정":
                                    html_td2 += f'<td style="{TD_P}background:{p_bg};color:{p_fc};font-weight:bold;">{v if pd.notna(v) else ""}</td>'
                                else:
                                    html_td2 += f'<td style="{TD_P}">{v if pd.notna(v) else ""}</td>'
                            html_td2 += '</tr>'
                        html_td2 += '</tbody></table>'
                        st.markdown(html_td2, unsafe_allow_html=True)

                        def export_trend_excel(data, prod_name, cat_name):
                            import openpyxl as _opx
                            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                            from openpyxl.utils import get_column_letter as gcl
                            wb2 = _opx.Workbook(); ws2 = wb2.active; ws2.title="품질트렌드"
                            thin2=Side(style="thin",color="B0C4D8"); bd2=Border(left=thin2,right=thin2,top=thin2,bottom=thin2)
                            hfill2=PatternFill("solid",start_color="D6E4F0"); tfill2=PatternFill("solid",start_color="1F4E79")
                            hfont2=Font(name="맑은 고딕",bold=True,color="1F4E79",size=10)
                            tfont2=Font(name="맑은 고딕",bold=True,color="FFFFFF",size=13)
                            nfont2=Font(name="맑은 고딕",size=10)
                            calign2=Alignment(horizontal="center",vertical="center")
                            for ci2, w in enumerate([15,12,10,30],1): ws2.column_dimensions[gcl(ci2)].width = w
                            ws2.merge_cells("A1:D1")
                            ws2["A1"] = f"[{prod_name}] {cat_name} 품질 데이터 보관 대장"
                            ws2["A1"].font = tfont2; ws2["A1"].fill = tfill2
                            ws2["A1"].alignment = calign2; ws2.row_dimensions[1].height = 30
                            for ci2, col in enumerate(data.columns, 1):
                                c = ws2.cell(row=2, column=ci2, value=col)
                                c.font = hfont2; c.fill = hfill2; c.border = bd2; c.alignment = calign2
                            ws2.row_dimensions[2].height = 20
                            pass_fill=PatternFill("solid",start_color="D4EDDA")
                            fail_fill=PatternFill("solid",start_color="F8D7DA")
                            for ri2, row in data.iterrows():
                                판정_x = str(row.get("판정",""))
                                for ci2, col in enumerate(data.columns,1):
                                    c = ws2.cell(row=ri2+3, column=ci2, value=str(row[col]) if pd.notna(row[col]) else "")
                                    c.font = nfont2; c.border = bd2; c.alignment = calign2
                                    if col == "판정":
                                        if 판정_x=="PASS": c.fill = pass_fill
                                        elif 판정_x=="FAIL": c.fill = fail_fill
                                ws2.row_dimensions[ri2+3].height = 18
                            out2 = io.BytesIO(); wb2.save(out2); return out2.getvalue()

                        st.markdown("<div style='margin-top:15px;'></div>", unsafe_allow_html=True)
                        st.download_button(
                            label=f"📥 {selected_cat} 기록 엑셀 다운로드",
                            data=export_trend_excel(display_df, selected_prod, selected_cat),
                            file_name=f"Hollys_{selected_prod}_{selected_cat}_{date.today().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True, type="primary")

# --- 4. 직원 관리 메뉴 ---
elif menu_selection == "직원 관리":
    if sub_menu == "조직도 및 인원 관리":
        st.markdown('<div class="section-title">회사 인원 및 조직도 관리</div>', unsafe_allow_html=True)
        st.write("사내 인원을 등록하면 보건증 현황 등 각종 관리 메뉴에 자동으로 연동되며, 조직도 엑셀 양식을 다운로드할 수 있습니다.")
        
        df_emp = load_employees()
        
        with st.expander("➕ 신규 인원 등록"):
            with st.form("form_employee"):
                e1, e2 = st.columns(2)
                with e1:
                    e_empno = st.text_input("사번 (예: H2026001)")
                    e_name = st.text_input("이름")
                with e2:
                    e_rank = st.text_input("직급 (예: 매니저, 대리, 사원, 팀장)")
                    e_contact = st.text_input("연락처")
                
                c_date, c_stat = st.columns(2)
                with c_date:
                    e_join_date = st.date_input("입사일")
                with c_stat:
                    e_status = st.selectbox("재직 상태", ["재직", "퇴사", "휴직"])

                c_haccp, c_ccp, c_etc = st.columns(3)
                with c_haccp:
                    e_haccp = st.selectbox("HACCP 직책", ["해당없음", "HACCP 팀장", "생산/시설관리팀", "품질관리팀", "업무지원팀"])
                with c_ccp:
                    e_ccp = st.multiselect("모니터링 담당 CCP", ["CCP-1", "CCP-2", "CCP-3", "CCP-4", "CCP-5", "CCP-6"])
                with c_etc:
                    e_etc = st.text_input("기타 (비고)")

                if st.form_submit_button("인원 등록"):
                    if e_name and e_empno:
                        if e_empno in df_emp['사번'].values:
                            st.error("이미 등록된 사번입니다. 다른 사번을 입력해 주세요.")
                        else:
                            ccp_str = ", ".join(e_ccp) if e_ccp else ""
                            new_emp = pd.DataFrame([[e_empno, e_rank, e_name, e_contact, str(e_join_date), e_status, e_haccp, ccp_str, e_etc]], columns=df_emp.columns)
                            df_emp = pd.concat([df_emp, new_emp], ignore_index=True)
                            df_emp.to_csv(EMPLOYEE_FILE, index=False, encoding='utf-8-sig')
                            st.success(f"{e_name} 님 등록이 완료되었습니다!")
                            st.rerun()
                    else:
                        st.error("사번과 이름을 모두 입력하세요.")

        st.markdown("**📋 전체 직원 목록** (아래 표에서 직접 수정할 수 있습니다.)")
        
        cfg_emp = {
            "재직상태": st.column_config.SelectboxColumn("재직상태", options=["재직", "퇴사", "휴직"]),
            "HACCP 직책": st.column_config.SelectboxColumn("HACCP 직책", options=["해당없음", "HACCP 팀장", "생산/시설관리팀", "품질관리팀", "업무지원팀"]),
            "모니터링 CCP": st.column_config.TextColumn("모니터링 CCP (쉼표로 구분하여 수정 가능)"),
            "기타": st.column_config.TextColumn("기타"), 
            "입사일": st.column_config.DateColumn("입사일", format="YYYY-MM-DD")
        }
        
        display_emp = df_emp.copy()
        display_emp['입사일'] = pd.to_datetime(display_emp['입사일'], errors='coerce').apply(lambda x: x.date() if pd.notna(x) else None)
        
        edited_emp = st.data_editor(display_emp, num_rows="dynamic", use_container_width=True, column_config=cfg_emp)
        
        def export_org_excel(df_view):
            df_view = df_view.fillna("") 
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
                workbook = writer.book
                worksheet = workbook.add_worksheet('HACCP조직도')
                
                worksheet.set_paper(9) 
                worksheet.fit_to_pages(1, 1) 
                worksheet.center_horizontally() 
                worksheet.set_margins(left=0.5, right=0.5, top=0.5, bottom=0.5) 
                
                title_fmt = workbook.add_format({'bold': True, 'font_size': 20, 'align': 'center', 'valign': 'vcenter'})
                doc_info_title_fmt = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bold': True, 'border': 1, 'bg_color': '#F2F2F2'})
                doc_info_val_fmt = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
                
                header_fmt = workbook.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})
                name_fmt = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
                section_fmt = workbook.add_format({'bg_color': '#E2EFDA', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True, 'font_size': 12})
                
                worksheet.set_column('A:A', 3)
                worksheet.set_column('B:G', 16) 
                
                worksheet.merge_range('B2:E4', "HACCP팀 조직도", title_fmt)
                
                worksheet.write('F2', "문서번호", doc_info_title_fmt)
                worksheet.write('G2', "HLSHS-101-A01", doc_info_val_fmt)
                worksheet.write('F3', "제정일자", doc_info_title_fmt)
                worksheet.write('G3', "2018-11-05", doc_info_val_fmt)
                worksheet.write('F4', "개정일자", doc_info_title_fmt)
                worksheet.write('G4', date.today().strftime('%Y-%m-%d'), doc_info_val_fmt)
                
                team_leaders = df_view[df_view['HACCP 직책'] == 'HACCP 팀장']['이름'].tolist()
                leader_name = ", ".join(team_leaders) if team_leaders else "(공란)"
                worksheet.merge_range('D6:E6', "HACCP 팀장", header_fmt)
                worksheet.merge_range('D7:E8', leader_name, name_fmt)
                
                teams = ["생산/시설관리팀", "품질관리팀", "업무지원팀"]
                col_map = {"생산/시설관리팀": 1, "품질관리팀": 3, "업무지원팀": 5} 
                
                max_team_members = 0
                for t in teams:
                    c_idx = col_map[t]
                    worksheet.merge_range(10, c_idx, 10, c_idx+1, t, header_fmt)
                    worksheet.set_row(10, 25)
                    
                    members = df_view[df_view['HACCP 직책'] == t]
                    team_heads = members[members['직급'].str.contains('팀장', na=False)]
                    team_normals = members[~members['직급'].str.contains('팀장', na=False)]
                    sorted_members = pd.concat([team_heads, team_normals])
                    
                    if len(sorted_members) > max_team_members: max_team_members = len(sorted_members)
                    
                    row_idx = 11
                    for _, row in sorted_members.iterrows():
                        info = f"{row['이름']}\n({row['직급']})"
                        worksheet.merge_range(row_idx, c_idx, row_idx, c_idx+1, info, name_fmt)
                        worksheet.set_row(row_idx, 35)
                        row_idx += 1
                    
                    while row_idx <= 11 + 5: 
                        worksheet.merge_range(row_idx, c_idx, row_idx, c_idx+1, "", name_fmt)
                        worksheet.set_row(row_idx, 35)
                        row_idx += 1

                ccp_start_row = 11 + max(max_team_members, 6) + 2
                
                worksheet.merge_range(ccp_start_row, 1, ccp_start_row, 6, "CCP 모니터링 담당", section_fmt)
                worksheet.set_row(ccp_start_row, 25)
                
                ccps = ["CCP-1", "CCP-2", "CCP-3", "CCP-4", "CCP-5", "CCP-6"]
                for i, ccp in enumerate(ccps):
                    c_idx = i + 1  
                    
                    worksheet.write(ccp_start_row + 1, c_idx, ccp, header_fmt)
                    worksheet.set_row(ccp_start_row + 1, 20)
                    
                    assigned = []
                    for _, row in df_view.iterrows():
                        if ccp in str(row['모니터링 CCP']):
                            assigned.append(f"{row['이름']}")
                
                    row_idx = ccp_start_row + 2
                    for name in assigned:
                        worksheet.write(row_idx, c_idx, name, name_fmt)
                        worksheet.set_row(row_idx, 25)
                        row_idx += 1
                    
                    end_row = ccp_start_row + 2 + 5 
                    while row_idx < end_row:
                        worksheet.write(row_idx, c_idx, "", name_fmt)
                        worksheet.set_row(row_idx, 25)
                        row_idx += 1

            return output.getvalue()

        c_e1, c_e2 = st.columns(2)
        with c_e1:
            if st.button("직원 정보 변경사항 저장"):
                edited_emp['입사일'] = edited_emp['입사일'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) else "")
                edited_emp.to_csv(EMPLOYEE_FILE, index=False, encoding='utf-8-sig')
                st.success("직원 정보가 성공적으로 업데이트 되었습니다.")
                st.rerun()
                
        with c_e2:
            st.download_button(
                label="HACCP 조직도 엑셀 다운로드 (인쇄 최적화)",
                data=export_org_excel(edited_emp),
                file_name=f"Hollys_OrgChart_{date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    elif sub_menu == "파트타이머 연차 관리":
        from dateutil.relativedelta import relativedelta as rdelta
        from datetime import date as dclass

        st.markdown('<div class="section-title">파트타이머 연차 관리</div>', unsafe_allow_html=True)
        st.write("직급이 **파트타이머**인 직원의 연차·무급휴가를 등록하고, 발생/사용/잔여 연차를 자동으로 계산합니다.")

        df_emp = load_employees()
        df_lv  = load_annual_leave()

        # 파트타이머만 필터
        pt_df = df_emp[
            df_emp["재직상태"].str.strip().str.upper().isin(["재직", "재직중", "Y", "TRUE", ""]) |
            df_emp["재직상태"].isna()
        ]
        pt_df = pt_df[pt_df["직급"].str.strip().str.contains("파트", na=False)]

        if pt_df.empty:
            st.info("직급에 '파트타이머'가 포함된 재직 직원이 없습니다. 조직도 및 인원 관리에서 직급을 '파트타이머'로 등록해 주세요.")
            st.stop()

        # ── 직원 선택 ─────────────────────────────────────────────────
        pt_options = {f"{r['이름']} ({r['사번']})": r['사번'] for _, r in pt_df.iterrows()}
        sel_label  = st.selectbox("👤 파트타이머 선택", list(pt_options.keys()), key="pt_sel")
        sel_empid  = pt_options[sel_label]
        sel_row    = pt_df[pt_df["사번"] == sel_empid].iloc[0]
        hire_date  = str(sel_row.get("입사일", "")).strip()
        today      = dclass.today()

        # ── 연차 발생 계산 ────────────────────────────────────────────
        accrued, used, remaining, detail = calc_annual_leave(hire_date, df_lv, sel_empid, today)

        # 근속 정보
        try:
            hire_dt = pd.to_datetime(hire_date).date()
            rd      = rdelta(today, hire_dt)
            tenure  = f"{rd.years}년 {rd.months}개월 {rd.days}일"
        except Exception:
            hire_dt = None
            tenure  = "입사일 정보 없음"

        # ── 요약 카드 ─────────────────────────────────────────────────
        st.markdown(f"""
        <div style="background:#EBF3FB;border-radius:10px;padding:14px 20px;margin:10px 0;">
          <span style="font-size:14px;color:#1F4E79;font-weight:bold;">📌 {sel_row['이름']}
          &nbsp;|&nbsp; 입사일: {hire_date} &nbsp;|&nbsp; 근속: {tenure}</span>
        </div>
        """, unsafe_allow_html=True)

        c1, c2, c3, c4 = st.columns(4)
        def stat_card(col, label, val, color, sub=""):
            col.markdown(f"""
            <div style="background:{color}18;border:2px solid {color};border-radius:10px;
                        padding:14px;text-align:center;">
              <div style="font-size:12px;color:#555;margin-bottom:4px;">{label}</div>
              <div style="font-size:28px;font-weight:bold;color:{color};">{val}</div>
              <div style="font-size:10px;color:#888;">{sub}</div>
            </div>""", unsafe_allow_html=True)

        stat_card(c1, "발생 연차",  accrued,   "#2E75B6", "누적 발생")
        stat_card(c2, "사용 연차",  used,      "#E67E22", "연차 사용")
        stat_card(c3, "잔여 연차",  remaining, "#27AE60" if remaining >= 0 else "#E74C3C", "사용 가능")
        unpaid_cnt = len(df_lv[(df_lv["사번"]==sel_empid) & (df_lv["유형"]=="무급휴가")])
        stat_card(c4, "무급 휴가",  unpaid_cnt,"#8E44AD", "사용 횟수")

        # 발생 상세 내역
        if detail:
            with st.expander("📊 연차 발생 상세 내역 보기"):
                for k, v in detail.items():
                    st.markdown(f"- **{k}**: {v}개")
                if hire_dt:
                    next_gen = None
                    rd_now = rdelta(today, hire_dt)
                    if rd_now.years == 0:
                        next_month_pt = hire_dt + rdelta(months=rd_now.months+1)
                        if next_month_pt > today:
                            next_gen = next_month_pt
                    if next_gen:
                        st.info(f"⏰ 다음 연차 발생 예정일: **{next_gen.strftime('%Y-%m-%d')}** (발생 조건: 해당 월 무급휴가 없을 시)")

        st.divider()

        # ── 탭: 기록 목록 | 추가 | 수정 ─────────────────────────────
        tab_list, tab_add = st.tabs(["📋 휴가 기록 목록", "➕ 새 기록 등록"])

        emp_records = df_lv[df_lv["사번"] == sel_empid].copy().reset_index(drop=True)

        # ── 탭1: 목록 + 수정/삭제 ────────────────────────────────────
        with tab_list:
            if emp_records.empty:
                st.info("등록된 휴가 기록이 없습니다. '새 기록 등록' 탭에서 추가하세요.")
            else:
                st.markdown(f"**총 {len(emp_records)}건** (연차: {used}개, 무급휴가: {unpaid_cnt}건)")

                # 날짜 컬럼을 datetime 타입으로 변환 (DateColumn 호환)
                display_records = emp_records[["날짜","유형","비고"]].copy()
                display_records["날짜"] = pd.to_datetime(display_records["날짜"], errors="coerce")

                edit_cfg = {
                    "날짜":  st.column_config.DateColumn("날짜", format="YYYY-MM-DD"),
                    "유형":  st.column_config.SelectboxColumn("유형", options=["연차","무급휴가"]),
                    "비고":  st.column_config.TextColumn("비고"),
                }
                edited = st.data_editor(
                    display_records,
                    column_config=edit_cfg,
                    use_container_width=True,
                    num_rows="dynamic",
                    key="lv_editor"
                )

                col_sv, col_dl = st.columns([2,1])
                with col_sv:
                    if st.button("수정 내용 저장", use_container_width=True, type="primary"):
                        # 기존 이 직원 레코드 제거 후 재삽입
                        df_lv_other = df_lv[df_lv["사번"] != sel_empid].copy()
                        new_recs = edited.copy()
                        new_recs["사번"] = sel_empid
                        new_recs["이름"] = sel_row["이름"]
                        new_recs["날짜"] = pd.to_datetime(new_recs["날짜"], errors="coerce").dt.strftime("%Y-%m-%d")
                        new_recs = new_recs.dropna(subset=["날짜"])
                        df_lv = pd.concat([df_lv_other, new_recs[["사번","이름","날짜","유형","비고"]]], ignore_index=True)
                        save_annual_leave(df_lv)
                        st.success("저장 완료!")
                        st.rerun()
                with col_dl:
                    # 엑셀 다운로드
                    def make_leave_excel(emp_row, records, acc, us, rem, unpd, detail_d):
                        import io as _io
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                        from openpyxl.worksheet.page import PageMargins

                        wb = Workbook(); ws = wb.active; ws.title = "연차기록"
                        BK = Side(style="thin", color="000000")
                        def bdr(): return Border(top=BK,bottom=BK,left=BK,right=BK)
                        def fill(h): return PatternFill("solid",fgColor=h)
                        def fnt(bold=False,sz=10,fc="000000"): return Font(name="맑은 고딕",bold=bold,size=sz,color=fc)
                        def aln(h="center",v="center",wrap=True): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
                        def wc(r,c,val="",bold=False,sz=10,fc="000000",bg=None,h="center"):
                            cell=ws.cell(row=r,column=c,value=str(val) if val is not None else "")
                            cell.font=fnt(bold,sz,fc); cell.alignment=aln(h); cell.border=bdr()
                            if bg: cell.fill=fill(bg)
                        def mc(r1,c1,r2,c2): ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)

                        # 제목
                        mc(1,1,1,6); wc(1,1,"파트타이머 연차 관리표",bold=True,sz=14,fc="FFFFFF",bg="1F4E79"); ws.row_dimensions[1].height=28
                        mc(2,1,2,6); wc(2,1,"커피클럽 로스팅센터  |  ㈜케이지할리스에프앤비",sz=9,fc="595959",bg="EBF3FB"); ws.row_dimensions[2].height=13

                        # 직원 정보
                        info_items = [("이름",emp_row["이름"]),("사번",emp_row["사번"]),
                                      ("직급",emp_row["직급"]),("입사일",emp_row.get("입사일",""))]
                        for i,(lbl,val) in enumerate(info_items):
                            c=i*2+1
                            wc(3,c,lbl,bold=True,bg="D6E4F0"); wc(3,c+1,val,h="left")
                        ws.row_dimensions[3].height=16

                        # 요약
                        mc(4,1,4,2); wc(4,1,"발생 연차",bold=True,bg="2E75B6",fc="FFFFFF")
                        mc(4,3,4,4); wc(4,3,"사용 연차",bold=True,bg="E67E22",fc="FFFFFF")
                        mc(4,5,4,6); wc(4,5,"잔여 연차",bold=True,bg="27AE60" if rem>=0 else "E74C3C",fc="FFFFFF")
                        mc(5,1,5,2); wc(5,1,acc,bold=True,sz=14)
                        mc(5,3,5,4); wc(5,3,us, bold=True,sz=14)
                        mc(5,5,5,6); wc(5,5,rem,bold=True,sz=14)
                        ws.row_dimensions[4].height=16; ws.row_dimensions[5].height=22

                        # 헤더
                        for ci,hdr in enumerate(["No","날짜","유형","비고"],1):
                            bg="1F4E79"
                            ws.cell(row=6,column=ci,value=hdr).font=fnt(bold=True,fc="FFFFFF")
                            ws.cell(row=6,column=ci).fill=fill(bg)
                            ws.cell(row=6,column=ci).border=bdr()
                            ws.cell(row=6,column=ci).alignment=aln()
                        ws.row_dimensions[6].height=16

                        r=7
                        for ni,(_, rec) in enumerate(records.iterrows(),1):
                            bg="F5F5F5" if r%2==0 else "FFFFFF"
                            wc(r,1,ni,bg=bg); wc(r,2,str(rec.get("날짜","")),bg=bg)
                            type_bg = "FCE4D6" if rec.get("유형")=="연차" else "EDE7F6"
                            wc(r,3,str(rec.get("유형","")),bg=type_bg)
                            wc(r,4,str(rec.get("비고","")),bg=bg,h="left")
                            ws.row_dimensions[r].height=18; r+=1

                        ws.column_dimensions["A"].width=5; ws.column_dimensions["B"].width=14
                        ws.column_dimensions["C"].width=12; ws.column_dimensions["D"].width=30

                        ws.page_setup.paperSize=9; ws.page_setup.fitToPage=True
                        ws.page_margins=PageMargins(left=0.5,right=0.5,top=0.7,bottom=0.7)

                        out=_io.BytesIO(); wb.save(out); return out.getvalue()

                    xl = make_leave_excel(sel_row, emp_records, accrued, used, remaining, unpaid_cnt, detail)
                    st.download_button(
                        "엑셀 다운로드",
                        data=xl,
                        file_name=f"연차기록_{sel_row['이름']}_{today.strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

        # ── 탭2: 새 기록 추가 ────────────────────────────────────────
        with tab_add:
            st.markdown("#### 새 휴가 기록 등록")
            with st.form("form_add_leave", clear_on_submit=True):
                fa1, fa2 = st.columns(2)
                with fa1:
                    f_date = st.date_input("날짜 *", value=today)
                with fa2:
                    f_type = st.selectbox("유형 *", ["연차", "무급휴가"])
                f_note = st.text_input("비고", placeholder="예: 개인 사정, 병가 등")

                # 무급휴가 선택 시 연차 미발생 안내
                if f_type == "무급휴가":
                    st.warning("⚠️ 무급휴가 등록 시, 해당 월의 연차 1개가 발생하지 않습니다.")

                submitted = st.form_submit_button("등록", type="primary", use_container_width=True)

            if submitted:
                new_rec = pd.DataFrame([{
                    "사번": sel_empid,
                    "이름": sel_row["이름"],
                    "날짜": f_date.strftime("%Y-%m-%d"),
                    "유형": f_type,
                    "비고": f_note
                }])
                df_lv = pd.concat([df_lv, new_rec], ignore_index=True)
                save_annual_leave(df_lv)
                st.success(f"✅ [{f_date}] {f_type} 등록 완료!")
                st.rerun()

        # ── 전체 파트타이머 현황 요약 ─────────────────────────────────
        st.divider()
        with st.expander("📊 전체 파트타이머 연차 현황 요약", expanded=False):
            rows = []
            for _, emp in pt_df.iterrows():
                eid = emp["사번"]
                acc, us, rem, _ = calc_annual_leave(str(emp.get("입사일","")), df_lv, eid, today)
                try:
                    rd2 = rdelta(today, pd.to_datetime(emp.get("입사일","")).date())
                    ten = f"{rd2.years}년 {rd2.months}개월"
                except:
                    ten = "-"
                upd = len(df_lv[(df_lv["사번"]==eid) & (df_lv["유형"]=="무급휴가")])
                rows.append({"이름":emp["이름"], "사번":eid, "입사일":emp.get("입사일",""),
                             "근속":ten, "발생":acc, "사용":us, "잔여":rem, "무급":upd})
            if rows:
                df_sum = pd.DataFrame(rows)
                st.dataframe(
                    df_sum.style
                        .applymap(lambda v: "background:#E2F0D9;font-weight:bold;" if isinstance(v,int) and v>0 and df_sum.columns[df_sum.isin([v]).any()].tolist()[0]=="잔여" else "", subset=["잔여"])
                        .applymap(lambda v: "color:#C00000;font-weight:bold;" if isinstance(v,int) and v<0 else "", subset=["잔여"]),
                    use_container_width=True,
                    hide_index=True
                )

    elif sub_menu == "보건증 현황관리":
        st.markdown('<div class="section-title">🩺 보건증 현황 관리표</div>', unsafe_allow_html=True)
        st.write("직원 보건증 현황을 등록합니다. **검진일로부터 1년 뒤가 만료일로 자동 계산**되며, 30일 전 대시보드 캘린더에 경고 뱃지가 뜹니다.")
        
        df_hc = load_health_cert()
        df_emp = load_employees()
        
        active_emps = df_emp[df_emp["재직상태"] == "재직"] if not df_emp.empty and "재직상태" in df_emp.columns else pd.DataFrame()
        
        with st.expander("➕ 신규 인원 보건증 등록"):
            if active_emps.empty:
                st.warning("등록된 재직 직원이 없습니다. 좌측 메뉴의 '직원 관리'에서 직원을 먼저 등록해 주세요.")
            else:
                with st.form("form_health"):
                    st.caption("아래 목록에서 직원을 선택하면 직급과 연락처가 자동으로 연동되어 저장됩니다.")
                    
                    emp_options = active_emps.apply(lambda x: f"{x['이름']} ({x['직급']}) - {x['사번']}", axis=1).tolist()
                    selected_emp_str = st.selectbox("직원 선택", emp_options)
                    
                    h_date = st.date_input("검진 일자")
                    
                    if st.form_submit_button("보건증 등록"):
                        if selected_emp_str:
                            emp_idx = emp_options.index(selected_emp_str)
                            sel_row = active_emps.iloc[emp_idx]
                            
                            h_rank = sel_row['직급']
                            h_name = sel_row['이름']
                            h_contact = sel_row['연락처']

                            new_df = pd.DataFrame([[h_rank, h_name, h_contact, str(h_date)]], columns=df_hc.columns)
                            df_hc = pd.concat([df_hc, new_df], ignore_index=True)
                            df_hc.to_csv(HEALTH_CERT_FILE, index=False, encoding='utf-8-sig')
                            st.success(f"[{h_name}] 보건증 등록 완료!")
                            st.rerun()

        display_hc = df_hc.copy()
        today = pd.to_datetime(date.today())
        
        exp_dates = []
        d_days = []
        statuses = []
        
        display_hc['검진일자'] = pd.to_datetime(display_hc['검진일자'], errors='coerce')
        
        for _, row in display_hc.iterrows():
            exam = row['검진일자']
            if pd.isna(exam):
                exp_dates.append("")
                d_days.append("")
                statuses.append("미등록")
                continue
            
            exp = exam + pd.DateOffset(years=1) - pd.Timedelta(days=1)
            days_left = (exp - today).days
            
            exp_dates.append(exp.strftime('%Y-%m-%d'))
            d_days.append(days_left)
            if days_left < 0: statuses.append("🔴 만료")
            elif days_left <= 30: statuses.append("🟠 갱신요망")
            else: statuses.append("🟢 정상")
                
        display_hc['만기일자 (자동)'] = exp_dates
        display_hc['D-Day'] = d_days
        display_hc['상태'] = statuses

        st.markdown("**📋 전체 직원 보건증 현황** (아래 표에서 직접 정보/검진일을 더블클릭해 수정할 수 있습니다.)")
        
        display_hc['검진일자'] = display_hc['검진일자'].apply(lambda x: x.date() if pd.notna(x) else None)
        
        cfg_hc = {
            "검진일자": st.column_config.DateColumn("검진일자 (더블클릭)", format="YYYY-MM-DD"),
            "만기일자 (자동)": st.column_config.Column("만기일자 (자동)", disabled=True),
            "D-Day": st.column_config.Column("D-Day", disabled=True),
            "상태": st.column_config.Column("상태", disabled=True)
        }

        edited_hc = st.data_editor(display_hc, num_rows="dynamic", use_container_width=True, column_config=cfg_hc)
        
        c_h1, c_h2 = st.columns(2)
        with c_h1:
            if st.button("보건증 데이터 저장 및 업데이트"):
                save_df = edited_hc[['직급', '이름', '연락처', '검진일자']]
                save_df['검진일자'] = save_df['검진일자'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) else "")
                save_df.to_csv(HEALTH_CERT_FILE, index=False, encoding='utf-8-sig')
                st.success("데이터가 저장되었습니다.")
                st.rerun()
                
        def export_health_excel(df_view):
            df_view = df_view.fillna("") 
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
                workbook = writer.book
                worksheet = workbook.add_worksheet('보건증현황')
                
                title_format = workbook.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
                header_format = workbook.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})
                normal_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
                alert_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_color': 'red', 'bold': True})
                
                worksheet.merge_range('A1:G1', "건강진단결과(보건증) 현황 관리표", title_format)
                worksheet.set_row(0, 40)
                worksheet.write('A2', f"출력일자: {date.today().strftime('%Y-%m-%d')}")
                
                headers = list(df_view.columns)
                for col_num, header in enumerate(headers):
                    worksheet.write(2, col_num, header, header_format)
                
                for row_num, row in df_view.iterrows():
                    for col_num, col_name in enumerate(headers):
                        val = row[col_name]
                        if isinstance(val, (date, datetime)): val = val.strftime('%Y-%m-%d')
                            
                        is_alert = False
                        if col_name in ['D-Day', '상태']:
                            if val != "" and row['D-Day'] != "":
                                try:
                                    if float(row['D-Day']) <= 30:
                                        is_alert = True
                                except ValueError:
                                    pass
                        
                        if is_alert:
                            worksheet.write(row_num + 3, col_num, val, alert_format)
                        else:
                            worksheet.write(row_num + 3, col_num, val, normal_format)
                
                worksheet.set_column('A:B', 15); worksheet.set_column('C:C', 20)
                worksheet.set_column('D:E', 18); worksheet.set_column('F:G', 15)
                
            return output.getvalue()
            
        with c_h2:
            st.download_button(
                label="보건증 현황 관리표 엑셀 다운로드",
                data=export_health_excel(display_hc),
                file_name=f"Hollys_HealthCert_{date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

# 🌟 5. 설비 관리 메뉴 
elif menu_selection == "설비 관리":
    
    if sub_menu == "제조위생설비이력관리":
        st.markdown('<div class="section-title">🏭 제조위생설비이력관리</div>', unsafe_allow_html=True)
        st.write("공장 내 설비를 등록하고 수리 이력을 관리합니다. 작성된 데이터는 엑셀 양식으로 바로 출력할 수 있습니다.")
        
        if 'selected_facility' not in st.session_state:
            st.session_state.selected_facility = None

        df_fac = load_facilities()
        df_rep = load_repairs()

        if st.session_state.selected_facility is None:
            
            with st.expander("➕ 신규 설비 이력카드 등록"):
                with st.form("form_facility"):
                    col1, col2 = st.columns(2)
                    with col1:
                        f_no = st.text_input("설비번호 (예: F-01)")
                        f_name = st.text_input("설비명")
                        f_usage = st.text_input("사용용도")
                        f_volt = st.text_input("전압")
                        f_year = st.text_input("구입년월")
                    with col2:
                        f_maker = st.text_input("제조회사명")
                        f_loc = st.text_input("설치장소")
                        f_dept = st.text_input("관리부서")
                        f_man_main = st.text_input("관리자 (정)")
                        f_man_sub = st.text_input("관리자 (부)")
                    
                    f_note = st.text_area("특이사항 (모델명, 규격, 안전·방호장치 등)")
                    
                    if st.form_submit_button("설비 등록 완료"):
                        if f_no and f_name:
                            if f_no in df_fac['설비번호'].values:
                                st.error("이미 존재하는 설비번호입니다.")
                            else:
                                new_f = pd.DataFrame([[f_no, f_name, f_usage, f_volt, f_year, f_maker, f_loc, f_dept, f_man_main, f_man_sub, f_note]], columns=df_fac.columns)
                                df_fac = pd.concat([df_fac, new_f], ignore_index=True)
                                df_fac.to_csv(FACILITY_FILE, index=False, encoding='utf-8-sig')
                                st.success(f"[{f_no}] {f_name} 설비 등록 완료!")
                                st.rerun()
                        else: 
                            st.error("설비번호와 설비명은 필수 항목입니다.")

            st.markdown("**📋 전체 등록 설비 목록 (DB)**")
            edited_fac = st.data_editor(df_fac, num_rows="dynamic", use_container_width=True)
            
            def export_facility_list(df_list):
                df_list = df_list.fillna("")
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
                    wb = writer.book
                    ws = wb.add_worksheet('설비목록표')
                    
                    ws.set_paper(9)
                    ws.fit_to_pages(1, 0)
                    
                    title_fmt = wb.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
                    header_fmt = wb.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})
                    val_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
                    
                    ws.merge_range('A1:K2', "제조위생설비 목록표", title_fmt)
                    
                    headers = list(df_list.columns)
                    for col_num, header in enumerate(headers):
                        ws.write(3, col_num, header, header_fmt)
                        ws.set_column(col_num, col_num, 12 if col_num != 10 else 35) 
                        
                    for r_idx, row in df_list.iterrows():
                        for c_idx, col_name in enumerate(headers):
                            ws.write(r_idx + 4, c_idx, str(row[col_name]), val_fmt)
                            ws.set_row(r_idx + 4, 25)
                            
                return output.getvalue()

            c_f1, c_f2 = st.columns(2)
            with c_f1:
                if st.button("설비 목록 정보 일괄 저장"):
                    edited_fac.to_csv(FACILITY_FILE, index=False, encoding='utf-8-sig')
                    st.success("설비 DB가 성공적으로 업데이트 되었습니다.")
                    st.rerun()
            with c_f2:
                st.download_button(
                    "설비 목록표 엑셀 다운로드", 
                    data=export_facility_list(edited_fac), 
                    file_name=f"Hollys_FacilityList_{date.today().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            st.markdown("---")
            st.markdown("**🔍 개별 설비 이력 카드 관리 (사진 첨부 및 수리내역 작성)**")
            st.write("아래 설비 카드의 버튼을 클릭하면 세부 이력 카드로 진입합니다.")
            
            if df_fac.empty:
                st.info("등록된 설비가 없습니다.")
            else:
                cols = st.columns(4)
                for idx, row in df_fac.iterrows():
                    f_no_val = row['설비번호']
                    f_name_val = row['설비명']
                    f_loc_val = row['설치장소']
                    with cols[idx % 4]:
                        with st.container(border=True):
                            st.markdown(f"<h4 style='color: #D11031; margin-bottom: 0;'>{f_no_val}</h4>", unsafe_allow_html=True)
                            st.markdown(f"**{f_name_val}**")
                            st.caption(f"장소: {f_loc_val}")
                            if st.button("이력카드 열람 📝", key=f"btn_fac_{f_no_val}", use_container_width=True):
                                st.session_state.selected_facility = f_no_val
                                st.rerun()

        else: 
            sel_fno = st.session_state.selected_facility
            filtered_fac = df_fac[df_fac['설비번호'] == sel_fno]
            
            if filtered_fac.empty:
                st.warning(f"선택한 설비({sel_fno}) 정보를 찾을 수 없습니다. 데이터가 삭제되었거나 번호가 변경되었습니다.")
                if st.button("⬅️ 설비 목록으로 돌아가기"):
                    st.session_state.selected_facility = None
                    st.rerun()
            else:
                f_row = filtered_fac.iloc[0]
                
                col_fb1, col_fb2 = st.columns([1, 6])
                with col_fb1:
                    if st.button("⬅️ 설비 목록으로", use_container_width=True):
                        st.session_state.selected_facility = None
                        st.rerun()
                with col_fb2:
                    st.markdown(f"<h3 style='margin-top: 0;'>[{sel_fno}] {f_row['설비명']} 제조위생설비 이력카드</h3>", unsafe_allow_html=True)
                    
                c_info, c_img = st.columns([1.5, 1])
                IMG_FILE = f"fac_photo_{sel_fno}.png"
                
                with c_info:
                    st.markdown('<div class="section-title">■ 1. 기계 기본 정보</div>', unsafe_allow_html=True)
                    info_df = pd.DataFrame({
                        "분류": ["설비명", "사용용도", "구입년월", "설치장소", "관리자(정)", "특이사항"],
                        "항목 내용": [f_row['설비명'], f_row['사용용도'], f_row['구입년월'], f_row['설치장소'], f_row['관리자_정'], f_row['특이사항']],
                        "분류_2": ["설비번호", "전압", "제조회사명", "관리부서", "관리자(부)", ""],
                        "항목 내용_2": [f_row['설비번호'], f_row['전압'], f_row['제조회사명'], f_row['관리부서'], f_row['관리자_부'], ""]
                    })
                    st.table(info_df)
                    
                with c_img:
                    st.markdown('<div class="section-title">■ 2. 기계 사진</div>', unsafe_allow_html=True)
                    upl_img = st.file_uploader("📥 기계 사진 등록 (PNG, JPG)", type=['png', 'jpg', 'jpeg'])
                    if upl_img:
                        with open(IMG_FILE, "wb") as f: f.write(upl_img.getbuffer())
                        st.rerun()
                    if os.path.exists(IMG_FILE):
                        st.image(IMG_FILE, use_container_width=True)
                        if st.button("사진 삭제", use_container_width=True):
                            os.remove(IMG_FILE)
                            st.rerun()
                            
                st.markdown('<div class="section-title">■ 3. 수리 이력 사항</div>', unsafe_allow_html=True)
                
                my_repairs = df_rep[df_rep['설비번호'] == sel_fno].copy()
                
                with st.form("form_repair"):
                    c_r1, c_r2, c_r3, c_r4 = st.columns([1, 2, 1, 1])
                    with c_r1: r_date = st.date_input("수리일자")
                    with c_r2: r_detail = st.text_input("수리사항 (내용)")
                    with c_r3: r_shop = st.text_input("수리처")
                    with c_r4: r_note = st.text_input("비고")
                    if st.form_submit_button("수리 내역 추가"):
                        if r_detail:
                            new_r = pd.DataFrame([[sel_fno, str(r_date), r_detail, r_shop, r_note]], columns=df_rep.columns)
                            df_rep = pd.concat([df_rep, new_r], ignore_index=True)
                            df_rep.to_csv(REPAIR_FILE, index=False, encoding='utf-8-sig')
                            st.success("수리 내역 추가 완료!")
                            st.rerun()
                        else:
                            st.error("수리사항 내용을 입력해 주세요.")
                            
                st.markdown("**수리 내역 열람 및 수정** (표에서 바로 수정 후 하단 저장 버튼 클릭)")
                
                display_rep = my_repairs.drop(columns=['설비번호']).copy()
                display_rep['수리일자'] = pd.to_datetime(display_rep['수리일자'], errors='coerce').apply(lambda x: x.date() if pd.notna(x) else None)
                
                cfg_rep = {
                    "수리일자": st.column_config.DateColumn("수리일자", format="YYYY-MM-DD")
                }
                
                edited_rep = st.data_editor(display_rep, num_rows="dynamic", use_container_width=True, column_config=cfg_rep)
                
                def export_facility_excel(f_data, r_data, img_path):
                    f_data = f_data.fillna("") 
                    r_data = r_data.fillna("")
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
                        wb = writer.book
                        ws = wb.add_worksheet('설비이력카드')
                        
                        title_fmt = wb.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
                        doc_fmt = wb.add_format({'align': 'right', 'valign': 'vcenter', 'bold': True})
                        header_fmt = wb.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})
                        val_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
                        val_left_fmt = wb.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter', 'text_wrap': True})
                        section_fmt = wb.add_format({'bold': True, 'font_size': 12, 'valign': 'vcenter'})
                        
                        ws.set_paper(9)
                        ws.fit_to_pages(1, 1)
                        ws.center_horizontally()
                        ws.set_margins(0.5, 0.5, 0.5, 0.5)
                        
                        ws.set_column('A:A', 12)
                        ws.set_column('B:B', 15)
                        ws.set_column('C:C', 12)
                        ws.set_column('D:D', 15)
                        ws.set_column('E:E', 12)
                        ws.set_column('F:F', 20)
                        
                        ws.merge_range('A1:F2', "제조위생설비 이력카드", title_fmt)
                        ws.write('F3', "문서번호", doc_fmt)
                        ws.write('F4', "HLSPP-203-F02", wb.add_format({'align': 'right', 'valign': 'vcenter'}))
                        
                        ws.merge_range('A6:D6', "1. 기계기본정보", section_fmt)
                        ws.merge_range('E6:F6', "2. 기계사진", section_fmt)
                        
                        ws.write('A7', '설비명', header_fmt)
                        ws.write('B7', f_data['설비명'], val_fmt)
                        ws.write('C7', '설비번호', header_fmt)
                        ws.write('D7', f_data['설비번호'], val_fmt)
                        
                        ws.write('A8', '사용용도', header_fmt)
                        ws.write('B8', f_data['사용용도'], val_fmt)
                        ws.write('C8', '전압', header_fmt)
                        ws.write('D8', f_data['전압'], val_fmt)
                        
                        ws.write('A9', '구입년월', header_fmt)
                        ws.write('B9', f_data['구입년월'], val_fmt)
                        ws.write('C9', '제조회사명', header_fmt)
                        ws.write('D9', f_data['제조회사명'], val_fmt)
                        
                        ws.write('A10', '설치장소', header_fmt)
                        ws.write('B10', f_data['설치장소'], val_fmt)
                        ws.write('C10', '관리부서', header_fmt)
                        ws.write('D10', f_data['관리부서'], val_fmt)
                        
                        ws.merge_range('A11:A12', '관리자', header_fmt)
                        ws.write('B11', '정', header_fmt)
                        ws.merge_range('C11:D11', f_data['관리자_정'], val_fmt)
                        
                        ws.write('B12', '부', header_fmt)
                        ws.merge_range('C12:D12', f_data['관리자_부'], val_fmt)
                        
                        ws.merge_range('A13:A16', '특이사항', header_fmt)
                        ws.merge_range('B13:D16', f_data['특이사항'], val_left_fmt)
                        
                        ws.merge_range('E7:F16', '', val_fmt)
                        if os.path.exists(img_path):
                            try: ws.insert_image('E7', img_path, {'x_offset': 5, 'y_offset': 5, 'object_position': 1, 'x_scale': 0.25, 'y_scale': 0.25})
                            except Exception: pass
                        
                        for r in range(6, 16): ws.set_row(r, 22)
                        
                        ws.merge_range('A18:F18', "3. 수리이력사항", section_fmt)
                        ws.write('A19', '수리일자', header_fmt)
                        ws.merge_range('B19:C19', '수리사항', header_fmt)
                        ws.merge_range('D19:E19', '수리처', header_fmt)
                        ws.write('F19', '비고', header_fmt)
                        
                        r_idx = 19
                        for _, row in r_data.iterrows():
                            val_date = row.get('수리일자', '')
                            if isinstance(val_date, (date, datetime)): val_date = val_date.strftime('%Y-%m-%d')
                            ws.write(r_idx, 0, str(val_date), val_fmt)
                            ws.merge_range(r_idx, 1, r_idx, 2, str(row.get('수리사항', '')), val_fmt)
                            ws.merge_range(r_idx, 3, r_idx, 4, str(row.get('수리처', '')), val_fmt)
                            ws.write(r_idx, 5, str(row.get('비고', '')), val_fmt)
                            ws.set_row(r_idx, 22)
                            r_idx += 1
                            
                        while r_idx < 35:
                            ws.write(r_idx, 0, "", val_fmt)
                            ws.merge_range(r_idx, 1, r_idx, 2, "", val_fmt)
                            ws.merge_range(r_idx, 3, r_idx, 4, "", val_fmt)
                            ws.write(r_idx, 5, "", val_fmt)
                            ws.set_row(r_idx, 22)
                            r_idx += 1
                            
                    return output.getvalue()
                    
                c_br1, c_br2 = st.columns(2)
                with c_br1:
                    if st.button("수리 이력 변경사항 저장", use_container_width=True):
                        df_rep = df_rep[df_rep['설비번호'] != sel_fno]
                        edited_rep['설비번호'] = sel_fno
                        edited_rep['수리일자'] = edited_rep['수리일자'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) else "")
                        df_rep = pd.concat([df_rep, edited_rep], ignore_index=True)
                        df_rep.to_csv(REPAIR_FILE, index=False, encoding='utf-8-sig')
                        st.success("저장 완료!")
                        st.rerun()
                        
                with c_br2:
                    st.download_button(
                        label="🖨️ 해당 설비이력카드 엑셀 다운로드 (인쇄 최적화)",
                        data=export_facility_excel(f_row, edited_rep, IMG_FILE),
                        file_name=f"Hollys_FacilityCard_{sel_fno}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

    elif sub_menu == "필터 점검관리":
        st.markdown('<div class="section-title">🚰 필터 점검 관리 (계획표)</div>', unsafe_allow_html=True)
        st.write("각종 필터(정수기, 공조기, 집진기 등)의 점검(교체/세척) 주기를 캘린더와 연동하여 관리하고 계획표를 출력합니다.")

        df_filter = load_filter_plan()

        with st.expander("➕ 새 필터 점검 항목 등록"):
            with st.form("form_filter"):
                f1, f2, f3 = st.columns(3)
                with f1:
                    flt_loc = st.text_input("설치장소 (예: 로스팅룸, 1층)")
                    flt_type = st.text_input("필터명 (예: 카본필터, 헤파필터)")
                with f2:
                    flt_content = st.selectbox("점검 종류", ["교체", "세척", "점검"])
                    flt_cycle_opt = st.selectbox("점검 주기", ["직접 입력(개월)", "파손시"])
                    if flt_cycle_opt == "직접 입력(개월)":
                        flt_cycle = st.number_input("주기 (개월)", min_value=1, step=1, value=6)
                    else:
                        flt_cycle = "파손시"
                with f3:
                    flt_date = st.date_input("최종 점검일자")
                    flt_note = st.text_input("비고 (특이사항)")

                if st.form_submit_button("필터 등록하기"):
                    if flt_loc and flt_type:
                        if flt_cycle == "파손시":
                            next_date_str = "파손시"
                        else:
                            next_date_str = str((flt_date + pd.DateOffset(months=int(flt_cycle))).date())
                        
                        new_df = pd.DataFrame([[flt_loc, flt_type, flt_content, str(flt_cycle), str(flt_date), next_date_str, "예정", flt_note]], columns=df_filter.columns)
                        df_filter = pd.concat([df_filter, new_df], ignore_index=True)
                        df_filter.to_csv(FILTER_PLAN_FILE, index=False, encoding='utf-8-sig')
                        log_history("필터 점검 항목 등록", "설비 관리", f"장소: {flt_loc}, 필터: {flt_type}")
                        st.success("필터 등록 완료! 대시보드 캘린더에 연동되었습니다.")
                        st.rerun()
                    else:
                        st.error("설치장소와 필터명을 입력하세요.")

        view_df_filter = df_filter.copy()
        view_df_filter['점검일자'] = pd.to_datetime(view_df_filter['점검일자'], errors='coerce')
        # [수정] 주기_개월을 숫자로 강제 변환하면 '파손시'가 유실되므로 문자열로 유지
        view_df_filter['주기_개월'] = view_df_filter['주기_개월'].astype(str).replace("nan", "").replace("None", "")

        def calc_next_filter(row):
            if pd.notna(row['점검일자']) and pd.notna(row['주기_개월']):
                try:
                    if str(row['주기_개월']) == "파손시":
                        return "파손시"
                    return row['점검일자'] + pd.DateOffset(months=int(float(row['주기_개월'])))
                except: return pd.NaT
            return pd.NaT

        view_df_filter['차기점검일자'] = view_df_filter.apply(calc_next_filter, axis=1)

        # UI 표시를 위해 날짜 객체로 변환 (파손시는 문자열 그대로 유지됨)
        def to_date_safe(val):
            if isinstance(val, (date, datetime, pd.Timestamp)):
                return val.date()
            return val

        view_df_filter['점검일자'] = view_df_filter['점검일자'].apply(to_date_safe)
        view_df_filter['차기점검일자'] = view_df_filter['차기점검일자'].apply(to_date_safe)

        cfg_filter = {
            "내용": st.column_config.SelectboxColumn("점검 종류", options=["교체", "세척", "점검"]),
            "점검일자": st.column_config.DateColumn("최종 점검일자", format="YYYY-MM-DD"),
            "차기점검일자": st.column_config.Column("차기 점검일자 (자동)", disabled=True),
            "주기_개월": st.column_config.TextColumn("주기 (개월)", help="'파손시' 또는 숫자 입력"),
            "상태": st.column_config.SelectboxColumn("상태", options=["예정", "완료"])
        }

        st.markdown("**📋 등록된 필터 점검 목록** (아래 표에서 텍스트와 날짜를 직접 수정 가능합니다)")
        st.caption("💡 점검 완료 시 이곳에서 **점검일자**를 갱신해주시면 대시보드 캘린더의 차기 점검 알람이 자동으로 연장됩니다.")
        edited_filter = st.data_editor(view_df_filter, num_rows="dynamic", use_container_width=True, column_config=cfg_filter)

        def export_filter_excel(df_view):
            df_view = df_view.fillna("")
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
                wb = writer.book
                ws = wb.add_worksheet('필터점검계획표')
                
                ws.set_paper(9) # A4
                ws.fit_to_pages(1, 0)
                
                title_fmt = wb.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
                header_fmt = wb.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})
                val_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
                
                ws.merge_range('A1:H2', f"{date.today().year}년 필터 점검 계획 및 기록표", title_fmt)
                
                headers = ["설치장소", "필터명", "내용(교체/세척)", "주기(개월)", "점검일자", "차기점검일자", "상태", "비고"]
                for c, h in enumerate(headers):
                    ws.write(3, c, h, header_fmt)
                
                ws.set_column('A:A', 18)
                ws.set_column('B:B', 20)
                ws.set_column('C:C', 15)
                ws.set_column('D:D', 12)
                ws.set_column('E:F', 15)
                ws.set_column('G:G', 10)
                ws.set_column('H:H', 20)
                
                r_idx = 4
                for _, row in df_view.iterrows():
                    ws.write(r_idx, 0, str(row.get('설치장소', '')), val_fmt)
                    ws.write(r_idx, 1, str(row.get('필터명', '')), val_fmt)
                    ws.write(r_idx, 2, str(row.get('내용', '')), val_fmt)
                    ws.write(r_idx, 3, str(row.get('주기_개월', '')), val_fmt)
                    
                    d1 = row.get('점검일자', '')
                    if isinstance(d1, (datetime, pd.Timestamp, date)): d1 = d1.strftime('%Y-%m-%d')
                    ws.write(r_idx, 4, str(d1), val_fmt)
                    
                    d2 = row.get('차기점검일자', '')
                    if isinstance(d2, (datetime, pd.Timestamp, date)): d2 = d2.strftime('%Y-%m-%d')
                    ws.write(r_idx, 5, str(d2), val_fmt)
                    
                    ws.write(r_idx, 6, str(row.get('상태', '')), val_fmt)
                    ws.write(r_idx, 7, str(row.get('비고', '')), val_fmt)
                    ws.set_row(r_idx, 25)
                    r_idx += 1
                    
            return output.getvalue()

        c_filt1, c_filt2 = st.columns(2)
        with c_filt1:
            if st.button("필터 점검표 저장 및 업데이트"):
                edited_filter['점검일자'] = pd.to_datetime(edited_filter['점검일자'], errors='coerce')
                
                def apply_next_calc(row):
                    res = calc_next_filter(row)
                    if isinstance(res, (date, datetime, pd.Timestamp)):
                        return res.strftime('%Y-%m-%d')
                    return str(res) if res else ""

                edited_filter['차기점검일자'] = edited_filter.apply(apply_next_calc, axis=1)
                edited_filter['점검일자'] = edited_filter['점검일자'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) else "")

                edited_filter.to_csv(FILTER_PLAN_FILE, index=False, encoding='utf-8-sig')
                st.success("성공적으로 저장 및 차기 점검일이 갱신되었습니다.")
                st.rerun()
                
        with c_filt2:
            st.download_button(
                label="필터점검계획표 엑셀 다운로드 (보관용)",
                data=export_filter_excel(edited_filter),
                file_name=f"Hollys_FilterPlan_{date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    elif sub_menu == "세척소독 기준":
        st.markdown('<div class="section-title">🧽 SSOP 위생·세척소독 기준표 관리</div>', unsafe_allow_html=True)
        st.write("위생표준작업절차(SSOP)에 따른 11가지 카테고리별 세척·소독 기준을 엑셀과 같은 형식으로 상세히 관리합니다. **식품 제조시설**은 설비 이력카드와 연동됩니다.")

        df_clean = load_cleaning_specs()
        df_fac = load_facilities()

        with st.expander("➕ 새 기준 및 세척방법 추가", expanded=False):
            selected_cat = st.selectbox("📌 대분류 선택", CLEAN_CATEGORIES)
            
            with st.form("form_clean_add"):
                c1, c2 = st.columns(2)
                with c1:
                    # 카테고리가 5. 식품 제조시설일 경우 설비 데이터 연동
                    if selected_cat == "5. 식품 제조시설":
                        if df_fac.empty:
                            st.warning("등록된 설비가 없습니다. 제조위생설비이력관리에서 먼저 등록해주세요.")
                            c_target = st.text_input("관리 대상/설비명 (임시 입력)")
                        else:
                            fac_options = df_fac.apply(lambda x: f"[{x['설비번호']}] {x['설비명']}", axis=1).tolist()
                            c_target = st.selectbox("연동할 대상 설비 선택", fac_options)
                    else:
                        c_target = st.text_input("관리 대상 (예: 바닥, 위생화, 손, 환풍기 등)")
                        
                    c_part = st.text_input("세부 관리 부위/항목 (예: 손가락 사이, 모터팬, 내부 등)")
                
                with c2:
                    c_cycle = st.text_input("청소/소독 주기 (예: 1회/일, 작업 전후)")
                    c_tool = st.text_input("사용 세제/도구 (예: 70% 알코올, 폼크린)")
                    c_manager = st.text_input("책임자 (예: 생산팀 담당자)")

                c_method = st.text_area("세척·소독 방법 (세부 작업 절차를 상세히 기재)")
                
                # 🔥 5분류(식품 제조시설)가 아닌 경우에만 폼 내부에서 사진 업로더 노출
                uploaded_photo = None
                if selected_cat != "5. 식품 제조시설":
                    uploaded_photo = st.file_uploader("📸 현장 사진 등록 (선택사항)", type=['jpg', 'jpeg', 'png'])

                if st.form_submit_button("위생 기준 등록"):
                    if c_target and c_part:
                        new_id = f"C-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                        
                        saved_img_path = ""
                        # 사진이 첨부되었으면 저장 처리
                        if uploaded_photo is not None:
                            saved_img_path = f"clean_photo_{new_id}.png"
                            with open(saved_img_path, "wb") as f:
                                f.write(uploaded_photo.getbuffer())

                        new_row = pd.DataFrame([[new_id, selected_cat, '일반구역', c_target, c_part, c_method, c_cycle, c_tool, c_manager, saved_img_path]], columns=df_clean.columns)
                        df_clean = pd.concat([df_clean, new_row], ignore_index=True)
                        df_clean.to_csv(CLEAN_FILE, index=False, encoding='utf-8-sig')
                        
                        st.success("위생 기준이 성공적으로 등록되었습니다!")
                        st.rerun()
                    else:
                        st.error("관리 대상과 부위 항목은 필수입니다.")

        st.markdown("### 📋 전체 세척소독 기준 텍스트 일괄 편집")
        st.caption("표에서 셀을 더블클릭하여 내용을 한 번에 수정할 수 있습니다.")
        cfg_clean = {
            "ID": st.column_config.Column(disabled=True, width="small"),
            "대분류": st.column_config.SelectboxColumn("대분류", options=CLEAN_CATEGORIES),
            "구역": st.column_config.TextColumn("구역", width="small"),
            "사진파일": st.column_config.Column("사진 경로", disabled=True)
        }
        edited_clean = st.data_editor(df_clean, num_rows="dynamic", use_container_width=True, column_config=cfg_clean)
        
        def export_cleaning_excel(df_view):
            df_view = df_view.fillna("")
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
                wb = writer.book
                ws = wb.add_worksheet('세척소독 기준표')

                title_fmt = wb.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
                header_fmt = wb.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})
                val_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
                left_fmt = wb.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter', 'text_wrap': True})

                ws.merge_range('A1:H2', "SSOP 11분류 세척소독 기준표", title_fmt)

                headers = ['대분류', '관리대상/설비명', '부위', '세척·소독 방법', '주기', '사용세제/도구', '책임자', '현장 사진']
                for c, h in enumerate(headers):
                    ws.write(3, c, h, header_fmt)

                ws.set_column('A:A', 15); ws.set_column('B:C', 15)
                ws.set_column('D:D', 35)
                ws.set_column('E:G', 15); ws.set_column('H:H', 22)

                r_idx = 4
                for _, row in df_view.sort_values(by="대분류").iterrows():
                    ws.write(r_idx, 0, str(row.get('대분류', '')), val_fmt)
                    ws.write(r_idx, 1, str(row.get('설비명', '')), val_fmt)
                    ws.write(r_idx, 2, str(row.get('부위', '')), val_fmt)
                    ws.write(r_idx, 3, str(row.get('세척소독방법', '')), left_fmt)
                    ws.write(r_idx, 4, str(row.get('주기', '')), val_fmt)
                    ws.write(r_idx, 5, str(row.get('사용도구', '')), val_fmt)
                    ws.write(r_idx, 6, str(row.get('책임자', '')), val_fmt)

                    img_path = str(row.get('사진파일', ''))
                    ws.write(r_idx, 7, "", val_fmt)
                    
                    # [에러수정] 할당 안되던 버그 완벽 수정
                    real_img_path = img_path
                    if row.get('대분류') == "5. 식품 제조시설":
                        match = re.search(r'\[(.*?)\]', str(row.get('설비명', '')))
                        if match:
                            f_no = match.group(1)
                            real_img_path = f"fac_photo_{f_no}.png"
                            
                    if real_img_path and os.path.exists(real_img_path):
                        try: ws.insert_image(r_idx, 7, real_img_path, {'x_offset': 5, 'y_offset': 5, 'x_scale': 0.12, 'y_scale': 0.12, 'positioning': 1})
                        except Exception: pass
                        
                    ws.set_row(r_idx, 75)
                    r_idx += 1
            return output.getvalue()

        col_ex_btn1, col_ex_btn2 = st.columns([1, 3])
        with col_ex_btn1:
            if st.button("세척기준 텍스트 변경사항 저장", use_container_width=True):
                edited_clean.to_csv(CLEAN_FILE, index=False, encoding='utf-8-sig')
                st.success("데이터베이스에 정상적으로 반영되었습니다.")
                st.rerun()
        with col_ex_btn2:
            st.download_button(
                label="SSOP 세척소독기준표 엑셀 다운로드 (보고서 제출용)",
                data=export_cleaning_excel(df_clean),
                file_name=f"Hollys_SSOP_Specs_{date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.divider()
        st.markdown("### 📸 카테고리별 상세 정보 및 현장 사진 관리")
        
        tab_names = [c.split('. ')[1] for c in CLEAN_CATEGORIES]
        tabs = st.tabs(tab_names)

        for i, tab in enumerate(tabs):
            cat_name = CLEAN_CATEGORIES[i]
            with tab:
                df_cat = df_clean[df_clean['대분류'] == cat_name]
                if df_cat.empty:
                    st.info(f"등록된 '{cat_name}' 기준이 없습니다.")
                else:
                    for idx, row in df_cat.iterrows():
                        with st.container(border=True):
                            c_img, c_desc = st.columns([1, 3])
                            part_id = row['ID']
                            target_name = row['설비명']
                            part_name = row['부위']
                            img_path = row['사진파일']

                            with c_img:
                                if cat_name == "5. 식품 제조시설":
                                    match = re.search(r'\[(.*?)\]', target_name)
                                    if match:
                                        f_no = match.group(1)
                                        fac_img_path = f"fac_photo_{f_no}.png"
                                        if os.path.exists(fac_img_path):
                                            st.image(fac_img_path, use_container_width=True)
                                            st.caption("※ 제조위생설비이력관리 사진 자동연동됨")
                                        else:
                                            st.info("이력카드 사진 미등록")
                                    else:
                                        st.warning("설비 연동 실패 (직접 입력값)")
                                else:
                                    if img_path and os.path.exists(str(img_path)):
                                        st.image(str(img_path), use_container_width=True)
                                        if st.button("등록된 사진 삭제", key=f"del_img_{part_id}"):
                                            os.remove(img_path)
                                            df_clean.loc[df_clean['ID'] == part_id, '사진파일'] = ""
                                            df_clean.to_csv(CLEAN_FILE, index=False, encoding='utf-8-sig')
                                            st.rerun()
                                    else:
                                        up_f = st.file_uploader(f"사진 추가 등록 ({part_name})", key=f"up_{part_id}", type=['jpg','png'])
                                        if up_f:
                                            new_img_path = f"clean_photo_{part_id}.png"
                                            with open(new_img_path, "wb") as f:
                                                f.write(up_f.getbuffer())
                                            df_clean.loc[df_clean['ID'] == part_id, '사진파일'] = new_img_path
                                            df_clean.to_csv(CLEAN_FILE, index=False, encoding='utf-8-sig')
                                            st.rerun()

                            with c_desc:
                                st.markdown(f"<h5 style='color: #D11031;'>▶ 관리대상: {target_name} &nbsp;|&nbsp; 부위: {part_name}</h5>", unsafe_allow_html=True)
                                st.markdown(f"**🔹 세척·소독 방법:** \n{row['세척소독방법']}")
                                st.markdown(f"**🔹 주기:** {row['주기']} &nbsp;|&nbsp; **🔹 세제/도구:** {row['사용도구']} &nbsp;|&nbsp; **🔹 담당/책임:** {row['책임자']}")

    elif sub_menu == "계측기기 검교정":
        st.markdown('<div class="section-title">⚖️ 계측기기 검교정 대장 및 성적서 관리</div>', unsafe_allow_html=True)
        
        tab1, tab2 = st.tabs(["📋 계측기기 검교정 대장 (Master)", "📑 사내 검교정 성적서 발행"])
        
        with tab1:
            st.write("사내에서 관리 중인 모든 계측기기의 목록 및 차기 검교정 일자를 관리합니다.")

            st.markdown("""
            <style>
            .calibration-header {
                background-color: white;
                padding: 20px;
                border-radius: 10px;
                box-shadow: 0 4px 6px rgba(0,0,0,0.1);
                margin-bottom: 20px;
                text-align: center;
            }
            </style>
            """, unsafe_allow_html=True)
            
            with st.container():
                 st.markdown('<div class="calibration-header">', unsafe_allow_html=True)
                 if os.path.exists("image_4588df.png"): st.image("image_4588df.png", width=150)
                 st.markdown('</div>', unsafe_allow_html=True)

            df_calib = load_calib_list()
            
            c_filter1, c_filter2 = st.columns([1, 3])
            with c_filter1:
                calib_type_filter = st.selectbox("검교정 구분 필터", ["전체", "사내(자체)", "사외(의뢰)"])
            
            view_df_calib = df_calib.copy()
            if calib_type_filter != "전체":
                view_df_calib = view_df_calib[view_df_calib["구분"] == calib_type_filter]

            view_df_calib['주기'] = pd.to_numeric(view_df_calib['주기'], errors='coerce')
            
            view_df_calib['검교정일자'] = pd.to_datetime(view_df_calib['검교정일자'], errors='coerce').apply(lambda x: x.date() if pd.notna(x) else None)

            def calc_next_calib(row):
                if pd.notna(row['검교정일자']) and pd.notna(row['주기']):
                    try: return row['검교정일자'] + pd.DateOffset(months=int(row['주기']))
                    except: return pd.NaT
                return pd.NaT

            view_df_calib['차기_검교정일자'] = view_df_calib.apply(calc_next_calib, axis=1)
            view_df_calib['차기_검교정일자'] = pd.to_datetime(view_df_calib['차기_검교정일자'], errors='coerce').apply(lambda x: x.date() if pd.notna(x) else None)
                
            cfg_calib = {
                "구분": st.column_config.SelectboxColumn("구분", options=["사내(자체)", "사외(의뢰)"]),
                "주기": st.column_config.NumberColumn("주기(개월)", min_value=1, step=1, format="%d"),
                "검교정일자": st.column_config.DateColumn("검교정일자", format="YYYY-MM-DD"),
                "차기_검교정일자": st.column_config.DateColumn("차기 검교정일자 (자동)", format="YYYY-MM-DD", disabled=True)
            }
            
            edited_calib = st.data_editor(view_df_calib, num_rows="dynamic", use_container_width=True, column_config=cfg_calib)
            
            if st.button("검교정 대장 서버에 저장", key="save_calib_master"):
                edited_calib['주기'] = pd.to_numeric(edited_calib['주기'], errors='coerce')
                edited_calib['검교정일자'] = pd.to_datetime(edited_calib['검교정일자'], errors='coerce')
                edited_calib['차기_검교정일자'] = edited_calib.apply(calc_next_calib, axis=1)

                edited_calib['검교정일자'] = edited_calib['검교정일자'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) else "")
                edited_calib['차기_검교정일자'] = edited_calib['차기_검교정일자'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) else "")
                edited_calib['주기'] = edited_calib['주기'].apply(lambda x: str(int(x)) if pd.notna(x) else "")

                if calib_type_filter == "전체":
                    edited_calib.to_csv(CALIB_LIST_FILE, index=False, encoding='utf-8-sig')
                else:
                    other_df = df_calib[df_calib["구분"] != calib_type_filter]
                    final_df = pd.concat([other_df, edited_calib], ignore_index=True)
                    final_df.to_csv(CALIB_LIST_FILE, index=False, encoding='utf-8-sig')
                
                st.success("계측기기 검교정 대장이 성공적으로 저장되었습니다.")
                st.rerun()
                
        with tab2:
            st.write("자체 검교정 진행 시 성적서 데이터를 기록하고 보관합니다.")
            df_calib_rep = load_calib_reports()
            
            df_calib_rep['교정일자'] = pd.to_datetime(df_calib_rep['교정일자'], errors='coerce').apply(lambda x: x.date() if pd.notna(x) else None)
            
            cfg_calib_rep = {
                "교정일자": st.column_config.DateColumn("교정일자", format="YYYY-MM-DD")
            }
            
            edited_calib_rep = st.data_editor(df_calib_rep, num_rows="dynamic", use_container_width=True, column_config=cfg_calib_rep)
            
            if st.button("자체 성적서 데이터 저장"):
                edited_calib_rep['교정일자'] = edited_calib_rep['교정일자'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) else "")
                edited_calib_rep.to_csv(CALIB_REPORT_FILE, index=False, encoding='utf-8-sig')
                st.success("성적서 기록이 안전하게 저장되었습니다.")
                st.rerun()

# --- 원·부자재 관리 메뉴 ---
elif menu_selection == "원·부자재 관리":

    # ── 파일 로드 헬퍼 ──────────────────────────────────────
    RM_COLS = ["원자재코드","원자재명","유형","규격","원산지","제조원","판매원","검사주기","비고",
               "최소_수분","최대_수분","최소_밀도","최대_밀도","관능_기준"]
    def load_rm():
        if os.path.exists(RM_FILE):
            try:
                df = pd.read_csv(RM_FILE, dtype=str)
                if "포장단위" in df.columns and "규격" not in df.columns:
                    df = df.rename(columns={"포장단위": "규격"})
                for c in RM_COLS:
                    if c not in df.columns: df[c] = ""
                return df[RM_COLS]
            except Exception: pass
        return pd.DataFrame(columns=RM_COLS)

    def rm_spec_path(code, kind):
        os.makedirs(RM_SPEC_DIR, exist_ok=True)
        safe = str(code).replace("/","_").replace("\\","_")
        return os.path.join(RM_SPEC_DIR, f"rm_{kind}_{safe}.csv")

    df_rm = load_rm()

    # ════════════════════════════════════════════════════
    # 1. 원자재 등록
    # ════════════════════════════════════════════════════
    if sub_menu == "원·부자재 등록":
        st.markdown('<div class="section-title">🌿 원·부자재 등록</div>', unsafe_allow_html=True)
        st.write("원·부자재 간편 규격을 등록합니다. 상세 규격은 규격서 마스터에서 관리하세요.")

        # 등록 폼
        with st.expander("➕ 새 원·부자재 등록", expanded=False):
            with st.form("rm_reg_form"):
                r1, r2, r3 = st.columns(3)
                with r1:
                    rm_code = st.text_input("원자재코드 *", placeholder="예: RM-001")
                    rm_name = st.text_input("원자재명 *",   placeholder="예: 커피콩_에티오피아")
                    rm_type = st.selectbox("유형", ["생두(원료)","스틱(원료)","포장재","부자재","첨가물","기타"])
                with r2:
                    rm_unit    = st.text_input("규격",  placeholder="예: 60kg")
                    rm_origin  = st.text_input("원산지",    placeholder="예: 에티오피아")
                    rm_maker   = st.text_input("제조원",    placeholder="예: -")
                    rm_seller  = st.text_input("판매원",    placeholder="예: ㈜케이지할리스에프앤비")
                with r3:
                    rm_cycle   = st.text_input("검사주기",  placeholder="예: 입고시")
                    rm_note    = st.text_input("비고",      placeholder="")
                st.markdown("**품질 규격 (간편)**")
                q1, q2, q3 = st.columns(3)
                with q1:
                    rm_moi_min = st.text_input("수분 최소 (%)", placeholder="예: 9.0")
                    rm_moi_max = st.text_input("수분 최대 (%)", placeholder="예: 12.0")
                with q2:
                    rm_den_min = st.text_input("밀도 최소",     placeholder="예: 670")
                    rm_den_max = st.text_input("밀도 최대",     placeholder="예: 750")
                with q3:
                    rm_organ   = st.text_input("관능 기준",     placeholder="예: 고유의 향미, 이미이취 없을 것")
                if st.form_submit_button("등록", type="primary"):
                    if not rm_code or not rm_name:
                        st.error("원자재코드와 원자재명은 필수입니다.")
                    elif rm_code in df_rm["원자재코드"].values:
                        st.error(f"코드 [{rm_code}]가 이미 존재합니다.")
                    else:
                        new_row = pd.DataFrame([[rm_code,rm_name,rm_type,rm_unit,rm_origin,rm_maker,rm_seller,
                                                  rm_cycle,rm_note,rm_moi_min,rm_moi_max,rm_den_min,rm_den_max,rm_organ]],
                                               columns=RM_COLS)
                        df_rm = pd.concat([df_rm, new_row], ignore_index=True)
                        df_rm.to_csv(RM_FILE, index=False, encoding='utf-8-sig')
                        st.success(f"[{rm_code}] {rm_name} 등록 완료!")
                        st.rerun()

        st.write("")
        if df_rm.empty:
            st.info("등록된 원·부자재가 없습니다.")
        else:
            st.markdown("**등록된 원·부자재 목록**")
            TH = "background:#D6E4F0;color:#1F4E79;font-weight:bold;padding:8px 6px;text-align:center;border:1px solid #b0c4d8;font-size:11px;"
            TD = "background:#fff;padding:7px 8px;border:1px solid #d0d7de;vertical-align:middle;font-size:11px;text-align:center;word-break:break-word;"
            html_rm = f'''<style>.rm-tbl{{width:100%;border-collapse:collapse;}}.rm-tbl tr:hover td{{background:#f5f8ff!important;}}</style>
<table class="rm-tbl"><thead><tr>
  <th style="{TH}width:80px;">코드</th>
  <th style="{TH}width:150px;">원자재명</th>
  <th style="{TH}width:70px;">유형</th>
  <th style="{TH}width:70px;">규격</th>
  <th style="{TH}width:80px;">원산지</th>
  <th style="{TH}width:70px;">검사주기</th>
  <th style="{TH}width:100px;">수분(%)</th>
  <th style="{TH}width:100px;">밀도</th>
  <th style="{TH}min-width:160px;">관능기준</th>
</tr></thead><tbody>'''
            for ri, row in df_rm.iterrows():
                moi_str = f'{row.get("최소_수분","") or ""} ~ {row.get("최대_수분","") or ""}' if (str(row.get("최소_수분","")).strip() or str(row.get("최대_수분","")).strip()) else "-"
                den_str = f'{row.get("최소_밀도","") or ""} ~ {row.get("최대_밀도","") or ""}' if (str(row.get("최소_밀도","")).strip() or str(row.get("최대_밀도","")).strip()) else (str(row.get("밀도_기준","")) or "-")
                html_rm += f'''<tr>
  <td style="{TD}font-weight:bold;color:#D11031;">{row["원자재코드"]}</td>
  <td style="{TD}text-align:left;font-weight:bold;">{row["원자재명"]}</td>
  <td style="{TD}">{row["유형"]}</td>
  <td style="{TD}">{row["규격"]}</td>
  <td style="{TD}">{row["원산지"]}</td>
  <td style="{TD}">{row["검사주기"]}</td>
  <td style="{TD}">{moi_str}</td>
  <td style="{TD}">{den_str}</td>
  <td style="{TD}text-align:left;">{row["관능_기준"]}</td>
</tr>'''
            html_rm += "</tbody></table>"
            st.markdown(html_rm, unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)

            # 수정/삭제
            with st.expander("✏️ 수정 / 🗑️ 삭제"):
                edited_rm = st.data_editor(df_rm, num_rows="dynamic", use_container_width=True, key="rm_editor")
                if st.button("변경사항 저장", type="primary"):
                    edited_rm.to_csv(RM_FILE, index=False, encoding='utf-8-sig')
                    st.success("저장 완료!")
                    st.rerun()

    # ════════════════════════════════════════════════════
    # 2. 원·부자재 규격서 마스터
    # ════════════════════════════════════════════════════
    elif sub_menu == "원·부자재 규격서 마스터":
        st.markdown('<div class="section-title">📋 원·부자재 규격서 마스터</div>', unsafe_allow_html=True)

        # session_state 초기화
        if "rm_sel_code" not in st.session_state: st.session_state.rm_sel_code = None
        if "rm_sel_name" not in st.session_state: st.session_state.rm_sel_name = None

        if st.session_state.rm_sel_code is None:
            # ── 목록 화면 ──────────────────────────────────
            st.write("상세 규격을 열람·수정할 원·부자재를 선택하세요.")
            search_rm = st.text_input("🔍 검색 (이름 또는 코드)", placeholder="원자재명 또는 코드 입력...")
            if df_rm.empty:
                st.info("등록된 원·부자재가 없습니다. '원자재 등록' 메뉴에서 먼저 등록하세요.")
            else:
                view_rm = df_rm[
                    df_rm["원자재명"].str.contains(search_rm, case=False, na=False) |
                    df_rm["원자재코드"].str.contains(search_rm, case=False, na=False)
                ] if search_rm else df_rm
                if view_rm.empty:
                    st.warning("일치하는 항목이 없습니다.")
                else:
                    cols_rm = st.columns(4)
                    for idx, row in view_rm.reset_index().iterrows():
                        with cols_rm[idx % 4]:
                            with st.container(border=True):
                                st.markdown(f"<h5 style='color:#D11031;margin-bottom:0;'>{row['원자재코드']}</h5>", unsafe_allow_html=True)
                                st.markdown(f"**{row['원자재명']}**")
                                st.caption(f"분류: {row['유형']}")
                                if st.button("규격서 열기 📝", key=f"rm_open_{row['원자재코드']}", use_container_width=True):
                                    st.session_state.rm_sel_code = row["원자재코드"]
                                    st.session_state.rm_sel_name = row["원자재명"]
                                    st.rerun()
        else:
            # ── 상세 규격서 화면 ────────────────────────────
            sel_code = st.session_state.rm_sel_code
            sel_name = st.session_state.rm_sel_name
            safe_c   = str(sel_code).replace("/","_").replace("\\","_")

            RM_BASIC_F  = rm_spec_path(safe_c, "basic")
            RM_SPEC_F   = rm_spec_path(safe_c, "spec")
            RM_SUPPLY_F = rm_spec_path(safe_c, "supply")
            RM_IMG_F    = os.path.join(RM_SPEC_DIR, f"rm_photo_{safe_c}.png")

            def load_rm_csv(fpath, cols):
                if os.path.exists(fpath):
                    try: return pd.read_csv(fpath, dtype=str)
                    except Exception: pass
                return pd.DataFrame(columns=cols)

            df_basic   = load_rm_csv(RM_BASIC_F,  ["항목","내용"])
            df_spec    = load_rm_csv(RM_SPEC_F,   ["검사항목","구분(법적/자가/사내)","규격","검사방법","검사주기","비고"])
            df_supply  = load_rm_csv(RM_SUPPLY_F, ["원료명","함량(%)","원산지","제조원","GMO/알러지"])

            # ── 원자재 등록 데이터로 초기값 자동 불러오기 ──
            rm_row = df_rm[df_rm["원자재코드"]==sel_code]
            if df_basic.empty and not rm_row.empty:
                r0 = rm_row.iloc[0]
                df_basic = pd.DataFrame([
                    ["유형",            r0.get("유형","")],
                    ["포장단위",        r0.get("규격","")],
                    ["원산지",          r0.get("원산지","")],
                    ["포장재질",        ""],
                    ["제조원",          r0.get("제조원","")],
                    ["판매원",          r0.get("판매원","")],
                    ["연락처",          ""],
                    ["소비기한/제조일자",""],
                    ["검사주기",        r0.get("검사주기","")],
                ], columns=["항목","내용"])
            if df_spec.empty and not rm_row.empty:
                r0 = rm_row.iloc[0]
                moi_rng = f'{r0.get("최소_수분","") or ""} ~ {r0.get("최대_수분","") or ""}'.strip(" ~") or ""
                den_rng = f'{r0.get("최소_밀도","") or ""} ~ {r0.get("최대_밀도","") or ""}'.strip(" ~") or ""
                org_val = r0.get("관능_기준","") or ""
                df_spec = pd.DataFrame([
                    ["성상",     "자가", org_val if org_val else "고유의 향미를 가지고 이미,이취가 없어야 한다","관능검사","입고시",""],
                    ["이물",     "자가", "불검출",   "육안검사","입고시",""],
                    ["포장상태", "사내", "뜯어지거나 오염된 흔적이 없을 것","육안검사","입고시",""],
                    ["수분",     "사내", moi_rng,   "수분밀도계측정","입고시",""],
                    ["밀도",     "사내", den_rng,   "수분밀도계측정","입고시",""],
                ], columns=["검사항목","구분(법적/자가/사내)","규격","검사방법","검사주기","비고"])

            # 뒤로가기
            hd1, hd2 = st.columns([1, 5])
            with hd1:
                if st.button("← 목록으로", use_container_width=True):
                    st.session_state.rm_sel_code = None
                    st.session_state.rm_sel_name = None
                    st.rerun()
            st.markdown(f"### 📄 원·부자재 규격서 — **{sel_name}** <span style='color:#888;font-size:14px;'>({sel_code})</span>", unsafe_allow_html=True)
            st.divider()

            tab_view, tab_basic, tab_spec, tab_supply, tab_img = st.tabs(
                ["📊 규격서 보기", "1. 기본정보", "2. 품질규격", "3. 원료정보", "4. 사진"])

            # ══════════════════════════════════════════
            # 탭1: 규격서 보기 + 엑셀 다운로드
            # ══════════════════════════════════════════
            with tab_view:
                TH2 = "background:#D6E4F0;color:#1F4E79;font-weight:bold;padding:8px 10px;border:1px solid #b0c4d8;font-size:11px;text-align:center;"
                TD2 = "background:#fff;padding:8px 10px;border:1px solid #d0d7de;vertical-align:top;font-size:11px;word-break:break-word;"
                TDG = "background:#EBF3FB;padding:8px 10px;border:1px solid #d0d7de;vertical-align:middle;font-size:11px;font-weight:bold;color:#1F4E79;"
                SEC = "background:#1F4E79;color:white;font-weight:bold;padding:6px 14px;font-size:12px;border-radius:4px;margin:12px 0 6px;"

                img_col, info_col = st.columns([1, 2])
                with img_col:
                    if os.path.exists(RM_IMG_F):
                        st.image(RM_IMG_F, caption="원자재 사진", use_container_width=True)
                    else:
                        st.markdown("<div style='border:2px dashed #ccc;height:160px;display:flex;align-items:center;justify-content:center;color:#aaa;border-radius:8px;font-size:13px;'>사진 없음<br>(4. 사진 탭에서 등록)</div>", unsafe_allow_html=True)
                with info_col:
                    st.markdown(f"<div style='{SEC}'>■ 기본 정보</div>", unsafe_allow_html=True)
                    html_b = '<table style="width:100%;border-collapse:collapse;">'
                    for _, br in df_basic.iterrows():
                        html_b += f'<tr><td style="{TDG}width:120px;">{br["항목"]}</td><td style="{TD2}">{br["내용"] if pd.notna(br["내용"]) else ""}</td></tr>'
                    html_b += "</table>"
                    st.markdown(html_b, unsafe_allow_html=True)

                st.markdown(f"<div style='{SEC}'>■ 품질 규격</div>", unsafe_allow_html=True)
                if not df_spec.empty:
                    html_s = '<table style="width:100%;border-collapse:collapse;"><thead><tr>'
                    for col in df_spec.columns:
                        html_s += f'<th style="{TH2}">{col}</th>'
                    html_s += "</tr></thead><tbody>"
                    for _, sr in df_spec.iterrows():
                        html_s += "<tr>"
                        for ci2, v in enumerate(sr):
                            bg2 = "#EBF3FB" if ci2==0 else "#fff"
                            fw2 = "bold" if ci2==0 else "normal"
                            html_s += f'<td style="{TD2}background:{bg2};font-weight:{fw2};">{v if pd.notna(v) else ""}</td>'
                        html_s += "</tr>"
                    html_s += "</tbody></table>"
                    st.markdown(html_s, unsafe_allow_html=True)

                if not df_supply.empty:
                    st.markdown(f"<div style='{SEC}'>■ 원료 정보</div>", unsafe_allow_html=True)
                    html_sup = '<table style="width:100%;border-collapse:collapse;"><thead><tr>'
                    for col in df_supply.columns:
                        html_sup += f'<th style="{TH2}">{col}</th>'
                    html_sup += "</tr></thead><tbody>"
                    for _, sr in df_supply.iterrows():
                        html_sup += "<tr>"
                        for v in sr:
                            html_sup += f'<td style="{TD2}">{v if pd.notna(v) else ""}</td>'
                        html_sup += "</tr>"
                    html_sup += "</tbody></table>"
                    st.markdown(html_sup, unsafe_allow_html=True)

                st.markdown("---")
                # ── 엑셀 출력 (보관용 최적화) ──────────────
                def export_rm_excel(code, name, df_b2, df_s2, df_sup2, img_p):
                    import io as _io2
                    import openpyxl as _opx
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter as gcl
                    wb2 = _opx.Workbook(); ws2 = wb2.active; ws2.title="원부자재규격서"
                    ws2.page_setup.paperSize = 9   # A4
                    ws2.page_setup.fitToWidth = 1
                    ws2.page_setup.fitToPage = True
                    ws2.print_options.horizontalCentered = True
                    ws2.page_margins.left  = 0.5; ws2.page_margins.right  = 0.5
                    ws2.page_margins.top   = 0.8; ws2.page_margins.bottom = 0.8

                    thin  = Side(style="thin",  color="B0C4D8")
                    med   = Side(style="medium", color="1F4E79")
                    bd    = Border(left=thin, right=thin, top=thin, bottom=thin)
                    bdm   = Border(left=med,  right=med,  top=med,  bottom=med)
                    hfill = PatternFill("solid", start_color="D6E4F0")
                    sfill = PatternFill("solid", start_color="D6E4F0")
                    tfill = PatternFill("solid", start_color="1F4E79")
                    hfont = Font(name="맑은 고딕", bold=True, color="1F4E79",  size=10)
                    sfont = Font(name="맑은 고딕", bold=True, color="1F4E79",  size=11)
                    tfont = Font(name="맑은 고딕", bold=True, color="FFFFFF",  size=14)
                    nfont = Font(name="맑은 고딕", size=10)
                    nalign = Alignment(horizontal="left",   vertical="center", wrap_text=True)
                    calign = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    # 컬럼 폭  A~F
                    for ci, w in enumerate([20, 28, 15, 28, 15, 15], 1):
                        ws2.column_dimensions[gcl(ci)].width = w

                    cur = 1
                    # ── 타이틀 ──
                    ws2.merge_cells(f"A{cur}:F{cur}")
                    ws2[f"A{cur}"] = f"원·부자재 규격서  |  {name}  ({code})"
                    ws2[f"A{cur}"].font = tfont; ws2[f"A{cur}"].fill = tfill
                    ws2[f"A{cur}"].alignment = calign; ws2.row_dimensions[cur].height = 36; cur += 1

                    # ── 기본정보 + 사진 side-by-side ──
                    ws2.merge_cells(f"A{cur}:D{cur}")
                    ws2[f"A{cur}"] = "■ 기본 정보"; ws2[f"A{cur}"].font = sfont
                    ws2[f"A{cur}"].fill = sfill; ws2[f"A{cur}"].alignment = calign
                    ws2.merge_cells(f"E{cur}:F{cur}")
                    ws2[f"E{cur}"] = "■ 제품 사진"; ws2[f"E{cur}"].font = sfont
                    ws2[f"E{cur}"].fill = sfill; ws2[f"E{cur}"].alignment = calign
                    ws2.row_dimensions[cur].height = 20; cur += 1

                    basic_start = cur
                    for _, br in df_b2.iterrows():
                        ws2[f"A{cur}"] = br["항목"]; ws2[f"A{cur}"].font = hfont
                        ws2[f"A{cur}"].fill = hfill; ws2[f"A{cur}"].border = bd; ws2[f"A{cur}"].alignment = calign
                        ws2.merge_cells(f"B{cur}:D{cur}")
                        ws2[f"B{cur}"] = str(br["내용"]) if pd.notna(br["내용"]) else ""
                        ws2[f"B{cur}"].font = nfont; ws2[f"B{cur}"].border = bd; ws2[f"B{cur}"].alignment = nalign
                        ws2.row_dimensions[cur].height = 20; cur += 1
                    basic_end = cur - 1

                    # 사진 영역 (E열 병합)
                    if basic_end >= basic_start:
                        ws2.merge_cells(f"E{basic_start}:F{basic_end}")
                        ws2[f"E{basic_start}"].border = bdm; ws2[f"E{basic_start}"].alignment = calign
                    if img_p and os.path.exists(img_p):
                        try:
                            from openpyxl.drawing.image import Image as OPXImg
                            img_o = OPXImg(img_p)
                            # 사진 셀 크기에 맞게 조정
                            row_h_px = 20 * (basic_end - basic_start + 1) * 1.33
                            img_o.width = 130; img_o.height = min(int(row_h_px), 200)
                            ws2.add_image(img_o, f"E{basic_start}")
                        except Exception: pass
                    cur += 1

                    # ── 품질규격 ──
                    ws2.merge_cells(f"A{cur}:F{cur}")
                    ws2[f"A{cur}"] = "■ 품질 규격"; ws2[f"A{cur}"].font = sfont
                    ws2[f"A{cur}"].fill = sfill; ws2[f"A{cur}"].alignment = calign
                    ws2.row_dimensions[cur].height = 20; cur += 1
                    spec_col_w = [2,1,2,2,1,1]  # A~F 상대비중
                    for ci2, col in enumerate(df_s2.columns, 1):
                        c = ws2.cell(row=cur, column=ci2, value=col)
                        c.font = hfont; c.fill = hfill; c.border = bd; c.alignment = calign
                    ws2.row_dimensions[cur].height = 20; cur += 1
                    for _, sr in df_s2.iterrows():
                        row_h = 20
                        for ci2, val in enumerate(sr, 1):
                            v_str = str(val) if pd.notna(val) else ""
                            if len(v_str) > 30: row_h = 35
                            c = ws2.cell(row=cur, column=ci2, value=v_str)
                            c.font = nfont; c.border = bd
                            c.alignment = calign if ci2 in (2,5,6) else nalign
                            if ci2 == 1: c.fill = hfill; c.font = hfont
                        ws2.row_dimensions[cur].height = row_h; cur += 1
                    cur += 1

                    # ── 원료정보 ──
                    if not df_sup2.empty:
                        ws2.merge_cells(f"A{cur}:F{cur}")
                        ws2[f"A{cur}"] = "■ 원료 정보"; ws2[f"A{cur}"].font = sfont
                        ws2[f"A{cur}"].fill = sfill; ws2[f"A{cur}"].alignment = calign
                        ws2.row_dimensions[cur].height = 20; cur += 1
                        for ci2, col in enumerate(df_sup2.columns, 1):
                            c = ws2.cell(row=cur, column=ci2, value=col)
                            c.font = hfont; c.fill = hfill; c.border = bd; c.alignment = calign
                        ws2.row_dimensions[cur].height = 20; cur += 1
                        for _, sr in df_sup2.iterrows():
                            for ci2, val in enumerate(sr, 1):
                                c = ws2.cell(row=cur, column=ci2, value=str(val) if pd.notna(val) else "")
                                c.font = nfont; c.border = bd; c.alignment = nalign
                            ws2.row_dimensions[cur].height = 20; cur += 1

                    out = _io2.BytesIO(); wb2.save(out); return out.getvalue()

                xl_data = export_rm_excel(sel_code, sel_name, df_basic, df_spec, df_supply, RM_IMG_F)
                st.download_button("📥 엑셀 다운로드 (보관용)",
                                   data=xl_data,
                                   file_name=f"원부자재규격서_{sel_name}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   type="primary")

            # ══════════════════════════════════════════
            # 다른 원·부자재 규격서 불러오기 공통 함수
            # ══════════════════════════════════════════
            # 현재 편집 중 코드를 제외한 다른 원자재 목록
            other_rm = df_rm[df_rm["원자재코드"] != sel_code]
            other_opts = ["선택 안 함"] + [
                f"{r['원자재코드']} — {r['원자재명']}"
                for _, r in other_rm.iterrows()
                if os.path.exists(rm_spec_path(str(r["원자재코드"]).replace("/","_").replace("\\","_"), "basic"))
                or os.path.exists(rm_spec_path(str(r["원자재코드"]).replace("/","_").replace("\\","_"), "spec"))
            ]

            # ══════════════════════════════════════════
            # 탭2: 기본정보 편집
            # ══════════════════════════════════════════
            with tab_basic:
                with st.expander("📥 다른 원·부자재 규격서에서 기본정보 불러오기", expanded=False):
                    if len(other_opts) > 1:
                        copy_src_b = st.selectbox("불러올 원자재 선택", other_opts, key="rm_copy_src_basic")
                        if copy_src_b != "선택 안 함":
                            src_code_b = copy_src_b.split(" — ")[0].strip()
                            safe_src_b = str(src_code_b).replace("/","_").replace("\\","_")
                            src_basic_f = rm_spec_path(safe_src_b, "basic")
                            if os.path.exists(src_basic_f):
                                preview_b = pd.read_csv(src_basic_f, dtype=str)
                                st.dataframe(preview_b, use_container_width=True, height=200)
                                if st.button("이 규격서의 기본정보 불러오기", type="primary", key="rm_do_copy_basic"):
                                    preview_b.to_csv(RM_BASIC_F, index=False, encoding='utf-8-sig')
                                    st.success(f"[{src_code_b}]의 기본정보를 불러왔습니다!"); st.rerun()
                            else:
                                st.warning("해당 원자재의 기본정보가 없습니다.")
                    else:
                        st.info("저장된 규격서가 있는 다른 원·부자재가 없습니다.")
                st.caption("항목명과 내용을 직접 편집하거나 행을 추가할 수 있습니다.")
                edited_basic = st.data_editor(df_basic, num_rows="dynamic", use_container_width=True, key="rm_basic_ed")
                if st.button("기본정보 저장", type="primary", key="rm_save_basic"):
                    edited_basic.to_csv(RM_BASIC_F, index=False, encoding='utf-8-sig')
                    st.success("저장 완료!"); st.rerun()

            # ══════════════════════════════════════════
            # 탭3: 품질규격 편집
            # ══════════════════════════════════════════
            with tab_spec:
                with st.expander("📥 다른 원·부자재 규격서에서 품질규격 불러오기", expanded=False):
                    if len(other_opts) > 1:
                        copy_src_s = st.selectbox("불러올 원자재 선택", other_opts, key="rm_copy_src_spec")
                        if copy_src_s != "선택 안 함":
                            src_code_s = copy_src_s.split(" — ")[0].strip()
                            safe_src_s = str(src_code_s).replace("/","_").replace("\\","_")
                            src_spec_f = rm_spec_path(safe_src_s, "spec")
                            if os.path.exists(src_spec_f):
                                preview_s = pd.read_csv(src_spec_f, dtype=str)
                                st.dataframe(preview_s, use_container_width=True, height=200)
                                if st.button("이 규격서의 품질규격 불러오기", type="primary", key="rm_do_copy_spec"):
                                    preview_s.to_csv(RM_SPEC_F, index=False, encoding='utf-8-sig')
                                    st.success(f"[{src_code_s}]의 품질규격을 불러왔습니다!"); st.rerun()
                            else:
                                st.warning("해당 원자재의 품질규격이 없습니다.")
                    else:
                        st.info("저장된 규격서가 있는 다른 원·부자재가 없습니다.")
                st.caption("항목을 추가·수정하세요. 법적/자가/사내 구분을 선택하고 규격 및 검사방법을 입력합니다.")
                edited_spec = st.data_editor(df_spec, num_rows="dynamic", use_container_width=True, key="rm_spec_ed",
                    column_config={
                        "구분(법적/자가/사내)": st.column_config.SelectboxColumn("구분", options=["법적","자가","사내"], width="small"),
                        "검사주기": st.column_config.SelectboxColumn("검사주기", options=["입고시","1회/년","분기별","반기별","월 1회","기타"]),
                    })
                if st.button("품질규격 저장", type="primary", key="rm_save_spec"):
                    edited_spec.to_csv(RM_SPEC_F, index=False, encoding='utf-8-sig')
                    st.success("저장 완료!"); st.rerun()

            # ══════════════════════════════════════════
            # 탭4: 원료정보 편집
            # ══════════════════════════════════════════
            with tab_supply:
                with st.expander("📥 다른 원·부자재 규격서에서 원료정보 불러오기", expanded=False):
                    other_sup_opts = ["선택 안 함"] + [
                        f"{r['원자재코드']} — {r['원자재명']}"
                        for _, r in other_rm.iterrows()
                        if os.path.exists(rm_spec_path(str(r["원자재코드"]).replace("/","_").replace("\\","_"), "supply"))
                    ]
                    if len(other_sup_opts) > 1:
                        copy_src_sup = st.selectbox("불러올 원자재 선택", other_sup_opts, key="rm_copy_src_supply")
                        if copy_src_sup != "선택 안 함":
                            src_code_sup = copy_src_sup.split(" — ")[0].strip()
                            safe_src_sup = str(src_code_sup).replace("/","_").replace("\\","_")
                            src_sup_f = rm_spec_path(safe_src_sup, "supply")
                            if os.path.exists(src_sup_f):
                                preview_sup = pd.read_csv(src_sup_f, dtype=str)
                                st.dataframe(preview_sup, use_container_width=True, height=180)
                                if st.button("이 규격서의 원료정보 불러오기", type="primary", key="rm_do_copy_supply"):
                                    preview_sup.to_csv(RM_SUPPLY_F, index=False, encoding='utf-8-sig')
                                    st.success(f"[{src_code_sup}]의 원료정보를 불러왔습니다!"); st.rerun()
                            else:
                                st.warning("해당 원자재의 원료정보가 없습니다.")
                    else:
                        st.info("저장된 원료정보가 있는 다른 원·부자재가 없습니다.")
                st.caption("원료명, 함량, 원산지, GMO·알러지 정보를 입력합니다.")
                edited_supply = st.data_editor(df_supply, num_rows="dynamic", use_container_width=True, key="rm_supply_ed")
                if st.button("원료정보 저장", type="primary", key="rm_save_supply"):
                    edited_supply.to_csv(RM_SUPPLY_F, index=False, encoding='utf-8-sig')
                    st.success("저장 완료!"); st.rerun()

            # ══════════════════════════════════════════
            # 탭5: 사진 업로드
            # ══════════════════════════════════════════
            with tab_img:
                st.caption("원자재 사진을 등록하면 규격서 보기 및 엑셀 출력에 자동 포함됩니다.")
                up_img = st.file_uploader("이미지 업로드 (JPG/PNG)", type=["jpg","jpeg","png"], key="rm_img_up")
                if up_img:
                    os.makedirs(RM_SPEC_DIR, exist_ok=True)
                    with open(RM_IMG_F, "wb") as f2: f2.write(up_img.read())
                    st.success("사진 저장 완료!"); st.rerun()
                if os.path.exists(RM_IMG_F):
                    ic1, ic2 = st.columns([1,3])
                    with ic1:
                        st.image(RM_IMG_F, caption="현재 등록된 사진", use_container_width=True)
                        if st.button("🗑️ 사진 삭제", key="rm_del_img"):
                            os.remove(RM_IMG_F); st.rerun()
                else:
                    st.info("등록된 사진이 없습니다.")


# --- 재고 관리 메뉴 ---
elif menu_selection == "재고 관리":
    if sub_menu == "현재고 현황 및 조정":
        st.markdown('<div class="section-title">📦 제품별 실시간 재고 현황 (영점 조정)</div>', unsafe_allow_html=True)
        st.write("데이터 히스토리에 입력된 `생산량`과 출고 관리에 등록된 `출고량`이 자동 합산되어 현재고가 실시간으로 노출됩니다.")
        
        df_data = load_data()
        df_out = load_outbound_records()
        df_adj = load_inventory_adj()
        
        inv_dict = {}
        # 1. 생산량 누적
        for _, row in df_data.iterrows():
            if str(row.get("제품명", "")).strip() in ["", "-", "None", "nan"]: continue
            key = (str(row.get("유형", "None")).strip(), str(row.get("제품명", "")).strip(), str(row.get("규격", "None")).strip())
            if key not in inv_dict: inv_dict[key] = {"총생산량": 0, "총출고량": 0, "조정수량": 0}
            try: qty = int(float(str(row.get("생산량", "0")).replace(",","")))
            except: qty = 0
            inv_dict[key]["총생산량"] += qty
            
        # 2. 출고량 차감
        for _, row in df_out.iterrows():
            if str(row.get("제품명", "")).strip() in ["", "-", "None", "nan"]: continue
            key = (str(row.get("유형", "None")).strip(), str(row.get("제품명", "")).strip(), str(row.get("규격", "None")).strip())
            if key not in inv_dict: inv_dict[key] = {"총생산량": 0, "총출고량": 0, "조정수량": 0}
            try: qty = int(float(str(row.get("수량", "0")).replace(",","")))
            except: qty = 0
            inv_dict[key]["총출고량"] += qty
            
        # 3. 누적 임의 조정치 적용
        for _, row in df_adj.iterrows():
            key = (str(row.get("유형", "None")).strip(), str(row.get("제품명", "")).strip(), str(row.get("규격", "None")).strip())
            if key not in inv_dict: inv_dict[key] = {"총생산량": 0, "총출고량": 0, "조정수량": 0}
            try: diff = int(float(str(row.get("차이", "0")).replace(",","")))
            except: diff = 0
            inv_dict[key]["조정수량"] += diff
            
        inv_list = []
        for k, v in inv_dict.items():
            curr_qty = v["총생산량"] - v["총출고량"] + v["조정수량"]
            inv_list.append({
                "유형": k[0] if k[0] not in ["", "nan"] else "-",
                "제품명": k[1],
                "규격": k[2] if k[2] not in ["", "nan"] else "-",
                "총생산량": v["총생산량"],
                "총출고량": v["총출고량"],
                "임의조정 누적분": v["조정수량"],
                "현재고 (P)": curr_qty
            })
            
        df_inv = pd.DataFrame(inv_list)
        if df_inv.empty:
            st.info("재고를 산정할 데이터(생산 또는 출고량)가 존재하지 않습니다.")
        else:
            df_inv = df_inv.sort_values(by=["유형", "제품명"]).reset_index(drop=True)
            st.dataframe(
                df_inv, 
                use_container_width=True, 
                hide_index=True,
                column_config={
                    "총생산량": st.column_config.NumberColumn(format="%d"),
                    "총출고량": st.column_config.NumberColumn(format="%d"),
                    "임의조정 누적분": st.column_config.NumberColumn(format="%d"),
                    "현재고 (P)": st.column_config.NumberColumn("현재고 (최종)", format="%d", help="생산량 - 출고량 + 임의조정 누적분")
                }
            )
            
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("##### 🚨 재고 강제 영점 / 수정 패널")
            st.error("주의사항: 전산 기록상 재고와 실제 물류창고 재고가 일치하지 않을 때, 전산상 숫자를 수정합니다. 관리자의 명백한 확인 후에만 사용 바랍니다.")
            
            with st.form("adj_inv_form", clear_on_submit=True):
                options = df_inv.apply(lambda r: f"[{r['유형']}] {r['제품명']} ({r['규격']}) — 전산재고: {r['현재고 (P)']}", axis=1).tolist()
                sel_item = st.selectbox("조정(수정)할 품목 지정", ["선택해주세요"] + options)
                
                c0, c1, c2 = st.columns([1,1,2])
                with c1:
                    actual_qty = st.number_input("조정 후 최종 실제재고 입력", min_value=0, step=1, value=0)
                with c2:
                    adj_reason = st.text_input("조정 사유", placeholder="예: 재고실사 불일치 (파손 2건), 누락 등")
                
                confirm_check = st.checkbox("해당 데이터 위변조/수정에 동의하며, 위 사유로 재고를 수정하겠습니다.")
                submitted = st.form_submit_button("🔁 현재고 즉시 수정(저장)")
                
                if submitted:
                    if sel_item == "선택해주세요":
                        st.error("조정 폼목을 올바르게 선택해주세요.")
                    elif not adj_reason.strip():
                        st.error("비고/사유는 필수 입력 사항입니다.")
                    elif not confirm_check:
                        st.error("재고 수정 사항에 동의함 체커에 체크해주셔야 수정이 가능합니다.")
                    else:
                        match_idx = options.index(sel_item)
                        target_row = df_inv.iloc[match_idx]
                        old_qty = target_row["현재고 (P)"]
                        
                        if actual_qty == old_qty:
                            st.info("입력하신 수량과 이미 계산된 전산 재고가 같습니다. 별도 수정이 필요 없습니다.")
                        else:
                            diff = actual_qty - old_qty
                            dir_str = "증가▲" if diff > 0 else "감소▼"
                            
                            new_adj = {
                                "조정일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "유형": target_row["유형"],
                                "제품명": target_row["제품명"],
                                "규격": target_row["규격"],
                                "기존재고": old_qty,
                                "변경재고": actual_qty,
                                "방향": dir_str,
                                "차이": diff,
                                "사유": adj_reason
                            }
                            df_adj = pd.concat([df_adj, pd.DataFrame([new_adj])], ignore_index=True)
                            df_adj.to_csv(INVENTORY_ADJ_FILE, index=False, encoding='utf-8-sig')
                            
                            log_history("재고 수동조정", "재고 관리", f"{target_row['제품명']} 수량수정( {old_qty} ➔ {actual_qty} ) 사유: {adj_reason}")
                            st.success(f"경고: {target_row['제품명']} 제품의 전산 재고가 강제 {dir_str} 조정되었습니다!")
                            st.rerun()

            with st.expander("📝 누적 재고 조정 히스토리 보기"):
                if df_adj.empty:
                    st.info("로딩된 임의조정 내역이 없습니다.")
                else:
                    st.dataframe(df_adj.sort_values(by="조정일시", ascending=False), use_container_width=True, hide_index=True)

# --- 출고 관리 메뉴 ---
elif menu_selection == "출고 관리":
    if sub_menu == "출고등록":
        st.markdown('<div class="section-title">🚚 출고 등록 및 배차 관리</div>', unsafe_allow_html=True)
        st.write("차량별 출고 내역을 등록합니다. 여러 제품을 한 번에 입력하여 효율적으로 관리할 수 있습니다.")

        df_out = load_outbound_records()
        df_p_list = load_specs()
        
        # 제품 목록 준비
        if df_p_list.empty:
            st.warning("등록된 제품이 없습니다. 제품 관리 메뉴에서 먼저 제품을 등록해주세요.")
            p_names = []
        else:
            p_names = df_p_list['제품명'].dropna().unique().tolist()
            
        if "outbound_row_count" not in st.session_state:
            st.session_state.outbound_row_count = 1

        # 1. 출고 등록 섹션
        with st.expander("➕ 새 출고 배차 등록", expanded=True):
            c1, c2 = st.columns(2)
            with c1:
                out_date = st.date_input("출고일자", value=date.today())
                out_time = st.time_input("출고시간", value=datetime.today().time()).strftime("%H:%M")
            with c2:
                out_dest = st.text_input("출하처 *", placeholder="예: 물류센터 / 대리점 등")
            
            st.write("---")
            
            # 상단 제목 영역에 제품 라인 추가 및 삭제 버튼 배치
            col_add1, col_add2, col_add3 = st.columns([1, 2, 2])
            with col_add1:
                st.markdown("**📦 출하 제품 및 수량**")
            with col_add2:
                if st.button("➕ 품목 1줄 추가", key="btn_add_ob_row"):
                    st.session_state.outbound_row_count += 1
                    st.rerun()
            with col_add3:
                if st.session_state.outbound_row_count > 1:
                    if st.button("➖ 맨 밑줄 삭제", key="btn_del_ob_row"):
                        st.session_state.outbound_row_count -= 1
                        st.rerun()

            out_items = []
            for i in range(st.session_state.outbound_row_count):
                ic0, ic1, ic2, ic3, ic4 = st.columns([2, 3, 2, 2, 2])
                with ic0:
                    type_opts = ["선택 안함"] + (df_p_list['유형'].dropna().unique().tolist() if not df_p_list.empty else [])
                    sel_t = st.selectbox(f"유형 ({i+1})", type_opts, key=f"ob_t_{i}")
                with ic1:
                    if sel_t != "선택 안함" and not df_p_list.empty:
                        p_opts = df_p_list[df_p_list['유형'] == sel_t]['제품명'].dropna().unique().tolist()
                    else:
                        p_opts = p_names
                    sel_p = st.selectbox(f"제품명 ({i+1})", ["선택 안함"] + p_opts, key=f"ob_p_{i}")
                with ic2:
                    if sel_p != "선택 안함" and not df_p_list.empty:
                        u_opts = df_p_list[df_p_list['제품명'] == sel_p]['규격'].dropna().unique().tolist()
                        
                        # 자동 유형 선택 보완 (제품 선택시 유형이 '선택 안함'이었을 경우 역추적)
                        if sel_t == "선택 안함":
                            guess_t = df_p_list[df_p_list['제품명'] == sel_p]['유형'].dropna().tolist()
                            if guess_t:
                                sel_t = guess_t[0]
                    else:
                        u_opts = []
                    
                    if u_opts:
                        sel_u = st.selectbox(f"규격 ({i+1})", u_opts, key=f"ob_u_{i}")
                    else:
                        sel_u = st.text_input(f"규격 직접입력 ({i+1})", key=f"ob_u_{i}")
                        
                with ic3:
                    qty = st.number_input(f"수량 ({i+1})", min_value=0, step=1, key=f"ob_q_{i}")
                with ic4:
                    note = st.text_input(f"비고 ({i+1})", key=f"ob_n_{i}")
                
                if sel_p != "선택 안함" and str(sel_p).strip() != "" and qty > 0:
                    out_items.append({
                        "유형": sel_t if sel_t != "선택 안함" else "-",
                        "제품명": sel_p,
                        "규격": sel_u,
                        "수량": str(qty),
                        "비고": note
                    })

            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🚀 출고 데이터 일괄 등록", type="primary", use_container_width=True):
                if not out_dest:
                    st.error("출하처는 필수 입력 사항입니다.")
                elif not out_items:
                    st.error("최소 하나 이상의 제품과 수량(0 초과)을 입력해야 합니다.")
                else:
                    new_rows = []
                    for item in out_items:
                        new_rows.append([
                            str(out_date), out_time, "", "", out_dest,
                            item["유형"], item["제품명"], item["규격"], item["수량"], item["비고"]
                        ])
                    
                    header_cols = ["출고일", "출고시간", "차량번호", "기사명", "출하처", "유형", "제품명", "규격", "수량", "비고"]
                    new_df = pd.DataFrame(new_rows, columns=header_cols)
                    
                    df_out = pd.concat([df_out, new_df], ignore_index=True)
                    df_out.to_csv(OUTBOUND_FILE, index=False, encoding='utf-8-sig')
                    
                    log_history("출고 등록", "출고 관리", f"출하처: {out_dest}, 제품 {len(new_rows)}건 등록")
                    st.success(f"총 {len(new_rows)}건의 출고 내역이 등록되었습니다.")
                    st.session_state.outbound_row_count = 1  # 폼 초기화
                    st.rerun()

        # 2. 조회 및 엑셀 출력 섹션
        st.markdown("### 📋 출고 내역 조회 및 엑셀 출력")

        view_c1, view_c2 = st.columns([1, 2])
        with view_c1:
            search_date = st.date_input("조회 기준일", value=date.today())
            
        if df_out.empty:
            st.info("등록된 출고 내역이 없습니다.")
        else:
            df_filtered = df_out[df_out["출고일"] == str(search_date)]
            
            if df_filtered.empty:
                st.warning(f"{search_date} 일자의 출고 내역이 없습니다.")
            else:
                # 당일 전체 / 개별 배차(차량) 모드 선택
                view_mode = st.radio("조회 모드 선택", ["당일 전체 출고조회", "개별 배차별(차량별) 조회"], horizontal=True)
                
                # 콤보박스나 필터 설정
                target_df = None
                file_name_prefix = ""
                doc_title = ""
                
                if view_mode == "당일 전체 출고조회":
                    target_df = df_filtered.copy()
                    file_name_prefix = f"Hollys_Outbound_Total_{search_date.strftime('%Y%m%d')}"
                    doc_title = f"출고 관리 대장 ({search_date})"
                else:
                    # 개별 배차 선택 (출고시간, 출하처 기준)
                    배차목록 = df_filtered.apply(lambda r: f"[{r['출고시간']}] {r['출하처']}", axis=1).unique().tolist()
                    if 배차목록:
                        sel_배차 = st.selectbox("배차 내역 선택", 배차목록)
                        분리 = sel_배차.split("] ")
                        s_time = 분리[0].replace("[", "")
                        s_dest = 분리[1].strip()
                        
                        target_df = df_filtered[
                            (df_filtered["출고시간"] == s_time) & 
                            (df_filtered["출하처"] == s_dest)
                        ]
                        file_name_prefix = f"Hollys_Outbound_Car_{search_date.strftime('%Y%m%d')}_{s_time.replace(':','')}"
                        doc_title = f"출고 명세서"

                if target_df is not None and not target_df.empty:
                    st.markdown(f"**검색 결과: {len(target_df)}건**")
                    
                    def get_excel_data(df_view, doc_title_str, mode):
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            wb = writer.book
                            ws = wb.add_worksheet('출고내역')
                            
                            title_fmt = wb.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#DDEBF7'})
                            header_fmt = wb.add_format({'bold': True, 'border': 1, 'align': 'center', 'bg_color': '#F2F2F2'})
                            data_fmt = wb.add_format({'border': 1, 'align': 'center'})
                            num_fmt = wb.add_format({'border': 1, 'align': 'right', 'num_format': '#,##0'})
                            text_left_fmt = wb.add_format({'align': 'left', 'valign': 'vcenter', 'font_size': 11})
                            text_bold_fmt = wb.add_format({'align': 'left', 'valign': 'vcenter', 'bold': True, 'font_size': 12})
                            
                            if mode == "당일 전체 출고조회":
                                ws.merge_range('A1:J2', doc_title_str, title_fmt)
                                headers = ["출고일", "출고시간", "차량번호", "기사명", "출하처", "유형", "제품명", "규격", "수량", "비고"]
                                for i, h in enumerate(headers):
                                    ws.write(3, i, h, header_fmt)
                                    
                                ws.set_column('A:B', 12); ws.set_column('C:F', 15); ws.set_column('G:G', 25); ws.set_column('H:J', 15)
                                
                                for ri, (_, row) in enumerate(df_view.iterrows(), 4):
                                    ws.write(ri, 0, str(row["출고일"]), data_fmt)
                                    ws.write(ri, 1, str(row["출고시간"]), data_fmt)
                                    ws.write(ri, 2, str(row.get("차량번호", "")), data_fmt)
                                    ws.write(ri, 3, str(row.get("기사명", "")), data_fmt)
                                    ws.write(ri, 4, str(row["출하처"]), data_fmt)
                                    ws.write(ri, 5, str(row["유형"]), data_fmt)
                                    ws.write(ri, 6, str(row["제품명"]), data_fmt)
                                    ws.write(ri, 7, str(row["규격"]), data_fmt)
                                    try: qty_val = int(float(row["수량"]))
                                    except: qty_val = 0
                                    ws.write(ri, 8, qty_val, num_fmt)
                                    ws.write(ri, 9, str(row["비고"]), data_fmt)
                            else:
                                # 개별 배차 모드 (수기 작성용 명세서 양식)
                                ws.set_column('A:A', 6); ws.set_column('B:C', 15); ws.set_column('D:D', 25)
                                ws.set_column('E:G', 12)
                                
                                ws.merge_range('A1:G2', "출고 명세서", title_fmt)
                                
                                # 상단부
                                r_date = df_view.iloc[0]["출고일"]
                                r_time = df_view.iloc[0]["출고시간"]
                                r_dest = df_view.iloc[0]["출하처"]
                                
                                ws.write('A4', f"출고일자 : {r_date}", text_bold_fmt)
                                ws.write('D4', f"출고시간 : {r_time}", text_bold_fmt)
                                
                                # 데이터 헤더
                                headers = ["No", "유형", "제품명", "규격", "수량", "비고"]
                                for i, h in enumerate(headers):
                                    # C와 D 사이에 제품명이 넓으므로 C D E F G 로 매핑 (A, B, C, D, E, F)
                                    ws.write(5, i, h, header_fmt)
                                
                                # 제품 데이터 쓰기
                                current_row = 6
                                for idx, (_, row) in enumerate(df_view.iterrows(), 1):
                                    ws.write(current_row, 0, idx, data_fmt)
                                    ws.write(current_row, 1, str(row["유형"]), data_fmt)
                                    ws.write(current_row, 2, str(row["제품명"]), data_fmt)
                                    ws.write(current_row, 3, str(row["규격"]), data_fmt)
                                    try: qty_val = int(float(row["수량"]))
                                    except: qty_val = 0
                                    ws.write(current_row, 4, qty_val, num_fmt)
                                    ws.write(current_row, 5, str(row["비고"]), data_fmt)
                                    current_row += 1
                                
                                # 하단부 여백 및 수기 입력 공간
                                current_row += 2
                                ws.write(current_row, 0, f"▶ 출하처 : {r_dest}", text_bold_fmt)
                                current_row += 2
                                ws.write(current_row, 0, "기사명 : _____________________", text_left_fmt)
                                ws.write(current_row, 3, "차량번호 : _____________________", text_left_fmt)
                                current_row += 2
                                ws.write(current_row, 0, "차량위생상태 : ( 양호 / 불량 )", text_left_fmt)
                                ws.write(current_row, 3, "인수자 서명 : ____________________", text_left_fmt)
                                
                        return output.getvalue()
                        
                    current_state = target_df.to_csv(index=False) + doc_title + view_mode
                    if st.session_state.get("outbound_cache_state") != current_state:
                        st.session_state["outbound_cache_blob"] = get_excel_data(target_df, doc_title, view_mode)
                        st.session_state["outbound_cache_state"] = current_state
                        
                    import base64
                    excel_blob = st.session_state["outbound_cache_blob"]
                    b64_data = base64.b64encode(excel_blob).decode("utf-8")
                    dl_link = (
                        f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_data}" '
                        f'download="{file_name_prefix}.xlsx" '
                        f'style="display: block; text-align: center; padding: 0.5rem; background-color: #f0f2f6; '
                        f'border-radius: 4px; color: #31333f; text-decoration: none; border: 1px solid #c2c8d1; '
                        f'font-weight: 500; font-family: sans-serif; margin-bottom: 1rem;">'
                        f'📥 {doc_title} 엑셀 바로 다운로드'
                        f'</a>'
                    )
                    st.markdown(dl_link, unsafe_allow_html=True)
                    
                    st.caption("✨ 아래 표를 바로 수정하거나 행을 추가 및 삭제한 뒤, **[변경사항 저장]** 버튼을 누르시면 데이터가 즉시 갱신됩니다.")
                    edited_target = st.data_editor(target_df, num_rows="dynamic", use_container_width=True, key="out_editor_inline")
                    
                    if st.button("📝 현재 조회된 내역 변경사항 서버 저장", type="primary", key="save_inline_ed"):
                        # 原 df_out 에서 조회되었던 데이터의 인덱스를 제외 (삭제)
                        df_out_remain = df_out.drop(index=target_df.index)
                        
                        # 수정된 데이터셋을 하단에 추가
                        df_out_new = pd.concat([df_out_remain, edited_target], ignore_index=True)
                        
                        # 날짜, 시간순 정렬 (데이터베이스를 깔끔하게 유지하기 위함)
                        if "출고일" in df_out_new.columns and "출고시간" in df_out_new.columns:
                            df_out_new = df_out_new.sort_values(by=["출고일", "출고시간"], ascending=[True, True])
                            
                        df_out_new.to_csv(OUTBOUND_FILE, index=False, encoding='utf-8-sig')
                        st.success("✔️ 데이터 갱신이 완료되었습니다.")
                        st.rerun()

# --- 10. HACCP 관리 메뉴 ---
elif menu_selection == "HACCP":
    if sub_menu == "HACCP 일지":
        st.markdown('<div class="section-title">🛡️ HACCP 일지 및 점검표 관리</div>', unsafe_allow_html=True)
        st.write("8대 분야별 HACCP 일지 및 점검표 양식을 업로드하고 관리할 수 있습니다.")

        # 8개 분류 탭 구성
        tabs = st.tabs([
            "1. 영업자 관리", "2. 위생관리", "3. 제조관리", "4. 용수관리", 
            "5. 보관운송관리", "6. 검사관리", "7. 회수관리", "8. HACCP"
        ])
        
        tab_names = ["1_영업자관리", "2_위생관리", "3_제조관리", "4_용수관리", "5_보관운송관리", "6_검사관리", "7_회수관리", "8_HACCP"]
        base_dir = "haccp_docs"
        os.makedirs(base_dir, exist_ok=True) # 폴더가 없으면 생성

        # 각 탭 안에 동일한 로직(업로드, 다운로드, 삭제) 반복
        for i, tab in enumerate(tabs):
            with tab:
                folder_name = os.path.join(base_dir, tab_names[i])
                os.makedirs(folder_name, exist_ok=True)
                
                st.markdown(f"#### 📂 {['영업자 관리', '위생관리', '제조관리', '용수관리', '보관운송관리', '검사관리', '회수관리', 'HACCP'][i]} 양식 관리")
                
                # 파일 업로더
                upl_file = st.file_uploader(f"새 문서 업로드", key=f"up_{i}", help="엑셀, 워드, 한글, PDF 등의 양식을 업로드하세요.")
                if upl_file:
                    file_path = os.path.join(folder_name, upl_file.name)
                    with open(file_path, "wb") as f:
                        f.write(upl_file.getbuffer())
                    st.success(f"'{upl_file.name}' 업로드 완료!")
                    st.rerun()
                
                st.markdown("##### 📄 등록된 문서 목록")
                files = os.listdir(folder_name)
                
                if not files:
                    st.info("등록된 문서가 없습니다.")
                else:
                    for f_name in files:
                        f_path = os.path.join(folder_name, f_name)
                        c_file, c_dl, c_del = st.columns([6, 1.5, 1.5])
                        
                        with c_file:
                            st.markdown(f"<div style='margin-top: 8px; font-weight:500;'>▪️ {f_name}</div>", unsafe_allow_html=True)
                            
                        with c_dl:
                            with open(f_path, "rb") as f_read:
                                st.download_button("다운로드", data=f_read, file_name=f_name, key=f"dl_{i}_{f_name}", use_container_width=True)
                                
                        with c_del:
                            if st.button("삭제", key=f"del_{i}_{f_name}", use_container_width=True):
                                os.remove(f_path)
                                st.rerun()

    elif sub_menu == "HACCP 기준서":
        st.markdown('<div class="section-title">📘 HACCP 기준서 및 별첨 관리</div>', unsafe_allow_html=True)
        st.write("문서를 업로드할 때 **개정번호와 사유를 반드시 입력**해야 저장되며, 내역은 '개정이력' 메뉴에 연동됩니다.")
        
        tab_std, tab_att = st.tabs(["HACCP 기준서", "별첨 자료"])
        
        # 파일 저장 폴더 생성
        base_dir_haccp = "haccp_standards"
        dir_std = os.path.join(base_dir_haccp, "standards")
        dir_att = os.path.join(base_dir_haccp, "attachments")
        os.makedirs(dir_std, exist_ok=True)
        os.makedirs(dir_att, exist_ok=True)
        
        def render_haccp_uploader(category, target_dir):
            st.markdown(f"#### 📂 {category} 업로드")
            with st.form(f"form_upload_{category}"):
                up_file = st.file_uploader("새 문서 파일 선택", key=f"file_{category}")
                
                c1, c2, c3 = st.columns([1, 1, 2])
                with c1: rev_date = st.date_input("개정일")
                with c2: rev_no = st.text_input("개정번호 (예: Rev.01)")
                with c3: rev_reason = st.text_input("개정사유 (필수 입력)")
                
                if st.form_submit_button("파일 업로드 및 개정이력 등록"):
                    # 💡 파일, 개정번호, 개정사유 3가지가 모두 있어야만 업로드 통과
                    if up_file and rev_no and rev_reason:
                        # 1. 파일 저장
                        file_path = os.path.join(target_dir, up_file.name)
                        with open(file_path, "wb") as f:
                            f.write(up_file.getbuffer())
                            
                        # 2. 개정이력 DB 업데이트
                        df_rev = load_haccp_revisions()
                        new_rev = pd.DataFrame([[str(rev_date), category, up_file.name, rev_no, rev_reason]], columns=df_rev.columns)
                        df_rev = pd.concat([df_rev, new_rev], ignore_index=True)
                        df_rev.to_csv(HACCP_REVISION_FILE, index=False, encoding='utf-8-sig')
                        
                        st.success(f"[{up_file.name}] 업로드 및 개정이력 연동 완료!")
                        st.rerun()
                    else:
                        st.error("⚠️ 파일 첨부, 개정번호, 개정사유를 모두 입력해야 업로드가 가능합니다.")
                        
            st.markdown(f"##### 📄 등록된 {category} 목록")
            files = os.listdir(target_dir)
            if not files:
                st.info("등록된 문서가 없습니다.")
            else:
                for f_name in files:
                    f_path = os.path.join(target_dir, f_name)
                    c_file, c_dl, c_del = st.columns([6, 1.5, 1.5])
                    with c_file:
                        st.markdown(f"<div style='margin-top: 8px; font-weight:500;'>▪️ {f_name}</div>", unsafe_allow_html=True)
                    with c_dl:
                        with open(f_path, "rb") as f_read:
                            st.download_button("다운로드", data=f_read, file_name=f_name, key=f"dl_{category}_{f_name}", use_container_width=True)
                    with c_del:
                        if st.button("삭제", key=f"del_{category}_{f_name}", use_container_width=True):
                            os.remove(f_path)
                            st.rerun()

        with tab_std:
            render_haccp_uploader("HACCP 기준서", dir_std)
        with tab_att:
            render_haccp_uploader("별첨", dir_att)

    elif sub_menu == "개정이력":
        st.markdown('<div class="section-title">🔄 HACCP 문서 개정 이력</div>', unsafe_allow_html=True)
        st.write("HACCP 기준서 및 별첨 자료에서 업로드된 파일들의 개정 이력이 최신순으로 자동 기록됩니다.")
        
        df_rev = load_haccp_revisions()
        
        if df_rev.empty:
            st.info("아직 등록된 개정 이력이 없습니다. 'HACCP 기준서' 메뉴에서 파일을 업로드해 주세요.")
        else:
            # 최신순 정렬
            df_rev['개정일자'] = pd.to_datetime(df_rev['개정일자'], errors='coerce')
            df_rev = df_rev.sort_values(by="개정일자", ascending=False).reset_index(drop=True)
            df_rev['개정일자'] = df_rev['개정일자'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) else "")
            
            cfg_rev = {
                "개정일자": st.column_config.TextColumn("개정일자", width="medium"),
                "분류": st.column_config.TextColumn("분류", width="small"),
                "문서명": st.column_config.TextColumn("문서명(파일명)"),
                "개정번호": st.column_config.TextColumn("개정번호", width="small"),
                "개정사유": st.column_config.TextColumn("개정사유")
            }
            
            st.dataframe(df_rev, use_container_width=True, column_config=cfg_rev, hide_index=True)
            
            # 엑셀 다운로드 버튼
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df_rev.to_excel(writer, index=False, sheet_name='개정이력')
            
            st.download_button(
                label="개정이력 엑셀 다운로드 (보관용)",
                data=output.getvalue(),
                file_name=f"HACCP_개정이력대장_{date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=False
            )
    elif sub_menu == "공정 흐름도":
        import plotly.graph_objects as go

        st.markdown('<div class="section-title">🔄 HACCP 제조공정 흐름도</div>', unsafe_allow_html=True)
        st.write("카테고리별로 독립된 공정 흐름도를 관리합니다. 레인(원재료/포장재/기타 등)을 직접 설정하여 병렬 흐름과 합류 화살표를 표현할 수 있습니다.")

        # ── 카테고리 관리 ────────────────────────────────────────────────
        df_cats = load_flow_categories()

        with st.expander("📁 공정 카테고리 관리 (추가·삭제)", expanded=False):
            st.caption("카테고리마다 독립된 흐름도 데이터를 저장합니다. 예: 캡슐커피, 스틱커피, 생두 선별 등")
            with st.form("form_add_cat"):
                c1, c2, c3 = st.columns([2, 3, 1])
                with c1:
                    new_cat_id = st.text_input("카테고리 ID (영문, 예: capsule)")
                with c2:
                    new_cat_name = st.text_input("카테고리명 (예: 캡슐커피 제조공정)")
                with c3:
                    st.markdown("<div style='margin-top:28px'></div>", unsafe_allow_html=True)
                    add_btn = st.form_submit_button("추가", use_container_width=True)
                if add_btn:
                    if new_cat_id and new_cat_name:
                        if new_cat_id in df_cats["cat_id"].values:
                            st.error("이미 존재하는 카테고리 ID입니다.")
                        else:
                            new_cat = pd.DataFrame([[new_cat_id, new_cat_name, ""]], columns=df_cats.columns)
                            df_cats = pd.concat([df_cats, new_cat], ignore_index=True)
                            df_cats.to_csv(FLOW_CATEGORY_FILE, index=False, encoding="utf-8-sig")
                            st.success(f"카테고리 [{new_cat_name}] 추가 완료!")
                            st.rerun()
                    else:
                        st.error("ID와 카테고리명을 모두 입력하세요.")

            st.markdown("**등록된 카테고리 목록**")
            for _, cat_row in df_cats.iterrows():
                cc1, cc2 = st.columns([6, 1])
                with cc1:
                    st.markdown(f"▪️ **{cat_row['cat_name']}** &nbsp; <span style='color:#999;font-size:0.85rem'>ID: {cat_row['cat_id']}</span>", unsafe_allow_html=True)
                with cc2:
                    if cat_row["cat_id"] != "main":
                        if st.button("삭제", key=f"del_cat_{cat_row['cat_id']}"):
                            df_cats = df_cats[df_cats["cat_id"] != cat_row["cat_id"]]
                            df_cats.to_csv(FLOW_CATEGORY_FILE, index=False, encoding="utf-8-sig")
                            st.rerun()
                    else:
                        st.caption("기본(삭제불가)")

        st.divider()

        # ── 카테고리 탭 ─────────────────────────────────────────────────
        if df_cats.empty:
            st.warning("카테고리가 없습니다. 위에서 먼저 추가해 주세요.")
        else:
            cat_tabs = st.tabs([row["cat_name"] for _, row in df_cats.iterrows()])

            for tab_i, (_, cat_row) in enumerate(df_cats.iterrows()):
                with cat_tabs[tab_i]:
                    cat_id   = cat_row["cat_id"]
                    cat_name = cat_row["cat_name"]
                    safe_id  = str(cat_id).replace("/", "_").replace("\\", "_").replace(" ", "_")
                    fpath    = f"haccp_flowchart_{safe_id}.csv"

                    df_flow = load_flowchart_by_cat(cat_id)

                    # ── 공정 단계 편집 테이블 ────────────────────────────
                    with st.expander("⚙️ 공정 단계 추가 / 수정 (표에서 직접 편집)", expanded=False):
                        st.markdown("""
**컬럼 안내**
- **lane**: 이 단계가 속한 흐름(레인) 이름. 예) `원재료`, `포장재`, `기타`  
  → 레인이 여러 개면 화면에 나란히 배치됩니다.
- **merge_from**: 이 단계로 합류하는 다른 레인 이름. 예) `포장재`  
  → 점선 빨간 화살표로 합류 표현됩니다.
- **merge_label**: 합류 화살표 위에 표시할 라벨. 예) `내포장재(PE)`
- **합류 레인 2/3**: 동일 단계에 2~3개 레인이 동시 합류할 때 사용
                        """)
                        for col_need in ["lane", "merge_from", "merge_label", "merge_from_2", "merge_label_2", "merge_from_3", "merge_label_3"]:
                            if col_need not in df_flow.columns:
                                df_flow[col_need] = ""

                        cfg_flow = {
                            "유형":          st.column_config.SelectboxColumn("유형", options=["공정", "검사", "판단", "보관", "출하"], width="small"),
                            "CCP여부":       st.column_config.SelectboxColumn("CCP 여부", options=["N", "Y"], width="small"),
                            "순서":          st.column_config.TextColumn("순서", width="small"),
                            "CCP번호":       st.column_config.TextColumn("CCP 번호", width="small"),
                            "lane":          st.column_config.TextColumn("레인(lane)", width="medium"),
                            "merge_from":    st.column_config.TextColumn("합류 레인1", width="medium"),
                            "merge_label":   st.column_config.TextColumn("합류 라벨1", width="medium"),
                            "merge_from_2":  st.column_config.TextColumn("합류 레인2", width="medium"),
                            "merge_label_2": st.column_config.TextColumn("합류 라벨2", width="medium"),
                            "merge_from_3":  st.column_config.TextColumn("합류 레인3", width="medium"),
                            "merge_label_3": st.column_config.TextColumn("합류 라벨3", width="medium"),
                        }
                        edited_flow = st.data_editor(
                            df_flow, num_rows="dynamic", use_container_width=True,
                            column_config=cfg_flow, hide_index=True,
                            key=f"editor_{cat_id}"
                        )
                        if st.button("공정 단계 저장", key=f"save_flow_{cat_id}"):
                            edited_flow.to_csv(fpath, index=False, encoding="utf-8-sig")
                            st.success("공정 흐름도 데이터가 저장되었습니다.")
                            st.rerun()

                    # ── 흐름도 시각화 ────────────────────────────────────
                    st.markdown(f"### 📊 {cat_name} 공정 흐름도")

                    df_flow = load_flowchart_by_cat(cat_id)
                    for col_need in ["lane", "merge_from", "merge_label", "merge_from_2", "merge_label_2", "merge_from_3", "merge_label_3"]:
                        if col_need not in df_flow.columns:
                            df_flow[col_need] = ""
                    df_flow = df_flow.fillna("")
                    df_flow["순서"] = pd.to_numeric(df_flow["순서"], errors="coerce")
                    df_flow = df_flow.dropna(subset=["순서"]).sort_values(["lane", "순서"]).reset_index(drop=True)

                    # ── 레인 배치 계산 ────────────────────────────────────
                    TYPE_COLOR = {
                        "공정":  "#4A90D9",
                        "검사":  "#27AE60",
                        "판단":  "#F39C12",
                        "보관":  "#8E44AD",
                        "출하":  "#2C3E50",
                    }
                    CCP_COLOR   = "#D11031"
                    ARROW_COLOR = "#555555"
                    MERGE_COLOR = "#D11031"
                    LANE_BG     = [
                        "rgba(74,144,217,0.07)",
                        "rgba(39,174,96,0.07)",
                        "rgba(243,156,18,0.07)",
                        "rgba(142,68,173,0.07)",
                        "rgba(44,62,80,0.07)",
                    ]

                    # 레인 순서: 첫 등장 순서 유지
                    all_lanes = []
                    for _, r in df_flow.iterrows():
                        l = str(r.get("lane", "")).strip()
                        if l and l not in all_lanes:
                            all_lanes.append(l)
                    if not all_lanes:
                        all_lanes = ["공정"]

                    n_lanes  = len(all_lanes)
                    lane_w   = 1.0 / (n_lanes + 1)
                    lane_xs  = {ln: (li + 1) * lane_w for li, ln in enumerate(all_lanes)}

                    BOX_W  = min(0.26, lane_w * 0.82)
                    BOX_H  = 0.050
                    GAP    = 0.016
                    STEP_H = BOX_H + GAP

                    # 메인 레인(첫 번째 레인) 기준 최대 행 수 → 그래프 높이 결정
                    main_lane = all_lanes[0]
                    main_rows = df_flow[df_flow["lane"] == main_lane]
                    n_main    = max(len(main_rows), 1)

                    fig_h = max(700, n_main * 78 + 150)

                    fig         = go.Figure()
                    shapes      = []
                    annotations = []

                    # 레인 배경 + 헤더
                    for li, ln in enumerate(all_lanes):
                        cx = lane_xs[ln]
                        shapes.append(dict(
                            type="rect",
                            x0=cx - lane_w / 2, x1=cx + lane_w / 2,
                            y0=-0.05, y1=1.06,
                            fillcolor=LANE_BG[li % len(LANE_BG)],
                            line=dict(width=0),
                            layer="below"
                        ))
                        annotations.append(dict(
                            x=cx, y=1.045,
                            text=f"<b>{ln}</b>",
                            showarrow=False,
                            font=dict(size=13, color="#333", family="Noto Sans KR"),
                            xanchor="center", yanchor="middle",
                            bgcolor="white",
                            bordercolor="#bbb",
                            borderwidth=1,
                            borderpad=5,
                        ))
                    # 레인별 독립 Y 인덱스 관리
                    lane_y_idx  = {ln: 0 for ln in all_lanes}
                    # 각 단계의 실제 y_mid를 저장 (합류 화살표 Y 계산용)
                    step_ymid   = {}
                    # 레인별 단계 목록 (순서 정렬)
                    lane_steps  = {ln: df_flow[df_flow["lane"] == ln].reset_index(drop=True) for ln in all_lanes}

                    for ln in all_lanes:
                        steps = lane_steps[ln]
                        cx    = lane_xs[ln]
                        prev_y_bot = None

                        for si, (_, row) in enumerate(steps.iterrows()):
                            y_top = 1.0 - si * STEP_H - 0.06
                            y_bot = y_top - BOX_H
                            y_mid = (y_top + y_bot) / 2
                            step_ymid[(ln, str(int(float(row["순서"]))))] = y_mid

                            is_ccp  = str(row.get("CCP여부", "N")).upper() == "Y"
                            유형     = str(row.get("유형", "공정"))
                            단계명   = str(row.get("단계명", ""))
                            설명     = str(row.get("설명", ""))
                            ccp_번호 = str(row.get("CCP번호", ""))

                            fill_c   = CCP_COLOR if is_ccp else TYPE_COLOR.get(유형, "#4A90D9")
                            border_c = "#A80D27" if is_ccp else "#2c2c2c"

                            if 유형 == "판단":
                                dx, dy = BOX_W / 2, BOX_H / 2
                                fig.add_trace(go.Scatter(
                                    x=[cx, cx + dx, cx, cx - dx, cx],
                                    y=[y_top, y_mid, y_bot, y_mid, y_top],
                                    fill="toself", fillcolor=fill_c,
                                    line=dict(color=border_c, width=2),
                                    mode="lines", hoverinfo="skip", showlegend=False
                                ))
                            else:
                                shapes.append(dict(
                                    type="rect",
                                    x0=cx - BOX_W / 2, x1=cx + BOX_W / 2,
                                    y0=y_bot, y1=y_top,
                                    fillcolor=fill_c,
                                    line=dict(color=border_c, width=2),
                                    layer="above"
                                ))

                            label = f"<b>{int(float(row['순서']))}. {단계명}</b>"
                            if is_ccp:
                                label += f" ⚠️{ccp_번호}"

                            annotations.append(dict(
                                x=cx, y=y_mid + 0.010,
                                text=label,
                                showarrow=False,
                                font=dict(size=12, color="white", family="Noto Sans KR"),
                                xanchor="center", yanchor="middle"
                            ))
                            annotations.append(dict(
                                x=cx, y=y_mid - 0.013,
                                text=f"<span style='font-size:9px'>{설명}</span>",
                                showarrow=False,
                                font=dict(size=9, color="rgba(255,255,255,0.88)"),
                                xanchor="center", yanchor="middle"
                            ))

                            # 같은 레인 내 수직 화살표
                            if prev_y_bot is not None:
                                shapes.append(dict(
                                    type="line",
                                    x0=cx, x1=cx,
                                    y0=prev_y_bot, y1=y_top,
                                    line=dict(color=ARROW_COLOR, width=2)
                                ))
                                annotations.append(dict(
                                    x=cx, y=y_top,
                                    ax=cx, ay=prev_y_bot,
                                    xref="x", yref="y", axref="x", ayref="y",
                                    showarrow=True, arrowhead=2,
                                    arrowsize=1.2, arrowcolor=ARROW_COLOR, arrowwidth=2
                                ))
                            prev_y_bot = y_bot

                    # ── 합류 화살표 그리기 (최대 3개 합류 레인 지원) ────
                    MERGE_PAIRS = [
                        ("merge_from",   "merge_label"),
                        ("merge_from_2", "merge_label_2"),
                        ("merge_from_3", "merge_label_3"),
                    ]

                    for _, row in df_flow.iterrows():
                        target_lane = str(row.get("lane", "")).strip()
                        순서str      = str(int(float(row["순서"])))
                        target_y    = step_ymid.get((target_lane, 순서str), None)
                        if target_y is None:
                            continue
                        to_cx = lane_xs.get(target_lane, 0.5)

                        for mf_col, ml_col in MERGE_PAIRS:
                            merge_from = str(row.get(mf_col, "")).strip()
                            merge_lbl  = str(row.get(ml_col, "")).strip()

                            if not merge_from or merge_from not in lane_xs:
                                continue

                            from_cx   = lane_xs[merge_from]
                            sub_steps = lane_steps.get(merge_from, pd.DataFrame())

                            if sub_steps.empty:
                                from_y = target_y
                            else:
                                last_sub_순서 = str(int(float(sub_steps.iloc[-1]["순서"])))
                                from_y = step_ymid.get((merge_from, last_sub_순서), target_y)

                            from_x_start = from_cx + BOX_W / 2 if from_cx < to_cx else from_cx - BOX_W / 2
                            to_x_end     = to_cx - BOX_W / 2  if from_cx < to_cx else to_cx + BOX_W / 2

                            # 수직 구간 (서브 레인 하단 → 합류 Y)
                            if abs(from_y - target_y) > 0.001:
                                shapes.append(dict(
                                    type="line",
                                    x0=from_cx, x1=from_cx,
                                    y0=from_y - BOX_H / 2, y1=target_y,
                                    line=dict(color=MERGE_COLOR, width=2, dash="dot")
                                ))

                            # 수평 구간
                            shapes.append(dict(
                                type="line",
                                x0=from_cx, x1=to_x_end,
                                y0=target_y, y1=target_y,
                                line=dict(color=MERGE_COLOR, width=2, dash="dot")
                            ))
                            # 화살표 머리
                            annotations.append(dict(
                                x=to_x_end, y=target_y,
                                ax=from_cx, ay=target_y,
                                xref="x", yref="y", axref="x", ayref="y",
                                showarrow=True, arrowhead=2,
                                arrowsize=1.1, arrowcolor=MERGE_COLOR, arrowwidth=2
                            ))
                            # 라벨
                            if merge_lbl:
                                mid_x = (from_cx + to_x_end) / 2
                                annotations.append(dict(
                                    x=mid_x, y=target_y + 0.014,
                                    text=f"<span style='font-size:9px;color:{MERGE_COLOR}'>{merge_lbl}</span>",
                                    showarrow=False,
                                    font=dict(size=9, color=MERGE_COLOR),
                                    xanchor="center"
                                ))

                    fig.update_layout(
                        shapes=shapes,
                        annotations=annotations,
                        xaxis=dict(visible=False, range=[-0.02, 1.02]),
                        yaxis=dict(visible=False, range=[-0.08, 1.12]),
                        margin=dict(l=10, r=10, t=55, b=10),
                        height=fig_h,
                        paper_bgcolor="#F8F9FA",
                        plot_bgcolor="#F8F9FA",
                        title=dict(
                            text=f"Hollys Roasting Center — {cat_name}",
                            x=0.5, xanchor="center",
                            font=dict(size=15, color="#212529")
                        )
                    )

                    st.plotly_chart(fig, use_container_width=True)

                    # 범례
                    st.markdown("""
                    <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:6px;">
                        <span style="background:#4A90D9;color:white;padding:3px 10px;border-radius:6px;font-size:0.82rem;">■ 공정</span>
                        <span style="background:#27AE60;color:white;padding:3px 10px;border-radius:6px;font-size:0.82rem;">■ 검사</span>
                        <span style="background:#F39C12;color:white;padding:3px 10px;border-radius:6px;font-size:0.82rem;">■ 판단</span>
                        <span style="background:#8E44AD;color:white;padding:3px 10px;border-radius:6px;font-size:0.82rem;">■ 보관</span>
                        <span style="background:#2C3E50;color:white;padding:3px 10px;border-radius:6px;font-size:0.82rem;">■ 출하</span>
                        <span style="background:#D11031;color:white;padding:3px 10px;border-radius:6px;font-size:0.82rem;">⚠️ CCP</span>
                        <span style="border:2px dashed #D11031;color:#D11031;padding:3px 10px;border-radius:6px;font-size:0.82rem;">‥▶ 합류 화살표</span>
                    </div>
                    """, unsafe_allow_html=True)

                    # CCP 요약표
                    df_ccp = df_flow[df_flow["CCP여부"].str.upper() == "Y"]
                    if not df_ccp.empty:
                        st.markdown("### ⚠️ CCP 지점 요약")
                        st.dataframe(
                            df_ccp[["순서", "CCP번호", "단계명", "설명", "lane"]].reset_index(drop=True),
                            use_container_width=True, hide_index=True
                        )

                    # ── 엑셀 보관용 출력 ─────────────────────────────────
                    def export_flowchart_excel(df_input, cat_nm):
                        import io as _io
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                        from openpyxl.utils import get_column_letter
                        from openpyxl.worksheet.page import PageMargins

                        df_f = df_input.copy().fillna("")
                        df_f["순서"] = pd.to_numeric(df_f["순서"], errors="coerce")
                        df_f = df_f.dropna(subset=["순서"]).sort_values(["lane","순서"]).reset_index(drop=True)

                        all_lanes = []
                        for _, r in df_f.iterrows():
                            l = str(r.get("lane","")).strip()
                            if l and l not in all_lanes:
                                all_lanes.append(l)
                        if not all_lanes:
                            return None

                        main_lane  = all_lanes[0]
                        sub_lanes  = all_lanes[1:]
                        lane_steps = {ln: df_f[df_f["lane"]==ln].sort_values("순서").reset_index(drop=True)
                                      for ln in all_lanes}
                        main_steps = lane_steps[main_lane]
                        n_main     = len(main_steps)

                        # 합류 정보: merge_map[메인_si] = [(from_lane, label), ...]
                        MKEYS = [("merge_from","merge_label"),
                                 ("merge_from_2","merge_label_2"),
                                 ("merge_from_3","merge_label_3")]
                        merge_map = {}
                        for si, (_, r) in enumerate(main_steps.iterrows()):
                            for mf, ml in MKEYS:
                                mfv = str(r.get(mf,"")).strip()
                                mlv = str(r.get(ml,"")).strip()
                                if mfv and mlv:
                                    merge_map.setdefault(si, []).append((mfv, mlv))

                        # ── 열 레이아웃 ──────────────────────────────────────────────
                        # 각 서브레인: 박스 4열 (파이프는 박스 내 2번째 열)
                        # 서브레인 사이 gap 1열
                        # MAIN_NO 1열 + MAIN 5열
                        SUB_W   = 4
                        SUB_GAP = 1
                        MAIN_NO_W = 1
                        MAIN_W  = 5

                        col = 1
                        sub_sc = {}    # 서브 박스 시작열
                        sub_ec = {}    # 서브 박스 끝열
                        sub_pc = {}    # 서브 파이프열 (박스 내 2번째 = sc+1)
                        for ln in sub_lanes:
                            sub_sc[ln] = col
                            sub_ec[ln] = col + SUB_W - 1
                            sub_pc[ln] = col + 1        # 박스 두번째 열 = 파이프
                            col += SUB_W + SUB_GAP

                        MAIN_NO = col
                        col    += MAIN_NO_W
                        MAIN_SC = col
                        MAIN_EC = col + MAIN_W - 1

                        # ── 행 레이아웃 ──────────────────────────────────────────────
                        R_T1    = 1
                        R_T2    = 2
                        R_LHDR  = 3
                        R_LDESC = 4
                        R_GAP   = 5
                        R_START = 6

                        BOX_H  = 2
                        ARR_H  = 1
                        STRIDE = BOX_H + ARR_H

                        def row_of(si):
                            return R_START + si * STRIDE

                        # ── 스타일 헬퍼 ──────────────────────────────────────────────
                        RED = "FF0000"   # 순수 빨강 (목표 파일과 동일)

                        def S(st="medium", c="000000"):
                            return Side(style=st, color=c.lstrip("#"))
                        def RS(st="thin"):   # 빨간 선
                            return Side(style=st, color=RED)
                        def fill(h):
                            return PatternFill("solid", fgColor=h.lstrip("#"))
                        def fnt(bold=False, sz=10, c="000000"):
                            return Font(name="맑은 고딕", bold=bold, size=sz, color=c.lstrip("#"))
                        def aln(h="center", v="center", wrap=False):
                            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

                        def sb(ws, r, c, top=None, bot=None, lft=None, rgt=None):
                            cell = ws.cell(row=r, column=c)
                            old  = cell.border
                            cell.border = Border(
                                top    = top if top is not None else old.top,
                                bottom = bot if bot is not None else old.bottom,
                                left   = lft if lft is not None else old.left,
                                right  = rgt if rgt is not None else old.right,
                            )

                        TYPE_BG = {"공정":"4A90D9","검사":"27AE60","판단":"F39C12",
                                   "보관":"8E44AD","출하":"2C3E50"}
                        CCP_BG  = "D11031"
                        LANE_C  = ["4A90D9","27AE60","F39C12","8E44AD","2C3E50"]
                        LANE_LT = ["BDD7EE","E2EFDA","FFF2CC","FCE4D6","F3E5F5"]
                        BK = S("medium")

                        wb = Workbook()
                        ws = wb.active
                        ws.title = cat_nm[:28]

                        # ── 열 너비 ─────────────────────────────────────────────────
                        for ln in sub_lanes:
                            for ci in range(sub_sc[ln], sub_ec[ln]+1):
                                ws.column_dimensions[get_column_letter(ci)].width = 8
                            # gap열
                            ws.column_dimensions[get_column_letter(sub_ec[ln]+1)].width = 1
                        ws.column_dimensions[get_column_letter(MAIN_NO)].width = 4
                        for ci in range(MAIN_SC, MAIN_EC+1):
                            ws.column_dimensions[get_column_letter(ci)].width = 9

                        # ── 행 높이 ─────────────────────────────────────────────────
                        ws.row_dimensions[R_T1].height    = 22
                        ws.row_dimensions[R_T2].height    = 13
                        ws.row_dimensions[R_LHDR].height  = 18
                        ws.row_dimensions[R_LDESC].height = 14
                        ws.row_dimensions[R_GAP].height   = 5

                        # ── 문서 제목 ────────────────────────────────────────────────
                        ws.merge_cells(start_row=R_T1, start_column=1, end_row=R_T1, end_column=MAIN_EC)
                        c = ws.cell(row=R_T1, column=1, value="제조공정흐름도")
                        c.font = fnt(bold=True, sz=18); c.alignment = aln("left")

                        ws.merge_cells(start_row=R_T2, start_column=1, end_row=R_T2, end_column=MAIN_EC)
                        c2 = ws.cell(row=R_T2, column=1, value=f"■ {cat_nm} | 커피클럽 로스팅센터")
                        c2.font = fnt(sz=9); c2.alignment = aln("left")

                        # ── 레인 헤더 ────────────────────────────────────────────────
                        for li, ln in enumerate(all_lanes):
                            if ln == main_lane:
                                sc2, ec2 = MAIN_SC, MAIN_EC
                            else:
                                sc2, ec2 = sub_sc[ln], sub_ec[ln]
                            bgc = LANE_C[li % len(LANE_C)]
                            ltc = LANE_LT[li % len(LANE_LT)]

                            ws.merge_cells(start_row=R_LHDR, start_column=sc2, end_row=R_LHDR, end_column=ec2)
                            ch = ws.cell(row=R_LHDR, column=sc2, value=ln)
                            ch.font = fnt(bold=True, sz=11, c="FFFFFF"); ch.alignment = aln()
                            ch.fill = fill(bgc)
                            ch.border = Border(left=BK, right=BK, top=BK, bottom=BK)

                            steps = lane_steps[ln]
                            desc  = str(steps.iloc[0].get("설명","")) if not steps.empty else ""
                            ws.merge_cells(start_row=R_LDESC, start_column=sc2, end_row=R_LDESC, end_column=ec2)
                            cd = ws.cell(row=R_LDESC, column=sc2, value=desc)
                            cd.font = fnt(sz=9); cd.alignment = aln(wrap=True)
                            cd.fill = fill(ltc)
                            cd.border = Border(left=BK, right=BK, top=BK, bottom=BK)

                        # ── 메인 공정 박스 ───────────────────────────────────────────
                        for si, (_, r) in enumerate(main_steps.iterrows()):
                            rs     = row_of(si)
                            is_ccp = str(r.get("CCP여부","N")).upper() == "Y"
                            type_s = str(r.get("유형","공정"))
                            name   = str(r.get("단계명",""))
                            ccp_no = str(r.get("CCP번호",""))
                            seq    = int(float(r["순서"]))
                            bgc    = CCP_BG if is_ccp else TYPE_BG.get(type_s,"4A90D9")

                            ws.row_dimensions[rs].height   = 15
                            ws.row_dimensions[rs+1].height = 15
                            ws.row_dimensions[rs+2].height = 8

                            # 번호 셀 (2행 병합)
                            ws.merge_cells(start_row=rs, start_column=MAIN_NO, end_row=rs+1, end_column=MAIN_NO)
                            cn = ws.cell(row=rs, column=MAIN_NO, value=seq)
                            cn.font = fnt(bold=True, sz=9, c="FFFFFF"); cn.alignment = aln()
                            cn.fill = fill(bgc)
                            cn.border = Border(left=BK, right=BK, top=BK, bottom=BK)

                            if is_ccp and ccp_no:
                                ws.merge_cells(start_row=rs, start_column=MAIN_SC, end_row=rs, end_column=MAIN_EC)
                                c1 = ws.cell(row=rs, column=MAIN_SC, value=name)
                                c1.font = fnt(bold=True, sz=11, c="FFFFFF"); c1.alignment = aln()
                                c1.fill = fill(bgc)
                                c1.border = Border(left=BK, right=BK, top=BK, bottom=BK)
                                ws.merge_cells(start_row=rs+1, start_column=MAIN_SC, end_row=rs+1, end_column=MAIN_EC)
                                c3 = ws.cell(row=rs+1, column=MAIN_SC, value=ccp_no)
                                c3.font = fnt(bold=True, sz=9, c="FFFFFF"); c3.alignment = aln()
                                c3.fill = fill(CCP_BG)
                                c3.border = Border(left=BK, right=BK, top=BK, bottom=BK)
                            else:
                                ws.merge_cells(start_row=rs, start_column=MAIN_SC, end_row=rs+1, end_column=MAIN_EC)
                                c1 = ws.cell(row=rs, column=MAIN_SC, value=name)
                                c1.font = fnt(bold=True, sz=11, c="FFFFFF"); c1.alignment = aln()
                                c1.fill = fill(bgc)
                                c1.border = Border(left=BK, right=BK, top=BK, bottom=BK)

                            if si < n_main - 1:
                                ws.merge_cells(start_row=rs+2, start_column=MAIN_NO, end_row=rs+2, end_column=MAIN_EC)
                                ca = ws.cell(row=rs+2, column=MAIN_NO, value="↓")
                                ca.font = fnt(sz=11); ca.alignment = aln()

                        # ── 서브레인 박스 ────────────────────────────────────────────
                        for ln in sub_lanes:
                            sc2   = sub_sc[ln]
                            ec2   = sub_ec[ln]
                            steps = lane_steps[ln]
                            li    = all_lanes.index(ln)
                            bgc   = LANE_C[li % len(LANE_C)]

                            for si2, (_, r2) in enumerate(steps.iterrows()):
                                rs2   = row_of(si2)
                                name2 = str(r2.get("단계명",""))
                                t2    = str(r2.get("유형","공정"))
                                is_c  = str(r2.get("CCP여부","N")).upper()=="Y"
                                seq2  = int(float(r2["순서"]))
                                bc2   = CCP_BG if is_c else TYPE_BG.get(t2, bgc)

                                ws.row_dimensions[rs2].height   = 15
                                ws.row_dimensions[rs2+1].height = 15

                                ws.merge_cells(start_row=rs2, start_column=sc2, end_row=rs2+1, end_column=ec2)
                                cb = ws.cell(row=rs2, column=sc2, value=f"{seq2}. {name2}")
                                cb.font = fnt(bold=True, sz=10, c="FFFFFF"); cb.alignment = aln(wrap=True)
                                cb.fill = fill(bc2)
                                cb.border = Border(left=BK, right=BK, top=BK, bottom=BK)

                                if si2 < len(steps)-1:
                                    ws.merge_cells(start_row=rs2+2, start_column=sc2, end_row=rs2+2, end_column=ec2)
                                    ca2 = ws.cell(row=rs2+2, column=sc2, value="↓")
                                    ca2.font = fnt(sz=11); ca2.alignment = aln()

                        # ══════════════════════════════════════════════════════════════
                        # ── 합류선 ────────────────────────────────────────────────────
                        #
                        # 목표 파일 분석 결과:
                        #   파이프 = 박스 내 2번째 열(sc+1)에 right=thin(FF0000) 세로선
                        #   박스 행 포함해서 처음부터 합류 직전 행까지 계속 내려옴
                        #   합류행에서 꺾임:
                        #     파이프열+1(=sc+2): left+bottom=thin(FF0000) + 라벨 텍스트
                        #     그 이후 ~ MAIN_NO 직전: bottom=thin(FF0000)
                        # ══════════════════════════════════════════════════════════════

                        sub_targets = {ln: [] for ln in sub_lanes}
                        for target_si, merges in merge_map.items():
                            for (from_lane, label) in merges:
                                if from_lane in sub_targets:
                                    sub_targets[from_lane].append((target_si, label))

                        for ln in sub_lanes:
                            targets = sub_targets[ln]
                            if not targets:
                                continue

                            pc       = sub_pc[ln]          # 파이프열 (sc+1)
                            label_c  = pc + 1              # 꺾임+라벨 열 (sc+2)
                            n_sub    = len(lane_steps[ln])
                            last_rs  = row_of(n_sub - 1)   # 서브 마지막 박스 시작행

                            # 여러 합류 대상이 있으면 마지막 합류까지 파이프 유지
                            max_target_si = max(t[0] for t in targets)
                            max_target_rs = row_of(max_target_si)

                            # 파이프 세로선: 서브 첫 박스 첫행 ~ 마지막 합류행 직전
                            # (박스 행도 포함 - 박스 테두리보다 나중에 그어 덮어씀)
                            pipe_start = row_of(0)                # 첫 서브 박스 시작행
                            pipe_end   = max_target_rs - 1        # 합류 직전 행까지

                            for pr in range(pipe_start, pipe_end + 1):
                                sb(ws, pr, pc, rgt=RS("thin"))

                            # 각 합류 대상마다 꺾임선 + 라벨
                            for (target_si, label) in targets:
                                target_rs = row_of(target_si)   # 합류 메인 박스 시작행

                                # 꺾임: 라벨열에 left+bottom, 라벨열+1 ~ MAIN_NO 직전까지 bottom
                                sb(ws, target_rs, label_c, lft=RS("thin"), bot=RS("thin"))
                                for ci in range(label_c + 1, MAIN_NO):
                                    sb(ws, target_rs, ci, bot=RS("thin"))

                                # 라벨 텍스트
                                cell_lbl = ws.cell(row=target_rs, column=label_c)
                                if not cell_lbl.value:   # 이미 값 있으면 덮어쓰기 않음
                                    cell_lbl.value = label
                                else:
                                    cell_lbl.value = str(cell_lbl.value) + " / " + label
                                cell_lbl.font = Font(name="맑은 고딕", size=8, bold=True, color=RED)
                                cell_lbl.alignment = Alignment(horizontal="left",
                                                               vertical="center", wrap_text=True)

                        # ── 하단 서명 ────────────────────────────────────────────────
                        foot_row = row_of(n_main) + 1
                        ws.row_dimensions[foot_row].height = 16
                        ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=MAIN_EC)
                        cf = ws.cell(row=foot_row, column=1, value="커피류 제조공정도")
                        cf.font = fnt(bold=True, sz=10); cf.alignment = aln("left")

                        # ── 인쇄 설정 ────────────────────────────────────────────────
                        ws.page_setup.paperSize   = 9
                        ws.page_setup.orientation = "landscape" if len(all_lanes) >= 3 else "portrait"
                        ws.page_setup.fitToPage   = True
                        ws.page_setup.fitToWidth  = 1
                        ws.page_setup.fitToHeight = 0
                        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
                        ws.print_title_rows = "1:5"

                        out = _io.BytesIO()
                        wb.save(out)
                        return out.getvalue()
                    st.divider()
                    _excel_data = export_flowchart_excel(df_flow, cat_name)
                    if _excel_data:
                        st.download_button(
                            label="공정흐름도 엑셀 다운로드 (보관용)",
                            data=_excel_data,
                            file_name=f"Hollys_공정흐름도_{cat_name}_{date.today().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )




    elif sub_menu == "공정별 CCP 결정도":
        st.markdown('<div class="section-title">공정별 CCP 결정도</div>', unsafe_allow_html=True)

        # ── CCP 판정 로직 ────────────────────────────────────────────
        def calc_ccp(q1, q2, q21, q3, q4, q5):
            if q1 == "예":       return "CP"
            if q1 == "아니오":
                if q2 == "아니오":
                    if q21 == "아니오":  return "CP"
                    if q21 == "예":      return "→Q2재검토"
                    return ""
                if q2 == "예":
                    if q3 == "예":       return "CCP"
                    if q3 == "아니오":
                        if q4 == "아니오": return "CP"
                        if q4 == "예":
                            if q5 == "예":      return "CP"
                            if q5 == "아니오":  return "CCP"
            return ""

        # 어떤 질문이 활성화되어야 하는지 계산
        def active_questions(q1, q2, q21, q3, q4):
            """현재 답변 상태에 따라 활성화되어야 할 질문 집합 반환"""
            active = {"Q1"}
            if q1 == "아니오":
                active.add("Q2")
                if q2 == "아니오":
                    active.add("Q21")
                elif q2 == "예":
                    active.add("Q3")
                    if q3 == "아니오":
                        active.add("Q4")
                        if q4 == "예":
                            active.add("Q5")
            return active

        # ── 카테고리 선택 ─────────────────────────────────────────────
        df_cats = load_flow_categories()
        if df_cats.empty:
            st.info("공정 카테고리가 없습니다. 공정 흐름도에서 카테고리를 먼저 등록하세요.")
        else:
            cat_options = {row["cat_name"]: row["cat_id"] for _, row in df_cats.iterrows()}
            sel_cat_name = st.selectbox("카테고리 선택", list(cat_options.keys()), key="ccp_cat")
            sel_cat_id   = cat_options[sel_cat_name]

            df_flow = load_flowchart_by_cat(sel_cat_id)
            df_dec  = load_ccp_decision(sel_cat_id)

            # ── [개선] 위해요소분석 결과와 자동 동기화 (강력한 매칭) ─────────
            sync_count = 0
            if os.path.exists(HAZARD_FILE):
                try:
                    df_haz_global = pd.read_csv(HAZARD_FILE, encoding="utf-8-sig").fillna("")
                    # 현재 카테고리 + 최종위해요소='Y' 필터
                    df_haz_sync = df_haz_global[
                        (df_haz_global["카테고리"].astype(str).str.strip() == sel_cat_name.strip()) & 
                        (df_haz_global["최종위해요소"].astype(str).str.strip().str.upper() == "Y")
                    ]
                    
                    for _, h_row in df_haz_sync.iterrows():
                        h_step = str(h_row["공정명"]).strip()
                        h_type = str(h_row["유형"]).strip()[:1]
                        h_hz   = str(h_row["위해요소"]).strip()
                        
                        exists = False
                        if not df_dec.empty:
                            exists = ((df_dec["단계명"].astype(str).str.strip() == h_step) & 
                                      (df_dec["위해유형"].astype(str).str.strip().str.startswith(h_type)) & 
                                      (df_dec["위해요소"].astype(str).str.strip() == h_hz)).any()
                        
                        if not exists:
                            new_dec_row = pd.DataFrame([{
                                "단계명": h_step, "위해유형": h_type, "위해요소": h_hz,
                                "Q1_답변": "", "Q1_비고": "", "Q2_답변": "", "Q2_비고": "",
                                "Q21_답변": "", "Q21_비고": "", "Q3_답변": "", "Q3_비고": "",
                                "Q4_답변": "", "Q4_비고": "", "Q5_답변": "", "Q5_비고": "",
                                "CCP결정": "", "CCP번호": "", "비고": ""
                            }])
                            df_dec = pd.concat([df_dec, new_dec_row], ignore_index=True)
                            sync_count += 1
                    
                    if sync_count > 0:
                        save_ccp_decision(sel_cat_id, df_dec)
                        st.info(f"🔄 위해요소분석 결과로부터 새로운 항목 {sync_count}건이 자동으로 추가되었습니다.")
                except Exception:
                    pass

            if df_flow.empty:
                st.info("선택한 카테고리의 공정 흐름도 데이터가 없습니다.")
                st.stop()

            # 메인 레인 + 유형=='공정' 단계 추출 (순서 컬럼 포함)
            all_lanes  = df_flow["lane"].dropna().unique().tolist()
            main_lane  = all_lanes[0] if all_lanes else None
            mask_main  = (df_flow["lane"] == main_lane) if main_lane else pd.Series([True]*len(df_flow))
            mask_type  = df_flow["유형"].str.strip() == "공정"
            df_steps   = df_flow[mask_main & mask_type].copy()
            df_steps["_순서_num"] = pd.to_numeric(df_steps["순서"], errors="coerce").fillna(9999)
            df_steps   = df_steps.sort_values("_순서_num").drop(columns=["_순서_num"]).reset_index(drop=True)
            step_names = df_steps["단계명"].tolist()
            # 공정흐름도 순서 그대로 No 매핑
            step_no    = {row["단계명"]: str(row["순서"]) for _, row in df_steps.iterrows()}

            if not step_names:
                st.warning("유형이 '공정'인 단계가 없습니다.")
                st.stop()

            # No 포함한 단계 표시 (순서 오름차순)
            steps_with_no = [f"{step_no.get(s,'?')}. {s}" for s in step_names]
            st.caption(f"연동 단계 ({len(step_names)}개): {' → '.join(steps_with_no)}")

            with st.expander("📋 CCP 결정 질문 흐름 보기", expanded=False):
                st.markdown("""
| 질문 | 내용 | 예 | 아니오 |
|------|------|----|--------|
| **질문1** | 확인된 위해요소를 관리하기 위한 선행요건이 있으며 잘 관리되고 있는가? | → CP | → 질문2 |
| **질문2** | 모든 공정(단계)에서 확인된 위해요소에 대한 조치방법이 있는가? | → 질문3 | → 질문2-1 |
| **질문2-1** | 이 공정(단계)에서 안전성을 위한 관리가 필요한가? | → 질문2 재검토 | → CP |
| **질문3** | 이 공정(단계)에서 발생가능성이 있는 위해요소를 제거하거나 허용수준까지 감소시킬 수 있는가? | **→ CCP** | → 질문4 |
| **질문4** | 확인된 위해요소의 오염이 허용수준을 초과하는가 또는 허용할 수 없는 수준으로 증가하는가? | → 질문5 | → CP |
| **질문5** | 확인된 위해요소를 제거하거나 그 발생을 허용수준으로 감소시킬 수 있는 이후의 공정이 있는가? | → CP | **→ CCP** |
                """)
            st.divider()

            # ── 세션 초기화 ──────────────────────────────────────────
            if "ccp_mode" not in st.session_state:
                st.session_state.ccp_mode = "list"
            if "ccp_edit_idx" not in st.session_state:
                st.session_state.ccp_edit_idx = None

            HZ_TYPES = ["", "B (생물학)", "C (화학)", "P (물리)"]
            ANS3     = ["", "예", "아니오", "해당없음"]  # 해당없음 포함

            # ════════════════════════════════════════════════════════
            # 목록 모드
            # ════════════════════════════════════════════════════════
            if st.session_state.ccp_mode == "list":
                col_add, _ = st.columns([1, 5])
                with col_add:
                    if st.button("항목 추가", type="primary", use_container_width=True):
                        st.session_state.ccp_mode = "add"
                        st.session_state.ccp_edit_idx = None
                        st.rerun()

                if df_dec.empty:
                    st.info("등록된 항목이 없습니다. '항목 추가' 버튼으로 등록하세요.")
                else:
                    df_dec_sorted = df_dec.copy()
                    df_dec_sorted["_no_sort"] = df_dec_sorted["단계명"].map(
                        lambda s: int(step_no[s]) if s in step_no and str(step_no[s]).lstrip('-').isdigit() else 9999
                    )
                    df_dec_sorted = df_dec_sorted.sort_values("_no_sort", kind="stable").drop(columns=["_no_sort"]).reset_index(drop=True)

                    # ── 공정명 rowspan 계산 ──
                    rows_ccp = []
                    for idx2, row in df_dec_sorted.iterrows():
                        rows_ccp.append({
                            "idx": idx2,
                            "step": str(row.get("단계명","")),
                            "htype": str(row.get("위해유형","")),
                            "hz":    str(row.get("위해요소","")),
                            "q1":  str(row.get("Q1_답변","")),
                            "q2":  str(row.get("Q2_답변","")),
                            "q21": str(row.get("Q21_답변","")),
                            "q3":  str(row.get("Q3_답변","")),
                            "q4":  str(row.get("Q4_답변","")),
                            "q5":  str(row.get("Q5_답변","")),
                            "q5b": str(row.get("Q5_비고","")),
                            "ccp_r": str(row.get("CCP결정","")),
                        })
                    # 연속 공정명 rowspan
                    n_ccp = len(rows_ccp)
                    step_span_arr = [0]*n_ccp
                    i2=0
                    while i2 < n_ccp:
                        j2=i2
                        while j2 < n_ccp and rows_ccp[j2]["step"]==rows_ccp[i2]["step"]: j2+=1
                        step_span_arr[i2]=j2-i2; i2=j2

                    TH_CCP = "background:#D6E4F0;color:#1F4E79;font-weight:bold;padding:8px 6px;text-align:center;border:1px solid #b0c4d8;font-size:11px;"
                    TH_Q   = "background:#2E75B6;color:white;font-weight:bold;padding:8px 4px;text-align:center;border:1px solid #1a5a9e;font-size:10px;"
                    TH_Q3  = "background:#1A3E6E;color:white;font-weight:bold;padding:8px 4px;text-align:center;border:1px solid #0d2a52;font-size:10px;"
                    TD_CCP = "background:#fff;padding:7px 6px;border:1px solid #d0d7de;vertical-align:middle;text-align:center;font-size:11px;word-break:break-word;"

                    html_ccp = f"""<style>
.ccp-tbl{{width:100%;border-collapse:collapse;}}
.ccp-tbl tr:hover td{{filter:brightness(0.97);}}
</style>
<table class="ccp-tbl">
<thead><tr>
  <th style="{TH_CCP}width:90px;">공정명</th>
  <th style="{TH_CCP}width:40px;">유형</th>
  <th style="{TH_CCP}width:140px;">위해요소</th>
  <th style="{TH_Q}width:65px;">질문1</th>
  <th style="{TH_Q}width:65px;">질문2</th>
  <th style="{TH_Q}width:65px;">질문2-1</th>
  <th style="{TH_Q3}width:65px;">질문3</th>
  <th style="{TH_Q}width:65px;">질문4</th>
  <th style="{TH_Q}width:65px;">질문5</th>
  <th style="{TH_CCP}width:70px;">CCP결정</th>
</tr></thead>
<tbody>"""
                    QBG = {"예":"#E2EFDA","아니오":"#FCE4D6","해당없음":"#F5F5F5"}
                    for ci2, d2 in enumerate(rows_ccp):
                        act2  = active_questions(d2["q1"],d2["q2"],d2["q21"],d2["q3"],d2["q4"])
                        ccp_r = d2["ccp_r"]
                        ccp_bg = "#D32F2F" if ccp_r=="CCP" else ("#2E7D32" if ccp_r=="CP" else ("#F57C00" if ccp_r else "#fff"))
                        ccp_fc = "white" if ccp_r in ("CCP","CP") else "#888"

                        def qtd(qk, qv):
                            if qk not in act2: return f"<td style='{TD_CCP}background:#eee;color:#aaa;'>—</td>"
                            bg2 = QBG.get(qv,"#fff"); fc2="#333"
                            return f"<td style='{TD_CCP}background:{bg2};'>{qv}</td>"

                        sp2 = step_span_arr[ci2]
                        if sp2 > 0:
                            no_v2 = step_no.get(d2["step"],"-")
                            step_td = f'<td rowspan="{sp2}" style="{TD_CCP}background:#f0f4f8;font-weight:bold;vertical-align:middle;">{no_v2}. {d2["step"]}</td>'
                        else:
                            step_td = ""

                        html_ccp += f"""<tr>
  {step_td}
  <td style="{TD_CCP}">{d2["htype"]}</td>
  <td style="{TD_CCP}text-align:left;">{d2["hz"]}</td>
  {qtd("Q1", d2["q1"])}
  {qtd("Q2", d2["q2"])}
  {qtd("Q21",d2["q21"])}
  {qtd("Q3", d2["q3"])}
  {qtd("Q4", d2["q4"])}
  {qtd("Q5", d2["q5"])}
  <td style="{TD_CCP}background:{ccp_bg};color:{ccp_fc};font-weight:bold;">{ccp_r}</td>
</tr>"""
                    html_ccp += "</tbody></table>"
                    st.markdown(html_ccp, unsafe_allow_html=True)
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.caption("수정·삭제할 항목을 선택하세요:")
                    for idx2, row in df_dec_sorted.iterrows():
                        step = str(row.get("단계명",""))
                        hz   = str(row.get("위해요소",""))
                        no_v = step_no.get(step, "-")
                        c1, c2, c3 = st.columns([5, 0.7, 0.7])
                        with c1:
                            st.caption(f"**{no_v}. {step}** — {hz}")
                        with c2:
                            if st.button("✏️", key=f"ed_{idx2}", use_container_width=True):
                                st.session_state.ccp_mode = "edit"
                                st.session_state.ccp_edit_idx = idx2
                                st.rerun()
                        with c3:
                            if st.button("🗑️", key=f"dl_{idx2}", use_container_width=True):
                                df_dec = df_dec.drop(index=idx2).reset_index(drop=True)
                                save_ccp_decision(sel_cat_id, df_dec)
                                st.rerun()
                    st.divider()

            # ════════════════════════════════════════════════════════
            # 추가 / 수정 모드 (폼 밖 → session_state로 실시간 조건부 질문)
            # ════════════════════════════════════════════════════════
            elif st.session_state.ccp_mode in ("add", "edit"):
                is_edit  = st.session_state.ccp_mode == "edit"
                edit_idx = st.session_state.ccp_edit_idx

                def gv(col):
                    if is_edit and edit_idx is not None and not df_dec.empty:
                        try:
                            v = df_dec.iloc[edit_idx][col]
                            return str(v) if pd.notna(v) else ""
                        except Exception: pass
                    return ""

                # session_state 키 prefix (카테고리+편집idx 기반으로 충돌 방지)
                PFX = f"ccp_form_{sel_cat_id}_{edit_idx}"

                # 초기값 세팅 (처음 진입 시만)
                init_key = f"{PFX}_init"
                if init_key not in st.session_state:
                    st.session_state[f"{PFX}_step"]    = gv("단계명")
                    st.session_state[f"{PFX}_htype"]   = gv("위해유형")
                    st.session_state[f"{PFX}_hz"]      = gv("위해요소")
                    st.session_state[f"{PFX}_Q1"]      = gv("Q1_답변")  or ""
                    st.session_state[f"{PFX}_Q2"]      = gv("Q2_답변")  or ""
                    st.session_state[f"{PFX}_Q21"]     = gv("Q21_답변") or ""
                    st.session_state[f"{PFX}_Q3"]      = gv("Q3_답변")  or ""
                    st.session_state[f"{PFX}_Q4"]      = gv("Q4_답변")  or ""
                    st.session_state[f"{PFX}_Q5"]      = gv("Q5_답변")  or ""
                    st.session_state[f"{PFX}_Q1_note"] = gv("Q1_비고")
                    st.session_state[f"{PFX}_Q2_note"] = gv("Q2_비고")
                    st.session_state[f"{PFX}_Q21_note"]= gv("Q21_비고")
                    st.session_state[f"{PFX}_Q3_note"] = gv("Q3_비고")
                    st.session_state[f"{PFX}_Q4_note"] = gv("Q4_비고")
                    st.session_state[f"{PFX}_Q5_note"] = gv("Q5_비고")
                    st.session_state[f"{PFX}_ccpno"]   = gv("CCP번호")
                    st.session_state[f"{PFX}_bigo"]    = gv("비고")
                    st.session_state[init_key] = True

                st.markdown(f"### {'✏️ 항목 수정' if is_edit else '항목 추가'}")

                # ── 기본 정보 ─────────────────────────────────────────
                fc1, fc2 = st.columns([2, 1])
                with fc1:
                    registered_steps = set(df_dec["단계명"].dropna().tolist()) if not df_dec.empty else set()
                    cur_step = st.session_state[f"{PFX}_step"]
                    if is_edit:
                        available_steps = step_names
                    else:
                        available_steps = [s for s in step_names if s not in registered_steps]
                    opts = [""] + available_steps
                    step_idx = opts.index(cur_step) if cur_step in opts else 0
                    f_step = st.selectbox("단계명 *", opts, index=step_idx, key=f"{PFX}_step_sel")
                    st.session_state[f"{PFX}_step"] = f_step
                with fc2:
                    ht_opts = HZ_TYPES
                    cur_ht  = st.session_state[f"{PFX}_htype"]
                    ht_idx  = next((i for i,v in enumerate(ht_opts) if v.startswith(cur_ht[:1])), 0) if cur_ht else 0
                    f_htype = st.selectbox("유형", ht_opts, index=ht_idx, key=f"{PFX}_htype_sel")
                    st.session_state[f"{PFX}_htype"] = f_htype

                # 위해요소 자동 도출 (단계+유형 기준)
                hz_final_list = []
                if os.path.exists(HAZARD_FILE):
                    _dh = pd.read_csv(HAZARD_FILE, encoding="utf-8-sig")
                    _cur_step2 = st.session_state.get(f"{PFX}_step", "")
                    _cur_ht2   = st.session_state.get(f"{PFX}_htype", "")
                    _ht_code   = _cur_ht2[:1] if _cur_ht2 else ""
                    _mask_fin  = _dh["카테고리"] == sel_cat_name
                    if _cur_step2:
                        _mask_fin = _mask_fin & (_dh["공정명"] == _cur_step2)
                    if _ht_code:
                        _mask_fin = _mask_fin & (_dh["유형"] == _ht_code)
                    _df_fin = _dh[_mask_fin & (_dh["최종위해요소"] == "Y")]
                    hz_final_list = _df_fin["위해요소"].tolist()

                # 수정 모드: 현재 위해요소 1개 표시
                # 추가 모드: 자동 도출된 목록 배지로 표시
                if f_step and f_htype:
                    if is_edit:
                        cur_hz_v = st.session_state[f"{PFX}_hz"]
                        st.info(f"수정 중인 위해요소: **{cur_hz_v}**")
                        f_hz = cur_hz_v
                    else:
                        if hz_final_list:
                            badges = "".join([
                                f"<span style='display:inline-block;background:#D6E8FF;color:#1A5FAD;"
                                f"border-radius:12px;padding:3px 10px;margin:2px 3px;font-size:12px;"
                                f"font-weight:bold;'>{h}</span>"
                                for h in hz_final_list
                            ])
                            st.markdown(
                                f"<div style='background:#f0f6ff;border:1px solid #b0ccee;"
                                f"border-radius:6px;padding:10px 14px;margin:6px 0;'>"
                                f"<div style='font-size:12px;color:#1A5FAD;font-weight:bold;margin-bottom:6px;'>"
                                f"🔵 최종위해요소 자동 도출 ({len(hz_final_list)}건) — 저장 시 일괄 등록됩니다</div>"
                                f"{badges}</div>",
                                unsafe_allow_html=True
                            )
                            f_hz = "__AUTO__"  # 저장 시 hz_final_list 전체 사용
                        else:
                            st.warning("해당 공정+유형의 최종위해요소가 없습니다. 위해요소분석을 먼저 완료하세요.")
                            f_hz = ""
                    st.session_state[f"{PFX}_hz"] = f_hz
                else:
                    f_hz = st.session_state[f"{PFX}_hz"]

                st.markdown("---")
                st.markdown("#### 📋 질문 답변")
                st.caption("이전 질문의 답변에 따라 다음 질문이 자동으로 활성화됩니다.")

                # 현재 답변 읽기
                q1  = st.session_state[f"{PFX}_Q1"]
                q2  = st.session_state[f"{PFX}_Q2"]
                q21 = st.session_state[f"{PFX}_Q21"]
                q3  = st.session_state[f"{PFX}_Q3"]
                q4  = st.session_state[f"{PFX}_Q4"]
                q5  = st.session_state[f"{PFX}_Q5"]

                act = active_questions(q1, q2, q21, q3, q4)

                def render_q(qkey, qlbl, qtxt, ss_key, ss_note_key, active, result_hint=""):
                    """단일 질문 렌더링. active=False면 비활성(회색) 표시"""
                    border_c = "#2E75B6" if active else "#cccccc"
                    bg_c     = "#EBF3FB" if active else "#F5F5F5"
                    fc_c     = "#000000" if active else "#aaaaaa"
                    hint_html = f' <span style="color:#D32F2F;font-weight:bold;">{result_hint}</span>' if result_hint and active else ""

                    st.markdown(
                        f"<div style='background:{bg_c};border-left:4px solid {border_c};"
                        f"padding:8px 12px;border-radius:4px;margin:6px 0;'>"
                        f"<span style='font-size:12px;font-weight:bold;color:{fc_c};'>{qlbl}</span>"
                        f"<span style='font-size:11px;color:{fc_c};'>&nbsp; {qtxt}</span>"
                        f"{hint_html}</div>",
                        unsafe_allow_html=True
                    )
                    if active:
                        qa, qn = st.columns([1, 3])
                        with qa:
                            cur = st.session_state.get(ss_key, "")
                            idx_a = ANS3.index(cur) if cur in ANS3 else 0
                            ans = st.radio("답변", ANS3[1:], index=max(0,idx_a-1),
                                           horizontal=True, key=f"radio_{ss_key}",
                                           label_visibility="collapsed")
                            st.session_state[ss_key] = ans
                        with qn:
                            note = st.text_input("비고/근거",
                                                  value=st.session_state.get(ss_note_key,""),
                                                  key=f"note_inp_{ss_key}",
                                                  placeholder="근거·출처 입력 (선택)")
                            st.session_state[ss_note_key] = note
                        return ans
                    else:
                        st.markdown(
                            "<div style='padding:4px 12px;color:#aaa;font-size:11px;'>"
                            "⏸ 이전 질문 답변에 따라 건너뜁니다.</div>",
                            unsafe_allow_html=True
                        )
                        # 비활성 질문 답변 초기화
                        st.session_state[ss_key] = ""
                        return ""

                # ── Q1 렌더 후 act 재계산하며 단계적으로 진행 ─────────
                q1 = render_q("Q1", "질문 1",
                    "확인된 위해요소를 관리하기 위한 선행요건이 있으며 잘 관리되고 있는가?",
                    f"{PFX}_Q1", f"{PFX}_Q1_note", True,
                    "→ CP (종료)" if q1=="예" else "→ 질문2로" if q1=="아니오" else "")
                act = active_questions(q1, q2, q21, q3, q4)

                # Q2 (Q1=아니오일 때만)
                q2 = render_q("Q2", "질문 2",
                    "모든 공정(단계)에서 확인된 위해요소에 대한 조치방법이 있는가?",
                    f"{PFX}_Q2", f"{PFX}_Q2_note", "Q2" in act,
                    "→ 질문3으로" if q2=="예" else "→ 질문2-1로" if q2=="아니오" else "")
                act = active_questions(q1, q2, q21, q3, q4)

                # Q2-1 (Q2=아니오일 때만)
                q21 = render_q("Q2-1", "질문 2-1",
                    "이 공정(단계)에서 안전성을 위한 관리가 필요한가?",
                    f"{PFX}_Q21", f"{PFX}_Q21_note", "Q21" in act,
                    "→ 질문2 재검토" if q21=="예" else "→ CP (종료)" if q21=="아니오" else "")
                act = active_questions(q1, q2, q21, q3, q4)

                # Q3 (Q2=예일 때만)
                q3 = render_q("Q3", "질문 3",
                    "이 공정(단계)에서 발생가능성이 있는 위해요소를 제거하거나 허용수준까지 감소시킬 수 있는가?",
                    f"{PFX}_Q3", f"{PFX}_Q3_note", "Q3" in act,
                    "🔴 → CCP 결정!" if q3=="예" else "→ 질문4로" if q3=="아니오" else "")
                act = active_questions(q1, q2, q21, q3, q4)

                # Q4 (Q3=아니오일 때만)
                q4 = render_q("Q4", "질문 4",
                    "확인된 위해요소의 오염이 허용수준을 초과하는가 또는 허용할 수 없는 수준으로 증가하는가?",
                    f"{PFX}_Q4", f"{PFX}_Q4_note", "Q4" in act,
                    "→ 질문5로" if q4=="예" else "→ CP (종료)" if q4=="아니오" else "")
                act = active_questions(q1, q2, q21, q3, q4)

                # Q5 (Q4=예일 때만)
                q5 = render_q("Q5", "질문 5",
                    "확인된 위해요소를 제거하거나 그 발생을 허용수준으로 감소시킬 수 있는 이후의 공정이 있는가?",
                    f"{PFX}_Q5", f"{PFX}_Q5_note", "Q5" in act,
                    "→ CP (종료)" if q5=="예" else "🔴 → CCP 결정!" if q5=="아니오" else "")

                # ── 자동 판정 배너 ────────────────────────────────────
                auto    = calc_ccp(q1, q2, q21, q3, q4, q5)
                badge_c = "#D32F2F" if auto=="CCP" else ("#2E7D32" if auto=="CP" else "#F57C00")
                st.markdown(
                    f'<div style="margin:14px 0;padding:12px 20px;border-radius:10px;'
                    f'border:2px solid {badge_c};background:#fafafa;display:flex;align-items:center;">'
                    f'<span style="font-size:13px;font-weight:bold;">📊 자동 판정:</span>'
                    f'<span style="background:{badge_c};color:white;border-radius:6px;'
                    f'padding:4px 20px;font-size:16px;font-weight:bold;margin-left:12px;">'
                    f'{auto if auto else "미결정 — 질문을 완성하세요"}</span></div>',
                    unsafe_allow_html=True
                )

                # ── CCP번호 / 비고 ────────────────────────────────────
                fc_a, fc_b = st.columns(2)
                with fc_a:
                    f_ccpno = st.text_input("CCP번호 (CCP인 경우)",
                                             value=st.session_state[f"{PFX}_ccpno"],
                                             placeholder="CCP-1",
                                             key=f"{PFX}_ccpno_inp")
                    st.session_state[f"{PFX}_ccpno"] = f_ccpno
                with fc_b:
                    f_bigo = st.text_input("비고",
                                            value=st.session_state[f"{PFX}_bigo"],
                                            key=f"{PFX}_bigo_inp")
                    st.session_state[f"{PFX}_bigo"] = f_bigo

                # ── 저장 / 취소 ───────────────────────────────────────
                sb1, sb2 = st.columns(2)
                with sb1:
                    if st.button("저장", type="primary", use_container_width=True, key="ccp_save"):
                        if not f_step:
                            st.error("단계명은 필수입니다.")
                        elif not is_edit and f_hz == "__AUTO__" and not hz_final_list:
                            st.error("최종위해요소가 없습니다. 위해요소분석을 먼저 완료하세요.")
                        elif not is_edit and not f_hz:
                            st.error("공정명과 유형을 선택하세요.")
                        else:
                            htype_s = f_htype[:1] if f_htype else ""
                            q_data = {
                                "Q1_답변":  st.session_state[f"{PFX}_Q1"],  "Q1_비고":  st.session_state[f"{PFX}_Q1_note"],
                                "Q2_답변":  st.session_state[f"{PFX}_Q2"],  "Q2_비고":  st.session_state[f"{PFX}_Q2_note"],
                                "Q21_답변": st.session_state[f"{PFX}_Q21"], "Q21_비고": st.session_state[f"{PFX}_Q21_note"],
                                "Q3_답변":  st.session_state[f"{PFX}_Q3"],  "Q3_비고":  st.session_state[f"{PFX}_Q3_note"],
                                "Q4_답변":  st.session_state[f"{PFX}_Q4"],  "Q4_비고":  st.session_state[f"{PFX}_Q4_note"],
                                "Q5_답변":  st.session_state[f"{PFX}_Q5"],  "Q5_비고":  st.session_state[f"{PFX}_Q5_note"],
                                "CCP결정": auto, "CCP번호": f_ccpno, "비고": f_bigo
                            }
                            if is_edit and edit_idx is not None and not df_dec.empty:
                                # 수정: 단일 행 업데이트
                                new_row = {"단계명": f_step, "위해유형": htype_s, "위해요소": f_hz, **q_data}
                                for k, v in new_row.items():
                                    if k not in df_dec.columns: df_dec[k] = ""
                                    df_dec.at[edit_idx, k] = v
                                save_ccp_decision(sel_cat_id, df_dec)
                                cnt = 1
                            else:
                                # 추가: 최종위해요소 일괄 등록
                                save_list = hz_final_list if f_hz == "__AUTO__" else [f_hz]
                                new_rows = [{"단계명": f_step, "위해유형": htype_s, "위해요소": h, **q_data} for h in save_list]
                                df_dec = pd.concat([df_dec, pd.DataFrame(new_rows)], ignore_index=True)
                                save_ccp_decision(sel_cat_id, df_dec)
                                cnt = len(save_list)
                            for k in list(st.session_state.keys()):
                                if k.startswith(PFX): del st.session_state[k]
                            st.session_state.ccp_mode = "list"
                            st.success(f"{cnt}건 저장 완료!")
                            st.rerun()
                with sb2:
                    if st.button("취소", use_container_width=True, key="ccp_cancel"):
                        for k in list(st.session_state.keys()):
                            if k.startswith(PFX): del st.session_state[k]
                        st.session_state.ccp_mode = "list"
                        st.rerun()

            # ── 엑셀 다운로드 ─────────────────────────────────────────
            if st.session_state.ccp_mode == "list" and not df_dec.empty:
                def export_ccp_excel(df_d, cat_nm):
                    import io as _io
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                    from openpyxl.utils import get_column_letter
                    from openpyxl.worksheet.page import PageMargins

                    wb = Workbook(); ws = wb.active; ws.title = "CCP결정도"
                    BK = Side(style="thin", color="000000")
                    def bdr(): return Border(top=BK,bottom=BK,left=BK,right=BK)
                    def fill(h): return PatternFill("solid", fgColor=h)
                    def fnt(bold=False,sz=9,fc="000000"): return Font(name="맑은 고딕",bold=bold,size=sz,color=fc)
                    def aln(h="center",v="center",wrap=True): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
                    def wc(r,c,val="",bold=False,sz=9,fc="000000",bg=None,h="center"):
                        cell=ws.cell(row=r,column=c,value=str(val) if val is not None else "")
                        cell.font=fnt(bold,sz,fc); cell.alignment=aln(h); cell.border=bdr()
                        if bg: cell.fill=fill(bg)
                    def mc(r1,c1,r2,c2): ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)

                    mc(1,1,1,18); wc(1,1,f"공정별 CCP 결정도  |  {cat_nm}",bold=True,sz=13,fc="FFFFFF",bg="1F4E79"); ws.row_dimensions[1].height=26
                    mc(2,1,2,18); wc(2,1,"커피클럽 로스팅센터  |  ㈜케이지할리스에프앤비",sz=9,fc="595959",bg="EBF3FB"); ws.row_dimensions[2].height=13

                    H1="1F4E79"; H2="2E75B6"; H3="1A3E6E"
                    mc(3,1,5,1);  wc(3,1,"No",bold=True,fc="FFFFFF",bg=H1)
                    mc(3,2,5,2);  wc(3,2,"구  분",bold=True,fc="FFFFFF",bg=H1)
                    mc(3,3,5,3);  wc(3,3,"유형",bold=True,sz=8,fc="FFFFFF",bg=H1)
                    mc(3,4,5,5);  wc(3,4,"위해요소\n(생물학:B 화학:C 물리:P)",bold=True,sz=8,fc="FFFFFF",bg=H1)
                    mc(3,18,5,18);wc(3,18,"CCP\n결정",bold=True,fc="FFFFFF",bg=H1)

                    QDEFS = [
                        ("질문1",  "확인된 위해요소를 관리하기 위한 선행요건이 있으며 잘 관리되고 있는가?",   "예 → CP관리",     "아니오 → 질문2로",    H2, 6, 7),
                        ("질문2",  "모든 공정(단계)에서 확인된 위해요소에 대한 조치방법이 있는가?",           "예 → 질문3로",    "아니오 → 질문2-1로",  H2, 8, 9),
                        ("질문2-1","이 공정(단계)에서 안전성을 위한 관리가 필요한가?",                        "예 → 질문2로",    "아니오 → CP관리",     H2, 10,11),
                        ("질문3",  "이 공정(단계)에서 발생가능성이 있는 위해요소를 제거하거나 허용수준까지 감소시킬 수 있는가?", "예 → CCP관리",  "아니오 → 질문4로",  H3, 12,13),
                        ("질문4",  "확인된 위해요소의 오염이 허용수준을 초과하는가 또는 허용할 수 없는 수준으로 증가하는가?",    "예 → 질문5로",  "아니오 → CP관리",   H2, 14,15),
                        ("질문5",  "확인된 위해요소를 제거하거나 그 발생을 허용수준으로 감소시킬 수 있는 이후의 공정이 있는가?","예 → CP관리",   "아니오 → CCP관리",  H2, 16,17),
                    ]
                    for qlbl,qtxt,ytxt,ntxt,qbg,ac,nc in QDEFS:
                        mc(3,ac,3,nc); wc(3,ac,qlbl,bold=True,fc="FFFFFF",bg=qbg)
                        mc(4,ac,4,nc); wc(4,ac,qtxt,sz=7,fc="FFFFFF",bg=qbg)
                        wc(5,ac,ytxt,sz=7,fc="FFFFFF",bg=qbg)
                        wc(5,nc,ntxt,sz=7,fc="FFFFFF",bg="C0392B" if "CCP" in ntxt else qbg)
                    ws.row_dimensions[3].height=16; ws.row_dimensions[4].height=45; ws.row_dimensions[5].height=14

                    QMAP = [("Q1_답변","Q1_비고",6,7),("Q2_답변","Q2_비고",8,9),("Q21_답변","Q21_비고",10,11),
                            ("Q3_답변","Q3_비고",12,13),("Q4_답변","Q4_비고",14,15),("Q5_답변","Q5_비고",16,17)]
                    r=6
                    for _, dr in df_d.iterrows():
                        bg="F5F5F5" if r%2==0 else "FFFFFF"
                        step=str(dr.get("단계명","")); no_v=step_no.get(step,str(r-5))
                        wc(r,1,no_v,bg=bg)
                        wc(r,2,step,bold=True,bg="D6E4F0",h="left")
                        wc(r,3,str(dr.get("위해유형","")),bg=bg)
                        mc(r,4,r,5); wc(r,4,str(dr.get("위해요소","")),bg=bg,h="left")

                        q1x=str(dr.get("Q1_답변","")); q2x=str(dr.get("Q2_답변",""))
                        q21x=str(dr.get("Q21_답변","")); q3x=str(dr.get("Q3_답변",""))
                        q4x=str(dr.get("Q4_답변",""))
                        act_xl = active_questions(q1x,q2x,q21x,q3x,q4x)
                        qact_map = {"Q1_답변":"Q1","Q2_답변":"Q2","Q21_답변":"Q21",
                                    "Q3_답변":"Q3","Q4_답변":"Q4","Q5_답변":"Q5"}

                        for acol,ncol,ac,nc in QMAP:
                            qk = qact_map.get(acol,"")
                            ans=str(dr.get(acol,"")); note=str(dr.get(ncol,""))
                            if qk and qk not in act_xl:
                                wc(r,ac,"—",bg="EEEEEE",fc="AAAAAA")
                                wc(r,nc,"",bg="EEEEEE")
                            else:
                                abg="E2EFDA" if ans=="예" else ("FCE4D6" if ans=="아니오" else bg)
                                wc(r,ac,ans,bg=abg); wc(r,nc,note,sz=8,bg=bg,h="left")

                        ccp_r=str(dr.get("CCP결정",""))
                        cbg="D32F2F" if ccp_r=="CCP" else ("2E7D32" if ccp_r=="CP" else "F57C00")
                        wc(r,18,ccp_r,bold=True,fc="FFFFFF",bg=cbg)
                        ws.row_dimensions[r].height=22; r+=1

                    ws.column_dimensions["A"].width=4; ws.column_dimensions["B"].width=13
                    ws.column_dimensions["C"].width=5; ws.column_dimensions["D"].width=13
                    ws.column_dimensions["E"].width=8
                    for _,_,ac,nc in QMAP:
                        ws.column_dimensions[get_column_letter(ac)].width=8
                        ws.column_dimensions[get_column_letter(nc)].width=16
                    ws.column_dimensions["R"].width=9
                    ws.page_setup.paperSize=9; ws.page_setup.orientation="landscape"
                    ws.page_setup.fitToPage=True; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
                    ws.page_margins=PageMargins(left=0.4,right=0.4,top=0.6,bottom=0.6)
                    ws.print_title_rows="1:5"
                    out=_io.BytesIO(); wb.save(out); return out.getvalue()

                excel_data = export_ccp_excel(df_dec, sel_cat_name)
                st.download_button(
                    label="공정별 CCP 결정도 엑셀 다운로드 (보관용)",
                    data=excel_data,
                    file_name=f"Hollys_CCP결정도_{sel_cat_name}_{date.today().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )


# ──────────────────────────────────────────────────────────────────
# 심각성 설정
# ──────────────────────────────────────────────────────────────────
    elif sub_menu == "심각성 설정":
        st.markdown('<div class="section-title">위해요소 심각성 설정</div>', unsafe_allow_html=True)

        SEVERITY_FILE = "severity_settings.csv"

        def load_severity():
            if os.path.exists(SEVERITY_FILE):
                try:
                    df = pd.read_csv(SEVERITY_FILE, encoding='utf-8-sig')
                    for c in ["유형","위해요소","심각성"]:
                        if c not in df.columns: df[c] = ""
                    return df
                except Exception: pass
            d = pd.DataFrame(columns=["유형","위해요소","심각성"])
            d.to_csv(SEVERITY_FILE, index=False, encoding='utf-8-sig')
            return d

        df_sev = load_severity()

        TYPE_COLORS = {"B":"#D6EAF8","C":"#D5F5E3","P":"#FAE5D3"}
        TYPE_LABEL  = {"B":"B (생물학적)","C":"C (화학적)","P":"P (물리적)"}

        tab_b, tab_c, tab_p, tab_add = st.tabs(["B (생물학적)","C (화학적)","P (물리적)","+ 항목 추가/수정"])

        def render_sev_table(typ):
            sub = df_sev[df_sev["유형"]==typ].reset_index(drop=True)
            if sub.empty:
                st.info(f"{typ} 유형 위해요소가 없습니다.")
                return
            bg = TYPE_COLORS.get(typ,"#fff")
            hdr1, hdr2, hdr3 = st.columns([3,1,1])
            for col, lbl in zip([hdr1,hdr2,hdr3],["위해요소명","심각성","관리"]):
                col.markdown(f"<div style='background:#1F4E79;color:white;padding:5px 8px;font-size:12px;font-weight:bold;border-radius:4px;text-align:center;'>{lbl}</div>", unsafe_allow_html=True)
            for i, row in sub.iterrows():
                c1,c2,c3 = st.columns([3,1,1])
                c1.markdown(f"<div style='background:{bg};padding:6px 8px;font-size:12px;border-bottom:1px solid #eee;border-radius:3px;'>{row['위해요소']}</div>", unsafe_allow_html=True)
                sev = int(row['심각성']) if str(row['심각성']).isdigit() else 0
                sev_color = "#C0392B" if sev==3 else ("#E67E22" if sev==2 else "#27AE60")
                c2.markdown(f"<div style='background:{sev_color};color:white;padding:6px 0;font-size:13px;font-weight:bold;text-align:center;border-radius:4px;'>{sev}</div>", unsafe_allow_html=True)
                orig_idx = df_sev[(df_sev["유형"]==typ)&(df_sev["위해요소"]==row["위해요소"])].index
                if not orig_idx.empty:
                    oi = orig_idx[0]
                    if c3.button("삭제", key=f"sev_del_{typ}_{i}", use_container_width=True):
                        st.session_state[f"sev_del_confirm_{oi}"] = True
                    if st.session_state.get(f"sev_del_confirm_{oi}", False):
                        st.warning(f"**{row['위해요소']}** 를 삭제하시겠습니까?")
                        da, db, _ = st.columns([1,1,4])
                        if da.button("확인", key=f"sev_delok_{oi}", type="primary"):
                            df_sev2 = load_severity()
                            df_sev2 = df_sev2.drop(index=oi).reset_index(drop=True)
                            df_sev2.to_csv(SEVERITY_FILE, index=False, encoding='utf-8-sig')
                            st.session_state.pop(f"sev_del_confirm_{oi}", None)
                            st.rerun()
                        if db.button("취소", key=f"sev_delcancel_{oi}"):
                            st.session_state.pop(f"sev_del_confirm_{oi}", None)
                            st.rerun()

        with tab_b: render_sev_table("B")
        with tab_c: render_sev_table("C")
        with tab_p: render_sev_table("P")

        with tab_add:
            st.markdown("#### 위해요소 추가 / 수정")
            edit_mode = st.radio("작업 선택", ["신규 추가","기존 수정"], horizontal=True, key="sev_edit_mode")

            if edit_mode == "신규 추가":
                with st.form("form_sev_add", clear_on_submit=True):
                    a1,a2,a3 = st.columns([1,3,1])
                    with a1: new_type = st.selectbox("유형", ["B","C","P"])
                    with a2: new_hz   = st.text_input("위해요소명 *")
                    with a3: new_sev  = st.selectbox("심각성", [1,2,3])
                    if st.form_submit_button("추가", type="primary", use_container_width=True):
                        if not new_hz.strip():
                            st.error("위해요소명을 입력하세요.")
                        elif ((df_sev["유형"]==new_type)&(df_sev["위해요소"]==new_hz.strip())).any():
                            st.error("이미 등록된 위해요소입니다.")
                        else:
                            new_row = pd.DataFrame([{"유형":new_type,"위해요소":new_hz.strip(),"심각성":new_sev}])
                            df_sev2 = pd.concat([df_sev, new_row], ignore_index=True)
                            df_sev2.to_csv(SEVERITY_FILE, index=False, encoding='utf-8-sig')
                            st.success(f"[{new_type}] {new_hz} (심각성:{new_sev}) 추가됐습니다.")
                            st.rerun()
            else:
                e1, e2 = st.columns([1,3])
                with e1: sel_type = st.selectbox("유형 선택", ["B","C","P"], key="sev_edit_type")
                sub_edit = df_sev[df_sev["유형"]==sel_type]
                with e2:
                    if sub_edit.empty:
                        st.info("등록된 항목 없음")
                    else:
                        sel_hz = st.selectbox("위해요소 선택", sub_edit["위해요소"].tolist(), key="sev_edit_hz")
                        sel_row = sub_edit[sub_edit["위해요소"]==sel_hz]
                        if not sel_row.empty:
                            cur_sev = int(sel_row.iloc[0]["심각성"]) if str(sel_row.iloc[0]["심각성"]).isdigit() else 1
                            new_sev_val = st.selectbox("심각성 변경", [1,2,3], index=cur_sev-1, key="sev_edit_val")
                            if st.button("저장", type="primary", key="sev_edit_save"):
                                idx_e = sel_row.index[0]
                                df_sev2 = load_severity()
                                df_sev2.at[idx_e,"심각성"] = new_sev_val
                                df_sev2.to_csv(SEVERITY_FILE, index=False, encoding='utf-8-sig')
                                st.success("수정됐습니다.")
                                st.rerun()

# ──────────────────────────────────────────────────────────────────
# 공정 위해요소분석
# ──────────────────────────────────────────────────────────────────
    elif sub_menu == "공정 위해요소분석":
        st.markdown('<div class="section-title">공정별 위해요소 분석</div>', unsafe_allow_html=True)

        HAZARD_FILE    = "hazard_analysis.csv"
        SEVERITY_FILE  = "severity_settings.csv"

        def load_hazard():
            if os.path.exists(HAZARD_FILE):
                try:
                    df = pd.read_csv(HAZARD_FILE, encoding='utf-8-sig')
                    for c in ["카테고리","No","공정명","유형","위해요소","발생원인","심각성","발생가능성","종합평가","최종위해요소","예방조치및관리방법"]:
                        if c not in df.columns: df[c] = ""
                    return df
                except Exception: pass
            d = pd.DataFrame(columns=["카테고리","No","공정명","유형","위해요소","발생원인","심각성","발생가능성","종합평가","예방조치및관리방법"])
            d.to_csv(HAZARD_FILE, index=False, encoding='utf-8-sig')
            return d

        def load_sev2():
            if os.path.exists(SEVERITY_FILE):
                try: return pd.read_csv(SEVERITY_FILE, encoding='utf-8-sig')
                except Exception: pass
            return pd.DataFrame(columns=["유형","위해요소","심각성"])

        df_haz  = load_hazard()
        df_sev2 = load_sev2()

        # 카테고리 선택
        from glob import glob as _glob
        cat_files = sorted(_glob("haccp_flowchart_*.csv"))
        cats_df   = pd.read_csv("haccp_flow_categories.csv", encoding='utf-8-sig') if os.path.exists("haccp_flow_categories.csv") else pd.DataFrame()

        if cats_df.empty:
            st.warning("공정 흐름도 카테고리가 없습니다. 공정 흐름도 메뉴에서 먼저 등록하세요.")
        else:
            cat_names = cats_df["cat_name"].tolist() if "cat_name" in cats_df.columns else []
            cat_ids   = cats_df["cat_id"].tolist()   if "cat_id"   in cats_df.columns else []
            sel_cat_name = st.selectbox("카테고리 선택", cat_names, key="haz_cat")
            sel_cat_idx  = cat_names.index(sel_cat_name) if sel_cat_name in cat_names else 0
            sel_cat_id   = cat_ids[sel_cat_idx]

            # 공정 목록 로드
            flow_file = f"haccp_flowchart_{sel_cat_id}.csv"
            if not os.path.exists(flow_file):
                st.warning("해당 카테고리의 공정 흐름도가 없습니다.")
            else:
                df_flow2 = pd.read_csv(flow_file, encoding='utf-8-sig')
                # lane == "공정" 인 행만 필터 (lane 컬럼 없으면 유형 == "공정" fallback)
                if "lane" in df_flow2.columns:
                    mask_m = df_flow2["lane"].str.strip() == "공정"
                elif "유형" in df_flow2.columns:
                    mask_m = df_flow2["유형"].str.strip() == "공정"
                else:
                    mask_m = pd.Series([True]*len(df_flow2))
                df_proc = df_flow2[mask_m].copy()
                df_proc["_no"] = pd.to_numeric(df_proc["순서"], errors="coerce").fillna(9999)
                df_proc  = df_proc.sort_values("_no").reset_index(drop=True)
                proc_list = df_proc["단계명"].tolist()
                proc_no_map = {row["단계명"]: str(int(row["_no"])) if row["_no"]!=9999 else "?" for _,row in df_proc.iterrows()}

                # 해당 카테고리 기존 데이터
                mask_cat = df_haz["카테고리"] == sel_cat_name
                df_haz_cat = df_haz[mask_cat].copy()

                # ── 탭: 목록 조회 | 항목 추가 | 엑셀 출력 ──
                t_view, t_add, t_excel = st.tabs(["목록 조회", "항목 추가/수정", "엑셀 출력"])
                PKX = f"haz_{sel_cat_id}"

                # 세션 상태 초기화 (수정용)
                if f"{PKX}_edit_proc" not in st.session_state: st.session_state[f"{PKX}_edit_proc"] = ""
                if f"{PKX}_edit_type" not in st.session_state: st.session_state[f"{PKX}_edit_type"] = ""

                with t_view:
                    if not proc_list:
                        st.info("해당 카테고리의 공정 흐름도에 '공정' 유형으로 등록된 단계가 없습니다.")
                    else:
                        # ── [개선] 공정 흐름도 단계 기반 자동 연동 목록 조회 ──
                        rows_data = []
                        # No 순으로 정렬된 공정 목록 사용
                        sorted_procs = sorted(proc_list, key=lambda x: int(proc_no_map.get(x, 9999)))

                        for p_nm in sorted_procs:
                            p_df_all = df_haz_cat[df_haz_cat["공정명"] == p_nm].copy()
                            if p_df_all.empty:
                                # 흐름도에는 있으나 분석 데이터가 없는 경우 -> 빈 행 표시 (자동 연동)
                                rows_data.append({
                                    "ri": -1, "proc": p_nm, "typ": "-", "hz": "<span style='color:#ccc;'>(분석 정보 없음)</span>",
                                    "sev": "-", "prob": "-", "total": "-", "cause": "-", "prev": "-",
                                    "is_final": False, "orig_idx": None
                                })
                            else:
                                p_df_all = p_df_all.sort_values("유형")
                                for ri, row in p_df_all.iterrows():
                                    sev  = str(row.get("심각성","")).strip()
                                    prob = str(row.get("발생가능성","")).strip()
                                    is_final = False
                                    total_v2 = ""
                                    try:
                                        if sev and prob and sev not in ["","None","nan"] and prob not in ["","None","nan"]:
                                            tv = int(float(sev))*int(float(prob))
                                            total_v2 = str(tv)
                                            is_final = tv >= 3
                                    except Exception: pass
                                    # df_haz[mask_cat] 에서의 인덱스가 필요함
                                    idx_in_cat = df_haz_cat[
                                        (df_haz_cat["공정명"] == p_nm) & 
                                        (df_haz_cat["유형"] == row["유형"]) & 
                                        (df_haz_cat["위해요소"] == row["위해요소"])
                                    ].index[0]
                                    rows_data.append({
                                        "ri": ri, "proc": p_nm, "typ": str(row.get("유형","")).strip(),
                                        "hz": str(row.get("위해요소","")).strip(),
                                        "sev": sev, "prob": prob, "total": total_v2, "cause": str(row.get("발생원인","")).strip(),
                                        "prev": str(row.get("예방조치및관리방법","")).strip(),
                                        "is_final": is_final, "orig_idx": idx_in_cat
                                    })

                        n = len(rows_data)
                        proc_rowspan_arr  = [0] * n
                        cause_rowspan_arr = [0] * n

                        i = 0
                        while i < n:
                            j = i
                            while j < n and rows_data[j]["proc"] == rows_data[i]["proc"]: j += 1
                            proc_rowspan_arr[i] = j - i
                            i = j

                        i = 0
                        while i < n:
                            j = i
                            while j < n and (
                                rows_data[j]["proc"]  == rows_data[i]["proc"] and
                                rows_data[j]["typ"]   == rows_data[i]["typ"]  and
                                rows_data[j]["cause"] == rows_data[i]["cause"]
                            ): j += 1
                            cause_rowspan_arr[i] = j - i
                            i = j

                        def fmt_cell(v):
                            return v.replace("\n","<br>").replace("* ","<br>• ").lstrip("<br>")

                        TH = "background:#D6E4F0;color:#1F4E79;font-weight:bold;padding:8px 6px;text-align:center;border:1px solid #b0c4d8;font-size:11px;"
                        TD = "background:#fff;padding:7px 8px;border:1px solid #d0d7de;vertical-align:middle;word-break:break-word;line-height:1.6;font-size:11px;"

                        html = f"""
<style>
.haz-tbl {{width:100%;border-collapse:collapse;}}
.haz-tbl tr:hover td {{background:#f5f8ff !important;}}
</style>
<table class="haz-tbl">
<thead>
<tr>
  <th style="{TH}width:90px;">공정명</th>
  <th style="{TH}width:38px;">유형</th>
  <th style="{TH}width:120px;">위해요소</th>
  <th style="{TH}width:48px;">심각성</th>
  <th style="{TH}width:52px;">발생<br>가능성</th>
  <th style="{TH}width:52px;">종합<br>평가</th>
  <th style="{TH}min-width:200px;">발생원인</th>
  <th style="{TH}min-width:200px;">예방조치 및 관리방법</th>
</tr>
</thead>
<tbody>
"""
                        seen_procs2     = set()
                        seen_proc_type2 = set()
                        del_buttons     = []

                        for idx, d in enumerate(rows_data):
                            proc     = d["proc"]
                            typ      = d["typ"]
                            sev      = d["sev"]
                            prob     = d["prob"]
                            hz       = d["hz"]
                            total_v  = d["total"]
                            is_final = d["is_final"]

                            sev_c  = "#C0392B" if sev=="3" else ("#E67E22" if sev=="2" else ("#27AE60" if sev=="1" else "#aaa"))
                            hz_bg  = "#D6E8FF" if is_final else "#fff"
                            hz_fc  = "#1A5FAD" if is_final else "#222"
                            hz_fw  = "bold"    if is_final else "normal"
                            tot_bg = "#1A5FAD" if is_final else "#fff"
                            tot_fc = "white"   if is_final else "#333"
                            tot_fw = "bold"    if is_final else "normal"

                            # 공정명 rowspan (연속 그룹 첫 행만)
                            pspan = proc_rowspan_arr[idx]
                            if pspan > 0:
                                proc_td = f'<td rowspan="{pspan}" style="{TD}background:#f0f4f8;font-weight:bold;text-align:center;vertical-align:middle;">{proc}</td>'
                            else:
                                proc_td = ""

                            # 발생원인·예방조치 rowspan (연속 그룹 첫 행만)
                            cspan = cause_rowspan_arr[idx]
                            if cspan > 0:
                                cause_td = f'<td rowspan="{cspan}" style="{TD}vertical-align:top;">{fmt_cell(d["cause"])}</td>'
                                prev_td  = f'<td rowspan="{cspan}" style="{TD}vertical-align:top;">{fmt_cell(d["prev"])}</td>'
                            else:
                                cause_td = ""
                                prev_td  = ""

                            html += f"""<tr>
  {proc_td}
  <td style="{TD}text-align:center;color:#555;font-weight:bold;">{typ}</td>
  <td style="{TD}background:{hz_bg};color:{hz_fc};font-weight:{hz_fw};">{hz}</td>
  <td style="background:{sev_c};color:white;font-weight:bold;text-align:center;padding:7px;border:1px solid #d0d7de;">{sev}</td>
  <td style="{TD}text-align:center;">{prob}</td>
  <td style="background:{tot_bg};color:{tot_fc};font-weight:{tot_fw};text-align:center;padding:7px;border:1px solid #d0d7de;">{total_v}</td>
  {cause_td}
  {prev_td}
</tr>
"""
                            del_buttons.append((d["ri"], proc, hz, d["orig_idx"]))

                        html += "</tbody></table>"
                        st.markdown(html, unsafe_allow_html=True)

                        # 삭제 버튼 (테이블 아래에 따로)
                        st.markdown("---")
                        st.markdown("**항목 삭제**")
                        del_opts = [f"{r[1]} / {r[2]}" for r in del_buttons]
                        del_sel  = st.selectbox("삭제할 항목 선택", ["선택하세요"] + del_opts, key=f"haz_del_sel_{sel_cat_id}")
                        if del_sel != "선택하세요":
                            sel_del_idx = del_opts.index(del_sel)
                            ri_d, proc_d, hz_d, orig_d = del_buttons[sel_del_idx]
                            st.warning(f"**{proc_d} / {hz_d}** 를 삭제하시겠습니까?")
                            da, db, _ = st.columns([1,1,4])
                            if da.button("삭제 확인", key=f"haz_delok2_{sel_cat_id}", type="primary"):
                                if orig_d is not None:
                                    df_haz = df_haz.drop(index=orig_d).reset_index(drop=True)
                                    df_haz.to_csv(HAZARD_FILE, index=False, encoding='utf-8-sig')
                                    st.rerun()
                            if db.button("취소", key=f"haz_delcancel2_{sel_cat_id}"):
                                st.rerun()

                        # ── 수정 버튼 (세션 상태 타겟팅) ──
                        st.markdown("---")
                        st.markdown("🔍 **항목 빠르게 수정하기**")
                        c1_edit, c2_edit, c3_edit = st.columns([2, 1, 1])
                        with c1_edit:
                            sel_edit_p = st.selectbox("수정할 공정 단계 선택", [""] + sorted_procs, key=f"haz_quick_p_{sel_cat_id}")
                        with c2_edit:
                            sel_edit_t = st.selectbox("유형 선택", ["", "B", "C", "P"], key=f"haz_quick_t_{sel_cat_id}")
                        with c3_edit:
                            st.markdown("<div style='margin-top:28px'></div>", unsafe_allow_html=True)
                            if st.button("수정 모드 전환", key=f"haz_quick_btn_{sel_cat_id}", type="primary"):
                                if sel_edit_p and sel_edit_t:
                                    st.session_state[f"{PKX}_edit_proc"] = sel_edit_p
                                    st.session_state[f"{PKX}_edit_type"] = sel_edit_t
                                    st.success(f"'{sel_edit_p} / {sel_edit_t}' 항목 수정 설정 완료! 상단 **'항목 추가/수정'** 탭을 클릭하세요.")
                                else:
                                    st.warning("공정과 유형을 모두 선택하세요.")

                with t_add:
                    st.markdown("#### 위해요소 항목 등록")
                    PKX = f"haz_{sel_cat_id}"
                    df_sev_fresh = load_sev2()

                    # ── STEP 1: 공정명 + 유형 선택 ──
                    st.markdown("**① 공정명과 유형 선택**")
                    h1, h2 = st.columns([2, 1])
                    
                    # 수정을 위해 넘어온 초기값 계산
                    def_p_idx = 0
                    if st.session_state[f"{PKX}_edit_proc"] in proc_list:
                        def_p_idx = proc_list.index(st.session_state[f"{PKX}_edit_proc"]) + 1
                    
                    def_t_idx = 0
                    if st.session_state[f"{PKX}_edit_type"] in ["B","C","P"]:
                        def_t_idx = ["B","C","P"].index(st.session_state[f"{PKX}_edit_type"]) + 1

                    with h1:
                        sel_proc = st.selectbox("공정명 *", [""] + proc_list, index=def_p_idx, key=f"{PKX}_proc")
                    with h2:
                        sel_type = st.selectbox("유형 *", ["", "B", "C", "P"], index=def_t_idx, key=f"{PKX}_type")

                    cur_type = st.session_state.get(f"{PKX}_type", "")
                    cur_proc = st.session_state.get(f"{PKX}_proc", "")

                    if cur_proc and cur_type:
                        # 기존 등록된 항목 (공정+유형 기준) - df_haz 전체에서 정확히 필터
                        existing_same = df_haz[
                            (df_haz["카테고리"]==sel_cat_name) &
                            (df_haz["공정명"]==cur_proc) &
                            (df_haz["유형"]==cur_type)
                        ]
                        # 기존 등록 위해요소 & 발생가능성 맵
                        already_map = {}  # {위해요소: 발생가능성}
                        for _, er in existing_same.iterrows():
                            hz_e  = str(er.get("위해요소","")).strip()
                            prob_e = str(er.get("발생가능성","")).strip()
                            if hz_e: already_map[hz_e] = prob_e

                        # 심각성 설정의 전체 위해요소 목록
                        if not df_sev_fresh.empty:
                            hz_all = df_sev_fresh[df_sev_fresh["유형"]==cur_type]["위해요소"].tolist()
                        else:
                            hz_all = []
                        # 미등록 항목
                        hz_new = [h for h in hz_all if h not in already_map]

                        # 발생원인/예방조치: 같은 공정+유형의 기존 값 정확히 불러오기
                        default_cause = existing_same.iloc[0]["발생원인"] if not existing_same.empty else ""
                        default_prev  = existing_same.iloc[0]["예방조치및관리방법"] if not existing_same.empty else ""

                        st.divider()

                        # ── STEP 2: 발생원인 / 예방조치 ──
                        st.markdown("**② 발생원인 · 예방조치 입력** *(같은 공정+유형 전체에 공유됩니다)*")
                        if not existing_same.empty:
                            st.info(f"💡 기존 [{cur_proc} / {cur_type}] 데이터의 발생원인·예방조치를 자동으로 불러왔습니다. 수정 가능합니다.")
                        sel_cause   = st.text_area("발생원인", value=default_cause, height=100, key=f"{PKX}_cause")
                        sel_prevent = st.text_area("예방조치 및 관리방법", value=default_prev, height=100, key=f"{PKX}_prevent")

                        st.divider()

                        # ── STEP 3: 위해요소 선택 + 발생가능성 ──
                        st.markdown("**③ 위해요소 선택 및 발생가능성 입력**")
                        st.caption("✅ 이미 등록된 항목은 자동 체크됩니다. 수정 후 저장하면 덮어씁니다.")

                        all_hz_display = list(already_map.keys()) + hz_new  # 기존 먼저, 신규 뒤

                        if not all_hz_display:
                            st.info(f"[{cur_proc} / {cur_type}] 심각성 설정에 위해요소 항목이 없습니다.")
                        else:
                            all_key = f"{PKX}_chk_all"
                            all_checked = st.checkbox("전체 선택", key=all_key)

                            hdr1, hdr2, hdr3, hdr4 = st.columns([0.3, 2, 1, 1])
                            hdr1.markdown("<div style='font-size:11px;font-weight:bold;color:#555;'>선택</div>", unsafe_allow_html=True)
                            hdr2.markdown("<div style='font-size:11px;font-weight:bold;color:#555;'>위해요소</div>", unsafe_allow_html=True)
                            hdr3.markdown("<div style='font-size:11px;font-weight:bold;color:#555;'>심각성</div>", unsafe_allow_html=True)
                            hdr4.markdown("<div style='font-size:11px;font-weight:bold;color:#555;'>발생가능성</div>", unsafe_allow_html=True)

                            selected_hz = []
                            prob_map    = {}

                            for hz_item in all_hz_display:
                                is_existing = hz_item in already_map
                                sev_val = ""
                                if not df_sev_fresh.empty:
                                    m2 = df_sev_fresh[df_sev_fresh["위해요소"]==hz_item]
                                    if not m2.empty:
                                        try: sev_val = str(int(m2.iloc[0]["심각성"]))
                                        except Exception: pass
                                sev_c2 = "#C0392B" if sev_val=="3" else ("#E67E22" if sev_val=="2" else ("#27AE60" if sev_val=="1" else "#aaa"))

                                # [개선] 모든 위해요소를 기본적으로 체크 상태로 표시 (사용자 신속 등록 요청)
                                default_chk  = True
                                default_prob_val = int(already_map[hz_item]) if is_existing and already_map[hz_item].isdigit() else 1

                                c1, c2, c3, c4 = st.columns([0.3, 2, 1, 1])
                                # 기존 항목 배경 구분
                                label_html = f"<div style='padding:6px 0;font-size:13px;'>{hz_item}"
                                if is_existing:
                                    label_html += " <span style='font-size:10px;color:#1A5FAD;background:#D6E8FF;padding:1px 5px;border-radius:3px;'>등록됨</span>"
                                label_html += "</div>"

                                chk = c1.checkbox("", value=default_chk, key=f"{PKX}_chk_{hz_item}", label_visibility="collapsed")
                                c2.markdown(label_html, unsafe_allow_html=True)
                                c3.markdown(f"<div style='background:{sev_c2};color:white;font-weight:bold;text-align:center;padding:4px;border-radius:4px;margin-top:2px;'>{sev_val}</div>", unsafe_allow_html=True)

                                if chk:
                                    prob_idx = default_prob_val - 1 if 1 <= default_prob_val <= 3 else 0
                                    prob_v2 = c4.selectbox("", [1,2,3], index=prob_idx, key=f"{PKX}_prob_{hz_item}", label_visibility="collapsed")
                                    selected_hz.append(hz_item)
                                    prob_map[hz_item] = (sev_val, prob_v2)
                                else:
                                    c4.markdown("<div style='color:#bbb;padding:6px 0;font-size:12px;'>—</div>", unsafe_allow_html=True)

                            # 저장 미리보기
                            if selected_hz:
                                st.markdown("---")
                                st.markdown("**저장 미리보기**")
                                prev_rows = []
                                for hz_item in selected_hz:
                                    sv, pv = prob_map[hz_item]
                                    tag = " *(수정)*" if hz_item in already_map else " *(신규)*"
                                    try:
                                        tv = int(sv)*int(pv)
                                        is_f = tv >= 3
                                        final_tag = " 🔵최종위해요소" if is_f else ""
                                        prev_rows.append(f"- **{hz_item}**{tag} | 심각성:{sv} × 발생가능성:{pv} = **{tv}**{final_tag}")
                                    except:
                                        prev_rows.append(f"- {hz_item}{tag}")
                                st.markdown("\n".join(prev_rows))

                            st.divider()
                            if st.button("저장", type="primary", use_container_width=True, key=f"{PKX}_save"):
                                if not sel_cause.strip():
                                    st.error("발생원인을 입력하세요.")
                                elif not selected_hz:
                                    st.error("위해요소를 1개 이상 선택하세요.")
                                else:
                                    no_v2 = proc_no_map.get(cur_proc, "")
                                    new_rows = []
                                    for hz_item in selected_hz:
                                        sv, pv = prob_map[hz_item]
                                        total_v2 = ""; is_fin2 = False
                                        try:
                                            tv2 = int(sv)*int(pv); total_v2=str(tv2); is_fin2=tv2>=3
                                        except Exception: pass
                                        new_rows.append({
                                            "카테고리": sel_cat_name,
                                            "No": no_v2, "공정명": cur_proc,
                                            "유형": cur_type, "위해요소": hz_item,
                                            "발생원인": sel_cause,
                                            "심각성": sv, "발생가능성": str(pv),
                                            "종합평가": total_v2,
                                            "최종위해요소": "Y" if is_fin2 else "",
                                            "예방조치및관리방법": sel_prevent,
                                        })
                                    # 기존 같은 공정+유형 행 제거 후 새로 저장 (덮어쓰기)
                                    mask_del = ~(
                                        (df_haz["카테고리"]==sel_cat_name) &
                                        (df_haz["공정명"]==cur_proc) &
                                        (df_haz["유형"]==cur_type)
                                    )
                                    df_haz = pd.concat([df_haz[mask_del], pd.DataFrame(new_rows)], ignore_index=True)
                                    df_haz.to_csv(HAZARD_FILE, index=False, encoding='utf-8-sig')
                                    finals = [r["위해요소"] for r in new_rows if r["최종위해요소"]=="Y"]
                                    st.success(f"{len(new_rows)}건 저장됐습니다!" + (f" ★ 최종위해요소: {', '.join(finals)}" if finals else ""))
                                    st.rerun()
                    else:
                        st.info("공정명과 유형을 선택하면 위해요소 입력 화면이 나타납니다.")


                with t_excel:
                    st.markdown("#### 엑셀 출력 (보관용)")
                    if df_haz_cat.empty:
                        st.info("출력할 데이터가 없습니다.")
                    else:
                        import openpyxl
                        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                        from openpyxl.utils import get_column_letter
                        import io as _io2

                        def export_hazard_excel(df_data, cat_name):
                            wb2 = openpyxl.Workbook()
                            ws2 = wb2.active
                            ws2.title = "공정별 위해요소분석"

                            thin = Side(style="thin")
                            bd   = Border(left=thin,right=thin,top=thin,bottom=thin)

                            def wc2(r,c,v,bold=False,bg=None,fc="000000",sz=9,h="center",vv="center",wrap=True):
                                cell=ws2.cell(row=r,column=c,value=v)
                                cell.font=Font(name="맑은 고딕",bold=bold,size=sz,color=fc)
                                cell.alignment=Alignment(horizontal=h,vertical=vv,wrap_text=wrap)
                                cell.border=bd
                                if bg: cell.fill=PatternFill("solid",fgColor=bg)

                            def mc2(r1,c1,r2,c2):
                                ws2.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)

                            # 제목
                            mc2(1,1,1,10); wc2(1,1,f"■ 공정별 위해분석 목록표 ({cat_name})",bold=True,sz=12,bg="1F4E79",fc="FFFFFF")
                            mc2(2,1,2,10); wc2(2,1,f"출력일: {date.today().strftime('%Y.%m.%d')}",sz=9,bg="D6E4F0")

                            # 헤더 — 병합 먼저 모두 처리 후 쓰기 (MergedCell 오류 방지)
                            # 3~4행 병합: No, 공정명, 유형, 위해요소명, 발생원인
                            for c in [1,2,3,4,5]:
                                mc2(3,c,4,c)
                            # 3행: 위해요소평가 (6~8 병합)
                            mc2(3,6,3,8)
                            # 3~4행: 예방조치 (9~10 병합)
                            mc2(3,9,4,10)

                            # 쓰기
                            for col,lbl in [(1,"No"),(2,"공정명"),(3,"위해요소\n유형"),(4,"위해요소명"),(5,"발생원인")]:
                                wc2(3,col,lbl,bold=True,bg="1F4E79",fc="FFFFFF",sz=9)
                            wc2(3,6,"위해요소평가",bold=True,bg="1F4E79",fc="FFFFFF",sz=9)
                            wc2(3,9,"예방조치 및 관리방법",bold=True,bg="1F4E79",fc="FFFFFF",sz=9)
                            # 4행 소헤더
                            wc2(4,6,"심각성",bold=True,bg="2E75B6",fc="FFFFFF",sz=8)
                            wc2(4,7,"발생가능성",bold=True,bg="2E75B6",fc="FFFFFF",sz=8)
                            wc2(4,8,"종합평가",bold=True,bg="2E75B6",fc="FFFFFF",sz=8)

                            # No 순 정렬
                            df_out = df_data.copy()
                            df_out["_ns"] = df_out["공정명"].map(lambda s: int(proc_no_map.get(s,9999)) if str(proc_no_map.get(s,9999)).isdigit() else 9999)
                            df_out = df_out.sort_values(["_ns","유형"]).drop(columns=["_ns"])

                            TYPE_BG_XL = {"B":"D6EAF8","C":"D5F5E3","P":"FAE5D3"}
                            row = 5
                            prev_proc = None
                            for _, dr in df_out.iterrows():
                                bg2 = TYPE_BG_XL.get(str(dr.get("유형","")),"FFFFFF")
                                no_v3 = str(dr.get("No",""))
                                proc_n = str(dr.get("공정명",""))
                                show_proc = proc_n if proc_n != prev_proc else ""
                                show_no   = no_v3 if proc_n != prev_proc else ""
                                prev_proc = proc_n

                                # 최종위해요소 판정 (종합평가 > 3)
                                total_v_xl = str(dr.get("종합평가",""))
                                is_final_xl = False
                                try: is_final_xl = int(float(total_v_xl)) > 3 if total_v_xl not in ["","None","nan"] else False
                                except Exception: pass
                                final_row_bg = "D6E8FF" if is_final_xl else bg2
                                final_txt_bg = "D6E8FF" if is_final_xl else "FAFAFA"

                                wc2(row,1,show_no,bg=final_row_bg)
                                wc2(row,2,show_proc,bold=(show_proc!=""),bg="D6E4F0" if show_proc else final_row_bg,h="left")
                                wc2(row,3,str(dr.get("유형","")),bg=final_row_bg)
                                wc2(row,4,str(dr.get("위해요소","")),bg=final_row_bg,h="left")
                                wc2(row,5,str(dr.get("발생원인","")),bg=final_txt_bg,h="left",wrap=True)
                                sev3 = str(dr.get("심각성",""))
                                sev_bg = "C0392B" if sev3=="3" else ("E67E22" if sev3=="2" else ("27AE60" if sev3=="1" else "FFFFFF"))
                                wc2(row,6,sev3,bg=sev_bg,fc="FFFFFF" if sev3 in ["1","2","3"] else "000000",bold=True)
                                wc2(row,7,str(dr.get("발생가능성","")),bg=final_row_bg)
                                wc2(row,8,total_v_xl,bg="1A5FAD" if is_final_xl else final_row_bg,fc="FFFFFF" if is_final_xl else "000000",bold=True)
                                mc2(row,9,row,10)
                                wc2(row,9,str(dr.get("예방조치및관리방법","")),bg=final_txt_bg,h="left",wrap=True)
                                ws2.row_dimensions[row].height = 40
                                row += 1

                            # 열 너비
                            for col, w2 in zip(range(1,11),[5,18,6,16,30,7,9,9,35,5]):
                                ws2.column_dimensions[get_column_letter(col)].width = w2

                            ws2.page_setup.paperSize=9
                            ws2.page_setup.orientation="landscape"
                            ws2.page_setup.fitToPage=True
                            ws2.page_setup.fitToWidth=1
                            ws2.print_title_rows="1:4"
                            out2=_io2.BytesIO(); wb2.save(out2); return out2.getvalue()

                        excel_bytes = export_hazard_excel(df_haz_cat, sel_cat_name)
                        st.download_button(
                            label="공정별 위해요소분석 엑셀 다운로드",
                            data=excel_bytes,
                            file_name=f"Hollys_위해요소분석_{sel_cat_name}_{date.today().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )
                        st.caption(f"총 {len(df_haz_cat)}건 | 카테고리: {sel_cat_name}")

# 원부자재 위해요소분석
# ──────────────────────────────────────────────────────────────────
    elif sub_menu == "원부자재 위해요소분석":
        st.markdown('<div class="section-title">원·부자재별 위해요소 분석</div>', unsafe_allow_html=True)

        _RM_HAZ_FILE   = "rm_hazard_analysis.csv"
        _SEVERITY_FILE = "severity_settings.csv"

        def load_rm_hazard():
            cols = ["원자재명","원자재코드","유형코드","No","유형","위해요소","발생원인",
                    "심각성","발생가능성","종합평가","최종위해요소","예방조치및관리방법"]
            if os.path.exists(_RM_HAZ_FILE):
                try:
                    df = pd.read_csv(_RM_HAZ_FILE, encoding='utf-8-sig')
                    for c in cols:
                        if c not in df.columns: df[c] = ""
                    return df
                except Exception: pass
            d = pd.DataFrame(columns=cols)
            d.to_csv(_RM_HAZ_FILE, index=False, encoding='utf-8-sig')
            return d

        def load_rm_list():
            RM_COLS2 = ["원자재코드","원자재명","유형","규격","원산지","제조원","판매원",
                        "검사주기","비고","최소_수분","최대_수분","최소_밀도","최대_밀도","관능_기준"]
            if os.path.exists(RM_FILE):
                try:
                    df = pd.read_csv(RM_FILE, dtype=str)
                    for c in RM_COLS2:
                        if c not in df.columns: df[c] = ""
                    return df[RM_COLS2]
                except Exception: pass
            return pd.DataFrame(columns=RM_COLS2)

        def load_sev_rm():
            if os.path.exists(_SEVERITY_FILE):
                try: return pd.read_csv(_SEVERITY_FILE, encoding='utf-8-sig')
                except Exception: pass
            return pd.DataFrame(columns=["유형","위해요소","심각성"])

        df_rm_haz  = load_rm_hazard()
        df_rm_list = load_rm_list()
        df_sev_rm  = load_sev_rm()

        if df_rm_list.empty:
            st.warning("등록된 원·부자재가 없습니다. '원·부자재 관리 > 원·부자재 등록'에서 먼저 등록하세요.")
        else:
            rm_names   = df_rm_list["원자재명"].tolist()
            rm_codes   = df_rm_list["원자재코드"].tolist()
            rm_types   = df_rm_list["유형"].tolist()   # 원·부자재 등록의 유형
            # 전체 데이터 (목록 조회는 전체 표시)
            df_haz_all = df_rm_haz.copy()
            # 항목추가 탭용 원자재 선택 상태
            rm_display = [f"{c} — {n}" for c, n in zip(rm_codes, rm_names)]

            TH = "background:#D6E4F0;color:#1F4E79;font-weight:bold;padding:8px 6px;text-align:center;border:1px solid #b0c4d8;font-size:11px;"
            TD = "background:#fff;padding:7px 8px;border:1px solid #d0d7de;vertical-align:middle;word-break:break-word;line-height:1.6;font-size:11px;"

            t_view, t_add, t_excel = st.tabs(["목록 조회", "항목 추가/수정", "엑셀 출력"])
            
            # 세션 상태 초기화 (수정용)
            PMX = "rmhaz_state"
            if f"{PMX}_edit_rm" not in st.session_state: st.session_state[f"{PMX}_edit_rm"] = ""
            if f"{PMX}_edit_type" not in st.session_state: st.session_state[f"{PMX}_edit_type"] = ""

            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 탭1: 목록 조회 (전체 원·부자재)
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            with t_view:
                    # ── [개선] 등록된 원부자재 기반 자동 연동 목록 조회 ──
                    rows_data = []
                    # 가나다순 정렬된 원자재 목록 사용
                    sorted_rms = df_rm_list.sort_values("원자재명")

                    for _, rm_row in sorted_rms.iterrows():
                        r_nm = rm_row["원자재명"]
                        r_cd = rm_row["원자재코드"]
                        r_df_all = df_rm_haz[df_rm_haz["원자재코드"] == r_cd].copy()
                        
                        if r_df_all.empty:
                            # 등록은 되어있으나 분석 데이터가 없는 경우 -> 빈 행 표시
                            rows_data.append({
                                "ri": -1, "rm_nm": r_nm, "typ": "-", "hz": "<span style='color:#ccc;'>(분석 정보 없음)</span>",
                                "sev": "-", "prob": "-", "total": "-", "cause": "-", "prev": "-",
                                "is_final": False, "orig_idx": None
                            })
                        else:
                            r_df_all = r_df_all.sort_values("유형")
                            for ri, row in r_df_all.iterrows():
                                sev = str(row.get("심각성","")).strip()
                                prob = str(row.get("발생가능성","")).strip()
                                is_final = False
                                total_v2 = ""
                                try:
                                    if sev and prob and sev not in ["","None","nan"] and prob not in ["","None","nan"]:
                                        tv = int(float(sev)) * int(float(prob))
                                        total_v2 = str(tv); is_final = tv >= 3
                                except Exception: pass
                                
                                idx_in_haz = df_rm_haz[
                                    (df_rm_haz["원자재코드"] == r_cd) & 
                                    (df_rm_haz["유형"] == row["유형"]) & 
                                    (df_rm_haz["위해요소"] == row["위해요소"])
                                ].index[0]
                                
                                rows_data.append({
                                    "ri": ri, "rm_nm": r_nm, "typ": str(row.get("유형","")).strip(),
                                    "hz": str(row.get("위해요소","")).strip(),
                                    "sev": sev, "prob": prob, "total": total_v2, 
                                    "cause": str(row.get("발생원인","")).strip(),
                                    "prev": str(row.get("예방조치및관리방법","")).strip(),
                                    "is_final": is_final, "orig_idx": idx_in_haz
                                })

                    n = len(rows_data)
                    nm_rowspan_arr    = [0]*n
                    typ_rowspan_arr   = [0]*n
                    cause_rowspan_arr = [0]*n

                    i = 0
                    while i < n:
                        j = i
                        while j < n and rows_data[j]["rm_nm"] == rows_data[i]["rm_nm"]: j += 1
                        nm_rowspan_arr[i] = j - i; i = j
                    i = 0
                    while i < n:
                        j = i
                        while j < n and (rows_data[j]["rm_nm"] == rows_data[i]["rm_nm"] and
                                         rows_data[j]["typ"]   == rows_data[i]["typ"]): j += 1
                        typ_rowspan_arr[i] = j - i; i = j
                    i = 0
                    while i < n:
                        j = i
                        while j < n and (rows_data[j]["rm_nm"]  == rows_data[i]["rm_nm"] and
                                         rows_data[j]["typ"]    == rows_data[i]["typ"] and
                                         rows_data[j]["cause"]  == rows_data[i]["cause"]): j += 1
                        cause_rowspan_arr[i] = j - i; i = j

                    def fmt_cell_rm(v):
                        return v.replace("\n","<br>").replace("* ","<br>• ").lstrip("<br>")

                    TH_v = "background:#D6E4F0;color:#1F4E79;font-weight:bold;padding:8px 6px;text-align:center;border:1px solid #b0c4d8;font-size:11px;"
                    TD_v = "background:#fff;padding:7px 8px;border:1px solid #d0d7de;vertical-align:middle;word-break:break-word;line-height:1.6;font-size:11px;"

                    # 원자재 유형 색상 (원·부자재 등록의 유형)
                    RM_TYPE_COLOR = {
                        "생두(원료)":"#217346","스틱(원료)":"#2E75B6","포장재":"#7B5EA7",
                        "부자재":"#D11031","첨가물":"#E67E22","기타":"#888"
                    }
                    # 위해요소 유형 배지 색
                    TYPE_BADGE = {"B":"#2E75B6","C":"#D11031","P":"#7B5EA7"}

                    html_rm = (
                        '<style>.rmhaz-tbl{width:100%;border-collapse:collapse;}'
                        '.rmhaz-tbl tr:hover td{background:#f5f8ff !important;}</style>'
                        '<table class="rmhaz-tbl"><thead><tr>'
                        f'<th style="{TH_v}min-width:110px;">원·부자재명</th>'
                        f'<th style="{TH_v}width:40px;">유형</th>'
                        f'<th style="{TH_v}min-width:130px;">위해요소</th>'
                        f'<th style="{TH_v}width:50px;">심각성</th>'
                        f'<th style="{TH_v}width:55px;">발생<br>가능성</th>'
                        f'<th style="{TH_v}width:55px;">종합<br>평가</th>'
                        f'<th style="{TH_v}min-width:200px;">발생원인</th>'
                        f'<th style="{TH_v}min-width:200px;">예방조치 및 관리방법</th>'
                        '</tr></thead><tbody>'
                    )
                    del_buttons = []
                    for idx, d in enumerate(rows_data):
                        rm_nm = d["rm_nm"]; typ = d["typ"]; sev = d["sev"]; prob = d["prob"]
                        hz = d["hz"]; total_v = d["total"]; is_final = d["is_final"]

                        sev_c  = "#C0392B" if sev=="3" else ("#E67E22" if sev=="2" else ("#27AE60" if sev=="1" else "#aaa"))
                        hz_bg  = "#D6E8FF" if is_final else "#fff"
                        hz_fc  = "#1A5FAD" if is_final else "#222"
                        hz_fw  = "bold"    if is_final else "normal"
                        tot_bg = "#1A5FAD" if is_final else "#fff"
                        tot_fc = "white"   if is_final else "#333"
                        tot_fw = "bold"    if is_final else "normal"
                        tc_b   = TYPE_BADGE.get(typ, "#555")

                        # 원자재 등록 유형 조회
                        rm_row_info = df_rm_list[df_rm_list["원자재명"]==rm_nm]
                        rm_reg_type = rm_row_info.iloc[0]["유형"] if not rm_row_info.empty else ""
                        rm_tc = RM_TYPE_COLOR.get(rm_reg_type, "#555")

                        nspan = nm_rowspan_arr[idx]
                        if nspan > 0:
                            nm_td = (
                                f'<td rowspan="{nspan}" style="{TD_v}text-align:center;vertical-align:middle;background:#f0f4f8;">'
                                f'<div style="font-weight:bold;color:#1F4E79;margin-bottom:4px;">{rm_nm}</div>'
                                f'<span style="background:{rm_tc}22;color:{rm_tc};padding:1px 7px;border-radius:8px;font-size:10px;font-weight:bold;">{rm_reg_type}</span>'
                                f'</td>'
                            )
                        else:
                            nm_td = ""

                        tspan = typ_rowspan_arr[idx]
                        if tspan > 0:
                            typ_td = (
                                f'<td rowspan="{tspan}" style="{TD_v}text-align:center;vertical-align:middle;">'
                                f'<span style="background:{tc_b}22;color:{tc_b};padding:2px 8px;'
                                f'border-radius:8px;font-size:11px;font-weight:bold;">{typ}</span></td>'
                            )
                        else:
                            typ_td = ""

                        cspan = cause_rowspan_arr[idx]
                        if cspan > 0:
                            cause_td = f'<td rowspan="{cspan}" style="{TD_v}vertical-align:top;">{fmt_cell_rm(d["cause"])}</td>'
                            prev_td  = f'<td rowspan="{cspan}" style="{TD_v}vertical-align:top;">{fmt_cell_rm(d["prev"])}</td>'
                        else:
                            cause_td = prev_td = ""

                        html_rm += (
                            f'<tr>{nm_td}{typ_td}'
                            f'<td style="{TD_v}background:{hz_bg};color:{hz_fc};font-weight:{hz_fw};">{hz}</td>'
                            f'<td style="background:{sev_c};color:white;font-weight:bold;text-align:center;padding:7px;border:1px solid #d0d7de;">{sev}</td>'
                            f'<td style="{TD_v}text-align:center;">{prob}</td>'
                            f'<td style="background:{tot_bg};color:{tot_fc};font-weight:{tot_fw};text-align:center;padding:7px;border:1px solid #d0d7de;">{total_v}</td>'
                            f'{cause_td}{prev_td}</tr>'
                        )
                        del_buttons.append((d["ri"], rm_nm, typ, hz, d["orig_idx"]))

                    html_rm += "</tbody></table>"
                    st.markdown(html_rm, unsafe_allow_html=True)

                    # 삭제
                    st.markdown("---")
                    st.markdown("**항목 삭제**")
                    del_opts = [f"{r[1]} / {r[2]} / {r[3]}" for r in del_buttons]
                    del_sel  = st.selectbox("삭제할 항목 선택", ["선택하세요"] + del_opts, key="rmhaz_del_sel")
                    if del_sel != "선택하세요":
                        sel_del_idx = del_opts.index(del_sel)
                        ri_d, nm_d, typ_d, hz_d, orig_d = del_buttons[sel_del_idx]
                        st.warning(f"**{nm_d} / {typ_d} / {hz_d}** 를 삭제하시겠습니까?")
                        da, db, _ = st.columns([1,1,4])
                        if da.button("삭제 확인", key="rmhaz_delok", type="primary"):
                            if orig_d is not None:
                                df_rm_haz = df_rm_haz.drop(index=orig_d).reset_index(drop=True)
                                df_rm_haz.to_csv(_RM_HAZ_FILE, index=False, encoding='utf-8-sig')
                                st.rerun()
                        if db.button("취소", key="rmhaz_delcancel"):
                            st.rerun()

                    # ── [개선] 항목 빠르게 수정하기 ──
                    st.markdown("---")
                    st.markdown("🔍 **항목 빠르게 수정하기**")
                    c1_em, c2_em, c3_em = st.columns([2, 1, 1])
                    with c1_em:
                        sel_edit_rm = st.selectbox("수정할 원·부자재 선택", [""] + rm_display, key="rmhaz_quick_rm")
                    with c2_em:
                        sel_edit_ty = st.selectbox("유형 선택", ["", "B", "C", "P"], key="rmhaz_quick_ty")
                    with c3_em:
                        st.markdown("<div style='margin-top:28px'></div>", unsafe_allow_html=True)
                        if st.button("수정 모드 전환", key="rmhaz_quick_btn", type="primary"):
                            if sel_edit_rm and sel_edit_ty:
                                st.session_state[f"{PMX}_edit_rm"] = sel_edit_rm
                                st.session_state[f"{PMX}_edit_type"] = sel_edit_ty
                                st.success(f"'{sel_edit_rm} / {sel_edit_ty}' 항목 수정 설정 완료! 상단 **'항목 추가/수정'** 탭을 클릭하세요.")
                            else:
                                st.warning("원부자재와 유형을 모두 선택하세요.")

            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 탭2: 항목 추가/수정
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            with t_add:
                st.markdown("#### 위해요소 항목 등록")

                # ① 원·부자재 선택
                st.markdown("**① 원·부자재 선택**")
                def_rm_idx = 0
                if st.session_state[f"{PMX}_edit_rm"] in rm_display:
                    def_rm_idx = rm_display.index(st.session_state[f"{PMX}_edit_rm"]) + 1
                
                sel_rm_disp = st.selectbox("원·부자재 선택 *", [""] + rm_display, index=def_rm_idx, key="rmhaz_add_rm_sel")
                if sel_rm_disp and sel_rm_disp in rm_display:
                    sel_rm_idx  = rm_display.index(sel_rm_disp)
                    sel_rm_code = rm_codes[sel_rm_idx]
                    sel_rm_name = rm_names[sel_rm_idx]
                    # 원·부자재 등록 유형 표시
                    rm_reg_type_add = rm_types[sel_rm_idx] if sel_rm_idx < len(rm_types) else ""
                    RM_TYPE_COLOR2  = {
                        "생두(원료)":"#217346","스틱(원료)":"#2E75B6","포장재":"#7B5EA7",
                        "부자재":"#D11031","첨가물":"#E67E22","기타":"#888"
                    }
                    rm_tc2 = RM_TYPE_COLOR2.get(rm_reg_type_add, "#555")
                    st.markdown(
                        f"<span style='background:{rm_tc2}22;color:{rm_tc2};padding:3px 12px;"
                        f"border-radius:10px;font-size:11px;font-weight:bold;'>{rm_reg_type_add}</span>"
                        f" <b>{sel_rm_name}</b> <span style='color:#888;'>({sel_rm_code})</span>",
                        unsafe_allow_html=True)

                    PKR = f"rmhaz_{sel_rm_code}"
                    df_sev_fresh_rm = load_sev_rm()

                    # ② 위해요소 유형(B/C/P) 선택
                    st.divider()
                    st.markdown("**② 위해요소 유형 선택**")
                    def_ty_idx = 0
                    if st.session_state[f"{PMX}_edit_type"] in ["B","C","P"]:
                        def_ty_idx = ["B","C","P"].index(st.session_state[f"{PMX}_edit_type"]) + 1

                    sel_type_rm = st.selectbox("위해요소 유형 *", ["", "B", "C", "P"], index=def_ty_idx, key=f"{PKR}_type")

                    if sel_type_rm:
                        existing_same_rm = df_rm_haz[
                            (df_rm_haz["원자재코드"] == sel_rm_code) &
                            (df_rm_haz["유형"] == sel_type_rm)
                        ]
                        already_map_rm = {}
                        for _, er in existing_same_rm.iterrows():
                            hz_e  = str(er.get("위해요소","")).strip()
                            prob_e = str(er.get("발생가능성","")).strip()
                            if hz_e: already_map_rm[hz_e] = prob_e

                        if not df_sev_fresh_rm.empty:
                            hz_all_rm = df_sev_fresh_rm[df_sev_fresh_rm["유형"]==sel_type_rm]["위해요소"].tolist()
                        else:
                            hz_all_rm = []
                        hz_new_rm = [h for h in hz_all_rm if h not in already_map_rm]

                        default_cause_rm = existing_same_rm.iloc[0]["발생원인"] if not existing_same_rm.empty else ""
                        default_prev_rm  = existing_same_rm.iloc[0]["예방조치및관리방법"] if not existing_same_rm.empty else ""

                        st.divider()
                        st.markdown("**③ 발생원인 · 예방조치 입력** *(같은 원·부자재+유형 전체에 공유됩니다)*")
                        if not existing_same_rm.empty:
                            st.info(f"💡 기존 [{sel_rm_name} / {sel_type_rm}] 데이터의 발생원인·예방조치를 자동으로 불러왔습니다.")
                        sel_cause_rm   = st.text_area("발생원인", value=default_cause_rm, height=100, key=f"{PKR}_cause")
                        sel_prevent_rm = st.text_area("예방조치 및 관리방법", value=default_prev_rm, height=100, key=f"{PKR}_prevent")

                        st.divider()
                        st.markdown("**④ 위해요소 선택 및 발생가능성 입력**")
                        st.caption("✅ 이미 등록된 항목은 자동 체크됩니다. 수정 후 저장하면 덮어씁니다.")

                        all_hz_display_rm = list(already_map_rm.keys()) + hz_new_rm

                        if not all_hz_display_rm:
                            st.info(f"[{sel_type_rm}] 유형의 위해요소가 심각성 설정에 없습니다. 'HACCP > 심각성 설정'에서 먼저 등록하세요.")
                        else:
                            all_key_rm = f"{PKR}_chk_all"
                            all_checked_rm = st.checkbox("전체 선택", key=all_key_rm)

                            hdr1, hdr2, hdr3, hdr4 = st.columns([0.3, 2, 1, 1])
                            hdr1.markdown("<div style='font-size:11px;font-weight:bold;color:#555;'>선택</div>", unsafe_allow_html=True)
                            hdr2.markdown("<div style='font-size:11px;font-weight:bold;color:#555;'>위해요소</div>", unsafe_allow_html=True)
                            hdr3.markdown("<div style='font-size:11px;font-weight:bold;color:#555;'>심각성</div>", unsafe_allow_html=True)
                            hdr4.markdown("<div style='font-size:11px;font-weight:bold;color:#555;'>발생가능성</div>", unsafe_allow_html=True)

                            selected_hz_rm = []; prob_map_rm = {}

                            for hz_item in all_hz_display_rm:
                                is_existing_rm = hz_item in already_map_rm
                                sev_val_rm = ""
                                if not df_sev_fresh_rm.empty:
                                    m2 = df_sev_fresh_rm[df_sev_fresh_rm["위해요소"]==hz_item]
                                    if not m2.empty:
                                        try: sev_val_rm = str(int(m2.iloc[0]["심각성"]))
                                        except Exception: pass
                                sev_c2 = "#C0392B" if sev_val_rm=="3" else ("#E67E22" if sev_val_rm=="2" else ("#27AE60" if sev_val_rm=="1" else "#aaa"))
                                # 모든 위해요소 기본 체크 상태로 표시
                                default_chk_rm = True
                                default_prob_v = int(already_map_rm[hz_item]) if is_existing_rm and already_map_rm[hz_item].isdigit() else 1

                                c1, c2, c3, c4 = st.columns([0.3, 2, 1, 1])
                                label_html = f"<div style='padding:6px 0;font-size:13px;'>{hz_item}"
                                if is_existing_rm:
                                    label_html += " <span style='font-size:10px;color:#1A5FAD;background:#D6E8FF;padding:1px 5px;border-radius:3px;'>등록됨</span>"
                                label_html += "</div>"
                                chk = c1.checkbox("", value=default_chk_rm, key=f"{PKR}_chk_{hz_item}", label_visibility="collapsed")
                                c2.markdown(label_html, unsafe_allow_html=True)
                                c3.markdown(f"<div style='background:{sev_c2};color:white;font-weight:bold;text-align:center;padding:4px;border-radius:4px;margin-top:2px;'>{sev_val_rm}</div>", unsafe_allow_html=True)
                                if chk:
                                    prob_idx_rm = default_prob_v - 1 if 1 <= default_prob_v <= 3 else 0
                                    prob_v_rm = c4.selectbox("", [1,2,3], index=prob_idx_rm, key=f"{PKR}_prob_{hz_item}", label_visibility="collapsed")
                                    selected_hz_rm.append(hz_item)
                                    prob_map_rm[hz_item] = (sev_val_rm, prob_v_rm)
                                else:
                                    c4.markdown("<div style='color:#bbb;padding:6px 0;font-size:12px;'>—</div>", unsafe_allow_html=True)

                            if selected_hz_rm:
                                st.markdown("---")
                                st.markdown("**저장 미리보기**")
                                prev_rows_rm = []
                                for hz_item in selected_hz_rm:
                                    sv, pv = prob_map_rm[hz_item]
                                    tag = " *(수정)*" if hz_item in already_map_rm else " *(신규)*"
                                    try:
                                        tv = int(sv)*int(pv); is_f = tv >= 3
                                        final_tag = " 🔵최종위해요소" if is_f else ""
                                        prev_rows_rm.append(f"- **{hz_item}**{tag} | 심각성:{sv} × 발생가능성:{pv} = **{tv}**{final_tag}")
                                    except:
                                        prev_rows_rm.append(f"- {hz_item}{tag}")
                                st.markdown("\n".join(prev_rows_rm))

                            st.divider()
                            if st.button("저장", type="primary", use_container_width=True, key=f"{PKR}_save"):
                                if not sel_cause_rm.strip():
                                    st.error("발생원인을 입력하세요.")
                                elif not selected_hz_rm:
                                    st.error("위해요소를 1개 이상 선택하세요.")
                                else:
                                    new_rows_rm = []
                                    for hz_item in selected_hz_rm:
                                        sv, pv = prob_map_rm[hz_item]
                                        total_v2 = ""; is_fin2 = False
                                        try:
                                            tv2 = int(sv)*int(pv); total_v2=str(tv2); is_fin2=tv2>=3
                                        except Exception: pass
                                        new_rows_rm.append({
                                            "원자재명": sel_rm_name,
                                            "원자재코드": sel_rm_code,
                                            "유형코드": rm_reg_type_add,
                                            "No": "",
                                            "유형": sel_type_rm,
                                            "위해요소": hz_item,
                                            "발생원인": sel_cause_rm,
                                            "심각성": sv,
                                            "발생가능성": str(pv),
                                            "종합평가": total_v2,
                                            "최종위해요소": "Y" if is_fin2 else "",
                                            "예방조치및관리방법": sel_prevent_rm,
                                        })
                                    mask_del_rm = ~(
                                        (df_rm_haz["원자재코드"]==sel_rm_code) &
                                        (df_rm_haz["유형"]==sel_type_rm)
                                    )
                                    df_rm_haz = pd.concat([df_rm_haz[mask_del_rm], pd.DataFrame(new_rows_rm)], ignore_index=True)
                                    df_rm_haz.to_csv(_RM_HAZ_FILE, index=False, encoding='utf-8-sig')
                                    finals_rm = [r["위해요소"] for r in new_rows_rm if r["최종위해요소"]=="Y"]
                                    st.success(f"{len(new_rows_rm)}건 저장됐습니다!" + (f" ★ 최종위해요소: {', '.join(finals_rm)}" if finals_rm else ""))
                                    st.rerun()
                                    st.info("위해요소 유형(B/C/P)을 선택하면 위해요소 입력 화면이 나타납니다.")
                            else:
                                st.info("분석할 원·부자재를 선택하세요.")
                        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                        # 탭3: 엑셀 출력
                        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            with t_excel:
                st.markdown("#### 엑셀 출력 (보관용)")
                if df_haz_all.empty:
                    st.info("출력할 데이터가 없습니다.")
                else:
                    import openpyxl as _opxl_rm
                    from openpyxl.styles import Font as _Font_rm, Alignment as _Align_rm
                    from openpyxl.styles import PatternFill as _Fill_rm, Border as _Border_rm, Side as _Side_rm
                    from openpyxl.utils import get_column_letter as _gcl_rm
                    import io as _io_rm

                    def export_rm_hazard_excel(df_data, rm_name, rm_code):
                        wb_r = _opxl_rm.Workbook()
                        ws_r = wb_r.active
                        ws_r.title = "원부자재 위해요소분석"
                        thin_r = _Side_rm(style="thin")
                        bd_r   = _Border_rm(left=thin_r,right=thin_r,top=thin_r,bottom=thin_r)

                        def wc_r(r,c,v,bold=False,bg=None,fc="000000",sz=9,h="center",vv="center",wrap=True):
                            cell = ws_r.cell(row=r,column=c,value=v)
                            cell.font      = _Font_rm(name="맑은 고딕",bold=bold,size=sz,color=fc)
                            cell.alignment = _Align_rm(horizontal=h,vertical=vv,wrap_text=wrap)
                            cell.border    = bd_r
                            if bg: cell.fill = _Fill_rm("solid",fgColor=bg)

                        def mc_r(r1,c1,r2,c2):
                            ws_r.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)

                        # 타이틀
                        mc_r(1,1,1,9); wc_r(1,1,f"■ 원·부자재 위해요소 분석표 ({rm_name} / {rm_code})",bold=True,sz=12,bg="1F4E79",fc="FFFFFF")
                        mc_r(2,1,2,9); wc_r(2,1,f"출력일: {date.today().strftime('%Y.%m.%d')}",sz=9,bg="D6E4F0")

                        # 헤더 병합 먼저
                        for c in [1,2,3,4]:
                            mc_r(3,c,4,c)
                        mc_r(3,5,3,7)   # 위해요소평가 (5~7 병합)
                        mc_r(3,8,4,9)   # 예방조치 (8~9 병합)

                        # 헤더 쓰기
                        for col, lbl in [(1,"원자재명"),(2,"유형"),(3,"위해요소명"),(4,"발생원인")]:
                            wc_r(3,col,lbl,bold=True,bg="1F4E79",fc="FFFFFF",sz=9)
                        wc_r(3,5,"위해요소평가",bold=True,bg="1F4E79",fc="FFFFFF",sz=9)
                        wc_r(3,8,"예방조치 및 관리방법",bold=True,bg="1F4E79",fc="FFFFFF",sz=9)
                        wc_r(4,5,"심각성",bold=True,bg="D6E4F0",fc="1F4E79",sz=8)
                        wc_r(4,6,"발생가능성",bold=True,bg="D6E4F0",fc="1F4E79",sz=8)
                        wc_r(4,7,"종합평가",bold=True,bg="D6E4F0",fc="1F4E79",sz=8)

                        df_out = df_data.sort_values("유형").reset_index(drop=True)
                        TYPE_BG_XL = {"B":"D6EAF8","C":"FFE8E8","P":"EFE8FF"}
                        row_r = 5
                        prev_typ = None
                        for _, dr in df_out.iterrows():
                            typ_r  = str(dr.get("유형",""))
                            bg_r   = TYPE_BG_XL.get(typ_r,"FFFFFF")
                            show_nm  = rm_name if typ_r != prev_typ else ""
                            show_typ = typ_r   if typ_r != prev_typ else ""
                            prev_typ = typ_r

                            total_r = str(dr.get("종합평가",""))
                            is_fin_r = False
                            try: is_fin_r = int(float(total_r)) >= 3 if total_r not in ["","None","nan"] else False
                            except Exception: pass
                            fin_bg = "D6E8FF" if is_fin_r else bg_r

                            wc_r(row_r,1,show_nm,bold=(show_nm!=""),bg="D6E4F0" if show_nm else fin_bg,h="left")
                            wc_r(row_r,2,show_typ,bg=fin_bg)
                            wc_r(row_r,3,str(dr.get("위해요소","")),bg=fin_bg,h="left")
                            wc_r(row_r,4,str(dr.get("발생원인","")),bg="FAFAFA" if not is_fin_r else "D6E8FF",h="left",wrap=True)
                            sev_r  = str(dr.get("심각성",""))
                            sev_bg_r = "C0392B" if sev_r=="3" else ("E67E22" if sev_r=="2" else ("27AE60" if sev_r=="1" else "FFFFFF"))
                            wc_r(row_r,5,sev_r,bg=sev_bg_r,fc="FFFFFF" if sev_r in ["1","2","3"] else "000000",bold=True)
                            wc_r(row_r,6,str(dr.get("발생가능성","")),bg=fin_bg)
                            wc_r(row_r,7,total_r,bg="1A5FAD" if is_fin_r else fin_bg,fc="FFFFFF" if is_fin_r else "000000",bold=True)
                            mc_r(row_r,8,row_r,9)
                            wc_r(row_r,8,str(dr.get("예방조치및관리방법","")),bg="FAFAFA" if not is_fin_r else "D6E8FF",h="left",wrap=True)
                            ws_r.row_dimensions[row_r].height = 40
                            row_r += 1

                        for col_r, w_r in zip(range(1,10),[18,6,16,30,7,9,9,35,5]):
                            ws_r.column_dimensions[_gcl_rm(col_r)].width = w_r

                        ws_r.page_setup.paperSize   = 9
                        ws_r.page_setup.orientation = "landscape"
                        ws_r.page_setup.fitToPage   = True
                        ws_r.page_setup.fitToWidth  = 1
                        ws_r.print_title_rows = "1:4"

                        out_r = _io_rm.BytesIO(); wb_r.save(out_r); return out_r.getvalue()

                    # 엑셀 출력: 원자재 필터 또는 전체
                    rm_names_excel = df_haz_all["원자재명"].unique().tolist()
                    sel_xl_nm = st.selectbox("출력할 원·부자재 선택 (전체 출력은 '전체')", ["전체"] + rm_names_excel, key="rmhaz_xl_nm")
                    if sel_xl_nm == "전체":
                        df_xl = df_haz_all.copy()
                        xl_label = "전체"
                    else:
                        df_xl = df_haz_all[df_haz_all["원자재명"]==sel_xl_nm].copy()
                        xl_label = sel_xl_nm
                    excel_bytes_rm = export_rm_hazard_excel(df_xl, xl_label, sel_xl_nm if sel_xl_nm != "전체" else "ALL")
                    st.download_button(
                        label=f"원·부자재 위해요소분석 엑셀 다운로드 ({xl_label})",
                        data=excel_bytes_rm,
                        file_name=f"Hollys_원부자재위해요소분석_{xl_label}_{date.today().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                    )
                    st.caption(f"총 {len(df_xl)}건 | {xl_label}")

# --- 9. 시스템 히스토리 메뉴 ---
elif menu_selection == "시스템 히스토리":
    st.markdown('<div class="section-title">📜 시스템 작업 히스토리</div>', unsafe_allow_html=True)
    st.write("누가 어떤 수정을 했는지, 시스템에서 자동으로 기록된 히스토리를 확인합니다.")
    
    if os.path.exists(HISTORY_LOG_FILE):
        df_hist = pd.read_csv(HISTORY_LOG_FILE, dtype=str).fillna("-")
        df_hist = df_hist.sort_values(by="일시", ascending=False)
        
        c1, c2, c3 = st.columns([2, 1, 1])
        with c1:
            search_h = st.text_input("🔍 내용 검색", placeholder="작업내용 또는 상세 입력...")
        with c2:
            target_filter = st.selectbox("메뉴 필터", ["전체"] + sorted(df_hist["대상"].unique().tolist()))
        with c3:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            if st.button("🔄 새로고침", use_container_width=True):
                st.rerun()
        
        if search_h:
            df_hist = df_hist[df_hist["작업내용"].str.contains(search_h, na=False) | df_hist["상세"].str.contains(search_h, na=False)]
        if target_filter != "전체":
            df_hist = df_hist[df_hist["대상"] == target_filter]
            
        st.dataframe(df_hist, use_container_width=True, hide_index=True)
        
        if st.button("로그 비우기 (주의)", type="secondary"):
            if st.checkbox("정말 삭제하시겠습니까?"):
                pd.DataFrame(columns=["일시", "작업자", "대상", "작업내용", "상세"]).to_csv(HISTORY_LOG_FILE, index=False, encoding='utf-8-sig')
                log_history("로그 초기화", "시스템", "히스토리 로그 파일을 초기화함")
                st.success("로그가 초기화되었습니다.")
                st.rerun()
    else:
        st.info("아직 기록된 히스토리가 없습니다.")
